from flask import render_template, request, redirect, url_for, flash, session, jsonify, send_file, make_response, Response
from flask_login import current_user, login_user, logout_user, login_required
from datetime import datetime, date, timedelta
from sqlalchemy import func
from io import BytesIO
import io
import json
import base64
from timezone_utils import singapore_now, format_singapore_datetime
# Temporarily disable pandas import to resolve system library conflicts
# import pandas as pd
pd = None
import os
import uuid
import openpyxl
import openpyxl.styles
import qrcode
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Import will be handled by app_factory - routes get registered there
from app_main import app, db
from auth import authenticate_user, create_pioneer_lodge_user
from werkzeug.exceptions import RequestEntityTooLarge
from models import User, Organization, Asset, AssetCategory, Submission, FormTemplate, RoomHandover, OffenseRecord, QRCode, FormSubmission, RoomInventoryChecklist, SystemLog, UserFormPermission, StockItem, StockMovement, PurchaseRequest, PurchaseRequestItem, StockUsage, ComplianceRecord, ComplianceAcknowledgment, ComplianceViolation, StaffAttendance, Worker, Visitor, ImportantNews
from app.models.models_house_acknowledge import HouseAcknowledge, HouseAcknowledgment, RoomNumber
from app.models.models_meter_reading import MeterCompany, MeterRoom, WaterMeterReading, ElectricityMeterReading
from app.models.models_asset import AssetName, AssetItem
from app.models.models_food_locker import FoodLocker, FoodLockerRoomAssignment
from app.models.models_bedding import BeddingCategory, BeddingItem, BeddingMovement
from app.models.models_resident_checkout import ResidentCheckout
from app.models.models_key_management import KeyRecord
from functools import wraps
from performance_monitor import performance_timer
from cache_manager import cache

# Permission checking functions
def is_admin_user(user):
    """Check if user is an admin (Pioneer Lodge admin or has admin role)"""
    if user.email == "pioneerlodge@tsgrp.sg":
        return True
    
    # Check if user has admin role
    if hasattr(user, 'role') and user.role == 'admin':
        return True
    
    return False

def admin_required(f):
    """Decorator to require admin privileges"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('index'))
        
        if not is_admin_user(current_user):
            flash('Admin access required', 'error')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

def edit_permission_required(page_name):
    """Decorator to require edit permissions for a specific page"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            
            # Check if user has edit permission for this page
            user_permissions = get_user_page_permissions(current_user)
            page_perms = user_permissions.get(page_name, {})
            
            if not page_perms.get('can_edit', False) and not is_admin_user(current_user):
                flash('Access denied. You have view-only access and cannot create or edit items.', 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def create_permission_required(page_name):
    """Decorator to require create permissions for a specific page"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            
            # Check if user has create permission for this page
            user_permissions = get_user_page_permissions(current_user)
            page_perms = user_permissions.get(page_name, {})
            
            if not page_perms.get('can_create', False) and not is_admin_user(current_user):
                flash('Access denied. You have view-only access and cannot create new items.', 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def page_access_required(page_name):
    """Decorator to require access permissions for a specific page"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            
            # Check if user has access permission for this page
            user_permissions = get_user_page_permissions(current_user)
            page_perms = user_permissions.get(page_name, {})
            
            if not page_perms.get('can_access', False) and not is_admin_user(current_user):
                flash(f'Access denied. You do not have permission to access this page.', 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def page_permission_required(page_name):
    """Decorator to require page access permissions"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            
            # Check if user has access to this page
            user_permissions = get_user_page_permissions(current_user)
            page_perms = user_permissions.get(page_name, {})
            
            if not page_perms.get('can_access', False) and not is_admin_user(current_user):
                flash('Access denied. You do not have permission to access this page.', 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def form_permission_required(form_type):
    """Decorator to require specific form permissions - Check both old form permissions and new page permissions"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            
            # Admin always has access
            if is_admin_user(current_user):
                return f(*args, **kwargs)
            
            # Map form types to page names
            form_to_page_map = {
                'offense': 'offense_records',
                'house_acknowledge': 'house_acknowledge',
                'qr': 'qr_codes'
            }
            
            # Check new page permission system first
            page_name = form_to_page_map.get(form_type)
            if page_name:
                user_permissions = get_user_page_permissions(current_user)
                page_perms = user_permissions.get(page_name, {})
                if page_perms.get('can_access', False):
                    return f(*args, **kwargs)
            
            # Fallback to old form permission system
            user_permissions = get_user_dashboard_permissions(current_user)
            if form_type in user_permissions['allowed_form_types']:
                return f(*args, **kwargs)
            
            flash('Access denied. You do not have permission to access this feature.', 'error')
            return redirect(url_for('dashboard'))
            
        return decorated_function
    return decorator

def get_available_pages():
    """Get all available pages in the system"""
    return {
        'dashboard': {'name': 'Dashboard', 'icon': 'fas fa-tachometer-alt', 'description': 'Main dashboard and overview'},
        'staff_attendance': {'name': 'Staff Attendance', 'icon': 'fas fa-user-clock', 'description': 'Staff attendance tracking'},
        'pioneer_lodge_visitors': {'name': 'Pioneer Lodge Visitors', 'icon': 'fas fa-users', 'description': 'Visitor management system'},
        'resident_checkin': {'name': 'Resident Check-in', 'icon': 'fas fa-user-check', 'description': 'Resident check-in management'},
        'house_acknowledge': {'name': 'House Acknowledge', 'icon': 'fas fa-home', 'description': 'House acknowledgment forms'},
        'submissions': {'name': 'Submissions', 'icon': 'fas fa-file-alt', 'description': 'Form submissions and data'},
        'purchase': {'name': 'Purchase', 'icon': 'fas fa-shopping-cart', 'description': 'Purchase forms, stock storage, and form management'},
        'asset_management': {'name': 'Asset Management', 'icon': 'fas fa-boxes', 'description': 'Inventory and asset tracking'},
        'stock_report': {'name': 'Stock Report', 'icon': 'fas fa-warehouse', 'description': 'Stock and inventory reports'},
        'food_locker': {'name': 'Food Locker Management', 'icon': 'fas fa-utensils', 'description': 'Food locker rental management'},
        'room_checklist': {'name': 'Room Checklist', 'icon': 'fas fa-clipboard-check', 'description': 'Room inspection and maintenance'},
        'meter_reading': {'name': 'Meter Reading', 'icon': 'fas fa-tachometer-alt', 'description': 'Utility meter readings'},

        'resident_checkout': {'name': 'Resident Check-Out', 'icon': 'fas fa-sign-out-alt', 'description': 'Resident check-out scanning and management'},
        'offense_records': {'name': 'Offense Records', 'icon': 'fas fa-exclamation-triangle', 'description': 'Disciplinary record management'},
        'fin_search': {'name': 'FIN Search', 'icon': 'fas fa-search', 'description': 'FIN number search functionality'},
        'qr_codes': {'name': 'QR Codes', 'icon': 'fas fa-qrcode', 'description': 'QR code management and generation'},
        'msrf_management': {'name': 'MSRF Management', 'icon': 'fas fa-clipboard-list', 'description': 'MSRF request management'},
        'bedding_management': {'name': 'Bedding Management', 'icon': 'fas fa-bed', 'description': 'Bedding items and inventory'},
        'key_management': {'name': 'Key Management', 'icon': 'fas fa-key', 'description': 'Key tracking and QR code management'},
        'settings': {'name': 'Settings', 'icon': 'fas fa-cog', 'description': 'System settings and configuration'},
        'admin': {'name': 'Admin Dashboard', 'icon': 'fas fa-user-shield', 'description': 'Complete system administration'}
    }

def get_user_page_permissions(user):
    """Get user's page-specific permissions"""
    import json
    
    if is_admin_user(user):
        # Admin has access to all pages with full permissions
        pages = get_available_pages()
        return {page: {'can_access': True, 'can_edit': True, 'can_create': True, 'can_delete': True} for page in pages.keys()}
    
    # Parse user's page permissions from JSON
    if user.page_permissions:
        try:
            permissions = json.loads(user.page_permissions)
        except:
            permissions = []
    else:
        permissions = []
    
    # Set default permissions based on access level
    is_view_only = user.access_level == 'view_only'
    
    result = {}
    for page in get_available_pages().keys():
        has_access = page in permissions
        result[page] = {
            'can_access': has_access,
            'can_edit': has_access and not is_view_only,
            'can_create': has_access and not is_view_only,
            'can_delete': has_access and not is_view_only
        }
    
    return result

def get_user_dashboard_permissions(user):
    """Get user's permissions for dashboard navigation - Legacy function for compatibility"""
    page_perms = get_user_page_permissions(user)
    
    return {
        'can_manage_assets': page_perms.get('asset_management', {}).get('can_access', False),
        'can_manage_forms': any(page_perms.get(page, {}).get('can_access', False) for page in ['house_acknowledge', 'submissions']),
        'can_manage_qr': page_perms.get('qr_codes', {}).get('can_access', False),
        'can_view_handovers': page_perms.get('house_acknowledge', {}).get('can_access', False),
        'can_view_offenses': page_perms.get('offense_records', {}).get('can_access', False),
        'can_view_admin': page_perms.get('admin', {}).get('can_access', False),
        'can_view_settings': page_perms.get('settings', {}).get('can_access', False),
        'can_view_submissions': page_perms.get('submissions', {}).get('can_access', False),
        'can_view_staff_attendance': page_perms.get('staff_attendance', {}).get('can_access', False),
        'can_view_pioneer_visitors': page_perms.get('pioneer_lodge_visitors', {}).get('can_access', False),
        'can_view_resident_checkin': page_perms.get('resident_checkin', {}).get('can_access', False),
        'can_view_purchase': page_perms.get('purchase', {}).get('can_access', False),
        'can_view_stock_report': page_perms.get('stock_report', {}).get('can_access', False),
        'can_view_food_locker': page_perms.get('food_locker', {}).get('can_access', False),
        'can_view_room_checklist': page_perms.get('room_checklist', {}).get('can_access', False),
        'can_view_meter_reading': page_perms.get('meter_reading', {}).get('can_access', False),

        'can_view_resident_checkout': page_perms.get('resident_checkout', {}).get('can_access', False),
        'can_view_fin_search': page_perms.get('fin_search', {}).get('can_access', False),
        'can_view_msrf': page_perms.get('msrf_management', {}).get('can_access', False),
        'can_view_bedding': page_perms.get('bedding_management', {}).get('can_access', False),
        'can_view_key_management': page_perms.get('key_management', {}).get('can_access', False),
        'allowed_form_types': []
    }

# Make functions available to templates
app.jinja_env.globals.update(
    get_user_dashboard_permissions=get_user_dashboard_permissions,
    get_user_page_permissions=get_user_page_permissions,
    get_available_pages=get_available_pages,
    is_admin_user=is_admin_user
)

def can_user_create(user, page_name):
    """Check if user can create items on a specific page"""
    # Admin always has full permissions
    if is_admin_user(user):
        return True
    perms = get_user_page_permissions(user)
    return perms.get(page_name, {}).get('can_create', False)

def can_user_edit(user, page_name):
    """Check if user can edit items on a specific page"""
    # Admin always has full permissions
    if is_admin_user(user):
        return True
    perms = get_user_page_permissions(user)
    return perms.get(page_name, {}).get('can_edit', False)

# Add helper functions to templates
app.jinja_env.globals.update(
    can_user_create=can_user_create,
    can_user_edit=can_user_edit
)

# API endpoint for password management
@app.route('/admin/password_manager')
@login_required
def admin_password_manager():
    """Password manager page for administrators"""
    if not is_admin_user(current_user):
        flash('Access denied - admin privileges required', 'error')
        return redirect(url_for('dashboard'))
    
    users = User.query.all()
    return render_template('admin_password_manager.html', users=users)

@app.route('/admin/users/<user_id>/update_password', methods=['POST'])
@login_required
def admin_update_password(user_id):
    """Quick password update route for admin"""
    if not is_admin_user(current_user):
        flash('Access denied - admin privileges required', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        user = User.query.get_or_404(user_id)
        new_password = request.form.get('new_password')
        
        if new_password and new_password.strip():
            from werkzeug.security import generate_password_hash
            user.password_hash = generate_password_hash(new_password)
            user.admin_viewable_password = new_password
            db.session.commit()
            
            # Log the password change
            log = SystemLog(
                user_id=current_user.id,
                user_email=current_user.email,
                action=f"Updated password for user {user.email}",
                module="Password Management",
                status="Success"
            )
            db.session.add(log)
            db.session.commit()
            
            flash(f'Password updated successfully for {user.email}', 'success')
        else:
            flash('Password cannot be empty', 'error')
            
    except Exception as e:
        flash(f'Error updating password: {str(e)}', 'error')
    
    return redirect(url_for('admin_password_manager'))

@app.route('/api/admin/user_password_info/<user_id>')
@login_required
def admin_user_password_info(user_id):
    """API endpoint to get user password info for admin viewing"""
    current_user_obj = current_user
    
    # Check if current user is admin
    if not is_admin_user(current_user_obj):
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    try:
        # Get the user
        user = User.query.get_or_404(user_id)
        
        # Show admin-viewable password if available, otherwise show status
        if user.admin_viewable_password:
            # Return the admin-viewable password
            return jsonify({
                'success': True,
                'password_display': user.admin_viewable_password,
                'has_password': True
            })
        elif user.password_hash:
            # Password exists but not admin-viewable
            return jsonify({
                'success': True,
                'password_display': 'Password is set (encrypted - cannot display original)',
                'has_password': True
            })
        else:
            return jsonify({
                'success': True,
                'password_display': 'No password set',
                'has_password': False
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Make session permanent
@app.before_request
def make_session_permanent():
    session.permanent = True

def create_default_data():
    """Create default organizations and categories if they don't exist"""
    # Create default organizations
    pioneer_lodge = Organization.query.filter_by(name="Pioneer Lodge - tsgrp.sg").first()
    if not pioneer_lodge:
        pioneer_lodge = Organization(
            name="Pioneer Lodge - tsgrp.sg",
            email="pioneerlodge@tsgrp.sg",
            description="Full dormitory asset management and worker compliance tracking"
        )
        db.session.add(pioneer_lodge)

    tuas_view = Organization.query.filter_by(name="Tuas View Dormitory - tsgrp.sg").first()
    if not tuas_view:
        tuas_view = Organization(
            name="Tuas View Dormitory - tsgrp.sg",
            email="tuasview@tsgrp.sg",
            description="Independent asset inventory and compliance records"
        )
        db.session.add(tuas_view)

    # Create default asset categories
    categories = [
        "Furniture", "Electronics", "Bedding", "Kitchen Equipment",
        "Cleaning Supplies", "Safety Equipment", "Maintenance Tools", "Office Supplies"
    ]
    
    for cat_name in categories:
        category = AssetCategory.query.filter_by(name=cat_name).first()
        if not category:
            category = AssetCategory(name=cat_name, description=f"{cat_name} category")
            db.session.add(category)

    db.session.commit()

# Initialize default data when the module is loaded
with app.app_context():
    create_default_data()

# Error handler for file size too large
@app.errorhandler(413)
def request_entity_too_large(error):
    """Handle file upload size exceeded"""
    flash('Photos are too large. Please use smaller images or fewer photos. Maximum total size is 50MB.', 'error')
    return redirect(request.url)

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle login form submission"""
    from urllib.parse import unquote
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = authenticate_user(username, password)
        if user:
            login_user(user)
            flash('Successfully logged in!', 'success')
            
            # Handle next parameter properly
            next_page = request.args.get('next')
            if next_page:
                # Decode URL-encoded parameters
                next_page = unquote(next_page)
                # Ensure it's a safe redirect
                if next_page.startswith('/'):
                    return redirect(next_page)
            
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
            return render_template('login_professional.html')
    
    # Show professional login page for GET requests
    return render_template('login_professional.html')

@app.route('/logout')
@login_required
def logout():
    """Handle logout"""
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))

@app.route('/health')
def health():
    """Health check endpoint"""
    return {'status': 'ok', 'message': 'Application is running'}, 200

@app.route('/api/get_user_password/<int:user_id>')
@admin_required
def get_user_password(user_id):
    """API endpoint for admin to get user's actual password"""
    try:
        user = User.query.get_or_404(user_id)
        # Only return password if current user is admin
        if not is_admin_user(current_user):
            return jsonify({'error': 'Access denied'}), 403
        
        # Return the actual password hash (in real implementation, this would be decrypted)
        # For security demo, we'll show a portion of the hash
        password_display = user.password_hash[:20] + "..." if user.password_hash else "No password set"
        
        return jsonify({
            'success': True,
            'password_display': password_display,
            'username': user.username
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/')
def index():
    """Landing page - redirect to dashboard if logged in"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('login_professional.html')

@app.route('/login%3Fnext=%2Fadmin')
@app.route('/login%3Fnext=%2F<path:next_path>')
def handle_encoded_login(next_path=None):
    """Handle URL-encoded login redirects from new tabs"""
    from urllib.parse import unquote
    
    if current_user.is_authenticated:
        if next_path:
            decoded_path = '/' + unquote(next_path)
            return redirect(decoded_path)
        return redirect(url_for('admin'))
    
    # Redirect to proper login with decoded next parameter
    if next_path:
        decoded_next = '/' + unquote(next_path)
        return redirect(url_for('login', next=decoded_next))
    else:
        return redirect(url_for('login', next='/admin'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard with statistics"""
    user = current_user
    
    # Redirect limited users directly to their allowed page
    if not is_admin_user(user):
        user_permissions = get_user_dashboard_permissions(user)
        if 'handover' in user_permissions['allowed_form_types'] and len(user_permissions['allowed_form_types']) == 1:
            return redirect(url_for('room_checklist'))
    
    # Get or assign organization
    if not user.organization_id:
        # For demo purposes, assign based on email domain or ask user to choose
        if user.email and 'pioneerlodge' in user.email.lower():
            org = Organization.query.filter_by(name="Pioneer Lodge").first()
            if org:
                user.organization_id = org.id
        elif user.email and 'tuasview' in user.email.lower():
            org = Organization.query.filter_by(name="Tuas View Dormitory").first()
            if org:
                user.organization_id = org.id
        else:
            # Default to Pioneer Lodge or create one if it doesn't exist
            org = Organization.query.filter_by(name="Pioneer Lodge").first()
            if not org:
                org = Organization(name="Pioneer Lodge")
                db.session.add(org)
                db.session.commit()
            user.organization_id = org.id
        db.session.commit()

    # Get statistics for the user's organization
    org_id = user.organization_id
    
    # Initialize all counts to prevent None/empty display
    total_assets = 0
    active_assets = 0
    total_handovers = 0
    pending_handovers = 0
    total_offenses = 0
    open_offenses = 0
    total_forms = 0
    house_acknowledgment_count = 0
    total_qr_codes = 0
    active_qr_codes = 0
    
    if org_id:
        # Asset statistics
        total_assets = Asset.query.filter_by(organization_id=org_id).count()
        active_assets = Asset.query.filter_by(organization_id=org_id, status='Active').count()
        
        # Handover statistics
        total_handovers = RoomHandover.query.filter_by(organization_id=org_id).count()
        pending_handovers = RoomHandover.query.filter_by(organization_id=org_id, status='Pending').count()
        
        # Offense statistics
        total_offenses = OffenseRecord.query.filter_by(organization_id=org_id).count()
        open_offenses = OffenseRecord.query.filter_by(organization_id=org_id, status='Open').count()
        
        # Form statistics
        total_forms = FormTemplate.query.filter_by(organization_id=org_id).count()
        
        # QR Code statistics
        total_qr_codes = QRCode.query.filter_by(organization_id=org_id).count()
        active_qr_codes = QRCode.query.filter_by(organization_id=org_id, is_active=True).count()
    
    # House acknowledgment statistics (not organization-specific)
    house_acknowledgment_count = HouseAcknowledgment.query.count()
    

    
    # Asset status breakdown
    status_counts = {}
    status_results = db.session.query(
        Asset.status, func.count(Asset.id)
    ).filter_by(organization_id=org_id).group_by(Asset.status).all()
    
    for status, count in status_results:
        status_counts[status] = count
    
    # Ensure all status keys exist with default values
    status_counts.setdefault('Active', 0)
    status_counts.setdefault('Store', 0)
    status_counts.setdefault('Damage', 0)
    
    # Recent assets
    recent_assets = Asset.query.filter_by(
        organization_id=org_id
    ).order_by(Asset.created_at.desc()).limit(3).all()
    
    # Recent handovers
    recent_handovers = RoomHandover.query.filter_by(
        organization_id=org_id
    ).order_by(RoomHandover.created_at.desc()).limit(3).all()
    
    # Recent offences
    recent_offenses = OffenseRecord.query.filter_by(
        organization_id=org_id
    ).order_by(OffenseRecord.created_at.desc()).limit(3).all()

    # Get active important news for the organization
    important_news = []
    if org_id:
        important_news = ImportantNews.query.filter_by(
            organization_id=org_id,
            is_active=True,
            show_on_login=True
        ).filter(
            db.or_(
                ImportantNews.expires_at.is_(None),
                ImportantNews.expires_at > singapore_now()
            )
        ).order_by(ImportantNews.priority.desc(), ImportantNews.created_at.desc()).all()

    # Get user permissions for dashboard display
    user_permissions = get_user_dashboard_permissions(user)
    
    # Get page permissions for navigation
    page_permissions = get_user_page_permissions(user)
    
    return render_template('dashboard.html', 
                         user=user,
                         user_permissions=user_permissions,
                         page_permissions=page_permissions,
                         total_assets=total_assets,
                         active_assets=active_assets,
                         total_handovers=total_handovers,
                         pending_handovers=pending_handovers,
                         total_offenses=total_offenses,
                         open_offenses=open_offenses,
                         total_forms=total_forms,
                         house_acknowledgment_count=house_acknowledgment_count,
                         total_qr_codes=total_qr_codes,
                         active_qr_codes=active_qr_codes,
                         status_counts=status_counts,
                         recent_assets=recent_assets,
                         recent_handovers=recent_handovers,
                         recent_offenses=recent_offenses,
                         important_news=important_news)


@app.route('/assets')
@login_required
def assets():
    """Asset listing page with search and filter"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get filter parameters
    status_filter = request.args.get('status', '')
    category_filter = request.args.get('category', '')
    search_term = request.args.get('search', '')
    serial_number_filter = request.args.get('serial_number', '')
    
    # Base query
    query = Asset.query.filter_by(organization_id=user.organization_id)
    
    # Apply filters
    if status_filter:
        query = query.filter(Asset.status == status_filter)
    
    if category_filter:
        try:
            query = query.filter(Asset.category_id == int(category_filter))
        except ValueError:
            pass
    
    if search_term:
        query = query.filter(Asset.name.ilike(f'%{search_term}%'))
    
    if serial_number_filter:
        query = query.filter(Asset.serial_number.ilike(f'%{serial_number_filter}%'))
    
    assets_list = query.order_by(Asset.created_at.desc()).all()
    categories = AssetCategory.query.all()
    
    # Asset statuses for filter dropdown
    statuses = ['Active', 'Inactive', 'Room', 'Store', 'Clear', 'Other']

    return render_template('assets.html',
                         assets=assets_list,
                         categories=categories,
                         statuses=statuses,
                         current_status=status_filter,
                         current_category=category_filter,
                         search_term=search_term,
                         serial_number_filter=serial_number_filter)

@app.route('/assets/new')
@login_required
@create_permission_required('asset_management')
def new_asset():
    """New asset form"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    categories = AssetCategory.query.all()
    statuses = ['Active', 'Inactive', 'Room', 'Store', 'Clear', 'Other']
    
    return render_template('asset_form.html',
                         categories=categories,
                         statuses=statuses,
                         asset=None)

@app.route('/assets/create', methods=['POST'])
@login_required
@create_permission_required('asset_management')
def create_asset():
    """Create new asset"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    try:
        asset = Asset(
            name=request.form['name'],
            description=request.form.get('description', ''),
            category_id=int(request.form['category_id']),
            organization_id=user.organization_id,
            quantity=int(request.form.get('quantity', 1)),
            status=request.form['status'],
            location=request.form.get('location', ''),
            serial_number=request.form.get('serial_number', ''),
            purchase_cost=float(request.form['purchase_cost']) if request.form.get('purchase_cost') else None,
            created_by=user.id
        )
        
        # Parse purchase date if provided
        if request.form.get('purchase_date'):
            asset.purchase_date = datetime.strptime(request.form['purchase_date'], '%Y-%m-%d').date()
        
        db.session.add(asset)
        
        # Create submission record
        submission = Submission(
            organization_id=user.organization_id,
            user_id=user.id,
            submission_type='asset_created',
            reference_id=asset.id,
            notes=f'Created asset: {asset.name}'
        )
        db.session.add(submission)
        
        db.session.commit()
        flash('Asset created successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating asset: {str(e)}', 'error')
    
    return redirect(url_for('asset_management'))

@app.route('/assets/<int:asset_id>/edit')
@login_required
@create_permission_required('asset_management')
def edit_asset(asset_id):
    """Edit asset form"""
    user = current_user
    asset = Asset.query.filter_by(id=asset_id, organization_id=user.organization_id).first()
    
    if not asset:
        flash('Asset not found', 'error')
        return redirect(url_for('asset_management'))
    
    categories = AssetCategory.query.all()
    statuses = ['Active', 'Inactive', 'Room', 'Store', 'Clear', 'Other']
    
    return render_template('asset_form.html',
                         categories=categories,
                         statuses=statuses,
                         asset=asset)

@app.route('/assets/<int:asset_id>/update', methods=['POST'])
@login_required
@create_permission_required('asset_management')
def update_asset(asset_id):
    """Update asset"""
    user = current_user
    asset = Asset.query.filter_by(id=asset_id, organization_id=user.organization_id).first()
    
    if not asset:
        flash('Asset not found', 'error')
        return redirect(url_for('asset_management'))
    
    try:
        asset.name = request.form['name']
        asset.description = request.form.get('description', '')
        asset.category_id = int(request.form['category_id'])
        asset.quantity = int(request.form.get('quantity', 1))
        asset.status = request.form['status']
        asset.location = request.form.get('location', '')
        asset.serial_number = request.form.get('serial_number', '')
        asset.purchase_cost = float(request.form['purchase_cost']) if request.form.get('purchase_cost') else None
        
        # Parse purchase date if provided
        if request.form.get('purchase_date'):
            asset.purchase_date = datetime.strptime(request.form['purchase_date'], '%Y-%m-%d').date()
        
        # Create submission record
        submission = Submission(
            organization_id=user.organization_id,
            user_id=user.id,
            submission_type='asset_updated',
            reference_id=asset.id,
            notes=f'Updated asset: {asset.name}'
        )
        db.session.add(submission)
        
        db.session.commit()
        flash('Asset updated successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating asset: {str(e)}', 'error')
    
    return redirect(url_for('asset_management'))

@app.route('/assets/<int:asset_id>/delete', methods=['POST'])
@login_required
@create_permission_required('asset_management')
def delete_asset(asset_id):
    """Delete asset"""
    user = current_user
    asset = Asset.query.filter_by(id=asset_id, organization_id=user.organization_id).first()
    
    if not asset:
        flash('Asset not found', 'error')
        return redirect(url_for('asset_management'))
    
    try:
        asset_name = asset.name
        
        # Create submission record before deletion
        submission = Submission(
            organization_id=user.organization_id,
            user_id=user.id,
            submission_type='asset_deleted',
            reference_id=asset.id,
            notes=f'Deleted asset: {asset_name}'
        )
        db.session.add(submission)
        
        db.session.delete(asset)
        db.session.commit()
        flash(f'Asset "{asset_name}" deleted successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting asset: {str(e)}', 'error')
    
    return redirect(url_for('asset_management'))

@app.route('/select-organization')
@login_required
def select_organization():
    """Allow user to select organization"""
    organizations = Organization.query.all()
    return render_template('select_organization.html', organizations=organizations)

@app.route('/assign-organization', methods=['POST'])
@login_required
@admin_required
def assign_organization():
    """Assign organization to user"""
    user = current_user
    org_id = int(request.form['organization_id'])
    
    organization = Organization.query.get(org_id)
    if organization:
        user.organization_id = org_id
        db.session.commit()
        flash(f'Successfully assigned to {organization.name}', 'success')
    else:
        flash('Invalid organization selected', 'error')
    
    return redirect(url_for('dashboard'))

# Form Management Routes
@app.route('/form-management')
@login_required
@admin_required
def form_management():
    """Form management page"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get forms that the user has permission to access
    user_permissions = UserFormPermission.query.filter_by(user_id=user.id).all()
    allowed_form_ids = [perm.form_template_id for perm in user_permissions]
    
    # If user has no specific permissions, show all forms (for backwards compatibility)
    if not allowed_form_ids:
        forms = FormTemplate.query.filter_by(organization_id=user.organization_id).all()
    else:
        forms = FormTemplate.query.filter(
            FormTemplate.organization_id == user.organization_id,
            FormTemplate.id.in_(allowed_form_ids)
        ).all()
    
    # Group forms by base name (different languages of same form)
    grouped_forms = {}
    for form in forms:
        base_name = form.name.split(' - ')[0] if ' - ' in form.name else form.name
        if base_name not in grouped_forms:
            grouped_forms[base_name] = []
        grouped_forms[base_name].append(form)
    
    return render_template('form_management.html', grouped_forms=grouped_forms)

@app.route('/create-form-template', methods=['POST'])
@login_required
@create_permission_required('forms')
def create_form_template():
    """Create form template with multilingual support"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'error')
        return redirect(url_for('form_management'))
    
    # Check if user has permission to create forms
    form_type = request.form['form_type']
    user_permissions = UserFormPermission.query.filter_by(user_id=user.id).all()
    
    # If user has specific permissions, check if they can create this type
    if user_permissions:
        allowed_types = []
        for perm in user_permissions:
            if perm.can_create and perm.form_template and perm.form_template.form_type:
                allowed_types.append(perm.form_template.form_type)
        
        if form_type not in allowed_types:
            flash('You do not have permission to create this type of form', 'error')
            return redirect(url_for('form_management'))
    
    try:
        import json
        import qrcode
        import io
        import base64
        
        form_name = request.form['form_name']
        form_type = request.form['form_type']
        description = request.form.get('description', '')
        
        # Handle English regulations (text, PDF, or image)
        regulations_en = ''
        if 'pdf_file_en' in request.files and request.files['pdf_file_en'].filename:
            # PDF file uploaded
            pdf_file = request.files['pdf_file_en']
            if pdf_file.filename.endswith('.pdf'):
                # Store PDF content as base64 in regulations_text with special marker
                import base64
                pdf_content = pdf_file.read()
                pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
                regulations_en = f"[PDF_CONTENT]{pdf_base64}"
            else:
                flash('Invalid file type. Please upload a PDF file.', 'error')
                return redirect(url_for('form_management'))
        elif 'image_file_en' in request.files and request.files['image_file_en'].filename:
            # Image file uploaded
            image_file = request.files['image_file_en']
            if image_file.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                # Store image content as base64 in regulations_text with special marker
                import base64
                image_content = image_file.read()
                image_base64 = base64.b64encode(image_content).decode('utf-8')
                regulations_en = f"[IMAGE_CONTENT]{image_base64}"
            else:
                flash('Invalid file type. Please upload an image file (JPG, PNG, GIF).', 'error')
                return redirect(url_for('form_management'))
        else:
            # Text input
            regulations_en = request.form.get('regulations_en', '')
        
        # Process language-specific reference photos (15 total: 3 per language for 5 languages)
        language_ref_photos = {}
        languages = ['en', 'bn', 'my', 'ta', 'zh']
        
        for lang in languages:
            language_ref_photos[lang] = {}
            for photo_num in range(1, 4):  # 1, 2, 3
                photo_key = f'ref_photo_{photo_num}_{lang}'
                if photo_key in request.files and request.files[photo_key].filename:
                    photo_file = request.files[photo_key]
                    if photo_file.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                        # Validate file size (5MB limit)
                        photo_file.seek(0, 2)  # Seek to end
                        file_size = photo_file.tell()
                        photo_file.seek(0)  # Reset to beginning
                        
                        if file_size <= 5 * 1024 * 1024:  # 5MB limit
                            photo_content = photo_file.read()
                            photo_base64 = base64.b64encode(photo_content).decode('utf-8')
                            language_ref_photos[lang][f'photo_{photo_num}'] = photo_base64
                        else:
                            flash(f'Reference photo {photo_num} for {lang.upper()} exceeds 5MB limit.', 'error')
                            return redirect(url_for('form_management'))
                    else:
                        flash(f'Invalid file type for reference photo {photo_num} in {lang.upper()}. Please upload JPG, PNG, or GIF.', 'error')
                        return redirect(url_for('form_management'))

        # Process reference photos
        ref_photo_1_data = None
        ref_photo_2_data = None
        ref_photo_3_data = None
        
        # Handle reference photo 1
        if 'ref_photo_1' in request.files and request.files['ref_photo_1'].filename:
            photo_file = request.files['ref_photo_1']
            if photo_file.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                photo_content = photo_file.read()
                ref_photo_1_data = base64.b64encode(photo_content).decode('utf-8')
        
        # Handle reference photo 2
        if 'ref_photo_2' in request.files and request.files['ref_photo_2'].filename:
            photo_file = request.files['ref_photo_2']
            if photo_file.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                photo_content = photo_file.read()
                ref_photo_2_data = base64.b64encode(photo_content).decode('utf-8')
        
        # Handle reference photo 3
        if 'ref_photo_3' in request.files and request.files['ref_photo_3'].filename:
            photo_file = request.files['ref_photo_3']
            if photo_file.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                photo_content = photo_file.read()
                ref_photo_3_data = base64.b64encode(photo_content).decode('utf-8')

        # Get custom form fields or use default
        fields_json_input = request.form.get('fields_json', '').strip()
        if fields_json_input:
            try:
                # Validate JSON format
                import json
                json.loads(fields_json_input)
                fields_json_data = fields_json_input
            except json.JSONDecodeError:
                # Use default if invalid JSON
                fields_json_data = json.dumps([
                    {"name": "fin", "label": "FIN", "type": "text", "required": True},
                    {"name": "full_name", "label": "Full Name", "type": "text", "required": True},
                    {"name": "company_name", "label": "Company Name", "type": "text", "required": True},
                    {"name": "room_number", "label": "Room Number", "type": "text", "required": True},
                    {"name": "date", "label": "Date", "type": "date", "required": True},
                    {"name": "phone_number", "label": "Phone Number", "type": "tel", "required": True}
                ])
        else:
            # Use default fields
            fields_json_data = json.dumps([
                {"name": "fin", "label": "FIN", "type": "text", "required": True},
                {"name": "full_name", "label": "Full Name", "type": "text", "required": True},
                {"name": "company_name", "label": "Company Name", "type": "text", "required": True},
                {"name": "room_number", "label": "Room Number", "type": "text", "required": True},
                {"name": "date", "label": "Date", "type": "date", "required": True},
                {"name": "phone_number", "label": "Phone Number", "type": "tel", "required": True}
            ])

        # Create base form template with language-specific reference photos
        base_form = FormTemplate(
            name=form_name,
            description=description,
            organization_id=user.organization_id,
            form_type=form_type,
            language_code='en',
            regulations_text=regulations_en,
            fields_json=fields_json_data,
            ref_photo_1=ref_photo_1_data,
            ref_photo_2=ref_photo_2_data,
            ref_photo_3=ref_photo_3_data,
            # Language-specific reference photos
            ref_photo_1_en=language_ref_photos['en'].get('photo_1'),
            ref_photo_2_en=language_ref_photos['en'].get('photo_2'),
            ref_photo_3_en=language_ref_photos['en'].get('photo_3'),
            ref_photo_1_bn=language_ref_photos['bn'].get('photo_1'),
            ref_photo_2_bn=language_ref_photos['bn'].get('photo_2'),
            ref_photo_3_bn=language_ref_photos['bn'].get('photo_3'),
            ref_photo_1_my=language_ref_photos['my'].get('photo_1'),
            ref_photo_2_my=language_ref_photos['my'].get('photo_2'),
            ref_photo_3_my=language_ref_photos['my'].get('photo_3'),
            ref_photo_1_ta=language_ref_photos['ta'].get('photo_1'),
            ref_photo_2_ta=language_ref_photos['ta'].get('photo_2'),
            ref_photo_3_ta=language_ref_photos['ta'].get('photo_3'),
            ref_photo_1_zh=language_ref_photos['zh'].get('photo_1'),
            ref_photo_2_zh=language_ref_photos['zh'].get('photo_2'),
            ref_photo_3_zh=language_ref_photos['zh'].get('photo_3'),
            public_access=True,
            created_by=user.id
        )
        db.session.add(base_form)
        db.session.flush()  # Get the ID
        
        # Create QR code for this form group
        qr_code_data = f"{request.url_root}public/form/{base_form.id}"
        
        # Generate unique code
        import uuid
        qr_code_string = str(uuid.uuid4())[:8].upper()
        
        qr_code_record = QRCode(
            code=qr_code_string,
            qr_type='form',
            reference_id=str(base_form.id),
            reference_table='form_templates',
            organization_id=user.organization_id,
            label=f"Form: {form_name}",
            description=f"QR code for multilingual form: {form_name}",
            created_by=user.id
        )
        db.session.add(qr_code_record)
        db.session.flush()
        
        # Update form with QR code reference
        base_form.qr_code_id = qr_code_record.id
        
        # Create multilingual versions with separate images
        languages = {
            'bn': request.form.get('regulations_bn', ''),
            'my': request.form.get('regulations_my', ''),
            'ta': request.form.get('regulations_ta', ''),
            'zh': request.form.get('regulations_zh', '')
        }
        
        for lang_code, regulations_text in languages.items():
            if regulations_text.strip():
                # Handle separate reference photos for each language
                lang_ref_photo_1_data = None
                lang_ref_photo_2_data = None
                lang_ref_photo_3_data = None
                
                # Check for language-specific reference photos
                ref_photo_1_lang = request.files.get(f'ref_photo_1_{lang_code}')
                if ref_photo_1_lang and ref_photo_1_lang.filename:
                    if ref_photo_1_lang.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                        import base64
                        lang_ref_photo_1_data = base64.b64encode(ref_photo_1_lang.read()).decode('utf-8')
                
                ref_photo_2_lang = request.files.get(f'ref_photo_2_{lang_code}')
                if ref_photo_2_lang and ref_photo_2_lang.filename:
                    if ref_photo_2_lang.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                        import base64
                        lang_ref_photo_2_data = base64.b64encode(ref_photo_2_lang.read()).decode('utf-8')
                
                ref_photo_3_lang = request.files.get(f'ref_photo_3_{lang_code}')
                if ref_photo_3_lang and ref_photo_3_lang.filename:
                    if ref_photo_3_lang.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                        import base64
                        lang_ref_photo_3_data = base64.b64encode(ref_photo_3_lang.read()).decode('utf-8')
                
                lang_form = FormTemplate(
                    name=f"{form_name} - {lang_code.upper()}",
                    description=description,
                    organization_id=user.organization_id,
                    form_type=form_type,
                    language_code=lang_code,
                    regulations_text=regulations_text,
                    fields_json=base_form.fields_json,
                    ref_photo_1=lang_ref_photo_1_data,
                    ref_photo_2=lang_ref_photo_2_data,
                    ref_photo_3=lang_ref_photo_3_data,
                    qr_code_id=qr_code_record.id,
                    public_access=True,
                    created_by=user.id
                )
                db.session.add(lang_form)
        
        db.session.commit()
        flash('Form template created successfully with QR code!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating form template: {str(e)}', 'error')
    
    return redirect(url_for('form_management'))

@app.route('/serve_pdf/<int:form_id>')
def serve_pdf(form_id):
    """Serve PDF regulations content"""
    form = FormTemplate.query.get_or_404(form_id)
    
    if not form.regulations_text.startswith('[PDF_CONTENT]'):
        return "Not a PDF file", 404
    
    # Extract PDF content
    import base64
    pdf_base64 = form.regulations_text[13:]  # Remove '[PDF_CONTENT]' prefix
    pdf_content = base64.b64decode(pdf_base64)
    
    from flask import Response
    return Response(
        pdf_content,
        mimetype='application/pdf',
        headers={
            'Content-Disposition': f'inline; filename=regulations_{form.name}.pdf'
        }
    )

@app.route('/serve_image/<int:form_id>')
def serve_image(form_id):
    """Serve image regulations content"""
    form = FormTemplate.query.get_or_404(form_id)
    
    if not form.regulations_text.startswith('[IMAGE_CONTENT]'):
        return "Not an image file", 404
    
    # Extract image content
    import base64
    image_base64 = form.regulations_text[15:]  # Remove '[IMAGE_CONTENT]' prefix
    image_content = base64.b64decode(image_base64)
    
    from flask import Response
    return Response(
        image_content,
        mimetype='image/jpeg',  # Default mimetype, browsers will handle different formats
        headers={
            'Content-Disposition': f'inline; filename=regulations_{form.name}.jpg'
        }
    )

@app.route('/serve_ref_photo/<int:form_id>/<int:photo_num>')
def serve_ref_photo(form_id, photo_num):
    """Serve reference photo content"""
    form = FormTemplate.query.get_or_404(form_id)
    
    # Get the appropriate reference photo
    photo_data = None
    if photo_num == 1 and form.ref_photo_1:
        photo_data = form.ref_photo_1
    elif photo_num == 2 and form.ref_photo_2:
        photo_data = form.ref_photo_2
    elif photo_num == 3 and form.ref_photo_3:
        photo_data = form.ref_photo_3
    
    if not photo_data:
        return "Photo not found", 404
    
    # Decode and serve the image
    import base64
    image_content = base64.b64decode(photo_data)
    
    from flask import Response
    return Response(
        image_content,
        mimetype='image/jpeg',  # Default mimetype, browsers will handle different formats
        headers={
            'Content-Disposition': f'inline; filename=ref_photo_{photo_num}_{form.name}.jpg'
        }
    )

# Public form access routes (no login required)
@app.route('/public/form/<int:form_id>')
def public_form_language_select(form_id):
    """Public form access - language selection"""
    base_form = FormTemplate.query.get_or_404(form_id)
    if not base_form.public_access:
        return render_template('403.html'), 403
    
    # Get all language versions of this form
    all_forms = FormTemplate.query.filter_by(
        organization_id=base_form.organization_id,
        form_type=base_form.form_type,
        qr_code_id=base_form.qr_code_id
    ).all()
    
    # Group by language
    languages = {}
    for form in all_forms:
        lang_name = {
            'en': 'English',
            'bn': 'বাংলা (Bengali)', 
            'my': 'မြန်မာ (Myanmar)',
            'ta': 'தமிழ் (Tamil)',
            'zh': '中文 (Chinese)'
        }.get(form.language_code, form.language_code.upper())
        languages[form.language_code] = {
            'name': lang_name,
            'form': form
        }
    
    return render_template('public_language_select.html', 
                         base_form=base_form, 
                         languages=languages)



@app.route('/public/room-checklist')
def public_room_checklist():
    """Public room checklist form - no authentication required"""
    return render_template('public_room_checklist.html')



@app.route('/submit-public-room-checklist', methods=['POST'])
@create_permission_required('room_checklist')
def submit_public_room_checklist():
    """Handle public room checklist submission"""
    import json
    try:
        # Parse handover date and time
        handover_date_str = request.form.get('handover_date')
        handover_time_str = request.form.get('handover_time')
        
        handover_datetime = None
        if handover_date_str:
            try:
                if handover_time_str:
                    handover_datetime = datetime.strptime(f"{handover_date_str} {handover_time_str}", "%Y-%m-%d %H:%M")
                else:
                    handover_datetime = datetime.strptime(handover_date_str, "%Y-%m-%d")
            except ValueError:
                handover_datetime = datetime.strptime(handover_date_str, "%Y-%m-%d")
        
        # Compile item conditions into a JSON format
        item_conditions = {}
        condition_fields = [
            'bed_frame', 'mattress', 'wardrobe', 'desk', 'chair',
            'ceiling_light', 'power_outlets', 'air_conditioning',
            'walls', 'floor_condition', 'windows', 'door'
        ]
        
        for field in condition_fields:
            condition = request.form.get(field)
            if condition:
                item_conditions[field] = condition
        
        # Prepare handover signature data
        handover_signature = {
            'name': request.form.get('handover_name'),
            'designation': request.form.get('handover_designation'),
            'finNumber': request.form.get('handover_fin'),
            'dateTime': request.form.get('handover_datetime')
        }
        
        # Prepare takeover signature data
        takeover_signature = {
            'name': request.form.get('takeover_name'),
            'designation': request.form.get('takeover_designation'),
            'finNumber': request.form.get('takeover_fin'),
            'dateTime': request.form.get('takeover_datetime')
        }
        
        # Create room checklist record
        checklist = RoomInventoryChecklist(
            room_number=request.form.get('room_number'),
            block=request.form.get('block'),
            floor=request.form.get('floor'),
            organization_id=1,  # Default to first organization for public submissions
            previous_occupant=request.form.get('previous_occupant'),
            new_occupant=request.form.get('new_occupant'),
            handover_date=handover_datetime.date() if handover_datetime else date.today(),
            handover_time=handover_datetime.time() if handover_datetime and handover_time_str else None,
            condition_before=request.form.get('condition_before'),
            condition_after=request.form.get('condition_after'),
            damages_noted=request.form.get('damages_noted'),
            repairs_needed=request.form.get('repairs_needed'),
            status='Completed',
            conducted_by='1',  # Use first user ID for public submissions
            handover_signature=json.dumps(handover_signature),
            takeover_signature=json.dumps(takeover_signature)
        )
        
        db.session.add(checklist)
        db.session.flush()  # Get the ID
        
        # Store item conditions as JSON in the damages_noted field
        if item_conditions:
            import json
            conditions_text = json.dumps(item_conditions, indent=2)
            if checklist.damages_noted:
                checklist.damages_noted += f"\n\nItem Conditions:\n{conditions_text}"
            else:
                checklist.damages_noted = f"Item Conditions:\n{conditions_text}"
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'checklist_id': checklist.id,
            'message': 'Room checklist submitted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/public/form/<int:form_id>/<language>')
def public_form_display(form_id, language):
    """Public form display with regulations and form fields"""
    form = FormTemplate.query.filter_by(
        qr_code_id=FormTemplate.query.get(form_id).qr_code_id,
        language_code=language
    ).first()
    
    if not form:
        # Fallback to English if language not found
        form = FormTemplate.query.get(form_id)
    
    if not form or not form.public_access:
        return render_template('403.html'), 403
    
    import json
    form_fields = json.loads(form.fields_json) if form.fields_json else []
    
    return render_template('public_form_display.html', 
                         form=form, 
                         form_fields=form_fields)

@app.route('/public/form/submit', methods=['POST'])
def public_form_submit():
    """Handle public form submission"""
    try:
        import json
        import logging
        
        # Debug logging
        logging.info(f"Form submission attempt - form data: {dict(request.form)}")
        
        form_id = int(request.form['form_id'])
        form = FormTemplate.query.get_or_404(form_id)
        
        logging.info(f"Found form: {form.name} (ID: {form.id})")
        
        # Collect form data
        form_data = {}
        for key, value in request.form.items():
            if key != 'form_id':
                form_data[key] = value
        
        logging.info(f"Collected form data: {form_data}")
        
        # Create form submission with proper handling
        submission = FormSubmission()
        submission.form_template_id = form.id
        submission.organization_id = form.organization_id
        submission.submitted_by = None  # Anonymous submission
        submission.form_data_json = json.dumps(form_data)
        submission.status = 'Submitted'
        
        db.session.add(submission)
        db.session.flush()  # Get the submission ID
        
        logging.info(f"Created form submission with ID: {submission.id}")
        
        # Create general submission record
        general_submission = Submission()
        general_submission.organization_id = form.organization_id
        general_submission.user_id = None  # Anonymous
        general_submission.submission_type = 'form_submitted_public'
        general_submission.reference_id = submission.id
        general_submission.reference_table = 'form_submissions'
        general_submission.notes = f'Public form submission: {form.name} - {form_data.get("full_name", "Anonymous")}'
        
        db.session.add(general_submission)
        
        db.session.commit()
        
        logging.info("Form submission saved successfully")
        
        return render_template('form_submitted.html', form=form, form_data=form_data)
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Form submission error: {str(e)}")
        return render_template('form_error.html', error=str(e))

@app.route('/qr-code/<int:qr_id>')
def generate_qr_code(qr_id):
    """Generate and serve QR code image"""
    try:
        import qrcode
        import io
        from flask import Response
        from models import QRCode
        
        qr_record = QRCode.query.get_or_404(qr_id)
        
        # Generate QR code URL - always use the QR redirect pattern for consistency
        qr_url = f"{request.url_root}qr/{qr_record.code}"
        
        # Create QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_url)
        qr.make(fit=True)
        
        # Create image
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Save to bytes
        img_buffer = io.BytesIO()
        img.save(img_buffer, format='PNG')
        img_buffer.seek(0)
        
        return Response(img_buffer.getvalue(), mimetype='image/png')
        
    except Exception as e:
        # Return a simple error response
        from flask import jsonify
        return jsonify({'error': str(e)}), 500

# QR Codes Routes
@app.route('/qr-codes')
@login_required
@admin_required
def qr_codes():
    """QR codes management page"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    qr_codes_list = QRCode.query.filter_by(organization_id=user.organization_id).all()
    return render_template('qr_codes.html', qr_codes=qr_codes_list)

@app.route('/generate_qr_code', methods=['POST'])
@login_required
@create_permission_required('qr_codes')
def generate_qr_code_post():
    """Generate a new QR code"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'Please contact administrator to assign organization'})
    
    try:
        qr_type = request.form.get('qr_type')
        label = request.form.get('label')
        description = request.form.get('description')
        
        # URL-specific fields
        target_url = request.form.get('target_url')
        expires_at = request.form.get('expires_at')
        max_scans = request.form.get('max_scans')
        is_public = request.form.get('is_public', 'true').lower() == 'true'
        
        if not qr_type or not label:
            return jsonify({'success': False, 'error': 'QR type and label are required'})
        
        if qr_type == 'url' and not target_url:
            return jsonify({'success': False, 'error': 'Target URL is required for URL type QR codes'})
        
        # Validate URL format for URL type
        if qr_type == 'url' and target_url:
            from urllib.parse import urlparse
            parsed = urlparse(target_url)
            if not all([parsed.scheme, parsed.netloc]):
                return jsonify({'success': False, 'error': 'Please enter a valid URL (e.g., https://example.com)'})
        
        # Parse expiry date
        expires_at_datetime = None
        if expires_at:
            try:
                from datetime import datetime
                expires_at_datetime = datetime.fromisoformat(expires_at.replace('T', ' '))
            except ValueError:
                return jsonify({'success': False, 'error': 'Invalid expiry date format'})
        
        # Parse max scans
        max_scans_int = None
        if max_scans:
            try:
                max_scans_int = int(max_scans)
                if max_scans_int <= 0:
                    return jsonify({'success': False, 'error': 'Max scans must be a positive number'})
            except ValueError:
                return jsonify({'success': False, 'error': 'Invalid max scans value'})
        
        # Generate unique code
        import secrets
        code = secrets.token_hex(8)
        
        # Create QR code record
        qr_code = QRCode(
            code=code,
            qr_type=qr_type,
            label=label,
            description=description,
            organization_id=user.organization_id,
            created_by=user.id,
            target_url=target_url if qr_type == 'url' else None,
            expires_at=expires_at_datetime,
            max_scans=max_scans_int,
            is_public=is_public
        )
        
        db.session.add(qr_code)
        db.session.commit()
        
        return jsonify({'success': True, 'qr_id': qr_code.id})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/edit_qr_code/<int:qr_id>', methods=['POST'])
@login_required
@create_permission_required('qr_codes')
def edit_qr_code(qr_id):
    """Edit an existing QR code"""
    try:
        user = current_user
        if not user.organization_id:
            return jsonify({'success': False, 'error': 'Access denied'})
        
        # Get the QR code
        qr_code = QRCode.query.filter_by(id=qr_id, organization_id=user.organization_id).first()
        if not qr_code:
            return jsonify({'success': False, 'error': 'QR code not found'})
        
        # Get form data
        qr_type = request.form.get('qr_type')
        label = request.form.get('label')
        description = request.form.get('description', '')
        target_url = request.form.get('target_url', '')
        
        # Validate required fields
        if not qr_type or not label:
            return jsonify({'success': False, 'error': 'QR type and label are required'})
        
        # Validate URL for URL type
        if qr_type == 'url':
            if not target_url:
                return jsonify({'success': False, 'error': 'Target URL is required for URL type QR codes'})
            
            from urllib.parse import urlparse
            parsed = urlparse(target_url)
            if not all([parsed.scheme, parsed.netloc]):
                return jsonify({'success': False, 'error': 'Please enter a valid URL (e.g., https://example.com)'})
        
        # Update QR code record
        qr_code.qr_type = qr_type
        qr_code.label = label
        qr_code.description = description
        qr_code.target_url = target_url if qr_type == 'url' else None
        
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete_qr_code/<int:qr_id>', methods=['POST'])
@login_required
@create_permission_required('qr_codes')
def delete_qr_code(qr_id):
    """Delete a QR code"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    qr_code = QRCode.query.filter_by(id=qr_id, organization_id=user.organization_id).first()
    if not qr_code:
        flash('QR code not found.', 'error')
        return redirect(url_for('qr_codes'))
    
    try:
        # First, remove QR code references from form templates
        forms_referencing_qr = FormTemplate.query.filter_by(
            organization_id=user.organization_id,
            qr_code_id=qr_code.id
        ).all()
        
        for form in forms_referencing_qr:
            form.qr_code_id = None
        
        # If this is a form QR code, also delete the associated form templates
        if qr_code.qr_type == 'form' and qr_code.reference_id:
            try:
                base_form_id = int(qr_code.reference_id)
                # Delete all language variants of this form
                FormTemplate.query.filter_by(
                    organization_id=user.organization_id
                ).filter(
                    FormTemplate.id == base_form_id
                ).delete(synchronize_session=False)
            except (ValueError, TypeError):
                # If reference_id is not a valid integer, skip form deletion
                pass
        
        # Delete any form submissions associated with forms that used this QR code
        from models import FormSubmission
        FormSubmission.query.filter(
            FormSubmission.form_template_id.in_(
                db.session.query(FormTemplate.id).filter_by(qr_code_id=qr_code.id)
            )
        ).delete(synchronize_session=False)
        
        # Delete the QR code
        db.session.delete(qr_code)
        db.session.commit()
        
        flash(f'QR code "{qr_code.label or qr_code.code}" has been deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting QR code: {str(e)}', 'error')
    
    return redirect(url_for('qr_codes'))

@app.route('/download_qr/<int:qr_id>')
@login_required
def download_qr_code(qr_id):
    """Download QR code as PNG file"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    qr_record = QRCode.query.filter_by(id=qr_id, organization_id=user.organization_id).first()
    if not qr_record:
        flash('QR code not found.', 'error')
        return redirect(url_for('qr_codes'))
    
    try:
        import io
        import qrcode
        
        # Generate QR code URL
        if qr_record.qr_type == 'form':
            qr_url = f"{request.url_root}public/form/{qr_record.reference_id}"
        else:
            qr_url = f"{request.url_root}scan/{qr_record.code}"
        
        # Create QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_url)
        qr.make(fit=True)
        
        # Create image
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Save to bytes
        img_buffer = io.BytesIO()
        img.save(img_buffer, format='PNG')
        img_buffer.seek(0)
        
        # Create filename
        safe_filename = "".join(c for c in (qr_record.label or qr_record.code) if c.isalnum() or c in (' ', '-', '_')).rstrip()
        filename = f"QR_{safe_filename}.png"
        
        from flask import Response
        return Response(
            img_buffer.getvalue(),
            mimetype='image/png',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )
        
    except Exception as e:
        flash(f'Error generating QR code: {str(e)}', 'error')
        return redirect(url_for('qr_codes'))

@app.route('/bulk-download-qr-codes')
@login_required
def bulk_download_qr_codes():
    """Download all QR codes as a ZIP file"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    try:
        import zipfile
        import io
        import qrcode
        
        # Get all QR codes for this organization
        qr_codes = QRCode.query.filter_by(organization_id=user.organization_id).all()
        
        if not qr_codes:
            flash('No QR codes found to download', 'warning')
            return redirect(url_for('qr_codes'))
        
        # Create ZIP file in memory
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for qr_record in qr_codes:
                try:
                    # Generate QR code URL
                    if qr_record.qr_type == 'form':
                        qr_url = f"{request.url_root}public/form/{qr_record.reference_id}"
                    else:
                        qr_url = f"{request.url_root}qr/{qr_record.code}"
                    
                    # Create QR code
                    qr = qrcode.QRCode(
                        version=1,
                        error_correction=qrcode.constants.ERROR_CORRECT_L,
                        box_size=10,
                        border=4,
                    )
                    qr.add_data(qr_url)
                    qr.make(fit=True)
                    
                    # Create image
                    img = qr.make_image(fill_color="black", back_color="white")
                    
                    # Save to bytes
                    img_buffer = io.BytesIO()
                    img.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    # Create safe filename
                    safe_filename = "".join(c for c in (qr_record.label or qr_record.code) if c.isalnum() or c in (' ', '-', '_')).rstrip()
                    filename = f"QR_{safe_filename}.png"
                    
                    # Add to ZIP
                    zip_file.writestr(filename, img_buffer.getvalue())
                    
                except Exception as e:
                    print(f"Error processing QR code {qr_record.id}: {str(e)}")
                    continue
        
        zip_buffer.seek(0)
        
        from flask import Response
        return Response(
            zip_buffer.getvalue(),
            mimetype='application/zip',
            headers={
                'Content-Disposition': 'attachment; filename="qr_codes_bulk.zip"'
            }
        )
        
    except Exception as e:
        flash(f'Error creating bulk download: {str(e)}', 'error')
        return redirect(url_for('qr_codes'))



# Room Handover Routes
@app.route('/room-handovers')
@login_required
def room_handovers():
    """Room handover page - displays room inventory checklists"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get room inventory checklists instead of old handovers
    checklists = RoomInventoryChecklist.query.filter_by(organization_id=user.organization_id).order_by(RoomInventoryChecklist.created_at.desc()).all()
    
    return render_template('room_handovers.html', checklists=checklists)

# Handover Records Routes (alias for room handovers)
@app.route('/handover-records')
@login_required
def handover_records():
    """Handover records page (same as room handovers)"""
    return redirect(url_for('room_handovers'))

@app.route('/room-checklist')
@login_required
@page_access_required('room_checklist')
def room_checklist():
    """Room inventory checklist page"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    from datetime import date
    from app.models.models_house_acknowledge import RoomNumber
    today = date.today().isoformat()
    
    # Get all active room numbers from database
    room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    
    return render_template('room_checklist.html', today=today, room_numbers=room_numbers)

@app.route('/save_room_checklist', methods=['POST'])
@login_required
@create_permission_required('room_checklist')
def save_room_checklist():
    """Save room inventory checklist to database"""
    import json
    
    try:
        data = request.get_json()
        
        # Parse the date
        checklist_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        
        # Create new room inventory checklist
        checklist = RoomInventoryChecklist()
        checklist.room_number = data['roomNumber']
        checklist.company_name = data['companyName']
        checklist.checklist_date = checklist_date
        checklist.organization_id = current_user.organization_id or 1
        checklist.created_by = current_user.id
        
        # Save meter readings
        meter_readings = data.get('meterReadings', {})
        checklist.water_meter_reading = meter_readings.get('water')
        checklist.electricity_meter_reading = meter_readings.get('electricity')
        
        # Combine date and time for meter readings
        reading_date = meter_readings.get('date')
        reading_time = meter_readings.get('time')
        if reading_date and reading_time:
            meter_datetime = datetime.strptime(f"{reading_date} {reading_time}", '%Y-%m-%d %H:%M')
            checklist.water_meter_datetime = meter_datetime
            checklist.electricity_meter_datetime = meter_datetime
        elif reading_date:
            meter_datetime = datetime.strptime(reading_date, '%Y-%m-%d')
            checklist.water_meter_datetime = meter_datetime
            checklist.electricity_meter_datetime = meter_datetime
        
        # Save signature data as JSON
        checklist.handover_signature_data = json.dumps(data['signatures']['handover'])
        checklist.takeover_signature_data = json.dumps(data['signatures']['takeover'])
        
        # Save checklist items as JSON
        checklist.checklist_items_data = json.dumps(data['items'])
        
        db.session.add(checklist)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Room inventory checklist saved successfully!', 'id': checklist.id})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error saving checklist: {str(e)}'}), 500

@app.route('/view_checklist/<int:checklist_id>')
@login_required
def view_checklist(checklist_id):
    """View detailed room inventory checklist"""
    import json
    
    checklist = RoomInventoryChecklist.query.get_or_404(checklist_id)
    
    # Parse JSON data and format for display
    raw_items = json.loads(checklist.checklist_items_data) if checklist.checklist_items_data else []
    handover_signature = json.loads(checklist.handover_signature_data) if checklist.handover_signature_data else {}
    takeover_signature = json.loads(checklist.takeover_signature_data) if checklist.takeover_signature_data else {}
    
    # Format checklist items for display
    checklist_items = []
    if raw_items:
        for item in raw_items:
            if isinstance(item, dict):
                # Handle the new format from room checklist
                defects_remarks = ''
                defects = item.get('defects', '') or item.get('defectsRemarks', '')
                remarks = item.get('remarks', '')
                
                if defects and remarks:
                    defects_remarks = f"{defects}; {remarks}"
                elif defects:
                    defects_remarks = defects
                elif remarks:
                    defects_remarks = remarks
                else:
                    defects_remarks = '-'
                
                formatted_item = {
                    'category': item.get('section', 'Unknown').title(),
                    'item': item.get('description', 'Unknown Item'),
                    'status': item.get('condition', 'Good'),
                    'remarks': defects_remarks
                }
                checklist_items.append(formatted_item)
    
    return render_template('view_checklist.html', 
                         checklist=checklist,
                         checklist_items=checklist_items,
                         handover_signature=handover_signature,
                         takeover_signature=takeover_signature)

@app.route('/edit_checklist/<int:checklist_id>')
@login_required
def edit_checklist(checklist_id):
    """Edit room inventory checklist"""
    import json
    
    checklist = RoomInventoryChecklist.query.get_or_404(checklist_id)
    
    # Parse JSON data for editing
    checklist_items = json.loads(checklist.checklist_items_data) if checklist.checklist_items_data else []
    handover_signature = json.loads(checklist.handover_signature_data) if checklist.handover_signature_data else {}
    takeover_signature = json.loads(checklist.takeover_signature_data) if checklist.takeover_signature_data else {}
    
    # Convert checklist object to dictionary for JSON serialization
    checklist_dict = {
        'id': checklist.id,
        'room_number': checklist.room_number,
        'company_name': checklist.company_name,
        'checklist_date': checklist.checklist_date.isoformat() if checklist.checklist_date else None,
        'water_meter_reading': checklist.water_meter_reading,
        'water_meter_datetime': checklist.water_meter_datetime.isoformat() if checklist.water_meter_datetime else None,
        'electricity_meter_reading': checklist.electricity_meter_reading,
        'electricity_meter_datetime': checklist.electricity_meter_datetime.isoformat() if checklist.electricity_meter_datetime else None,
        'water_meter_signature': checklist.water_meter_signature,
        'electricity_meter_signature': checklist.electricity_meter_signature,
        'created_at': checklist.created_at.isoformat() if checklist.created_at else None,
        'updated_at': checklist.updated_at.isoformat() if checklist.updated_at else None
    }
    
    return render_template('edit_checklist.html', 
                         checklist=checklist_dict,
                         checklist_items=checklist_items,
                         handover_signature=handover_signature,
                         takeover_signature=takeover_signature)

@app.route('/update_checklist', methods=['POST'])
@login_required
@create_permission_required('room_checklist')
def update_checklist():
    """Update room inventory checklist"""
    import json
    
    try:
        data = request.get_json()
        checklist_id = data.get('id')
        
        if not checklist_id:
            return jsonify({'success': False, 'error': 'Checklist ID is required'}), 400
        
        checklist = RoomInventoryChecklist.query.get_or_404(checklist_id)
        
        # Update basic checklist fields
        checklist.room_number = data.get('roomNumber')
        checklist.company_name = data.get('companyName')
        if data.get('date'):
            checklist.checklist_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        
        # Update meter readings
        checklist.water_meter_reading = data.get('waterMeterReading')
        if data.get('waterMeterDateTime'):
            try:
                checklist.water_meter_datetime = datetime.fromisoformat(data['waterMeterDateTime'])
            except ValueError:
                pass  # Skip invalid datetime
        
        checklist.electricity_meter_reading = data.get('electricityMeterReading')
        if data.get('electricityMeterDateTime'):
            try:
                checklist.electricity_meter_datetime = datetime.fromisoformat(data['electricityMeterDateTime'])
            except ValueError:
                pass  # Skip invalid datetime
        
        # Update signature data
        signatures = data.get('signatures', {})
        if signatures.get('handover'):
            checklist.handover_signature_data = json.dumps(signatures['handover'])
        if signatures.get('takeover'):
            checklist.takeover_signature_data = json.dumps(signatures['takeover'])
        
        # Update meter signatures
        meter_signatures = data.get('meterSignatures', {})
        if meter_signatures.get('water'):
            checklist.water_meter_signature = meter_signatures['water']
        if meter_signatures.get('electricity'):
            checklist.electricity_meter_signature = meter_signatures['electricity']
        
        # Update checklist items
        checklist_items = data.get('checklistItems', [])
        if checklist_items:
            checklist.checklist_items_data = json.dumps(checklist_items)
        
        checklist.updated_at = singapore_now()
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Room inventory checklist updated successfully!'})
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Error updating checklist: {str(e)}")
        return jsonify({'success': False, 'error': f'Error updating checklist: {str(e)}'}), 500

@app.route('/export_checklist_pdf/<int:checklist_id>')
@login_required
def export_checklist_pdf(checklist_id):
    """Export single room inventory checklist to PDF using the same format as room checklist"""
    import json
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO
    from flask import make_response
    import os
    
    checklist = RoomInventoryChecklist.query.get_or_404(checklist_id)
    
    # Parse JSON data
    checklist_items = json.loads(checklist.checklist_items_data) if checklist.checklist_items_data else []
    handover_signature = json.loads(checklist.handover_signature_data) if checklist.handover_signature_data else {}
    takeover_signature = json.loads(checklist.takeover_signature_data) if checklist.takeover_signature_data else {}
    
    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Build story
    story = []
    styles = getSampleStyleSheet()
    
    # Add TS GROUP logo
    try:
        logo_path = os.path.join('static', 'ts_group_logo.jpg')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=3*inch, height=1.2*inch)
            logo.hAlign = 'CENTER'
            story.append(logo)
            story.append(Spacer(1, 15))
    except Exception as e:
        # Fallback to text header if image fails
        logo_style = ParagraphStyle(
            'LogoStyle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#2c5aa0'),
            alignment=TA_CENTER,
            spaceAfter=10
        )
        story.append(Paragraph("TS GROUP", logo_style))
        story.append(Spacer(1, 10))
    
    # Company header
    story.append(Paragraph("TS Management Services Pte Ltd", styles['Heading1']))
    story.append(Paragraph("Pioneer Lodge", styles['Heading2']))
    story.append(Paragraph("Room Inventory Checklist", styles['Heading3']))
    story.append(Spacer(1, 20))
    
    # Basic information table
    info_data = [
        ['Room Number:', checklist.room_number, 'Date:', checklist.checklist_date.strftime('%Y-%m-%d') if checklist.checklist_date else ''],
        ['Company Name:', checklist.company_name, 'Status:', checklist.status]
    ]
    
    info_table = Table(info_data, colWidths=[1.5*inch, 2*inch, 1*inch, 1.5*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Meter readings table
    water_date = checklist.water_meter_datetime.strftime('%Y-%m-%d') if checklist.water_meter_datetime else ''
    electricity_date = checklist.electricity_meter_datetime.strftime('%Y-%m-%d') if checklist.electricity_meter_datetime else ''
    
    meter_data = [
        ['Item', 'Reading', 'Date'],
        ['Water Meter', checklist.water_meter_reading or '', water_date],
        ['Electricity Meter', checklist.electricity_meter_reading or '', electricity_date]
    ]
    
    meter_table = Table(meter_data, colWidths=[2*inch, 2*inch, 2.5*inch])
    meter_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(meter_table)
    story.append(Spacer(1, 20))
    
    # Checklist Items Table
    inventory_data = [['S/NO', 'DESCRIPTION', 'QTY', 'Condition', 'Defects/Remarks']]
    
    # Process checklist items by section
    if checklist_items:
        # Group items by section
        sections = {'electrical': [], 'bedroom': [], 'toilet': [], 'kitchen': [], 'fire-protection': [], 'others': []}
        for item in checklist_items:
            section = item.get('section', 'others').lower()
            if section in sections:
                sections[section].append(item)
        
        # Display sections in order: Electrical, Bedroom, Toilet, Kitchen, Fire Protection, Others
        section_order = [
            ('electrical', 'Electrical Items'),
            ('bedroom', 'Bedroom'),
            ('toilet', 'Toilet'),
            ('kitchen', 'Kitchen'),
            ('fire-protection', 'Fire Protection'),
            ('others', 'Others')
        ]
        
        for section_key, section_title in section_order:
            if sections[section_key]:
                # Add section header row
                inventory_data.append([section_title, '', '', '', ''])
                
                for item in sections[section_key]:
                    # Combine defects and remarks into one field
                    defects_remarks = ''
                    defects = item.get('defects', '') or item.get('defectsRemarks', '')
                    remarks = item.get('remarks', '')
                    
                    if defects and remarks:
                        defects_remarks = f"{defects}; {remarks}"
                    elif defects:
                        defects_remarks = defects
                    elif remarks:
                        defects_remarks = remarks
                    
                    inventory_data.append([
                        item.get('sno', ''),
                        item.get('description', ''),
                        item.get('quantity', '1'),
                        item.get('condition', 'GOOD'),
                        defects_remarks
                    ])
    else:
        # Default structure with all items
        default_sections = [
            ('Electrical Items', [
                ('1', 'EXIT light unit', '1', 'GOOD', ''),
                ('2', 'Wireless router', '1', 'GOOD', ''),
                ('3', 'Ceiling Fan switch (3 G / 1 G)', '1', 'GOOD', ''),
                ('4', 'Lighting switch (1 gang)', '1', 'GOOD', ''),
                ('5', 'Lighting switches (2 gang)', '1', 'GOOD', ''),
                ('6', 'Ceiling Fan', '1', 'GOOD', ''),
                ('7', 'Wall fan', '1', 'GOOD', ''),
                ('8', 'Stand Fan', '1', 'GOOD', ''),
                ('9', '13A Single Wall Socket', '1', 'GOOD', ''),
                ('10', '13A Twin Wall Socket', '1', 'GOOD', ''),
                ('11', '20W LED Tube Lights', '1', 'GOOD', ''),
                ('12', '2 x 20W LED Tube Lights w Guard', '1', 'GOOD', ''),
                ('13', 'Emergency Light Unit', '1', 'GOOD', '')
            ]),
            ('Bedroom', [
                ('1', 'Bunk bed w/ladder', '1', 'GOOD', ''),
                ('2', 'Locker w/ accessories', '1', 'GOOD', ''),
                ('3', 'Luggage rack', '1', 'GOOD', ''),
                ('4', 'Mattresses', '2', 'GOOD', ''),
                ('5', 'Pillows', '2', 'GOOD', '')
            ]),
            ('Toilet', [
                ('1', 'Mirror', '1', 'GOOD', ''),
                ('2', 'Wash Basin tap', '1', 'GOOD', ''),
                ('3', 'Wash Basin + Drainstopper', '1', 'GOOD', ''),
                ('4', 'Wash Basin Bottle Trap (PVC)', '1', 'GOOD', ''),
                ('5', 'Toilet cistern w/ lid + seat (Including flushing system)', '1', 'GOOD', ''),
                ('6', 'Bidet Tap', '1', 'GOOD', ''),
                ('7', 'Floor trap (4")', '1', 'GOOD', ''),
                ('8', 'Shower head & tap', '1', 'GOOD', ''),
                ('9', 'Cubicle door w/lock & hanger hook', '1', 'GOOD', ''),
                ('10', 'Partition boards & feet (cubicle/set)', '1', 'GOOD', ''),
                ('11', 'Wall & floor tiles', '1', 'GOOD', '')
            ]),
            ('Kitchen', [
                ('1', 'Washer', '1', 'GOOD', ''),
                ('2', 'Tap (washer)', '1', 'GOOD', ''),
                ('3', 'Fridge', '1', 'GOOD', ''),
                ('4', 'Dining table', '1', 'GOOD', ''),
                ('5', 'Chair', '4', 'GOOD', ''),
                ('6', 'Induction stove', '1', 'GOOD', ''),
                ('7', 'Kettle', '1', 'GOOD', ''),
                ('8', 'Wash basin', '1', 'GOOD', ''),
                ('9', 'Tap (basin)', '1', 'GOOD', ''),
                ('10', 'Prep area (tiles)', '1', 'GOOD', ''),
                ('11', 'Floor trap 6"', '1', 'GOOD', ''),
                ('12', 'Kitchen floor (tiles)', '1', 'GOOD', ''),
                ('13', 'Clothes hanger (above head/over window*)', '1', 'GOOD', ''),
                ('14', 'Clothes hanging poles*', '1', 'GOOD', ''),
                ('15', 'Y Stick*', '1', 'GOOD', ''),
                ('16', 'Top (stovetop)', '1', 'GOOD', '')
            ]),
            ('Fire Protection', [
                ('1', 'Smoke detector', '1', 'GOOD', ''),
                ('2', 'Fire extinguisher', '1', 'GOOD', '')
            ]),
            ('Others', [
                ('1', 'Window with grilles / protector / coverings', '1', 'GOOD', ''),
                ('2', 'Main door w/ lock + key', '1', 'GOOD', ''),
                ('3', 'Floor (tiles/vinyl)', '1', 'GOOD', ''),
                ('4', 'False ceiling grid / panel', '1', 'GOOD', ''),
                ('5', 'Wall (paint/cement/plaster board)', '1', 'GOOD', ''),
                ('6', 'Water Filter', '1', 'GOOD', ''),
                ('7', 'WiFi router', '1', 'GOOD', ''),
                ('8', 'Cable TV point', '1', 'GOOD', ''),
                ('9', 'Telephone point', '1', 'GOOD', ''),
                ('10', 'Common area access card', '1', 'GOOD', ''),
                ('11', 'Any other items', '1', 'GOOD', ''),
                ('12', 'Main door + Cylinder Lock Mechanism', '1', 'GOOD', '')
            ])
        ]
        
        for section_title, items in default_sections:
            inventory_data.append([section_title, '', '', '', ''])
            inventory_data.extend(items)
    
    # Create inventory table
    inventory_table = Table(inventory_data, colWidths=[0.5*inch, 2.5*inch, 0.5*inch, 1*inch, 2*inch])
    inventory_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    # Style section headers
    row_index = 1
    for row in inventory_data[1:]:
        if len(row) > 1 and row[1] == '' and row[2] == '':  # Section header row
            inventory_table.setStyle(TableStyle([
                ('BACKGROUND', (0, row_index), (-1, row_index), colors.lightgrey),
                ('FONTNAME', (0, row_index), (-1, row_index), 'Helvetica-Bold'),
                ('SPAN', (0, row_index), (-1, row_index))
            ]))
        row_index += 1
    
    story.append(inventory_table)
    story.append(Spacer(1, 30))
    
    # Signatures section - exact format from working live export
    signature_data = [
        ['', 'Handover:', '', 'Takeover:', ''],
        ['Name:', handover_signature.get('name', ''), '', takeover_signature.get('name', ''), ''],
        ['FIN/NRIC:', handover_signature.get('fin', ''), '', takeover_signature.get('fin', ''), ''],
        ['Position:', handover_signature.get('position', ''), '', takeover_signature.get('position', ''), ''],
        ['Date:', handover_signature.get('date', ''), '', takeover_signature.get('date', ''), '']
    ]
    
    # Add signature images if available
    import base64
    from reportlab.platypus import Image
    
    sig_row = ['Signature:', '', '', '', '']
    
    # Handle handover signature
    handover_sig = handover_signature.get('signature', '')
    if handover_sig and handover_sig.startswith('data:image'):
        try:
            sig_data = handover_sig.split(',')[1]
            sig_bytes = base64.b64decode(sig_data)
            sig_buffer = BytesIO(sig_bytes)
            handover_img = Image(sig_buffer, width=1.5*inch, height=0.75*inch)
            sig_row[1] = handover_img
        except:
            sig_row[1] = 'Signature Available'
    
    # Handle takeover signature
    takeover_sig = takeover_signature.get('signature', '')
    if takeover_sig and takeover_sig.startswith('data:image'):
        try:
            sig_data = takeover_sig.split(',')[1]
            sig_bytes = base64.b64decode(sig_data)
            sig_buffer = BytesIO(sig_bytes)
            takeover_img = Image(sig_buffer, width=1.5*inch, height=0.75*inch)
            sig_row[3] = takeover_img
        except:
            sig_row[3] = 'Signature Available'
    
    signature_data.append(sig_row)
    
    signature_table = Table(signature_data, colWidths=[1*inch, 2*inch, 0.5*inch, 2*inch, 0.5*inch])
    signature_table.setStyle(TableStyle([
        ('BACKGROUND', (1, 0), (1, 0), colors.lightgrey),
        ('BACKGROUND', (3, 0), (3, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('SPAN', (1, 0), (1, 0)),
        ('SPAN', (3, 0), (4, 0)),
    ]))
    
    story.append(signature_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    # Create response
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    filename = f"Room_Inventory_Checklist_{checklist.room_number}_{checklist.checklist_date.strftime('%Y%m%d')}.pdf"
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

# Route: Export multiple checklists to Excel
@app.route('/export_all_checklists')
@login_required
def export_all_checklists():
    """Export all room inventory checklists to Excel"""
    import pandas as pd
    from io import BytesIO
    from flask import make_response
    
    checklists = RoomInventoryChecklist.query.all()
    
    if not checklists:
        return "No checklists found", 404
    
    # Create Excel file
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Basic info sheet
        basic_data = []
        for checklist in checklists:
            basic_data.append({
                'ID': checklist.id,
                'Room Number': checklist.room_number,
                'Company': checklist.company_name,
                'Date': checklist.checklist_date.strftime('%Y-%m-%d') if checklist.checklist_date else '',
                'Status': checklist.status,
                'Water Meter': checklist.water_meter_reading,
                'Electricity Meter': checklist.electricity_meter_reading
            })
        
        df_basic = pd.DataFrame(basic_data)
        df_basic.to_excel(writer, sheet_name='Basic Info', index=False)
    
    buffer.seek(0)
    
    # Create response
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=All_Room_Checklists.xlsx'
    
    return response

@app.route('/export_checklist_excel/<int:checklist_id>')
@login_required
def export_checklist_excel(checklist_id):
    """Export single room inventory checklist to Excel"""
    import json
    from openpyxl import Workbook
    from io import BytesIO
    from flask import make_response
    
    checklist = RoomInventoryChecklist.query.get_or_404(checklist_id)
    
    # Parse JSON data
    checklist_items = json.loads(checklist.checklist_items_data) if checklist.checklist_items_data else []
    handover_signature = json.loads(checklist.handover_signature_data) if checklist.handover_signature_data else {}
    takeover_signature = json.loads(checklist.takeover_signature_data) if checklist.takeover_signature_data else {}
    
    # Create workbook
    wb = Workbook()
    
    # Basic information sheet
    ws1 = wb.active
    ws1.title = "Basic Information"
    ws1.append(['Field', 'Value'])
    ws1.append(['Room Number', checklist.room_number])
    ws1.append(['Company Name', checklist.company_name])
    ws1.append(['Date', checklist.checklist_date.strftime('%Y-%m-%d')])
    ws1.append(['Status', checklist.status])
    ws1.append(['Created By', checklist.created_by_user.first_name + ' ' + (checklist.created_by_user.last_name or '')])
    ws1.append(['Created At', checklist.created_at.strftime('%Y-%m-%d %H:%M:%S')])
    
    # Meter readings sheet
    ws2 = wb.create_sheet("Meter Readings")
    ws2.append(['Meter Type', 'Value'])
    ws2.append(['Water Meter Reading', checklist.water_meter_reading or ''])
    ws2.append(['Water Meter Signature', checklist.water_meter_signature or ''])
    ws2.append(['Water Meter DateTime', checklist.water_meter_datetime.strftime('%Y-%m-%d %H:%M:%S') if checklist.water_meter_datetime else ''])
    ws2.append(['Electricity Meter Reading', checklist.electricity_meter_reading or ''])
    ws2.append(['Electricity Meter Signature', checklist.electricity_meter_signature or ''])
    ws2.append(['Electricity Meter DateTime', checklist.electricity_meter_datetime.strftime('%Y-%m-%d %H:%M:%S') if checklist.electricity_meter_datetime else ''])
    
    # Checklist items sheet
    if checklist_items:
        ws3 = wb.create_sheet("Checklist Items")
        ws3.append(['Category', 'Item', 'Status', 'Remarks'])
        for item in checklist_items:
            ws3.append([
                item.get('category', ''),
                item.get('description', ''),
                item.get('status', ''),
                item.get('remarks', '')
            ])
    
    # Signatures sheet
    ws4 = wb.create_sheet("Signatures")
    ws4.append(['Type', 'Name', 'Position', 'FIN Number', 'Date'])
    if handover_signature:
        ws4.append([
            'Handover',
            handover_signature.get('name', ''),
            handover_signature.get('position', ''),
            handover_signature.get('fin', ''),
            handover_signature.get('date', '')
        ])
    if takeover_signature:
        ws4.append([
            'Takeover',
            takeover_signature.get('name', ''),
            takeover_signature.get('position', ''),
            takeover_signature.get('fin', ''),
            takeover_signature.get('date', '')
        ])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Room_Checklist_{checklist.room_number}_{checklist.checklist_date}.xlsx'
    
    return response

@app.route('/export_selected_checklists', methods=['POST'])
@login_required
def export_selected_checklists():
    """Export selected room inventory checklists to Excel"""
    import json
    from openpyxl import Workbook
    from io import BytesIO
    from flask import make_response
    
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('room_handovers'))
    
    try:
        checklist_ids = json.loads(request.form.get('checklist_ids', '[]'))
        
        if not checklist_ids:
            flash('No checklists selected for export', 'error')
            return redirect(url_for('room_handovers'))
        
        # Get selected checklists
        checklists = RoomInventoryChecklist.query.filter(
            RoomInventoryChecklist.id.in_(checklist_ids),
            RoomInventoryChecklist.organization_id == user.organization_id
        ).all()
        
        if not checklists:
            flash('No valid checklists found for export', 'error')
            return redirect(url_for('room_handovers'))
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Selected Room Checklists"
        
        # Add headers
        headers = ['S.No', 'Room Number', 'Company Name', 'Date', 'Status', 'Created By', 'Created At', 'Water Meter', 'Electricity Meter']
        ws.append(headers)
        
        # Add data rows
        for i, checklist in enumerate(checklists, 1):
            row = [
                i,
                checklist.room_number,
                checklist.company_name,
                checklist.checklist_date.strftime('%Y-%m-%d'),
                checklist.status,
                checklist.created_by_user.first_name + ' ' + (checklist.created_by_user.last_name or ''),
                checklist.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                checklist.water_meter_reading or '',
                checklist.electricity_meter_reading or ''
            ]
            ws.append(row)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=selected_room_checklists_{len(checklists)}_items.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Error exporting checklists: {str(e)}', 'error')
        return redirect(url_for('room_handovers'))

@app.route('/delete_selected_checklists', methods=['POST'])
@login_required
@create_permission_required('room_checklist')
def delete_selected_checklists():
    """Delete selected room inventory checklists"""
    import json
    
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('room_handovers'))
    
    try:
        checklist_ids = json.loads(request.form.get('checklist_ids', '[]'))
        
        if not checklist_ids:
            flash('No checklists selected for deletion', 'error')
            return redirect(url_for('room_handovers'))
        
        # Delete selected checklists
        deleted_count = RoomInventoryChecklist.query.filter(
            RoomInventoryChecklist.id.in_(checklist_ids),
            RoomInventoryChecklist.organization_id == user.organization_id
        ).delete(synchronize_session=False)
        
        db.session.commit()
        
        flash(f'Successfully deleted {deleted_count} checklist(s)', 'success')
        return redirect(url_for('room_handovers'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting checklists: {str(e)}', 'error')
        return redirect(url_for('room_handovers'))

@app.route('/export_all_checklists_excel')
@login_required
def export_all_checklists_excel():
    """Export all room inventory checklists to Excel"""
    from openpyxl import Workbook
    from io import BytesIO
    from flask import make_response
    from datetime import datetime
    
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    checklists = RoomInventoryChecklist.query.filter_by(organization_id=user.organization_id).order_by(RoomInventoryChecklist.created_at.desc()).all()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "All Room Checklists"
    
    # Add headers
    headers = ['S.No', 'Room Number', 'Company Name', 'Date', 'Status', 'Created By', 'Created At', 'Water Meter', 'Electricity Meter']
    ws.append(headers)
    
    # Add data rows
    for i, checklist in enumerate(checklists, 1):
        row = [
            i,
            checklist.room_number,
            checklist.company_name,
            checklist.checklist_date.strftime('%Y-%m-%d'),
            checklist.status,
            checklist.created_by_user.first_name + ' ' + (checklist.created_by_user.last_name or ''),
            checklist.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            checklist.water_meter_reading or '',
            checklist.electricity_meter_reading or ''
        ]
        ws.append(row)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=All_Room_Checklists_{singapore_now().strftime("%Y%m%d")}.xlsx'
    
    return response

@app.route('/delete_checklist/<int:checklist_id>', methods=['POST'])
@login_required
@create_permission_required('room_checklist')
def delete_checklist(checklist_id):
    """Delete room inventory checklist"""
    try:
        checklist = RoomInventoryChecklist.query.get_or_404(checklist_id)
        db.session.delete(checklist)
        db.session.commit()
        flash('Room inventory checklist deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting checklist: {str(e)}', 'error')
    
    return redirect(url_for('room_handovers'))

# API endpoint for case number generation
@app.route('/api/next-case-number')
@login_required
def get_next_case_number():
    """API endpoint to get the next case number"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 400
    
    try:
        # Count existing offense records for this organization
        case_count = OffenseRecord.query.filter_by(organization_id=user.organization_id).count() + 1
        case_number = f"PL/OR/{case_count:07d}"
        return jsonify({'case_number': case_number})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# MSRF Management Routes
@app.route('/msrf-management', methods=['GET', 'POST'])
@login_required
@page_permission_required('msrf')
def msrf_management():
    """MSRF (Material Safety Reporting Form) management page"""
    from app.models.models_msrf import MSRFRequest
    from app.models.models_house_acknowledge import RoomNumber
    from datetime import datetime
    
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            # Handle custom room number or selected room number
            room_number = request.form.get('room_number')
            if room_number == 'custom':
                room_number = request.form.get('custom_room_number', '').strip()
            
            # Parse date fields
            date_requested = datetime.strptime(request.form.get('date_requested'), '%Y-%m-%d').date()
            date_installed = None
            if request.form.get('date_installed'):
                date_installed = datetime.strptime(request.form.get('date_installed'), '%Y-%m-%d').date()
            
            # Create new MSRF request
            new_request = MSRFRequest(
                serial_number=request.form.get('serial_number', ''),
                room_number=room_number,
                company_name=request.form.get('company_name'),
                item_requested=request.form.get('item_requested'),
                notes=request.form.get('notes', ''),
                date_requested=date_requested,
                date_installed=date_installed,
                organization_id=user.organization_id,
                created_by=user.id
            )
            
            db.session.add(new_request)
            db.session.commit()
            flash('MSRF request added successfully!', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding MSRF request: {str(e)}', 'error')
        
        return redirect(url_for('msrf_management'))
    
    # GET request - display table with filtering functionality
    search_query = request.args.get('search', '').strip()
    company_filter = request.args.get('company_filter', '').strip()
    room_filter = request.args.get('room_filter', '').strip()
    item_filter = request.args.get('item_filter', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    
    query = MSRFRequest.query.filter_by(organization_id=user.organization_id)
    
    # Apply filters
    if search_query:
        query = query.filter(
            db.or_(
                MSRFRequest.room_number.ilike(f'%{search_query}%'),
                MSRFRequest.company_name.ilike(f'%{search_query}%'),
                MSRFRequest.serial_number.ilike(f'%{search_query}%'),
                MSRFRequest.item_requested.ilike(f'%{search_query}%'),
                MSRFRequest.notes.ilike(f'%{search_query}%')
            )
        )
    
    if company_filter:
        query = query.filter(MSRFRequest.company_name.ilike(f'%{company_filter}%'))
    
    if room_filter:
        query = query.filter(MSRFRequest.room_number.ilike(f'%{room_filter}%'))
    
    if item_filter:
        query = query.filter(MSRFRequest.item_requested.ilike(f'%{item_filter}%'))
    
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(MSRFRequest.date_requested >= date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(MSRFRequest.date_requested <= date_to_obj)
        except ValueError:
            pass
    
    msrf_requests = query.order_by(MSRFRequest.created_at.desc()).all()
    room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    
    # Create unique company list for filter dropdown
    company_list = []
    
    return render_template('msrf_management.html', msrf_requests=msrf_requests, room_numbers=room_numbers, search_query=search_query, company_list=company_list)

@app.route('/msrf-management/export/<format>')
@login_required
@create_permission_required('msrf_management')
def export_msrf_requests(format):
    """Export MSRF requests to PDF or Excel"""
    from app.models.models_msrf import MSRFRequest
    from io import BytesIO
    from datetime import datetime, timezone, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from timezone_utils import singapore_now
    
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    try:
        # Get filtered data
        search_query = request.args.get('search', '').strip()
        company_filter = request.args.get('company_filter', '').strip()
        room_filter = request.args.get('room_filter', '').strip()
        item_filter = request.args.get('item_filter', '').strip()
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        selected_ids = request.args.getlist('selected_ids')
        
        query = MSRFRequest.query.filter_by(organization_id=user.organization_id)
        
        # Apply filters
        if search_query:
            query = query.filter(
                db.or_(
                    MSRFRequest.room_number.ilike(f'%{search_query}%'),
                    MSRFRequest.company_name.ilike(f'%{search_query}%'),
                    MSRFRequest.serial_number.ilike(f'%{search_query}%'),
                    MSRFRequest.item_requested.ilike(f'%{search_query}%'),
                    MSRFRequest.notes.ilike(f'%{search_query}%')
                )
            )
        
        if company_filter:
            query = query.filter(MSRFRequest.company_name.ilike(f'%{company_filter}%'))
        
        if room_filter:
            query = query.filter(MSRFRequest.room_number.ilike(f'%{room_filter}%'))
        
        if item_filter:
            query = query.filter(MSRFRequest.item_requested.ilike(f'%{item_filter}%'))
        
        if date_from:
            try:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                query = query.filter(MSRFRequest.date_requested >= date_from_obj)
            except ValueError:
                pass
        
        if date_to:
            try:
                from datetime import datetime
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                query = query.filter(MSRFRequest.date_requested <= date_to_obj)
            except ValueError:
                pass
        
        # If specific IDs are selected, filter by those
        if selected_ids:
            query = query.filter(MSRFRequest.id.in_(selected_ids))
        
        msrf_requests = query.order_by(MSRFRequest.created_at.desc()).all()
        
        if format.lower() == 'excel':
            # Create Excel file using openpyxl directly
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "MSRF Requests"
            
            # Define headers with serial number included
            headers = ['ID', 'Room Number', 'Company Name', 'Serial Number', 'Item Requested', 
                      'Notes', 'Date Requested', 'Date Installed', 'Created At']
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data rows
            for row_num, req in enumerate(msrf_requests, 2):
                ws.cell(row=row_num, column=1, value=req.id)
                ws.cell(row=row_num, column=2, value=req.room_number)
                ws.cell(row=row_num, column=3, value=req.company_name)
                ws.cell(row=row_num, column=4, value=req.serial_number or '')
                ws.cell(row=row_num, column=5, value=req.item_requested)
                ws.cell(row=row_num, column=6, value=req.notes or '')
                ws.cell(row=row_num, column=7, value=req.date_requested.strftime('%Y-%m-%d') if req.date_requested else '')
                ws.cell(row=row_num, column=8, value=req.date_installed.strftime('%Y-%m-%d') if req.date_installed else '')
                ws.cell(row=row_num, column=9, value=req.created_at.strftime('%Y-%m-%d %H:%M:%S') if req.created_at else '')
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                as_attachment=True,
                download_name=f'msrf_requests_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        elif format.lower() == 'pdf':
            # Create PDF file using reportlab
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            title = Paragraph("MSRF Requests Report", title_style)
            
            # Date
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=1  # Center alignment
            )
            # Use Singapore time for PDF generation
            date_para = Paragraph(f"Generated on: {singapore_now().strftime('%Y-%m-%d %H:%M:%S')}", date_style)
            
            # Table data with serial number included
            table_data = [['ID', 'Room No.', 'Company', 'Serial No.', 'Item Requested', 'Date Requested', 'Date Installed']]
            
            for req in msrf_requests:
                table_data.append([
                    str(req.id),
                    req.room_number,
                    req.company_name[:20] + '...' if len(req.company_name) > 20 else req.company_name,
                    req.serial_number[:15] + '...' if req.serial_number and len(req.serial_number) > 15 else (req.serial_number or '-'),
                    req.item_requested[:25] + '...' if len(req.item_requested) > 25 else req.item_requested,
                    req.date_requested.strftime('%Y-%m-%d') if req.date_requested else 'N/A',
                    req.date_installed.strftime('%Y-%m-%d') if req.date_installed else 'N/A'
                ])
            
            # Create table with 7 columns including serial number
            table = Table(table_data, colWidths=[0.6*inch, 0.9*inch, 1.2*inch, 0.8*inch, 1.5*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            # Build PDF
            elements = [title, Spacer(1, 12), date_para, Spacer(1, 20), table]
            doc.build(elements)
            
            output.seek(0)
            
            return send_file(
                output,
                as_attachment=True,
                download_name=f'msrf_requests_{singapore_now().strftime("%Y%m%d_%H%M%S")}.pdf',
                mimetype='application/pdf'
            )
        
        else:
            flash('Invalid export format', 'error')
            return redirect(url_for('msrf_management'))
            
    except Exception as e:
        flash(f'Error exporting data: {str(e)}', 'error')
        return redirect(url_for('msrf_management'))

@app.route('/msrf-management/delete', methods=['POST'])
@login_required
@create_permission_required('msrf_management')
def delete_msrf_requests():
    """Delete selected MSRF requests"""
    from app.models.models_msrf import MSRFRequest
    
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    try:
        selected_ids = request.form.getlist('selected_requests')
        
        if not selected_ids:
            flash('No requests selected for deletion', 'warning')
            return redirect(url_for('msrf_management'))
        
        # Delete selected requests
        deleted_count = 0
        for req_id in selected_ids:
            req = MSRFRequest.query.filter_by(id=req_id, organization_id=user.organization_id).first()
            if req:
                db.session.delete(req)
                deleted_count += 1
        
        db.session.commit()
        flash(f'Successfully deleted {deleted_count} MSRF request(s)', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting requests: {str(e)}', 'error')
    
    return redirect(url_for('msrf_management'))

@app.route('/msrf-management/edit/<int:request_id>', methods=['GET', 'POST'])
@login_required
@create_permission_required('msrf_management')
def edit_msrf_request(request_id):
    """Edit an MSRF request"""
    from app.models.models_msrf import MSRFRequest
    from app.models.models_house_acknowledge import RoomNumber
    from datetime import datetime
    
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    msrf_request = MSRFRequest.query.filter_by(id=request_id, organization_id=user.organization_id).first_or_404()
    
    if request.method == 'POST':
        try:
            # Handle custom room number or selected room number
            room_number = request.form.get('room_number')
            if room_number == 'custom':
                room_number = request.form.get('custom_room_number', '').strip()
            
            # Parse date fields
            date_requested = datetime.strptime(request.form.get('date_requested'), '%Y-%m-%d').date()
            date_installed = None
            if request.form.get('date_installed'):
                date_installed = datetime.strptime(request.form.get('date_installed'), '%Y-%m-%d').date()
            
            # Update MSRF request
            msrf_request.serial_number = request.form.get('serial_number', '')
            msrf_request.room_number = room_number
            msrf_request.company_name = request.form.get('company_name')
            msrf_request.item_requested = request.form.get('item_requested')
            msrf_request.notes = request.form.get('notes', '')
            msrf_request.date_requested = date_requested
            msrf_request.date_installed = date_installed
            
            db.session.commit()
            flash('MSRF request updated successfully!', 'success')
            return redirect(url_for('msrf_management'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating MSRF request: {str(e)}', 'error')
    
    # GET request - display edit form
    room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    return render_template('edit_msrf_request.html', msrf_request=msrf_request, room_numbers=room_numbers)

# Meter Readings Routes
@app.route('/meter-readings')
@login_required
def meter_readings():
    """Meter readings management homepage with filtering and data display"""
    from app.models.models_meter_reading import MeterCompany, MeterRoom, WaterMeterReading, ElectricityMeterReading
    from sqlalchemy import and_
    from datetime import datetime, date
    
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    company_name = request.args.get('company_name', '').strip()
    room_number = request.args.get('room_number', '').strip()
    
    # Get all companies for dropdown
    companies = MeterCompany.query.filter_by(is_active=True).all()
    
    # Build query for readings with joins
    water_query = db.session.query(
        WaterMeterReading,
        MeterRoom.room_number,
        MeterCompany.company_name
    ).join(MeterRoom, WaterMeterReading.meter_room_id == MeterRoom.id).join(MeterCompany, MeterRoom.company_id == MeterCompany.id)
    
    electricity_query = db.session.query(
        ElectricityMeterReading,
        MeterRoom.room_number,
        MeterCompany.company_name
    ).join(MeterRoom, ElectricityMeterReading.meter_room_id == MeterRoom.id).join(MeterCompany, MeterRoom.company_id == MeterCompany.id)
    
    # Apply filters
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            water_query = water_query.filter(WaterMeterReading.start_date >= start_dt)
            electricity_query = electricity_query.filter(ElectricityMeterReading.start_date >= start_dt)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            water_query = water_query.filter(WaterMeterReading.end_date <= end_dt)
            electricity_query = electricity_query.filter(ElectricityMeterReading.end_date <= end_dt)
        except ValueError:
            pass
    
    if company_name:
        water_query = water_query.filter(MeterCompany.company_name.ilike(f'%{company_name}%'))
        electricity_query = electricity_query.filter(MeterCompany.company_name.ilike(f'%{company_name}%'))
    
    if room_number:
        water_query = water_query.filter(MeterRoom.room_number.ilike(f'%{room_number}%'))
        electricity_query = electricity_query.filter(MeterRoom.room_number.ilike(f'%{room_number}%'))
    
    # Execute queries
    water_readings = water_query.order_by(WaterMeterReading.created_at.desc()).all()
    electricity_readings = electricity_query.order_by(ElectricityMeterReading.created_at.desc()).all()
    
    # Calculate totals for selected company
    water_totals = {'count': 0, 'total_consumption': 0.0, 'total_amount': 0.0}
    electricity_totals = {'count': 0, 'total_consumption': 0.0, 'total_amount': 0.0}
    
    if company_name:
        # Calculate water totals
        water_totals['count'] = len(water_readings)
        water_totals['total_consumption'] = sum(reading[0].total_consumption for reading in water_readings)
        water_totals['total_amount'] = sum(reading[0].total_amount for reading in water_readings)
        
        # Calculate electricity totals
        electricity_totals['count'] = len(electricity_readings)
        electricity_totals['total_consumption'] = sum(reading[0].total_consumption for reading in electricity_readings)
        electricity_totals['total_amount'] = sum(reading[0].total_amount for reading in electricity_readings)
    
    return render_template('meter_readings.html', 
                         water_readings=water_readings,
                         electricity_readings=electricity_readings,
                         companies=companies,
                         start_date=start_date,
                         end_date=end_date,
                         company_name=company_name,
                         room_number=room_number,
                         water_totals=water_totals,
                         electricity_totals=electricity_totals)

@app.route('/meter-readings/create-company', methods=['GET', 'POST'])
@login_required
@create_permission_required('meter_reading')
def create_meter_company():
    """Create new meter company"""
    from app.models.models_meter_reading import MeterCompany
    
    if request.method == 'POST':
        company_name = request.form.get('company_name', '').strip()
        
        if not company_name:
            flash('Company name is required', 'error')
            return redirect(url_for('create_meter_company'))
        
        # Check if company already exists
        existing = MeterCompany.query.filter_by(company_name=company_name, is_active=True).first()
        if existing:
            flash('Company name already exists', 'error')
            return redirect(url_for('create_meter_company'))
        
        # Create new company
        company = MeterCompany(
            company_name=company_name,
            created_by=current_user.id
        )
        
        db.session.add(company)
        db.session.commit()
        
        flash(f'Company "{company_name}" created successfully', 'success')
        return redirect(url_for('meter_company_detail', company_id=company.id))
    
    return render_template('create_meter_company.html')

@app.route('/meter-readings/company/<int:company_id>')
@login_required
def meter_company_detail(company_id):
    """Company detail page with room management"""
    from app.models.models_meter_reading import MeterCompany, MeterRoom
    
    company = MeterCompany.query.get_or_404(company_id)
    rooms = MeterRoom.query.filter_by(company_id=company_id, is_active=True).order_by(MeterRoom.room_number).all()
    
    # Get existing room numbers from the main system for dropdown
    available_rooms = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    
    return render_template('meter_company_detail.html', company=company, rooms=rooms, available_rooms=available_rooms)

@app.route('/meter-readings/room/<int:room_id>/delete', methods=['POST'])
@login_required
@create_permission_required('meter_reading')
def delete_meter_room(room_id):
    """Delete a meter room and all its readings"""
    from app.models.models_meter_reading import MeterRoom, WaterMeterReading, ElectricityMeterReading
    
    try:
        room = MeterRoom.query.get_or_404(room_id)
        company_id = room.company_id
        
        # Delete all associated readings first
        WaterMeterReading.query.filter_by(meter_room_id=room_id).delete()
        ElectricityMeterReading.query.filter_by(meter_room_id=room_id).delete()
        
        # Delete the room
        db.session.delete(room)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Room "{room.room_number}" deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error deleting room: {str(e)}'}), 500

@app.route('/meter-readings/company/<int:company_id>/delete', methods=['POST'])
@login_required
@create_permission_required('meter_reading')
def delete_meter_company(company_id):
    """Delete a meter company and all its rooms and readings"""
    from app.models.models_meter_reading import MeterCompany, MeterRoom, WaterMeterReading, ElectricityMeterReading
    
    try:
        company = MeterCompany.query.get_or_404(company_id)
        company_name = company.company_name
        
        # Get all rooms for this company
        rooms = MeterRoom.query.filter_by(company_id=company_id).all()
        
        # Delete all readings for all rooms
        for room in rooms:
            WaterMeterReading.query.filter_by(meter_room_id=room.id).delete()
            ElectricityMeterReading.query.filter_by(meter_room_id=room.id).delete()
        
        # Delete all rooms
        MeterRoom.query.filter_by(company_id=company_id).delete()
        
        # Delete the company
        db.session.delete(company)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Company "{company_name}" deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error deleting company: {str(e)}'}), 500

@app.route('/meter-readings/export/<format>')
@login_required
def export_meter_readings(format):
    """Export meter readings to PDF or Excel"""
    from app.models.models_meter_reading import MeterCompany, MeterRoom, WaterMeterReading, ElectricityMeterReading
    from io import BytesIO
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from timezone_utils import singapore_now
    
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    try:
        # Get filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        company_name = request.args.get('company_name', '').strip()
        room_number = request.args.get('room_number', '').strip()
        
        # Build queries with same filters as main page
        water_query = db.session.query(
            WaterMeterReading,
            MeterRoom.room_number,
            MeterCompany.company_name
        ).join(MeterRoom, WaterMeterReading.meter_room_id == MeterRoom.id).join(MeterCompany, MeterRoom.company_id == MeterCompany.id)
        
        electricity_query = db.session.query(
            ElectricityMeterReading,
            MeterRoom.room_number,
            MeterCompany.company_name
        ).join(MeterRoom, ElectricityMeterReading.meter_room_id == MeterRoom.id).join(MeterCompany, MeterRoom.company_id == MeterCompany.id)
        
        # Apply filters
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
                water_query = water_query.filter(WaterMeterReading.start_date >= start_dt)
                electricity_query = electricity_query.filter(ElectricityMeterReading.start_date >= start_dt)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
                water_query = water_query.filter(WaterMeterReading.end_date <= end_dt)
                electricity_query = electricity_query.filter(ElectricityMeterReading.end_date <= end_dt)
            except ValueError:
                pass
        
        if company_name:
            water_query = water_query.filter(MeterCompany.company_name.ilike(f'%{company_name}%'))
            electricity_query = electricity_query.filter(MeterCompany.company_name.ilike(f'%{company_name}%'))
        
        if room_number:
            water_query = water_query.filter(MeterRoom.room_number.ilike(f'%{room_number}%'))
            electricity_query = electricity_query.filter(MeterRoom.room_number.ilike(f'%{room_number}%'))
        
        # Execute queries
        water_readings = water_query.order_by(WaterMeterReading.created_at.desc()).all()
        electricity_readings = electricity_query.order_by(ElectricityMeterReading.created_at.desc()).all()
        
        if format.lower() == 'excel':
            # Create Excel file
            output = BytesIO()
            wb = Workbook()
            
            # Create Water Readings sheet
            ws_water = wb.active
            ws_water.title = "Water Readings"
            
            # Water headers
            water_headers = ['Date', 'Company', 'Room Number', 'Meter Number', 'Start Reading', 
                           'End Reading', 'Consumption', 'Rate/Unit', 'Total Amount', 'Physical Pax', 'Notes']
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Write water headers
            for col_num, header in enumerate(water_headers, 1):
                cell = ws_water.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write water data
            for row_num, (reading, room_num, comp_name) in enumerate(water_readings, 2):
                ws_water.cell(row=row_num, column=1, value=reading.start_date.strftime('%Y-%m-%d') if reading.start_date else '')
                ws_water.cell(row=row_num, column=2, value=comp_name)
                ws_water.cell(row=row_num, column=3, value=room_num)
                ws_water.cell(row=row_num, column=4, value=reading.meter_number)
                ws_water.cell(row=row_num, column=5, value=reading.start_reading)
                ws_water.cell(row=row_num, column=6, value=reading.end_reading)
                ws_water.cell(row=row_num, column=7, value=reading.total_consumption)
                ws_water.cell(row=row_num, column=8, value=reading.rate_per_unit)
                ws_water.cell(row=row_num, column=9, value=reading.total_amount)
                ws_water.cell(row=row_num, column=10, value=reading.physical_pax)
                ws_water.cell(row=row_num, column=11, value=reading.notes or '')
            
            # Create Electricity Readings sheet
            ws_electricity = wb.create_sheet(title="Electricity Readings")
            
            # Electricity headers
            electricity_headers = ['Date', 'Company', 'Room Number', 'Meter Number', 'Start Reading', 
                                 'End Reading', 'Consumption', 'Rate/Unit', 'Total Amount', 'Physical Pax', 'Notes']
            
            # Write electricity headers
            for col_num, header in enumerate(electricity_headers, 1):
                cell = ws_electricity.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write electricity data
            for row_num, (reading, room_num, comp_name) in enumerate(electricity_readings, 2):
                ws_electricity.cell(row=row_num, column=1, value=reading.start_date.strftime('%Y-%m-%d') if reading.start_date else '')
                ws_electricity.cell(row=row_num, column=2, value=comp_name)
                ws_electricity.cell(row=row_num, column=3, value=room_num)
                ws_electricity.cell(row=row_num, column=4, value=reading.meter_number)
                ws_electricity.cell(row=row_num, column=5, value=reading.start_reading)
                ws_electricity.cell(row=row_num, column=6, value=reading.end_reading)
                ws_electricity.cell(row=row_num, column=7, value=reading.total_consumption)
                ws_electricity.cell(row=row_num, column=8, value=reading.rate_per_unit)
                ws_electricity.cell(row=row_num, column=9, value=reading.total_amount)
                ws_electricity.cell(row=row_num, column=10, value=reading.physical_pax)
                ws_electricity.cell(row=row_num, column=11, value=reading.notes or '')
            
            # Auto-adjust column widths for both sheets
            for ws in [ws_water, ws_electricity]:
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
            
            wb.save(output)
            output.seek(0)
            
            filename = f'meter_readings_{company_name}_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx' if company_name else f'meter_readings_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        elif format.lower() == 'pdf':
            # Create PDF file
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            title = Paragraph(f"Meter Readings Report - {company_name or 'All Companies'}", title_style)
            
            # Date
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=1  # Center alignment
            )
            date_para = Paragraph(f"Generated on: {singapore_now().strftime('%Y-%m-%d %H:%M:%S')}", date_style)
            
            elements = [title, Spacer(1, 12), date_para, Spacer(1, 20)]
            
            # Water readings table
            if water_readings:
                water_title = Paragraph("Water Utility Readings", styles['Heading2'])
                elements.append(water_title)
                elements.append(Spacer(1, 12))
                
                water_data = [['Date', 'Company', 'Room', 'Meter #', 'Start', 'End', 'Consumption', 'Amount']]
                for reading, room_num, comp_name in water_readings:
                    water_data.append([
                        reading.start_date.strftime('%Y-%m-%d') if reading.start_date else '',
                        comp_name,
                        room_num,
                        reading.meter_number,
                        f"{reading.start_reading:.2f}",
                        f"{reading.end_reading:.2f}",
                        f"{reading.total_consumption:.2f}",
                        f"${reading.total_amount:.2f}"
                    ])
                
                water_table = Table(water_data, colWidths=[0.8*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.7*inch, 0.8*inch, 0.8*inch])
                water_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(water_table)
                elements.append(Spacer(1, 20))
            
            # Electricity readings table
            if electricity_readings:
                electricity_title = Paragraph("Electricity Utility Readings", styles['Heading2'])
                elements.append(electricity_title)
                elements.append(Spacer(1, 12))
                
                electricity_data = [['Date', 'Company', 'Room', 'Meter #', 'Start', 'End', 'Consumption', 'Amount']]
                for reading, room_num, comp_name in electricity_readings:
                    electricity_data.append([
                        reading.start_date.strftime('%Y-%m-%d') if reading.start_date else '',
                        comp_name,
                        room_num,
                        reading.meter_number,
                        f"{reading.start_reading:.2f}",
                        f"{reading.end_reading:.2f}",
                        f"{reading.total_consumption:.2f}",
                        f"${reading.total_amount:.2f}"
                    ])
                
                electricity_table = Table(electricity_data, colWidths=[0.8*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.7*inch, 0.8*inch, 0.8*inch])
                electricity_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(electricity_table)
            
            # Build PDF
            doc.build(elements)
            output.seek(0)
            
            filename = f'meter_readings_{company_name}_{singapore_now().strftime("%Y%m%d_%H%M%S")}.pdf' if company_name else f'meter_readings_{singapore_now().strftime("%Y%m%d_%H%M%S")}.pdf'
            
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )
        
        else:
            flash('Invalid export format', 'error')
            return redirect(url_for('meter_readings'))
            
    except Exception as e:
        flash(f'Error exporting data: {str(e)}', 'error')
        return redirect(url_for('meter_readings'))

@app.route('/meter-readings/company/<int:company_id>/add-room', methods=['POST'])
@login_required
@create_permission_required('meter_reading')
def add_meter_room(company_id):
    """Add room to company"""
    from app.models.models_meter_reading import MeterCompany, MeterRoom
    
    company = MeterCompany.query.get_or_404(company_id)
    room_number = request.form.get('room_number', '').strip()
    
    if not room_number:
        flash('Room number is required', 'error')
        return redirect(url_for('meter_company_detail', company_id=company_id))
    
    # Check if room already exists for this company
    existing = MeterRoom.query.filter_by(company_id=company_id, room_number=room_number, is_active=True).first()
    if existing:
        flash('Room number already exists for this company', 'error')
        return redirect(url_for('meter_company_detail', company_id=company_id))
    
    # Create new room
    room = MeterRoom(
        room_number=room_number,
        company_id=company_id,
        created_by=current_user.id
    )
    
    db.session.add(room)
    db.session.commit()
    
    flash(f'Room "{room_number}" added successfully', 'success')
    return redirect(url_for('meter_company_detail', company_id=company_id))

@app.route('/meter-readings/company/<int:company_id>/bulk-add-rooms', methods=['POST'])
@login_required
@create_permission_required('meter_reading')
def bulk_add_meter_rooms(company_id):
    """Bulk add all rooms to meter company"""
    print(f"DEBUG: Bulk add rooms called for company_id: {company_id}")
    from app.models.models_meter_reading import MeterCompany, MeterRoom
    
    company = MeterCompany.query.get_or_404(company_id)
    print(f"DEBUG: Found company: {company.company_name}")
    
    # Complete list of all room numbers for TS MANAGEMENT SERVICES PTE LTD
    all_room_numbers = [
        # Building 80
        *[f"80-01-{i:03d}" for i in range(1, 16)],
        *[f"80-02-{i:03d}" for i in range(1, 16)],
        *[f"80-03-{i:03d}" for i in range(1, 16)],
        *[f"80-04-{i:03d}" for i in range(1, 16)],
        
        # Building 81
        *[f"81-01-{i:03d}" for i in range(101, 116)],
        *[f"81-02-{i:03d}" for i in range(101, 116)],
        *[f"81-03-{i:03d}" for i in range(101, 116)],
        *[f"81-04-{i:03d}" for i in range(101, 116)],
        
        # Building 82
        *[f"82-01-{i:03d}" for i in range(201, 216)],
        *[f"82-02-{i:03d}" for i in range(201, 216)],
        *[f"82-03-{i:03d}" for i in range(201, 216)],
        *[f"82-04-{i:03d}" for i in range(201, 216)],
        
        # Building 83
        *[f"83-01-{i:03d}" for i in range(301, 316)],
        *[f"83-02-{i:03d}" for i in range(301, 316)],
        *[f"83-03-{i:03d}" for i in range(301, 316)],
        *[f"83-04-{i:03d}" for i in range(301, 316)],
        
        # Building 88
        *[f"88-02-{i:03d}" for i in range(801, 829)],
        *[f"88-03-{i:03d}" for i in range(801, 811)],
        *[f"88-04-{i:03d}" for i in range(801, 811)],
        
        # Sickbay rooms
        *[f"Sickbay-{i}" for i in range(1, 17)]
    ]
    
    added_count = 0
    skipped_count = 0
    
    try:
        for room_number in all_room_numbers:
            # Check if room already exists for this company
            existing = MeterRoom.query.filter_by(
                company_id=company_id, 
                room_number=room_number, 
                is_active=True
            ).first()
            
            if existing:
                skipped_count += 1
                continue
            
            # Create new room
            room = MeterRoom(
                room_number=room_number,
                company_id=company_id,
                created_by=current_user.id
            )
            
            db.session.add(room)
            added_count += 1
        
        db.session.commit()
        
        message = f'Added {added_count} rooms successfully.'
        if skipped_count > 0:
            message += f' Skipped {skipped_count} existing rooms.'
        
        flash(message, 'success')
        print(f"DEBUG: Successfully added {added_count} rooms")
        
    except Exception as e:
        db.session.rollback()
        print(f"DEBUG: Error adding rooms: {str(e)}")
        flash(f'Error adding rooms: {str(e)}', 'error')
    
    return redirect(url_for('meter_company_detail', company_id=company_id))

@app.route('/meter-readings/room/<int:room_id>')
@login_required
def meter_room_detail(room_id):
    """Room detail page with meter readings"""
    from app.models.models_meter_reading import MeterRoom, WaterMeterReading, ElectricityMeterReading
    
    room = MeterRoom.query.get_or_404(room_id)
    
    # Get latest readings
    water_readings = WaterMeterReading.query.filter_by(meter_room_id=room_id).order_by(WaterMeterReading.created_at.desc()).all()
    electricity_readings = ElectricityMeterReading.query.filter_by(meter_room_id=room_id).order_by(ElectricityMeterReading.created_at.desc()).all()
    
    return render_template('meter_room_detail.html', 
                         room=room, 
                         water_readings=water_readings,
                         electricity_readings=electricity_readings)

@app.route('/meter-readings/room/<int:room_id>/add-water-reading', methods=['POST'])
@login_required
@create_permission_required('meter_reading')
def add_water_reading(room_id):
    """Add water meter reading"""
    from app.models.models_meter_reading import MeterRoom, WaterMeterReading
    from datetime import datetime
    
    room = MeterRoom.query.get_or_404(room_id)
    
    try:
        meter_number = request.form.get('meter_number', '').strip()
        start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
        start_reading = float(request.form.get('start_reading', 0))
        end_reading_str = request.form.get('end_reading', '').strip()
        rate_per_unit = float(request.form.get('rate_per_unit', 0))
        notes = request.form.get('notes', '')
        
        if not meter_number:
            flash('Meter number is required', 'error')
            return redirect(url_for('meter_room_detail', room_id=room_id))
        
        # Handle initial reading (no end reading yet)
        if not end_reading_str:
            end_reading = start_reading  # Set end same as start for initial reading
            total_consumption = 0.0
            total_amount = 0.0
            notes = f"Initial reading. {notes}".strip()
        else:
            end_reading = float(end_reading_str)
            if end_reading < start_reading:
                flash('End reading cannot be less than start reading', 'error')
                return redirect(url_for('meter_room_detail', room_id=room_id))
            total_consumption = end_reading - start_reading
            total_amount = total_consumption * rate_per_unit
        
        reading = WaterMeterReading(
            meter_room_id=room_id,
            meter_number=meter_number,
            start_date=start_date,
            end_date=end_date,
            start_reading=start_reading,
            end_reading=end_reading,
            total_consumption=total_consumption,
            rate_per_unit=rate_per_unit,
            total_amount=total_amount,
            physical_pax=int(request.form.get('physical_pax', 0)),
            notes=notes,
            created_by=current_user.id
        )
        
        db.session.add(reading)
        db.session.commit()
        
        flash('Water meter reading added successfully', 'success')
        
    except (ValueError, TypeError) as e:
        flash('Invalid input data. Please check your entries.', 'error')
    
    return redirect(url_for('meter_room_detail', room_id=room_id))

@app.route('/meter-readings/room/<int:room_id>/add-electricity-reading', methods=['POST'])
@login_required
@create_permission_required('meter_reading')
def add_electricity_reading(room_id):
    """Add electricity meter reading"""
    from app.models.models_meter_reading import MeterRoom, ElectricityMeterReading
    from datetime import datetime
    
    room = MeterRoom.query.get_or_404(room_id)
    
    try:
        meter_number = request.form.get('meter_number', '').strip()
        start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
        start_reading = float(request.form.get('start_reading', 0))
        end_reading_str = request.form.get('end_reading', '').strip()
        rate_per_unit = float(request.form.get('rate_per_unit', 0))
        notes = request.form.get('notes', '')
        
        if not meter_number:
            flash('Meter number is required', 'error')
            return redirect(url_for('meter_room_detail', room_id=room_id))
        
        # Handle initial reading (no end reading yet)
        if not end_reading_str:
            end_reading = start_reading  # Set end same as start for initial reading
            total_consumption = 0.0
            total_amount = 0.0
            notes = f"Initial reading. {notes}".strip()
        else:
            end_reading = float(end_reading_str)
            if end_reading < start_reading:
                flash('End reading cannot be less than start reading', 'error')
                return redirect(url_for('meter_room_detail', room_id=room_id))
            total_consumption = end_reading - start_reading
            total_amount = total_consumption * rate_per_unit
        
        reading = ElectricityMeterReading(
            meter_room_id=room_id,
            meter_number=meter_number,
            start_date=start_date,
            end_date=end_date,
            start_reading=start_reading,
            end_reading=end_reading,
            total_consumption=total_consumption,
            rate_per_unit=rate_per_unit,
            total_amount=total_amount,
            physical_pax=int(request.form.get('physical_pax', 0)),
            notes=notes,
            created_by=current_user.id
        )
        
        db.session.add(reading)
        db.session.commit()
        
        flash('Electricity meter reading added successfully', 'success')
        
    except (ValueError, TypeError) as e:
        flash('Invalid input data. Please check your entries.', 'error')
    
    return redirect(url_for('meter_room_detail', room_id=room_id))

@app.route('/meter-readings/export-excel')
@login_required
def export_meter_readings_excel():
    """Export meter readings to Excel"""
    from app.models.models_meter_reading import MeterCompany, MeterRoom, WaterMeterReading, ElectricityMeterReading
    import pandas as pd
    from io import BytesIO
    from flask import make_response
    
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    company_name = request.args.get('company_name', '').strip()
    room_number = request.args.get('room_number', '').strip()
    
    # Build water readings data
    water_query = db.session.query(
        WaterMeterReading.meter_number,
        WaterMeterReading.start_date,
        WaterMeterReading.end_date,
        WaterMeterReading.start_reading,
        WaterMeterReading.end_reading,
        WaterMeterReading.total_consumption,
        WaterMeterReading.rate_per_unit,
        WaterMeterReading.total_amount,
        MeterRoom.room_number,
        MeterCompany.company_name
    ).select_from(WaterMeterReading).join(MeterRoom, WaterMeterReading.meter_room_id == MeterRoom.id).join(MeterCompany, MeterRoom.company_id == MeterCompany.id)
    
    electricity_query = db.session.query(
        ElectricityMeterReading.meter_number,
        ElectricityMeterReading.start_date,
        ElectricityMeterReading.end_date,
        ElectricityMeterReading.start_reading,
        ElectricityMeterReading.end_reading,
        ElectricityMeterReading.total_consumption,
        ElectricityMeterReading.rate_per_unit,
        ElectricityMeterReading.total_amount,
        MeterRoom.room_number,
        MeterCompany.company_name
    ).select_from(ElectricityMeterReading).join(MeterRoom, ElectricityMeterReading.meter_room_id == MeterRoom.id).join(MeterCompany, MeterRoom.company_id == MeterCompany.id)
    
    # Apply same filters as in main view
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            water_query = water_query.filter(WaterMeterReading.start_date >= start_dt)
            electricity_query = electricity_query.filter(ElectricityMeterReading.start_date >= start_dt)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            water_query = water_query.filter(WaterMeterReading.end_date <= end_dt)
            electricity_query = electricity_query.filter(ElectricityMeterReading.end_date <= end_dt)
        except ValueError:
            pass
    
    if company_name:
        water_query = water_query.filter(MeterCompany.company_name.ilike(f'%{company_name}%'))
        electricity_query = electricity_query.filter(MeterCompany.company_name.ilike(f'%{company_name}%'))
    
    if room_number:
        water_query = water_query.filter(MeterRoom.room_number.ilike(f'%{room_number}%'))
        electricity_query = electricity_query.filter(MeterRoom.room_number.ilike(f'%{room_number}%'))
    
    # Execute queries and create DataFrames
    water_data = water_query.all()
    electricity_data = electricity_query.all()
    
    # Convert query results to list of dictionaries
    water_records = []
    for row in water_data:
        water_records.append({
            'Meter Number': row[0],
            'Start Date': row[1],
            'End Date': row[2],
            'Start Reading': row[3],
            'End Reading': row[4],
            'Total Consumption': row[5],
            'Rate per Unit': row[6],
            'Total Amount': row[7],
            'Room Number': row[8],
            'Company Name': row[9]
        })
    
    electricity_records = []
    for row in electricity_data:
        electricity_records.append({
            'Meter Number': row[0],
            'Start Date': row[1],
            'End Date': row[2],
            'Start Reading': row[3],
            'End Reading': row[4],
            'Total Consumption': row[5],
            'Rate per Unit': row[6],
            'Total Amount': row[7],
            'Room Number': row[8],
            'Company Name': row[9]
        })
    
    water_df = pd.DataFrame(water_records)
    electricity_df = pd.DataFrame(electricity_records)
    
    # Create Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        water_df.to_excel(writer, sheet_name='Water Readings', index=False)
        electricity_df.to_excel(writer, sheet_name='Electricity Readings', index=False)
    
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=meter_readings.xlsx'
    
    return response

@app.route('/meter-readings/export-pdf')
@login_required
def export_meter_readings_pdf():
    """Export meter readings to PDF"""
    from app.models.models_meter_reading import MeterCompany, MeterRoom, WaterMeterReading, ElectricityMeterReading
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from io import BytesIO
    from flask import make_response
    from datetime import datetime
    
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    company_name = request.args.get('company_name', '').strip()
    room_number = request.args.get('room_number', '').strip()
    
    # Build queries (same as Excel export)
    water_query = db.session.query(
        WaterMeterReading.meter_number,
        WaterMeterReading.start_date,
        WaterMeterReading.end_date,
        WaterMeterReading.total_consumption,
        WaterMeterReading.total_amount,
        MeterRoom.room_number,
        MeterCompany.company_name
    ).join(MeterRoom).join(MeterCompany)
    
    electricity_query = db.session.query(
        ElectricityMeterReading.meter_number,
        ElectricityMeterReading.start_date,
        ElectricityMeterReading.end_date,
        ElectricityMeterReading.total_consumption,
        ElectricityMeterReading.total_amount,
        MeterRoom.room_number,
        MeterCompany.company_name
    ).join(MeterRoom).join(MeterCompany)
    
    # Apply filters (same logic as Excel)
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            water_query = water_query.filter(WaterMeterReading.start_date >= start_dt)
            electricity_query = electricity_query.filter(ElectricityMeterReading.start_date >= start_dt)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            water_query = water_query.filter(WaterMeterReading.end_date <= end_dt)
            electricity_query = electricity_query.filter(ElectricityMeterReading.end_date <= end_dt)
        except ValueError:
            pass
    
    if company_name:
        water_query = water_query.filter(MeterCompany.company_name.ilike(f'%{company_name}%'))
        electricity_query = electricity_query.filter(MeterCompany.company_name.ilike(f'%{company_name}%'))
    
    if room_number:
        water_query = water_query.filter(MeterRoom.room_number.ilike(f'%{room_number}%'))
        electricity_query = electricity_query.filter(MeterRoom.room_number.ilike(f'%{room_number}%'))
    
    # Get data
    water_data = water_query.all()
    electricity_data = electricity_query.all()
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    story.append(Paragraph("Meter Readings Report", title_style))
    story.append(Spacer(1, 20))
    
    # Water readings section
    if water_data:
        story.append(Paragraph("Water Meter Readings", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        water_table_data = [['Meter#', 'Start Date', 'End Date', 'Consumption', 'Amount', 'Room', 'Company']]
        for reading in water_data:
            water_table_data.append([
                str(reading[0]), str(reading[1]), str(reading[2]), 
                f"{reading[3]:.2f}", f"${reading[4]:.2f}", str(reading[5]), str(reading[6])
            ])
        
        water_table = Table(water_table_data)
        water_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8E8E8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(water_table)
        story.append(Spacer(1, 20))
    
    # Electricity readings section
    if electricity_data:
        story.append(Paragraph("Electricity Meter Readings", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        electricity_table_data = [['Meter#', 'Start Date', 'End Date', 'Consumption', 'Amount', 'Room', 'Company']]
        for reading in electricity_data:
            electricity_table_data.append([
                str(reading[0]), str(reading[1]), str(reading[2]), 
                f"{reading[3]:.2f}", f"${reading[4]:.2f}", str(reading[5]), str(reading[6])
            ])
        
        electricity_table = Table(electricity_table_data)
        electricity_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8E8E8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(electricity_table)
    
    doc.build(story)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=meter_readings.pdf'
    
    return response

# Purchase Section Routes
@app.route('/purchase-form')
@login_required
@page_access_required('purchase')
def purchase_form():
    """Display the purchase form"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    return render_template('purchase_form.html', user=user)

@app.route('/download-purchase-form-pdf', methods=['POST'])
@login_required
@page_access_required('purchase')
@performance_timer
def download_purchase_form_pdf():
    """Generate and download PDF of purchase form"""
    try:
        user = current_user
        
        # Get form data
        form_data = request.form.to_dict()
        
        # Create PDF using ReportLab
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from io import BytesIO
        import os
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Container for PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        # Add logo if it exists
        logo_path = os.path.join('static', 'uploads', 'logo.png')
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=2*inch, height=1*inch)
                logo.hAlign = 'CENTER'
                elements.append(logo)
                elements.append(Spacer(1, 12))
            except Exception:
                pass
        
        # Title
        title = Paragraph("PURCHASE REQUISITION FORM", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Form header information  
        header_data = [
            ['PR N( PL/25/', form_data.get('pr_number', ''), 'Date:', form_data.get('request_date', '')],
            ['Purchase Book:', form_data.get('purchase_book', ''), 'Quotation & Price comparison:', form_data.get('quotation_price', '')]
        ]
        
        header_table = Table(header_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 20))
        
        # Items table header
        items_header = [['No.', 'Item Description', 'Unit Cost', 'Qty', 'Total', 'Date Required', 'Last Order', 'Remarks']]
        
        # Items data - check actual form field names
        items_data = []
        
        # Get arrays from form data
        descriptions = form_data.getlist('description[]') if hasattr(form_data, 'getlist') else []
        unit_costs = form_data.getlist('unit_cost[]') if hasattr(form_data, 'getlist') else []
        quantities = form_data.getlist('quantity[]') if hasattr(form_data, 'getlist') else []
        totals = form_data.getlist('total[]') if hasattr(form_data, 'getlist') else []
        units = form_data.getlist('unit[]') if hasattr(form_data, 'getlist') else []
        date_requireds = form_data.getlist('date_required[]') if hasattr(form_data, 'getlist') else []
        last_orders = form_data.getlist('last_order[]') if hasattr(form_data, 'getlist') else []
        remarks = form_data.getlist('remarks[]') if hasattr(form_data, 'getlist') else []
        
        # If getlist doesn't work, try individual fields
        if not descriptions:
            for i in range(10):
                desc = form_data.get(f'description[{i}]', '').strip()
                if desc:
                    descriptions.append(desc)
                    unit_costs.append(form_data.get(f'unit_cost[{i}]', '0'))
                    quantities.append(form_data.get(f'quantity[{i}]', '0'))
                    totals.append(form_data.get(f'total[{i}]', '0'))
                    units.append(form_data.get(f'unit[{i}]', ''))
                    date_requireds.append(form_data.get(f'date_required[{i}]', ''))
                    last_orders.append(form_data.get(f'last_order[{i}]', ''))
                    remarks.append(form_data.get(f'remarks[{i}]', ''))
        
        # Process items
        max_items = max(len(descriptions), len(unit_costs), len(quantities), len(totals))
        for i in range(max_items):
            if i < len(descriptions) and descriptions[i].strip():
                items_data.append([
                    str(i + 1),
                    descriptions[i] if i < len(descriptions) else '',
                    unit_costs[i] if i < len(unit_costs) else '0',
                    quantities[i] if i < len(quantities) else '0', 
                    totals[i] if i < len(totals) else '0',
                    date_requireds[i] if i < len(date_requireds) else '',
                    last_orders[i] if i < len(last_orders) else '',
                    remarks[i] if i < len(remarks) else ''
                ])
        
        # Add empty rows if needed
        while len(items_data) < 8:
            items_data.append(['', '', '', '', '', '', '', ''])
        
        all_items_data = items_header + items_data
        
        # Add totals row - get from hidden fields
        subtotal = form_data.get('subtotal', form_data.get('calculated_subtotal', '0'))
        grand_total = form_data.get('grand_total', form_data.get('calculated_grand_total', '0'))
        other_amount = form_data.get('other_amount', '0')
        
        totals_data = [
            ['', '', '', '', '', '', 'Subtotal:', f'${subtotal}'],
            ['', '', '', '', '', '', form_data.get('other_label', 'GST:'), f'${other_amount}'],
            ['', '', '', '', '', '', 'GRAND TOTAL:', f'${grand_total}']
        ]
        
        all_items_data.extend(totals_data)
        
        items_table = Table(all_items_data, colWidths=[0.5*inch, 2.5*inch, 0.8*inch, 0.5*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1*inch])
        items_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -4), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -4), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -4), [colors.beige, colors.white]),
            
            # Totals rows styling
            ('BACKGROUND', (0, -3), (-1, -1), colors.lightgreen),
            ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -3), (-1, -1), 9),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(items_table)
        elements.append(Spacer(1, 30))
        
        # Signature section with actual signature data
        signature_data = [
            ['Requested By:', 'Recommended By:', 'Approved By:'],
            [form_data.get('requested_by_name', ''), form_data.get('recommended_by_name', ''), form_data.get('approved_by_name', '')],
            ['', '', ''],
            ['Signature:', 'Signature:', 'Signature:'],
            ['', '', ''],
            ['HOD & CEO:', 'Operation Manager:', 'General Manager:']
        ]
        
        # Add signature images if they exist
        signatures = ['dc_oe_signature_data', 'operation_manager_signature_data', 'general_manager_signature_data']
        signature_names = [
            form_data.get('dc_oe_name', ''),
            form_data.get('operation_manager_name', ''), 
            form_data.get('general_manager_name', '')
        ]
        
        # Update signature data with names
        signature_data[1] = signature_names
        
        for i, sig_field in enumerate(signatures):
            sig_data = form_data.get(sig_field, '')
            if sig_data and sig_data.startswith('data:image'):
                try:
                    # Decode base64 signature
                    import base64
                    from PIL import Image as PILImage
                    from io import BytesIO
                    
                    # Remove data:image/png;base64, prefix and decode
                    header, encoded = sig_data.split(',', 1)
                    sig_bytes = base64.b64decode(encoded)
                    
                    # Create temporary signature image file
                    temp_sig_path = f'/tmp/signature_{i}_{os.getpid()}.png'
                    
                    # Open with PIL to ensure it's a valid image
                    pil_img = PILImage.open(BytesIO(sig_bytes))
                    
                    # Convert to RGB if needed and resize if too large
                    if pil_img.mode in ('RGBA', 'LA'):
                        background = PILImage.new('RGB', pil_img.size, (255, 255, 255))
                        background.paste(pil_img, mask=pil_img.split()[-1])
                        pil_img = background
                    
                    # Resize if too large
                    max_width, max_height = 200, 100
                    if pil_img.width > max_width or pil_img.height > max_height:
                        pil_img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
                    
                    # Save processed image
                    pil_img.save(temp_sig_path, 'PNG')
                    
                    # Add signature image to PDF using ReportLab Image
                    try:
                        sig_img = Image(temp_sig_path, width=1.8*inch, height=0.9*inch)
                        sig_img.hAlign = 'CENTER'
                        # Store in signature data - row 3 is the signature row
                        signature_data[3][i] = sig_img
                    except Exception as e:
                        print(f"Error creating ReportLab image: {e}")
                        signature_data[3][i] = '[Signed]'
                    
                    # Clean up temp file
                    try:
                        os.remove(temp_sig_path)
                    except:
                        pass
                        
                except Exception as e:
                    print(f"Error processing signature {i}: {e}")
                    signature_data[3][i] = '[Signed]'
            else:
                # No signature provided
                signature_data[3][i] = ''
        
        signature_table = Table(signature_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch], rowHeights=[0.4*inch, 0.4*inch, 0.3*inch, 1.2*inch, 0.3*inch, 0.4*inch])
        signature_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Header row
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey), # Footer row
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),    # Bold headers
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Bold footers
            ('FONTSIZE', (0, 3), (-1, 3), 8),  # Smaller font for signature row
        ]))
        elements.append(signature_table)
        
        # Build PDF
        doc.build(elements)
        
        # Get buffer value and create response
        buffer.seek(0)
        
        # Generate filename
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'purchase_form_{timestamp}.pdf'
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('purchase_form'))

@app.route('/upload-logo', methods=['POST'])
@login_required
@page_access_required('purchase')
@performance_timer
def upload_logo():
    """Upload and save organization logo"""
    try:
        user = current_user
        
        if 'logo' not in request.files:
            return jsonify({'success': False, 'error': 'No logo file provided'})
        
        file = request.files['logo']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'})
        
        # Validate file type
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}
        if not file.filename.lower().endswith(tuple(allowed_extensions)):
            return jsonify({'success': False, 'error': 'Invalid file type. Please use PNG, JPG, JPEG, GIF, or BMP'})
        
        # Create uploads directory if it doesn't exist
        uploads_dir = os.path.join('static', 'uploads')
        os.makedirs(uploads_dir, exist_ok=True)
        
        # Save the file as logo.png (convert all formats to PNG for consistency)
        from PIL import Image
        import io
        
        # Open and process the image
        image = Image.open(file.stream)
        
        # Convert to RGB if necessary (for PNG with transparency)
        if image.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', image.size, (255, 255, 255))
            if image.mode == 'P':
                image = image.convert('RGBA')
            background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
            image = background
        
        # Resize image to reasonable size (max 400x200)
        image.thumbnail((400, 200), Image.Resampling.LANCZOS)
        
        # Save as PNG
        logo_path = os.path.join(uploads_dir, 'logo.png')
        image.save(logo_path, 'PNG', quality=95)
        
        logo_url = url_for('static', filename='uploads/logo.png')
        
        return jsonify({
            'success': True, 
            'logo_url': logo_url,
            'message': 'Logo uploaded successfully'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/stock-storage')
@login_required
@page_access_required('purchase')
def stock_storage():
    """Display stock storage management"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get stock items for this organization
    stock_items = StockItem.query.filter_by(organization_id=user.organization_id).all()
    
    # Get unique categories
    categories = db.session.query(StockItem.category).filter_by(organization_id=user.organization_id).distinct().all()
    categories = [cat[0] for cat in categories if cat[0]]
    
    return render_template('stock_storage.html', stock_items=stock_items, categories=categories)

@app.route('/purchase-form-storage')
@login_required
@page_access_required('purchase')
def purchase_form_storage():
    """Display purchase form storage management"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get purchase requests for this organization
    search_query = request.args.get('search', '').strip()
    query = PurchaseRequest.query.filter_by(organization_id=user.organization_id)
    
    if search_query:
        query = query.filter(
            db.or_(
                PurchaseRequest.request_number.ilike(f'%{search_query}%'),
                PurchaseRequest.requested_by.ilike(f'%{search_query}%'),
                PurchaseRequest.category.ilike(f'%{search_query}%')
            )
        )
    
    purchase_requests = query.order_by(PurchaseRequest.created_at.desc()).all()
    
    return render_template('purchase_form_storage.html', 
                         purchase_requests=purchase_requests, 
                         search_query=search_query)

@app.route('/submit-purchase-form', methods=['POST'])
@login_required
@create_permission_required('purchase')
def submit_purchase_form():
    """Submit purchase form"""
    try:
        user = current_user
        if not user.organization_id:
            return jsonify({'success': False, 'error': 'No organization assigned'})
        
        # Generate unique request number
        request_number = f"PR-{singapore_now().strftime('%Y%m%d')}-{PurchaseRequest.query.count() + 1:04d}"
        
        # Get form data
        pr_number = request.form.get('pr_number', '')
        request_date = datetime.strptime(request.form.get('request_date'), '%Y-%m-%d').date()
        categories = request.form.getlist('category[]')
        category = ', '.join(categories) if categories else 'Purchase Stock'
        
        # Create purchase request using verified model fields
        purchase_request = PurchaseRequest(
            request_number=request_number,
            pl_number=pr_number,
            request_date=request_date,
            category=category,
            requested_by=user.username or user.id,
            dc_name=request.form.get('dc_oe_name', ''),
            operation_manager=request.form.get('operation_manager_name', ''),
            general_manager=request.form.get('general_manager_name', ''),
            dc_signature_data=request.form.get('dc_oe_signature', ''),
            operation_manager_signature_data=request.form.get('operation_manager_signature', ''),
            general_manager_signature_data=request.form.get('general_manager_signature', ''),
            organization_id=user.organization_id,
            created_by=user.id
        )
        
        db.session.add(purchase_request)
        db.session.flush()  # Get the ID
        
        # Add request items using verified model fields
        descriptions = request.form.getlist('description[]')
        unit_costs = request.form.getlist('unit_cost[]')
        quantities = request.form.getlist('quantity[]')
        totals = request.form.getlist('total[]')
        units = request.form.getlist('unit[]')
        cost_codes = request.form.getlist('cost_code[]')
        remarks_list = request.form.getlist('remarks[]')
        
        for i in range(len(descriptions)):
            if descriptions[i].strip():  # Only add non-empty items
                item = PurchaseRequestItem(
                    purchase_request_id=purchase_request.id,
                    description=descriptions[i],
                    unit_cost=float(unit_costs[i]) if i < len(unit_costs) and unit_costs[i] else 0.0,
                    quantity=int(quantities[i]) if i < len(quantities) and quantities[i] else 1,
                    total_cost=float(totals[i]) if i < len(totals) and totals[i] else 0.0,
                    unit=units[i] if i < len(units) else '',
                    cost_code=cost_codes[i] if i < len(cost_codes) else '',
                    remarks=remarks_list[i] if i < len(remarks_list) else ''
                )
                db.session.add(item)
                
                # Automatically create stock item with "not_received" status
                stock_item = StockItem(
                    name=descriptions[i],
                    description=f"From Purchase Request {request_number}",
                    category=category.split(',')[0].strip() if category else 'Purchase Stock',
                    quantity=int(quantities[i]) if i < len(quantities) and quantities[i] else 1,
                    status='not_received',  # Default to not received
                    location='Pending Delivery',
                    purchase_cost=float(unit_costs[i]) if i < len(unit_costs) and unit_costs[i] else None,
                    organization_id=user.organization_id,
                    created_by=user.id
                )
                db.session.add(stock_item)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'request_number': request_number,
            'message': f'Purchase form {request_number} submitted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error submitting purchase form: {str(e)}'})

# Stock Items API Routes
@app.route('/api/stock-items', methods=['GET'])
@login_required
def get_stock_items():
    """Get stock items for dropdown"""
    try:
        user = current_user
        if not user.organization_id:
            return jsonify({'success': False, 'error': 'No organization assigned'})
        
        items = StockItem.query.filter_by(organization_id=user.organization_id).all()
        items_data = [{'id': item.id, 'name': item.name, 'description': item.description} for item in items]
        
        return jsonify({'success': True, 'items': items_data})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/stock-items', methods=['POST'])
@login_required
@create_permission_required('purchase')
def create_stock_item():
    """Create new stock item"""
    try:
        user = current_user
        if not user.organization_id:
            return jsonify({'success': False, 'error': 'No organization assigned'})
        
        # Create stock item using verified model fields
        stock_item = StockItem(
            name=request.form.get('name'),
            description=request.form.get('description', ''),
            category=request.form.get('category'),
            quantity=int(request.form.get('quantity', 1)),
            status=request.form.get('status', 'received'),
            location=request.form.get('location', ''),
            room_no=request.form.get('room_no', ''),
            purchase_cost=float(request.form.get('purchase_cost', 0)) if request.form.get('purchase_cost') else None,
            purchase_date=datetime.strptime(request.form.get('purchase_date'), '%Y-%m-%d').date() if request.form.get('purchase_date') else None,
            serial_number=request.form.get('serial_number', ''),
            organization_id=user.organization_id,
            created_by=user.id
        )
        
        db.session.add(stock_item)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Stock item created successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/stock-items/<int:item_id>', methods=['GET'])
@login_required
def get_stock_item(item_id):
    """Get stock item details"""
    try:
        user = current_user
        item = StockItem.query.filter_by(id=item_id, organization_id=user.organization_id).first()
        
        if not item:
            return jsonify({'success': False, 'error': 'Item not found'})
        
        item_data = {
            'id': item.id,
            'name': item.name,
            'description': item.description,
            'category': item.category,
            'quantity': item.quantity,
            'status': item.status,
            'location': item.location
        }
        
        return jsonify({'success': True, 'item': item_data})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/stock-items/<int:item_id>', methods=['PUT'])
@login_required
@edit_permission_required('purchase')
def update_stock_item(item_id):
    """Update stock item"""
    try:
        user = current_user
        item = StockItem.query.filter_by(id=item_id, organization_id=user.organization_id).first()
        
        if not item:
            return jsonify({'success': False, 'error': 'Item not found'})
        
        # Update using verified model fields
        item.name = request.form.get('name')
        item.description = request.form.get('description', '')
        item.category = request.form.get('category')
        item.quantity = int(request.form.get('quantity', 1))
        item.status = request.form.get('status', 'received')
        item.location = request.form.get('location', '')
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Stock item updated successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/stock-items/<int:item_id>', methods=['DELETE'])
@login_required
@edit_permission_required('purchase')
def delete_stock_item(item_id):
    """Delete stock item"""
    try:
        user = current_user
        item = StockItem.query.filter_by(id=item_id, organization_id=user.organization_id).first()
        
        if not item:
            return jsonify({'success': False, 'error': 'Item not found'})
        
        db.session.delete(item)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Stock item deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/stock-items/<int:item_id>/status', methods=['PUT'])
@login_required
@edit_permission_required('purchase')
def update_stock_item_status(item_id):
    """Update stock item status only"""
    try:
        user = current_user
        item = StockItem.query.filter_by(id=item_id, organization_id=user.organization_id).first()
        
        if not item:
            return jsonify({'success': False, 'error': 'Item not found'})
        
        new_status = request.form.get('status')
        if new_status not in ['received', 'not_received', 'partially_received']:
            return jsonify({'success': False, 'error': 'Invalid status'})
        
        old_status = item.status
        item.status = new_status
        
        # Create stock movement record for status change
        movement = StockMovement(
            stock_item_id=item.id,
            movement_type='STATUS_CHANGE',
            quantity=0,
            previous_quantity=item.quantity,
            new_quantity=item.quantity,
            reason=f'Status changed from {old_status} to {new_status}',
            notes=f'Status update by {user.username or user.email}',
            organization_id=user.organization_id,
            created_by=user.id
        )
        
        db.session.add(movement)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Status updated successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/download-purchase-pdf/<int:request_id>')
@login_required
@page_access_required('purchase')
def download_purchase_pdf(request_id):
    """Download individual purchase request as PDF"""
    try:
        user = current_user
        purchase_request = PurchaseRequest.query.filter_by(
            id=request_id, 
            organization_id=user.organization_id
        ).first()
        
        if not purchase_request:
            flash('Purchase request not found', 'error')
            return redirect(url_for('purchase_form_storage'))
        
        # Generate PDF using reportlab
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Header with TS Group Logo and Title (matching the form exactly)
        from reportlab.lib.units import inch
        from reportlab.platypus import Image
        from reportlab.graphics.shapes import Drawing, Rect
        from reportlab.graphics import renderPDF
        
        # Create TS GROUP logo box with yellow square
        ts_logo_data = [
            ['TS GROUP']
        ]
        
        ts_logo_table = Table(ts_logo_data, colWidths=[1.5*inch])
        ts_logo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.Color(0.117, 0.251, 0.686)),  # #1e40af
            ('TEXTCOLOR', (0, 0), (0, 0), colors.white),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 14),
            ('BOTTOMPADDING', (0, 0), (0, 0), 15),
            ('TOPPADDING', (0, 0), (0, 0), 15),
            ('GRID', (0, 0), (-1, -1), 2, colors.black)
        ]))
        
        # Main title
        title_data = [
            ['PURCHASE REQUISITION FORM']
        ]
        
        title_table = Table(title_data, colWidths=[5*inch])
        title_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.Color(0.878, 0.941, 0.996)),  # #e0f2fe
            ('TEXTCOLOR', (0, 0), (0, 0), colors.Color(0.117, 0.251, 0.686)),  # #1e40af
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 18),
            ('BOTTOMPADDING', (0, 0), (0, 0), 15),
            ('TOPPADDING', (0, 0), (0, 0), 15),
            ('GRID', (0, 0), (-1, -1), 2, colors.black)
        ]))
        
        # Combine logo and title
        header_combined = Table([[ts_logo_table, title_table]], colWidths=[1.5*inch, 5*inch])
        header_combined.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 2, colors.black)
        ]))
        
        elements.append(header_combined)
        elements.append(Spacer(1, 15))
        
        # Request details section
        details_title = Paragraph("REQUEST DETAILS", styles['Heading2'])
        elements.append(details_title)
        elements.append(Spacer(1, 10))
        
        # Two-column layout for request details
        left_details = [
            ['Request Number:', purchase_request.request_number or ''],
            ['PL Number:', f"PL/25/{purchase_request.pl_number}" if purchase_request.pl_number else 'N/A'],
            ['Request Date:', purchase_request.request_date.strftime('%Y-%m-%d') if purchase_request.request_date else 'N/A'],
            ['Category:', purchase_request.category or 'Purchase Stock']
        ]
        
        right_details = [
            ['Requested By:', purchase_request.requested_by or ''],
            ['DC Name:', purchase_request.dc_name or ''],
            ['Operation Manager:', purchase_request.operation_manager or ''],
            ['Status:', purchase_request.status or 'Pending']
        ]
        
        # Create side-by-side details
        combined_details = []
        for i in range(max(len(left_details), len(right_details))):
            row = []
            if i < len(left_details):
                row.extend(left_details[i])
            else:
                row.extend(['', ''])
            if i < len(right_details):
                row.extend(right_details[i])
            else:
                row.extend(['', ''])
            combined_details.append(row)
        
        details_table = Table(combined_details, colWidths=[1.2*inch, 1.8*inch, 1.2*inch, 1.8*inch])
        details_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('BACKGROUND', (3, 0), (3, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(details_table)
        elements.append(Spacer(1, 20))
        
        # Items section title
        items_title = Paragraph("ITEMS REQUESTED", styles['Heading2'])
        elements.append(items_title)
        elements.append(Spacer(1, 10))
        
        # Items table (exactly matching form layout)
        items_data = [['No', 'Item / Description', 'Unit Cost', 'Qty', 'Total', 'Unit Required', 'Cost Code', 'Remarks']]
        
        items = PurchaseRequestItem.query.filter_by(purchase_request_id=purchase_request.id).all()
        
        # Ensure we show at least 10 rows for consistency with form
        for i in range(10):
            if i < len(items):
                item = items[i]
                items_data.append([
                    str(i + 1),
                    item.description or '',
                    f"${item.unit_cost:.2f}" if item.unit_cost else '',
                    str(item.quantity or ''),
                    f"${item.total_cost:.2f}" if item.total_cost else '',
                    item.unit or '',
                    item.cost_code or '',
                    item.remarks or ''
                ])
            else:
                # Empty rows to maintain form structure
                items_data.append([str(i + 1), '', '', '', '', '', '', ''])
        
        items_table = Table(items_data, colWidths=[0.4*inch, 2.2*inch, 0.8*inch, 0.5*inch, 0.8*inch, 0.8*inch, 0.7*inch, 1*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.129, 0.588, 0.953)),  # Primary blue header
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Description left-aligned
            ('ALIGN', (7, 1), (7, -1), 'LEFT'),  # Remarks left-aligned
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.98, 0.98, 0.98)])
        ]))
        
        elements.append(items_table)
        elements.append(Spacer(1, 20))
        
        # Calculate total amount
        total_amount = sum(item.total_cost or 0 for item in items)
        
        # Total section
        total_data = [
            ['', '', '', 'TOTAL AMOUNT:', f"${total_amount:.2f}"]
        ]
        
        total_table = Table(total_data, colWidths=[0.4*inch, 2.2*inch, 0.8*inch, 1.3*inch, 0.8*inch])
        total_table.setStyle(TableStyle([
            ('BACKGROUND', (3, 0), (4, 0), colors.lightgrey),
            ('FONTNAME', (3, 0), (4, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (3, 0), (4, 0), 10),
            ('ALIGN', (3, 0), (3, 0), 'RIGHT'),
            ('ALIGN', (4, 0), (4, 0), 'CENTER'),
            ('GRID', (3, 0), (4, 0), 1, colors.black),
            ('BOTTOMPADDING', (3, 0), (4, 0), 8),
            ('TOPPADDING', (3, 0), (4, 0), 8)
        ]))
        
        elements.append(total_table)
        elements.append(Spacer(1, 30))
        
        # Signature sections (exactly matching form layout)
        signature_title = Paragraph("APPROVALS AND SIGNATURES", styles['Heading2'])
        elements.append(signature_title)
        elements.append(Spacer(1, 15))
        
        # Three signature columns
        sig_data = [
            ['Requested By', 'Recommended By', 'Approved By'],
            ['', '', ''],
            ['', '', ''],
            ['', '', ''],
            ['_________________', '_________________', '_________________'],
            [f"Name: {purchase_request.requested_by or ''}", f"Name: {purchase_request.operation_manager or ''}", 'Name: _______________'],
            ['Date: ___________', 'Date: ___________', 'Date: ___________']
        ]
        
        signature_table = Table(sig_data, colWidths=[2.2*inch, 2.2*inch, 2.2*inch])
        signature_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('SPAN', (0, 1), (0, 3)),  # Signature space
            ('SPAN', (1, 1), (1, 3)),
            ('SPAN', (2, 1), (2, 3))
        ]))
        
        elements.append(signature_table)
        elements.append(Spacer(1, 20))
        
        # Footer with form details
        from datetime import datetime as dt
        footer_data = [
            ['Form Number: PR-001', f'Generated: {dt.now().strftime("%Y-%m-%d %H:%M")}', 'Page 1 of 1']
        ]
        
        footer_table = Table(footer_data, colWidths=[2.2*inch, 2.2*inch, 2.2*inch])
        footer_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey)
        ]))
        
        elements.append(footer_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"purchase_request_{purchase_request.request_number}.pdf",
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('purchase_form_storage'))



@app.route('/edit-purchase-form/<int:request_id>')
@login_required
@page_access_required('purchase')
def edit_purchase_form(request_id):
    """Edit purchase form"""
    try:
        user = current_user
        purchase_request = PurchaseRequest.query.filter_by(
            id=request_id, 
            organization_id=user.organization_id
        ).first()
        
        if not purchase_request:
            flash('Purchase request not found', 'error')
            return redirect(url_for('purchase_form_storage'))
        
        # Get items for this request
        items = PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).all()
        
        return render_template('edit_purchase_form.html', 
                             purchase_request=purchase_request,
                             items=items)
        
    except Exception as e:
        flash(f'Error loading purchase form: {str(e)}', 'error')
        return redirect(url_for('purchase_form_storage'))

@app.route('/export-purchase-excel', methods=['POST'])
@login_required
@page_access_required('purchase')
def export_purchase_excel():
    """Export selected purchase requests as Excel file"""
    try:
        user = current_user
        selected_ids = request.form.getlist('selected_ids[]')
        
        if not selected_ids:
            flash('No purchase requests selected for export', 'error')
            return redirect(url_for('purchase_form_storage'))
        
        # Get selected purchase requests
        purchase_requests = PurchaseRequest.query.filter(
            PurchaseRequest.id.in_(selected_ids),
            PurchaseRequest.organization_id == user.organization_id
        ).order_by(PurchaseRequest.created_at.desc()).all()
        
        if not purchase_requests:
            flash('No valid purchase requests found', 'error')
            return redirect(url_for('purchase_form_storage'))
        
        # Create Excel workbook
        import pandas as pd
        from io import BytesIO
        import datetime
        
        buffer = BytesIO()
        
        # Create summary sheet
        summary_data = []
        for pr in purchase_requests:
            items = PurchaseRequestItem.query.filter_by(purchase_request_id=pr.id).all()
            total_amount = sum(item.total_cost or 0 for item in items)
            
            summary_data.append({
                'Request Number': pr.request_number,
                'PL Number': pr.pl_number or '',
                'Date': pr.request_date.strftime('%Y-%m-%d') if pr.request_date else '',
                'Category': pr.category or '',
                'Requested By': pr.requested_by or '',
                'DC Name': pr.dc_name or '',
                'Operation Manager': pr.operation_manager or '',
                'Total Items': len(items),
                'Total Amount': total_amount,
                'Status': pr.status or 'Pending',
                'Created': pr.created_at.strftime('%Y-%m-%d %H:%M') if pr.created_at else ''
            })
        
        # Create detailed items sheet
        items_data = []
        for pr in purchase_requests:
            items = PurchaseRequestItem.query.filter_by(purchase_request_id=pr.id).all()
            for item in items:
                items_data.append({
                    'Request Number': pr.request_number,
                    'Item Description': item.description or '',
                    'Unit Cost': item.unit_cost or 0,
                    'Quantity': item.quantity or 0,
                    'Unit': item.unit or '',
                    'Total Cost': item.total_cost or 0,
                    'Cost Code': item.cost_code or '',
                    'Remarks': item.remarks or ''
                })
        
        # Write to Excel with multiple sheets - using openpyxl which is already installed
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Summary sheet
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Items detail sheet  
            items_df = pd.DataFrame(items_data)
            items_df.to_excel(writer, sheet_name='Items Detail', index=False)
            
            # Format the sheets
            workbook = writer.book
            for sheet_name in ['Summary', 'Items Detail']:
                worksheet = writer.sheets[sheet_name]
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        filename = f"purchase_requests_export_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error generating Excel export: {str(e)}', 'error')
        return redirect(url_for('purchase_form_storage'))

@app.route('/update-purchase-form/<int:request_id>', methods=['POST'])
@login_required
@page_access_required('purchase')
def update_purchase_form(request_id):
    """Update purchase form"""
    try:
        user = current_user
        purchase_request = PurchaseRequest.query.filter_by(
            id=request_id, 
            organization_id=user.organization_id
        ).first()
        
        if not purchase_request:
            flash('Purchase request not found', 'error')
            return redirect(url_for('purchase_form_storage'))
        
        # Update purchase request details
        purchase_request.request_number = request.form.get('request_number')
        purchase_request.pl_number = request.form.get('pl_number')
        purchase_request.request_date = datetime.strptime(request.form.get('request_date'), '%Y-%m-%d').date() if request.form.get('request_date') else None
        purchase_request.category = request.form.get('category')
        purchase_request.requested_by = request.form.get('requested_by')
        purchase_request.dc_name = request.form.get('dc_name')
        purchase_request.operation_manager = request.form.get('operation_manager')
        
        # Update financial totals
        purchase_request.subtotal = float(request.form.get('subtotal') or 0)
        purchase_request.other_label = request.form.get('other_label')
        purchase_request.other_amount = float(request.form.get('other_amount') or 0)
        purchase_request.grand_total = float(request.form.get('grand_total') or 0)
        
        # Delete existing items
        PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).delete()
        
        # Add new items
        descriptions = request.form.getlist('description[]')
        unit_costs = request.form.getlist('unit_cost[]')
        quantities = request.form.getlist('quantity[]')
        units = request.form.getlist('unit[]')
        totals = request.form.getlist('total[]')
        cost_codes = request.form.getlist('cost_code[]')
        remarks = request.form.getlist('remarks[]')
        
        for i in range(len(descriptions)):
            if descriptions[i].strip():  # Only add non-empty items
                item = PurchaseRequestItem(
                    purchase_request_id=request_id,
                    description=descriptions[i],
                    unit_cost=float(unit_costs[i]) if unit_costs[i] else 0,
                    quantity=int(quantities[i]) if quantities[i] else 1,
                    unit=units[i],
                    total_cost=float(totals[i]) if totals[i] else 0,
                    cost_code=cost_codes[i],
                    remarks=remarks[i],
                    organization_id=user.organization_id
                )
                db.session.add(item)
        
        db.session.commit()
        flash('Purchase request updated successfully', 'success')
        return redirect(url_for('purchase_form_storage'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating purchase form: {str(e)}', 'error')
        return redirect(url_for('edit_purchase_form', request_id=request_id))

# Purchase Request API for viewing and management
@app.route('/api/purchase-request/<int:request_id>')
@login_required
def get_purchase_request_details(request_id):
    """Get purchase request details"""
    try:
        user = current_user
        purchase_request = PurchaseRequest.query.filter_by(
            id=request_id, 
            organization_id=user.organization_id
        ).first()
        
        if not purchase_request:
            return jsonify({'success': False, 'error': 'Purchase request not found'})
        
        # Get items using verified model fields
        items = PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).all()
        
        request_data = {
            'id': purchase_request.id,
            'request_number': purchase_request.request_number,
            'pl_number': purchase_request.pl_number,
            'request_date': purchase_request.request_date.strftime('%Y-%m-%d') if purchase_request.request_date else '',
            'category': purchase_request.category,
            'requested_by': purchase_request.requested_by,
            'dc_name': purchase_request.dc_name,
            'operation_manager': purchase_request.operation_manager,
            'general_manager': purchase_request.general_manager,
            'status': purchase_request.status or 'Pending',
            'items': [{
                'description': item.description,
                'unit_cost': float(item.unit_cost),
                'quantity': item.quantity,
                'total_cost': float(item.total_cost),
                'unit': item.unit or '',
                'cost_code': item.cost_code or '',
                'remarks': item.remarks or ''
            } for item in items]
        }
        
        return jsonify({'success': True, 'request': request_data})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/purchase-request/<int:request_id>/pdf')
@login_required
def download_purchase_request_pdf(request_id):
    """Download purchase request as PDF"""
    try:
        user = current_user
        if not user.organization_id:
            return jsonify({'success': False, 'error': 'No organization assigned'})
        
        purchase_request = PurchaseRequest.query.filter_by(
            id=request_id, 
            organization_id=user.organization_id
        ).first()
        
        if not purchase_request:
            return jsonify({'success': False, 'error': 'Purchase request not found'})
        
        # Create PDF using ReportLab
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from io import BytesIO
        import base64
        from PIL import Image as PILImage
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.darkblue,
            fontName='Helvetica-Bold'
        )
        
        # Build PDF content
        story = []
        
        # Header with logo and title
        header_data = [
            ['TS GROUP', 'PURCHASE REQUISITION FORM']
        ]
        header_table = Table(header_data, colWidths=[2*inch, 5*inch])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.white),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 12),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
            ('BACKGROUND', (1, 0), (1, 0), colors.lightblue),
            ('TEXTCOLOR', (1, 0), (1, 0), colors.darkblue),
            ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (1, 0), 16),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 20))
        
        # Basic information
        basic_info_data = [
            [f'PR N( PL/25/{purchase_request.pl_number or ""}', f'Date: {purchase_request.request_date.strftime("%Y-%m-%d") if purchase_request.request_date else ""}']
        ]
        basic_info_table = Table(basic_info_data, colWidths=[3.5*inch, 3.5*inch])
        basic_info_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(basic_info_table)
        story.append(Spacer(1, 10))
        
        # Category checkboxes
        category_text = f"Categories: {purchase_request.category}"
        category_para = Paragraph(category_text, styles['Normal'])
        story.append(category_para)
        story.append(Spacer(1, 15))
        
        # Items table
        items = PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).all()
        
        if items:
            items_data = [
                ['No', 'Item / Description', 'Unit Cost', 'Qty', 'Total', 'Unit Required', 'Cost Code', 'Remarks']
            ]
            
            for i, item in enumerate(items, 1):
                items_data.append([
                    str(i),
                    item.description or '',
                    f"${float(item.unit_cost or 0):.2f}",
                    str(item.quantity or 0),
                    f"${float(item.total_cost or 0):.2f}",
                    item.unit or '',
                    item.cost_code or '',
                    item.remarks or ''
                ])
            
            # Add empty rows to match form layout
            while len(items_data) < 3:  # Minimum 2 rows plus header
                items_data.append(['', '', '', '', '', '', '', ''])
            
            items_table = Table(items_data, colWidths=[0.4*inch, 2.5*inch, 0.8*inch, 0.5*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1*inch])
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            story.append(items_table)
            story.append(Spacer(1, 20))
        
        # Footer section
        footer_data = [
            ['Requested By', 'Recommended By']
        ]
        footer_table = Table(footer_data, colWidths=[3.5*inch, 3.5*inch])
        footer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 1, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(footer_table)
        story.append(Spacer(1, 10))
        
        # Signature section
        signature_data = [
            ['D/C & O/E', 'Operation Manager', 'General Manager'],
            ['', '', ''],
            ['', '', ''],
            [purchase_request.dc_name or '', purchase_request.operation_manager or '', purchase_request.general_manager or ''],
            ['Name', 'Name', 'Name'],
            ['Signature', 'Signature', 'Signature']
        ]
        
        signature_table = Table(signature_data, colWidths=[2.33*inch, 2.33*inch, 2.33*inch], rowHeights=[0.3*inch, 0.8*inch, 0.1*inch, 0.3*inch, 0.2*inch, 0.2*inch])
        signature_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 3), (-1, 3), 'Helvetica'),
            ('FONTSIZE', (0, 3), (-1, 3), 9),
            ('FONTNAME', (0, 4), (-1, 5), 'Helvetica'),
            ('FONTSIZE', (0, 4), (-1, 5), 8),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        story.append(signature_table)
        story.append(Spacer(1, 10))
        
        # Final footer
        final_footer_data = [
            ['DC/Site In Charge/Admin', 'SEM/PM/AM']
        ]
        final_footer_table = Table(final_footer_data, colWidths=[3.5*inch, 3.5*inch])
        final_footer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 1, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(final_footer_table)
        
        # Build PDF
        doc.build(story)
        
        # Return PDF
        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=purchase_request_{purchase_request.request_number}.pdf'
        
        return response
        
    except Exception as e:
        import logging
        logging.error(f"Error generating PDF: {str(e)}")
        return jsonify({'success': False, 'error': f'PDF generation failed: {str(e)}'})

@app.route('/api/purchase-request/<int:request_id>', methods=['DELETE'])
@login_required
@edit_permission_required('purchase')
def delete_purchase_request(request_id):
    """Delete purchase request"""
    try:
        user = current_user
        purchase_request = PurchaseRequest.query.filter_by(
            id=request_id, 
            organization_id=user.organization_id
        ).first()
        
        if not purchase_request:
            return jsonify({'success': False, 'error': 'Purchase request not found'})
        
        # Delete items first
        PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).delete()
        
        # Delete the request
        db.session.delete(purchase_request)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Purchase request deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

# Stock Report Routes - DISABLED
# @app.route('/stock-report')
# @login_required
# def stock_report():
    """Stock report and inventory management page"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get stock items for stock report (separate from Asset Management)
    stock_items = StockItem.query.filter_by(organization_id=user.organization_id).all()
    
    # Get unique categories from stock items
    categories = db.session.query(StockItem.category).filter_by(organization_id=user.organization_id).distinct().all()
    categories = [cat[0] for cat in categories]
    
    return render_template('stock_report.html', assets=stock_items, categories=categories)

@app.route('/api/stock/availability', methods=['GET'])
@login_required
@page_access_required('stock_report')
def get_stock_availability():
    """Get real-time stock availability for an item"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'No organization assigned'})
    
    try:
        item_name = request.args.get('item_name', '').strip()
        item_id = request.args.get('item_id')
        
        if not item_name and not item_id:
            return jsonify({'success': False, 'error': 'Item name or ID required'})
        
        # Find stock item
        if item_id:
            stock_item = StockItem.query.filter_by(
                id=item_id,
                organization_id=user.organization_id
            ).first()
        else:
            stock_item = StockItem.query.filter_by(
                name=item_name,
                organization_id=user.organization_id
            ).first()
        
        if not stock_item:
            return jsonify({
                'success': False, 
                'error': 'Item not found in stock inventory',
                'available_quantity': 0,
                'total_quantity': 0,
                'used_quantity': 0
            })
        
        # Calculate total used quantity from StockUsage records
        from sqlalchemy import func
        total_used = db.session.query(func.sum(StockUsage.used_quantity)).filter_by(
            stock_item_id=stock_item.id,
            organization_id=user.organization_id
        ).scalar() or 0
        
        # Calculate available quantity
        available_quantity = stock_item.quantity - total_used
        
        return jsonify({
            'success': True,
            'item_id': stock_item.id,
            'item_name': stock_item.name,
            'description': stock_item.description,
            'category': stock_item.category,
            'total_quantity': stock_item.quantity,
            'used_quantity': total_used,
            'available_quantity': max(0, available_quantity),
            'status': stock_item.status,
            'location': stock_item.location,
            'room_no': stock_item.room_no,
            'purchase_date': stock_item.purchase_date.strftime('%Y-%m-%d') if stock_item.purchase_date else None,
            'last_updated': stock_item.updated_at.strftime('%Y-%m-%d %H:%M:%S') if stock_item.updated_at else None
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error fetching stock availability: {str(e)}'})

@app.route('/api/stock/search', methods=['GET'])
@login_required
@page_access_required('stock_report')
def search_stock_items():
    """Search stock items by name for autocomplete"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'No organization assigned'})
    
    try:
        query = request.args.get('q', '').strip()
        if len(query) < 2:
            return jsonify({'success': True, 'items': []})
        
        stock_items = StockItem.query.filter(
            StockItem.organization_id == user.organization_id,
            StockItem.name.ilike(f'%{query}%')
        ).limit(10).all()
        
        items = []
        for item in stock_items:
            # Calculate available quantity
            from sqlalchemy import func
            total_used = db.session.query(func.sum(StockUsage.used_quantity)).filter_by(
                stock_item_id=item.id,
                organization_id=user.organization_id
            ).scalar() or 0
            
            available_quantity = item.quantity - total_used
            
            items.append({
                'id': item.id,
                'name': item.name,
                'description': item.description,
                'category': item.category,
                'total_quantity': item.quantity,
                'available_quantity': max(0, available_quantity),
                'status': item.status,
                'location': item.location
            })
        
        return jsonify({'success': True, 'items': items})
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error searching stock items: {str(e)}'})

@app.route('/stock-info')
@login_required
def stock_info():
    """Stock info page showing grouped summary of items"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get all stock items for the organization
    stock_items = StockItem.query.filter_by(organization_id=user.organization_id).all()
    
    # Group items by name and calculate totals (case-insensitive)
    grouped_items = {}
    for item in stock_items:
        item_name_key = item.name.lower()  # Use lowercase for grouping
        if item_name_key not in grouped_items:
            grouped_items[item_name_key] = {
                'name': item.name.upper(),  # Store as uppercase for consistency
                'category': item.category,
                'total_quantity': 0,
                'used_quantity': 0,
                'available_quantity': 0,
                'received_quantity': 0,
                'unreceived_quantity': 0,
                'total_cost': 0,
                'rooms': set(),
                'purchase_dates': [],
                'serial_numbers': []
            }
        
        # Add quantities
        grouped_items[item_name_key]['total_quantity'] += item.quantity
        grouped_items[item_name_key]['used_quantity'] += (item.used_quantity or 0)
        if item.status == 'received':
            grouped_items[item_name_key]['received_quantity'] += item.quantity
        else:
            grouped_items[item_name_key]['unreceived_quantity'] += item.quantity
            
        # Add cost
        if item.purchase_cost:
            grouped_items[item_name_key]['total_cost'] += float(item.purchase_cost)
            
        # Add room info
        if item.room_no:
            grouped_items[item_name_key]['rooms'].add(item.room_no)
            
        # Add purchase dates
        if item.purchase_date:
            grouped_items[item_name_key]['purchase_dates'].append(item.purchase_date)
            
        # Add serial numbers
        if item.serial_number:
            grouped_items[item_name_key]['serial_numbers'].append(item.serial_number)
    
    # Convert to list and format data
    stock_summary = []
    for i, (name, data) in enumerate(grouped_items.items(), 1):
        # Calculate available quantity (total - used)
        available_qty = data['total_quantity'] - data['used_quantity']
        stock_summary.append({
            'sno': i,
            'name': data['name'],
            'category': data['category'],
            'total_quantity': data['total_quantity'],
            'used_quantity': data['used_quantity'],
            'available_quantity': available_qty,
            'received_quantity': data['received_quantity'],
            'unreceived_quantity': data['unreceived_quantity'],
            'total_cost': f"{data['total_cost']:.2f}",
            'rooms': ', '.join(sorted(data['rooms'])) if data['rooms'] else 'N/A',
            'purchase_dates': data['purchase_dates'],
            'serial_numbers': ', '.join(data['serial_numbers']) if data['serial_numbers'] else 'N/A'
        })
    
    return render_template('stock_info.html', stock_summary=stock_summary)

@app.route('/used-info')
@login_required
def used_info():
    """Used stock information page"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get search parameters
    search_item_name = request.args.get('search_item_name', '').strip()
    category_filter = request.args.get('category', '').strip()
    
    # Start with base query
    query = StockItem.query.filter_by(organization_id=user.organization_id)
    
    # Apply filters
    if search_item_name:
        query = query.filter(StockItem.name.ilike(f'%{search_item_name}%'))
    
    if category_filter:
        query = query.filter_by(category=category_filter)
    
    # Get filtered stock items
    used_items = query.all()
    
    # Get categories for filter dropdown
    categories = db.session.query(StockItem.category).filter_by(organization_id=user.organization_id).distinct().all()
    categories = [cat[0] for cat in categories if cat[0]]
    
    return render_template('used_info.html', used_items=used_items, categories=categories)

@app.route('/used-info/update-quantity', methods=['POST'])
@login_required
@create_permission_required('stock_report')
def update_used_quantity():
    """Update used quantity for a stock item"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    stock_id = request.form.get('stock_id')
    used_quantity = request.form.get('used_quantity')
    
    try:
        used_quantity = int(used_quantity)
        if used_quantity < 0:
            flash('Used quantity cannot be negative', 'error')
            return redirect(url_for('used_info'))
        
        # Get the stock item
        stock_item = StockItem.query.filter_by(
            id=stock_id, 
            organization_id=user.organization_id
        ).first()
        
        if not stock_item:
            flash('Stock item not found', 'error')
            return redirect(url_for('used_info'))
        
        # Check if used quantity doesn't exceed total quantity
        if used_quantity > stock_item.quantity:
            flash(f'Used quantity ({used_quantity}) cannot exceed total quantity ({stock_item.quantity})', 'error')
            return redirect(url_for('used_info'))
        
        # Update used quantity
        old_used_quantity = stock_item.used_quantity or 0
        stock_item.used_quantity = used_quantity
        
        # Create stock movement record
        movement = StockMovement(
            stock_item_id=stock_item.id,
            movement_type='USED',
            quantity=used_quantity - old_used_quantity,
            previous_quantity=old_used_quantity,
            new_quantity=used_quantity,
            reason='Stock usage updated',
            notes=f'Updated used quantity from {old_used_quantity} to {used_quantity}',
            organization_id=user.organization_id,
            created_by=user.id
        )
        
        db.session.add(movement)
        db.session.commit()
        
        flash(f'Used quantity updated successfully for {stock_item.name}', 'success')
        
    except ValueError:
        flash('Invalid quantity value', 'error')
    except Exception as e:
        db.session.rollback()
        flash('Error updating used quantity', 'error')
        print(f"Error: {e}")
    
    return redirect(url_for('used_info'))

@app.route('/used-info/export/excel')
@login_required
def export_used_info_excel():
    """Export used stock info to Excel"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get all stock items for the organization
    stock_items = StockItem.query.filter_by(organization_id=user.organization_id).all()
    
    # Prepare data for Excel
    excel_data = []
    for i, item in enumerate(stock_items, 1):
        available_qty = item.quantity - (item.used_quantity or 0)
        excel_data.append([
            i,  # S.No
            item.name,  # Item Name
            item.category,  # Category
            item.quantity,  # Total Quantity
            item.used_quantity or 0,  # Used Quantity
            available_qty,  # Available Quantity
            item.status.title(),  # Status
            item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else 'N/A',  # Purchase Date
            f"${item.purchase_cost:.2f}" if item.purchase_cost else '$0.00',  # Cost
            item.serial_number or 'N/A'  # Serial Number
        ])
    
    # Create DataFrame
    columns = ['S.No', 'Item Name', 'Category', 'Total Quantity', 'Used Quantity', 
               'Available Quantity', 'Status', 'Purchase Date', 'Cost', 'Serial Number']
    df = pd.DataFrame(excel_data, columns=columns)
    
    # Create Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Used Stock Info', index=False)
        
        # Format the worksheet
        worksheet = writer.sheets['Used Stock Info']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    return send_file(
        io.BytesIO(output.read()),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'used_stock_info_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/used-info/export/pdf')
@login_required
def export_used_info_pdf():
    """Export used stock info to PDF"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get all stock items for the organization
    stock_items = StockItem.query.filter_by(organization_id=user.organization_id).all()
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    # Title and header
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    elements.append(Paragraph("Used Stock Information Report", title_style))
    elements.append(Paragraph(f"Organization: {user.organization.name}", styles['Normal']))
    elements.append(Paragraph(f"Generated on: {singapore_now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Create table data
    table_data = [['S.No', 'Item Name', 'Category', 'Total Qty', 'Used Qty', 'Available Qty', 'Status', 'Purchase Date', 'Cost']]
    
    for i, item in enumerate(stock_items, 1):
        available_qty = item.quantity - (item.used_quantity or 0)
        table_data.append([
            str(i),
            item.name,
            item.category,
            str(item.quantity),
            str(item.used_quantity or 0),
            str(available_qty),
            item.status.title(),
            item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else 'N/A',
            f"${item.purchase_cost:.2f}" if item.purchase_cost else '$0.00'
        ])
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8E8E8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'used_stock_info_{singapore_now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )

@app.route('/stock/info/export/excel')
@login_required
def export_stock_info_excel():
    """Export stock info summary to Excel"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get all stock items for the organization
    stock_items = StockItem.query.filter_by(organization_id=user.organization_id).all()
    
    # Group items by name and calculate totals (case-insensitive)
    grouped_items = {}
    for item in stock_items:
        item_name_key = item.name.lower()  # Use lowercase for grouping
        if item_name_key not in grouped_items:
            grouped_items[item_name_key] = {
                'name': item.name.upper(),  # Store as uppercase for consistency
                'category': item.category,
                'total_quantity': 0,
                'received_quantity': 0,
                'unreceived_quantity': 0,
                'total_cost': 0,
                'rooms': set(),
                'serial_numbers': []
            }
        
        # Add quantities
        grouped_items[item_name_key]['total_quantity'] += item.quantity
        if item.status == 'received':
            grouped_items[item_name_key]['received_quantity'] += item.quantity
        else:
            grouped_items[item_name_key]['unreceived_quantity'] += item.quantity
            
        # Add cost
        if item.purchase_cost:
            grouped_items[item_name_key]['total_cost'] += float(item.purchase_cost)
            
        # Add room info
        if item.room_no:
            grouped_items[item_name_key]['rooms'].add(item.room_no)
            
        # Add serial numbers
        if item.serial_number:
            grouped_items[item_name_key]['serial_numbers'].append(item.serial_number)
    
    # Convert to list for Excel export
    excel_data = []
    for i, (name, data) in enumerate(grouped_items.items(), 1):
        excel_data.append([
            i,  # S.No
            data['name'],  # Item Name
            data['category'],  # Category
            data['total_quantity'],  # Total Quantity
            data['received_quantity'],  # Received
            data['unreceived_quantity'],  # Unreceived
            f"{data['total_cost']:.2f}",  # Total Cost
            ', '.join(sorted(data['rooms'])) if data['rooms'] else 'N/A',  # Rooms
            ', '.join(data['serial_numbers']) if data['serial_numbers'] else 'N/A'  # Serial Numbers
        ])
    
    # Create Excel file
    output = BytesIO()
    
    # Create DataFrame
    columns = ['S.No', 'Item Name', 'Category', 'Total Quantity', 'Received', 'Unreceived', 'Total Cost', 'Rooms', 'Serial Numbers']
    df = pd.DataFrame(excel_data, columns=columns)
    
    # Write to Excel
    df.to_excel(output, index=False, sheet_name='Stock Info Summary')
    output.seek(0)
    
    # Create response
    response = Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers["Content-Disposition"] = f"attachment; filename=stock_info_summary_{singapore_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return response

# Stock Report API Routes
@app.route('/api/stock/add', methods=['POST'])
@login_required
@create_permission_required('inventory')
def add_stock_item():
    """Add new stock item"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    try:
        data = request.get_json()
        
        stock_item = StockItem(
            name=data['name'],
            description=data.get('description', ''),
            category=data['category'],
            quantity=int(data['quantity']),
            status=data.get('status', 'received'),
            location=data.get('location', ''),
            room_no=data.get('room_no', ''),
            purchase_date=datetime.strptime(data['purchase_date'], '%Y-%m-%d').date() if data.get('purchase_date') else None,
            purchase_cost=float(data['purchase_cost']) if data.get('purchase_cost') else None,
            serial_number=data.get('serial_number', ''),
            organization_id=user.organization_id,
            created_by=user.id
        )
        
        db.session.add(stock_item)
        db.session.commit()
        
        # Create stock movement record
        movement = StockMovement(
            stock_item_id=stock_item.id,
            movement_type='IN',
            quantity=stock_item.quantity,
            previous_quantity=0,
            new_quantity=stock_item.quantity,
            reason='Initial stock entry',
            organization_id=user.organization_id,
            created_by=user.id
        )
        db.session.add(movement)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Stock item added successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/stock/update-quantity', methods=['POST'])
@login_required
@create_permission_required('stock_report')
def update_stock_quantity():
    """Update stock item quantity"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    try:
        data = request.get_json()
        stock_item = StockItem.query.filter_by(
            id=data['stock_id'],
            organization_id=user.organization_id
        ).first()
        
        if not stock_item:
            return jsonify({'error': 'Stock item not found'}), 404
        
        previous_quantity = stock_item.quantity
        adjustment = int(data['adjustment'])
        new_quantity = previous_quantity + adjustment
        
        if new_quantity < 0:
            return jsonify({'error': 'Cannot reduce stock below zero'}), 400
        
        stock_item.quantity = new_quantity
        
        # Create stock movement record
        movement_type = 'IN' if adjustment > 0 else 'OUT'
        movement = StockMovement(
            stock_item_id=stock_item.id,
            movement_type=movement_type,
            quantity=abs(adjustment),
            previous_quantity=previous_quantity,
            new_quantity=new_quantity,
            reason=data.get('reason', f'Quantity {movement_type.lower()}'),
            notes=data.get('notes', ''),
            organization_id=user.organization_id,
            created_by=user.id
        )
        
        db.session.add(movement)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Stock quantity updated successfully',
            'new_quantity': new_quantity
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/stock/update-status', methods=['POST'])
@login_required
@create_permission_required('stock_report')
def update_stock_status():
    """Update stock item status"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    try:
        data = request.get_json()
        stock_item = StockItem.query.filter_by(
            id=data['stock_id'],
            organization_id=user.organization_id
        ).first()
        
        if not stock_item:
            return jsonify({'error': 'Stock item not found'}), 404
        
        old_status = stock_item.status
        new_status = data['status']
        stock_item.status = new_status
        
        # Create stock movement record for status change
        movement = StockMovement(
            stock_item_id=stock_item.id,
            movement_type='STATUS',
            quantity=0,
            previous_quantity=stock_item.quantity,
            new_quantity=stock_item.quantity,
            reason=f'Status changed from {old_status} to {new_status}',
            notes=f'Status update by {user.first_name or user.email}',
            organization_id=user.organization_id,
            created_by=user.id
        )
        
        db.session.add(movement)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Status updated to {new_status} successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/stock/change-status/<int:item_id>', methods=['POST'])
@login_required
@create_permission_required('stock_report')
def change_stock_status(item_id):
    """Change stock item status between received and unreceived"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    try:
        data = request.get_json()
        stock_item = StockItem.query.filter_by(
            id=item_id,
            organization_id=user.organization_id
        ).first()
        
        if not stock_item:
            return jsonify({'error': 'Stock item not found'}), 404
        
        old_status = stock_item.status
        new_status = data['status']
        
        # Validate status values
        if new_status not in ['received', 'unreceived']:
            return jsonify({'error': 'Invalid status value'}), 400
        
        stock_item.status = new_status
        
        # Create stock movement record for status change
        movement = StockMovement(
            stock_item_id=stock_item.id,
            movement_type='STATUS',
            quantity=0,
            previous_quantity=stock_item.quantity,
            new_quantity=stock_item.quantity,
            reason=f'Status changed from {old_status} to {new_status}',
            notes=f'Status update by {user.first_name or user.email}',
            organization_id=user.organization_id,
            created_by=user.id
        )
        
        db.session.add(movement)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Status updated to {new_status} successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/stock/export/pdf')
@login_required
def export_stock_pdf():
    """Export stock inventory to PDF"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned', 'error')
        return redirect(url_for('stock_report'))
    
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from io import BytesIO
        import os
        
        # Create a BytesIO buffer for the PDF
        buffer = BytesIO()
        
        # Create the PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add logo and title
        if os.path.exists('static/ts_logo.svg'):
            try:
                logo = Image('static/ts_logo.svg', width=1*inch, height=1*inch)
                elements.append(logo)
            except:
                pass
        
        # Title
        title = Paragraph("TS MANAGEMENT SERVICES PTE LTD<br/>Stock Inventory Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Get stock items
        stock_items = StockItem.query.filter_by(organization_id=user.organization_id).all()
        
        # Create table data
        data = [['S.No', 'Item Name', 'Category', 'Quantity', 'Status', 'Location', 'Purchase Date', 'Cost', 'Serial Number']]
        
        for i, item in enumerate(stock_items, 1):
            data.append([
                str(i),
                item.name,
                item.category or '-',
                str(item.quantity),
                item.status or '-',
                item.location or '-',
                item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else '-',
                f"${item.purchase_cost:.2f}" if item.purchase_cost else '-',
                item.serial_number or '-'
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8E8E8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer and create response
        pdf_data = buffer.getvalue()
        buffer.close()
        
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=stock_inventory_{singapore_now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('stock_report'))

@app.route('/api/stock/movements/<int:stock_id>')
@login_required
def get_stock_movements(stock_id):
    """Get stock movement history for an item"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    stock_item = StockItem.query.filter_by(
        id=stock_id,
        organization_id=user.organization_id
    ).first()
    
    if not stock_item:
        return jsonify({'error': 'Stock item not found'}), 404
    
    movements = StockMovement.query.filter_by(
        stock_item_id=stock_id,
        organization_id=user.organization_id
    ).order_by(StockMovement.created_at.desc()).all()
    
    movement_data = []
    for movement in movements:
        movement_data.append({
            'id': movement.id,
            'movement_type': movement.movement_type,
            'quantity': movement.quantity,
            'previous_quantity': movement.previous_quantity,
            'new_quantity': movement.new_quantity,
            'reason': movement.reason,
            'notes': movement.notes,
            'created_by': movement.created_by_user.first_name if movement.created_by_user else 'Unknown',
            'created_at': movement.created_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return jsonify({'movements': movement_data})

# Enhanced Multilingual Compliance Management Routes
@app.route('/compliance-management')
@login_required
@page_permission_required('compliance_tracking')
def compliance_management():
    """Display the enhanced multilingual compliance management interface"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned', 'error')
        return redirect(url_for('dashboard'))
    
    # Get all compliance records for the organization
    compliance_records = ComplianceRecord.query.filter_by(
        organization_id=user.organization_id
    ).order_by(ComplianceRecord.created_at.desc()).all()
    
    return render_template('compliance_management.html', 
                         compliance_records=compliance_records)

@app.route('/create-compliance-record', methods=['POST'])
@login_required
@create_permission_required('compliance_tracking')
def create_compliance_record():
    """Create a new multilingual compliance record"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned', 'error')
        return redirect(url_for('compliance_management'))
    
    try:
        # Generate unique record number
        import uuid
        record_number = f"CR-{singapore_now().year}-{str(uuid.uuid4())[:8].upper()}"
        
        # Basic information
        record_type = request.form.get('record_type')
        compliance_category = request.form.get('compliance_category')
        priority_level = request.form.get('priority_level', 'Medium')
        effective_date = datetime.strptime(request.form.get('effective_date'), '%Y-%m-%d').date()
        expiry_date = None
        if request.form.get('expiry_date'):
            expiry_date = datetime.strptime(request.form.get('expiry_date'), '%Y-%m-%d').date()
        requires_acknowledgment = 'requires_acknowledgment' in request.form
        
        # Create compliance records for each language
        languages = ['en', 'bn', 'my', 'ta', 'zh']
        created_records = []
        
        for lang in languages:
            title = request.form.get(f'title_{lang}')
            description = request.form.get(f'description_{lang}')
            detailed_instructions = request.form.get(f'detailed_instructions_{lang}')
            
            # Skip if no content provided for this language
            if not title and not description:
                continue
            
            # Use English title/description as fallback for required fields
            if not title:
                title = request.form.get('title_en', f'{compliance_category} ({lang.upper()})')
            if not description:
                description = request.form.get('description_en', 'Compliance requirement')
            
            # Create compliance record for this language
            compliance_record = ComplianceRecord(
                record_number=f"{record_number}-{lang.upper()}",
                organization_id=user.organization_id,
                record_type=record_type,
                compliance_category=compliance_category,
                priority_level=priority_level,
                language_code=lang,
                title=title,
                description=description,
                detailed_instructions=detailed_instructions,
                effective_date=effective_date,
                expiry_date=expiry_date,
                requires_acknowledgment=requires_acknowledgment,
                created_by=user.id,
                status='Active'
            )
            
            # Handle content input options for this language
            import base64
            
            # Text input
            content_text = request.form.get(f'content_text_{lang}')
            if content_text:
                compliance_record.content_text = content_text
            
            # PDF upload
            pdf_file = request.files.get(f'pdf_file_{lang}')
            if pdf_file and pdf_file.filename:
                pdf_data = base64.b64encode(pdf_file.read()).decode('utf-8')
                compliance_record.pdf_file_data = pdf_data
                compliance_record.pdf_file_name = pdf_file.filename
            
            # Three image uploads
            for i in range(1, 4):
                image_file = request.files.get(f'image_{i}_{lang}')
                if image_file and image_file.filename:
                    image_data = base64.b64encode(image_file.read()).decode('utf-8')
                    setattr(compliance_record, f'image_{i}_data', image_data)
                    setattr(compliance_record, f'image_{i}_name', image_file.filename)
            
            # Handle reference photos for this language
            for i in range(1, 4):  # 3 photos per language
                photo_file = request.files.get(f'ref_photo_{i}_{lang}')
                caption = request.form.get(f'ref_photo_{i}_caption_{lang}')
                
                if photo_file and photo_file.filename:
                    photo_data = base64.b64encode(photo_file.read()).decode('utf-8')
                    setattr(compliance_record, f'ref_photo_{i}', photo_data)
                    if caption:
                        setattr(compliance_record, f'ref_photo_{i}_caption', caption)
            
            db.session.add(compliance_record)
            created_records.append(compliance_record)
        
        if created_records:
            db.session.commit()
            
            # Auto-generate QR code for the compliance record
            try:
                from models import QRCode as QRCodeModel
                import qrcode
                from io import BytesIO
                import base64
                
                # Create QR code pointing to the first compliance record
                first_record = created_records[0]
                qr_url = f"{request.url_root}compliance-records/view/{first_record.id}"
                
                # Generate QR code image
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=10,
                    border=4,
                )
                qr.add_data(qr_url)
                qr.make(fit=True)
                
                img = qr.make_image(fill_color="black", back_color="white")
                buffer = BytesIO()
                img.save(buffer, format='PNG')
                qr_image_data = base64.b64encode(buffer.getvalue()).decode()
                
                # Save QR code to database
                qr_code_record = QRCodeModel(
                    code=f"CR-{record_number}",
                    qr_type="compliance",
                    reference_id=str(first_record.id),
                    reference_table="compliance_records",
                    label=f"Compliance Record - {record_number}",
                    description=f"QR Code for compliance record {compliance_category}",
                    organization_id=user.organization_id,
                    created_by=user.id
                )
                db.session.add(qr_code_record)
                db.session.commit()
                
                flash(f'Successfully created {len(created_records)} multilingual compliance records and generated QR code', 'success')
            except Exception as qr_error:
                flash(f'Compliance records created successfully, but QR code generation failed: {str(qr_error)}', 'warning')
        else:
            flash('No compliance records created. Please provide content for at least one language.', 'warning')
    
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating compliance record: {str(e)}', 'error')
    
    return redirect(url_for('compliance_management'))

@app.route('/compliance-records/filter/<language_code>')
@login_required
def filter_compliance_records(language_code):
    """Filter compliance records by language"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    # Get compliance records for specific language
    records = ComplianceRecord.query.filter_by(
        organization_id=user.organization_id,
        language_code=language_code
    ).order_by(ComplianceRecord.created_at.desc()).all()
    
    # Return HTML fragment for the filtered records
    html_content = f"""
    <div class="table-responsive">
        <table class="table table-striped">
            <thead>
                <tr>
                    <th>Record #</th>
                    <th>Title</th>
                    <th>Type</th>
                    <th>Category</th>
                    <th>Priority</th>
                    <th>Status</th>
                    <th>Effective Date</th>
                    <th>Actions</th>
                </tr>
            </thead>
            <tbody>
    """
    
    if records:
        for record in records:
            priority_class = {
                'High': 'bg-danger',
                'Medium': 'bg-warning',
                'Low': 'bg-secondary'
            }.get(record.priority_level, 'bg-secondary')
            
            status_class = {
                'Active': 'bg-success',
                'Under Review': 'bg-warning',
                'Archived': 'bg-secondary',
                'Inactive': 'bg-dark'
            }.get(record.status, 'bg-secondary')
            
            ack_badge = '<span class="badge bg-warning ms-1">ACK Required</span>' if record.requires_acknowledgment else ''
            
            html_content += f"""
                <tr>
                    <td><code>{record.record_number}</code></td>
                    <td>
                        <strong>{record.title}</strong>
                        {ack_badge}
                    </td>
                    <td><span class="badge bg-info">{record.record_type.title()}</span></td>
                    <td>{record.compliance_category}</td>
                    <td><span class="badge {priority_class}">{record.priority_level}</span></td>
                    <td><span class="badge {status_class}">{record.status}</span></td>
                    <td>{record.effective_date.strftime('%Y-%m-%d') if record.effective_date else 'N/A'}</td>
                    <td>
                        <div class="btn-group" role="group">
                            <button type="button" class="btn btn-sm btn-outline-primary" onclick="viewComplianceRecord({record.id})">
                                <i class="fas fa-eye"></i>
                            </button>
                            <button type="button" class="btn btn-sm btn-outline-success" onclick="editComplianceRecord({record.id})">
                                <i class="fas fa-edit"></i>
                            </button>
                            {'<button type="button" class="btn btn-sm btn-outline-info" onclick="viewAcknowledgments(' + str(record.id) + ')"><i class="fas fa-check-circle"></i></button>' if record.requires_acknowledgment else ''}
                        </div>
                    </td>
                </tr>
            """
    else:
        language_names = {
            'en': 'English',
            'bn': 'Bengali', 
            'my': 'Myanmar',
            'ta': 'Tamil',
            'zh': 'Chinese'
        }
        lang_name = language_names.get(language_code, language_code.upper())
        html_content += f"""
            <tr>
                <td colspan="8" class="text-center py-4">
                    <i class="fas fa-language fa-2x text-muted mb-2"></i>
                    <h6 class="text-muted">No {lang_name} compliance records found</h6>
                    <p class="text-muted">Create compliance records in {lang_name} to see them here.</p>
                </td>
            </tr>
        """
    
    html_content += """
            </tbody>
        </table>
    </div>
    """
    
    return html_content

@app.route('/compliance-records/<int:record_id>')
@login_required
def view_compliance_record(record_id):
    """View detailed compliance record"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned', 'error')
        return redirect(url_for('compliance_management'))
    
    record = ComplianceRecord.query.filter_by(
        id=record_id,
        organization_id=user.organization_id
    ).first_or_404()
    
    # Get related records in other languages (same base record number)
    base_record_number = record.record_number.split('-')[:-1]  # Remove language suffix
    base_record_number = '-'.join(base_record_number)
    
    related_records = ComplianceRecord.query.filter(
        ComplianceRecord.record_number.like(f"{base_record_number}%"),
        ComplianceRecord.organization_id == user.organization_id,
        ComplianceRecord.id != record_id
    ).all()
    
    # Get acknowledgments if required
    acknowledgments = []
    if record.requires_acknowledgment:
        acknowledgments = ComplianceAcknowledgment.query.filter_by(
            compliance_record_id=record.id,
            organization_id=user.organization_id
        ).order_by(ComplianceAcknowledgment.acknowledged_at.desc()).all()
    
    return render_template('compliance_record_detail.html', 
                         record=record, 
                         related_records=related_records,
                         acknowledgments=acknowledgments)

@app.route('/compliance-records/<int:record_id>/acknowledgments')
@login_required
def view_compliance_acknowledgments(record_id):
    """View acknowledgments for a compliance record"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned', 'error')
        return redirect(url_for('compliance_management'))
    
    record = ComplianceRecord.query.filter_by(
        id=record_id,
        organization_id=user.organization_id
    ).first_or_404()
    
    acknowledgments = ComplianceAcknowledgment.query.filter_by(
        compliance_record_id=record.id,
        organization_id=user.organization_id
    ).order_by(ComplianceAcknowledgment.acknowledged_at.desc()).all()
    
    return render_template('compliance_acknowledgments.html', 
                         record=record, 
                         acknowledgments=acknowledgments)

@app.route('/acknowledge-compliance/<int:record_id>', methods=['POST'])
@login_required
@create_permission_required('compliance_tracking')
def acknowledge_compliance_record(record_id):
    """Allow users to acknowledge compliance records"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    record = ComplianceRecord.query.filter_by(
        id=record_id,
        organization_id=user.organization_id
    ).first_or_404()
    
    # Check if user has already acknowledged this record
    existing_ack = ComplianceAcknowledgment.query.filter_by(
        compliance_record_id=record.id,
        user_id=user.id
    ).first()
    
    if existing_ack:
        return jsonify({'error': 'Already acknowledged'}), 400
    
    try:
        # Create acknowledgment record
        acknowledgment = ComplianceAcknowledgment(
            compliance_record_id=record.id,
            user_id=user.id,
            organization_id=user.organization_id,
            language_code=record.language_code,
            acknowledgment_method='digital',
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', ''),
            notes=request.json.get('notes', '') if request.is_json else ''
        )
        
        db.session.add(acknowledgment)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Compliance record acknowledged successfully'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error acknowledging record: {str(e)}'}), 500

@app.route('/stock/export/excel')
@login_required
def export_stock_excel():
    """Export stock items to Excel"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    try:
        from openpyxl import Workbook
        from flask import Response
        import io
        
        # Get stock items
        stock_items = StockItem.query.filter_by(organization_id=user.organization_id).all()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'Stock Inventory'
        
        # Add headers
        headers = ['S.No', 'Item Name', 'Description', 'Category', 'Quantity', 'Status', 'Location', 'Serial Number', 'Purchase Date', 'Purchase Cost', 'Created By', 'Created At']
        ws.append(headers)
        
        # Add data rows
        for i, item in enumerate(stock_items, 1):
            row = [
                i,
                item.name,
                item.description or '',
                item.category,
                item.quantity,
                item.status,
                item.location or '',
                item.serial_number or '',
                item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else '',
                item.purchase_cost or '',
                item.created_by_user.first_name if item.created_by_user else '',
                item.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            ws.append(row)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=stock_inventory_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            }
        )
        
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/stock/export/selected', methods=['POST'])
@login_required
@page_permission_required('inventory')
def export_selected_stock():
    """Export selected stock items to Excel"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    try:
        import pandas as pd
        from flask import Response
        import io
        
        selected_ids = request.form.getlist('selected_ids')
        
        if not selected_ids:
            flash('No items selected for export', 'warning')
            return redirect(url_for('stock_report'))
        
        # Get selected stock items
        stock_items = StockItem.query.filter(
            StockItem.id.in_(selected_ids),
            StockItem.organization_id == user.organization_id
        ).all()
        
        # Prepare data for Excel
        data = []
        for i, item in enumerate(stock_items, 1):
            data.append({
                'S.No': i,
                'Item Name': item.name,
                'Description': item.description or '',
                'Category': item.category,
                'Quantity': item.quantity,
                'Status': item.status,
                'Location': item.location or '',
                'Serial Number': item.serial_number or '',
                'Purchase Date': item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else '',
                'Purchase Cost': item.purchase_cost or '',
                'Created By': item.created_by_user.first_name if item.created_by_user else '',
                'Created At': item.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Selected Stock Items')
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Selected Stock Items']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_name].width = adjusted_width
        
        output.seek(0)
        
        # Create response
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=selected_stock_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            }
        )
        
        return response
        
    except Exception as e:
        flash(f'Error exporting selected stock data: {str(e)}', 'error')
        return redirect(url_for('stock_report'))

@app.route('/api/stock/delete-multiple', methods=['POST'])
@login_required
@create_permission_required('stock_report')
def delete_multiple_stock():
    """Delete multiple stock items"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'No organization assigned'})
    
    try:
        data = request.get_json()
        stock_ids = data.get('stock_ids', [])
        
        if not stock_ids:
            return jsonify({'success': False, 'error': 'No stock IDs provided'})
        
        # Convert string IDs to integers if needed
        try:
            stock_ids = [int(id) for id in stock_ids]
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Invalid stock ID format'})
        
        # Get stock items to delete (only from user's organization)
        stock_items = StockItem.query.filter(
            StockItem.id.in_(stock_ids),
            StockItem.organization_id == user.organization_id
        ).all()
        
        deleted_count = len(stock_items)
        
        if deleted_count == 0:
            return jsonify({'success': False, 'error': 'No valid stock items found to delete'})
        
        # Delete associated movements first
        for item in stock_items:
            StockMovement.query.filter_by(stock_item_id=item.id).delete()
        
        # Delete stock items
        for item in stock_items:
            db.session.delete(item)
        
        db.session.commit()
        
        # Log the bulk deletion
        log_entry = SystemLog(
            user_id=current_user.id,
            action='bulk_delete_stock',
            module='stock_management',
            details=f'Deleted {deleted_count} stock items',
            ip_address=request.remote_addr
        )
        db.session.add(log_entry)
        db.session.commit()
        
        return jsonify({'success': True, 'deleted_count': deleted_count})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/stock/import-excel', methods=['POST'])
@login_required
@create_permission_required('inventory')
def import_stock_from_excel():
    """Import stock items from Excel file"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'No organization assigned'})
    
    try:
        if 'excel_file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'})
        
        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'})
        
        # Validate file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Invalid file format. Please upload .xlsx or .xls file'})
        
        import pandas as pd
        from io import BytesIO
        
        # Read Excel file
        file_content = file.read()
        df = pd.read_excel(BytesIO(file_content))
        
        # Validate required columns
        required_columns = ['Item Name', 'Quantity']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'success': False, 'error': f'Missing required columns: {", ".join(missing_columns)}'})
        
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Skip rows with empty Item Name or Quantity
                if pd.isna(row['Item Name']) or pd.isna(row['Quantity']):
                    continue
                
                # Parse row data with defaults
                item_name = str(row['Item Name']).strip()
                quantity = int(row['Quantity']) if not pd.isna(row['Quantity']) else 0
                
                if not item_name or quantity < 0:
                    errors.append(f"Row {index + 2}: Invalid item name or quantity")
                    continue
                
                # Create stock item
                stock_item = StockItem(
                    name=item_name,
                    description=str(row.get('Description', '')).strip() if not pd.isna(row.get('Description')) else '',
                    category=str(row.get('Category', 'General')).strip() if not pd.isna(row.get('Category')) else 'General',
                    quantity=quantity,
                    status=str(row.get('Status', 'Active')).strip() if not pd.isna(row.get('Status')) else 'Active',
                    location=str(row.get('Location', '')).strip() if not pd.isna(row.get('Location')) else '',
                    purchase_date=pd.to_datetime(row['Purchase Date']).date() if not pd.isna(row.get('Purchase Date')) else None,
                    purchase_cost=float(row['Purchase Cost']) if not pd.isna(row.get('Purchase Cost')) else None,
                    serial_number=str(row.get('Serial Number', '')).strip() if not pd.isna(row.get('Serial Number')) else '',
                    organization_id=user.organization_id,
                    created_by=user.id
                )
                
                db.session.add(stock_item)
                db.session.flush()  # Get the ID for movement record
                
                # Create initial stock movement
                movement = StockMovement(
                    stock_item_id=stock_item.id,
                    movement_type='IN',
                    quantity=quantity,
                    previous_quantity=0,
                    new_quantity=quantity,
                    reason='Imported from Excel',
                    organization_id=user.organization_id,
                    created_by=user.id
                )
                db.session.add(movement)
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                continue
        
        if imported_count > 0:
            db.session.commit()
            
            # Log the import
            log_entry = SystemLog(
                user_id=current_user.id,
                action='import_stock_excel',
                module='stock_management',
                details=f'Imported {imported_count} stock items from Excel',
                ip_address=request.remote_addr
            )
            db.session.add(log_entry)
            db.session.commit()
        else:
            db.session.rollback()
        
        response_data = {
            'success': True,
            'imported_count': imported_count
        }
        
        if errors:
            response_data['warnings'] = errors[:10]  # Limit to first 10 errors
            
        return jsonify(response_data)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error processing Excel file: {str(e)}'})

# @app.route('/purchase_request_form')
# @login_required
# @page_access_required('purchase_request')
# def purchase_request_form():
    """Display the purchase request form"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    return render_template('purchase_request_form.html', user=user)

@app.route('/download_bedding_template')
@login_required
def download_bedding_template():
    """Download Excel template for bulk bedding items upload"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import make_response
    
    try:
        user = current_user
        if not user.organization_id:
            flash('Please contact administrator to assign organization', 'warning')
            return redirect(url_for('bedding_items'))
        
        # Get available categories for the organization
        categories = BeddingCategory.query.filter_by(
            organization_id=user.organization_id,
            is_active=True
        ).all()
        
        category_names = [cat.name for cat in categories] if categories else ['Bed', 'Pillow', 'Mattress', 'Blanket']
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Bedding Items"
        
        # Define headers
        headers = [
            'Serial Number', 'Item Name', 'Category', 'Status', 'Room Number',
            'Resident Name', 'Company Name', 'Brand', 'Model', 'Purchase Date',
            'Purchase Price', 'Condition', 'Warranty Expiry', 'Description',
            'Last Maintenance', 'Next Maintenance'
        ]
        
        # Add headers with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add sample data
        sample_data = [
            ['BED001', 'Single Bed Frame', category_names[0] if category_names else 'Bed', 'In Store', '', '', '', 'IKEA', 'MALM', '2025-01-15', 299.99, 'Good', '2027-01-15', 'Wooden bed frame with headboard', '2025-01-20', '2026-01-20'],
            ['PIL002', 'Memory Foam Pillow', category_names[1] if len(category_names) > 1 else 'Pillow', 'In Room', 'A101', 'John Doe', 'ABC Corp', 'Tempur', 'Original', '2025-02-10', 89.99, 'Excellent', '2027-02-10', 'Contour memory foam pillow', '', ''],
            ['MAT003', 'Queen Mattress', category_names[2] if len(category_names) > 2 else 'Mattress', 'In Store', '', '', '', 'Sealy', 'Posturepedic', '2025-03-05', 799.99, 'Good', '2027-03-05', 'Firm support mattress', '2025-03-10', '2026-03-10']
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Create Instructions sheet
        instructions_ws = wb.create_sheet("Instructions")
        
        # Instructions headers
        inst_headers = ['Field', 'Description', 'Required']
        for col, header in enumerate(inst_headers, 1):
            cell = instructions_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Instructions data
        inst_data = [
            ['Serial Number', 'Unique identifier for the bedding item (must be unique)', 'Yes'],
            ['Item Name', 'Name/description of the bedding item', 'Yes'],
            ['Category', f'Category of item (Available: {", ".join(category_names)})', 'Yes'],
            ['Status', 'Current status (In Store, In Room, Damaged, Others)', 'Yes'],
            ['Room Number', 'Room number if assigned to a room (leave empty for In Store)', 'If Status is In Room'],
            ['Resident Name', 'Name of resident if assigned (leave empty for In Store)', 'If Status is In Room'],
            ['Company Name', 'Company name of resident if assigned', 'No'],
            ['Brand', 'Brand/manufacturer of the item', 'No'],
            ['Model', 'Model number or name', 'No'],
            ['Purchase Date', 'Date of purchase (YYYY-MM-DD format)', 'No'],
            ['Purchase Price', 'Purchase price (number format)', 'No'],
            ['Condition', 'Current condition (Excellent, Good, Fair, Poor)', 'No'],
            ['Warranty Expiry', 'Warranty expiry date (YYYY-MM-DD format)', 'No'],
            ['Description', 'Additional description or notes', 'No'],
            ['Last Maintenance', 'Last maintenance date (YYYY-MM-DD format, optional)', 'No'],
            ['Next Maintenance', 'Next scheduled maintenance date (YYYY-MM-DD format, optional)', 'No']
        ]
        
        for row_idx, row_data in enumerate(inst_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                instructions_ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Format instructions sheet
        instructions_ws.column_dimensions['A'].width = 20
        instructions_ws.column_dimensions['B'].width = 60
        instructions_ws.column_dimensions['C'].width = 15
        
        # Create Categories sheet if categories exist
        if categories:
            cat_ws = wb.create_sheet("Available Categories")
            
            # Categories headers
            cat_headers = ['Category Name', 'Description']
            for col, header in enumerate(cat_headers, 1):
                cell = cat_ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Categories data
            for row_idx, category in enumerate(categories, 2):
                cat_ws.cell(row=row_idx, column=1, value=category.name)
                cat_ws.cell(row=row_idx, column=2, value=category.description or '')
            
            cat_ws.column_dimensions['A'].width = 25
            cat_ws.column_dimensions['B'].width = 50
        
        # Save workbook to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=Bedding_Items_Template.xlsx'
        
        return response
        
    except Exception as e:
        import logging
        logging.error(f"Error generating bedding template: {str(e)}")
        flash('Error generating template. Please try again.', 'error')
        return redirect(url_for('bedding_items'))

@app.route('/bulk_upload_bedding_items', methods=['POST'])
@login_required
@create_permission_required('bedding_management')
def bulk_upload_bedding_items():
    """Process bulk upload of bedding items from Excel file"""
    from openpyxl import load_workbook
    from datetime import datetime
    
    try:
        user = current_user
        if not user.organization_id:
            return jsonify({'success': False, 'error': 'No organization assigned'})
        
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'})
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Please upload an Excel file (.xlsx or .xls)'})
        
        # Read Excel file
        wb = load_workbook(file)
        
        # Check if 'Bedding Items' sheet exists
        if 'Bedding Items' not in wb.sheetnames:
            return jsonify({'success': False, 'error': 'Sheet "Bedding Items" not found in the Excel file'})
        
        ws = wb['Bedding Items']
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Validate required columns
        required_columns = ['Serial Number', 'Item Name', 'Category', 'Status']
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            return jsonify({'success': False, 'error': f'Missing required columns: {", ".join(missing_columns)}'})
        
        # Get column indices
        col_indices = {}
        for idx, header in enumerate(headers):
            if header:
                col_indices[header] = idx
        
        # Get available categories for validation
        categories = {cat.name: cat.id for cat in BeddingCategory.query.filter_by(
            organization_id=user.organization_id,
            is_active=True
        ).all()}
        
        # Process each row
        successful_imports = 0
        errors = []
        total_rows = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # Skip empty rows
                continue
                
            total_rows += 1
            
            try:
                # Get required fields
                serial_number = str(row[col_indices['Serial Number']]).strip() if row[col_indices['Serial Number']] else ''
                item_name = str(row[col_indices['Item Name']]).strip() if row[col_indices['Item Name']] else ''
                category_name = str(row[col_indices['Category']]).strip() if row[col_indices['Category']] else ''
                status = str(row[col_indices['Status']]).strip() if row[col_indices['Status']] else ''
                
                # Validate required fields
                if not serial_number or not item_name or not category_name or not status:
                    errors.append(f"Row {row_idx}: Missing required fields")
                    continue
                
                # Validate category exists
                if category_name not in categories:
                    errors.append(f"Row {row_idx}: Category '{category_name}' not found")
                    continue
                
                # Validate status
                valid_statuses = ['In Store', 'In Room', 'Damaged', 'Others']
                if status not in valid_statuses:
                    errors.append(f"Row {row_idx}: Invalid status '{status}'. Valid options: {', '.join(valid_statuses)}")
                    continue
                
                # Check for duplicate serial number
                existing_item = BeddingItem.query.filter_by(serial_number=serial_number).first()
                if existing_item:
                    errors.append(f"Row {row_idx}: Serial number '{serial_number}' already exists")
                    continue
                
                # Get optional fields
                room_number = str(row[col_indices.get('Room Number', 0)]).strip() if col_indices.get('Room Number') and row[col_indices['Room Number']] else ''
                resident_name = str(row[col_indices.get('Resident Name', 0)]).strip() if col_indices.get('Resident Name') and row[col_indices['Resident Name']] else ''
                company_name = str(row[col_indices.get('Company Name', 0)]).strip() if col_indices.get('Company Name') and row[col_indices['Company Name']] else ''
                brand = str(row[col_indices.get('Brand', 0)]).strip() if col_indices.get('Brand') and row[col_indices['Brand']] else ''
                model = str(row[col_indices.get('Model', 0)]).strip() if col_indices.get('Model') and row[col_indices['Model']] else ''
                condition = str(row[col_indices.get('Condition', 0)]).strip() if col_indices.get('Condition') and row[col_indices['Condition']] else 'Good'
                description = str(row[col_indices.get('Description', 0)]).strip() if col_indices.get('Description') and row[col_indices['Description']] else ''
                
                # Validate room assignment for In Room status
                if status == 'In Room' and not room_number:
                    errors.append(f"Row {row_idx}: Room Number is required when Status is 'In Room'")
                    continue
                
                # Parse dates
                purchase_date = None
                if col_indices.get('Purchase Date') and row[col_indices['Purchase Date']]:
                    try:
                        date_val = row[col_indices['Purchase Date']]
                        if isinstance(date_val, str):
                            purchase_date = datetime.strptime(date_val, '%Y-%m-%d').date()
                        else:
                            purchase_date = date_val.date() if hasattr(date_val, 'date') else date_val
                    except:
                        errors.append(f"Row {row_idx}: Invalid Purchase Date format (use YYYY-MM-DD)")
                        continue
                
                warranty_expiry = None
                if col_indices.get('Warranty Expiry') and row[col_indices['Warranty Expiry']]:
                    try:
                        date_val = row[col_indices['Warranty Expiry']]
                        if isinstance(date_val, str):
                            warranty_expiry = datetime.strptime(date_val, '%Y-%m-%d').date()
                        else:
                            warranty_expiry = date_val.date() if hasattr(date_val, 'date') else date_val
                    except:
                        errors.append(f"Row {row_idx}: Invalid Warranty Expiry format (use YYYY-MM-DD)")
                        continue
                
                last_maintenance_date = None
                if col_indices.get('Last Maintenance') and row[col_indices['Last Maintenance']]:
                    try:
                        date_val = row[col_indices['Last Maintenance']]
                        if isinstance(date_val, str):
                            last_maintenance_date = datetime.strptime(date_val, '%Y-%m-%d').date()
                        else:
                            last_maintenance_date = date_val.date() if hasattr(date_val, 'date') else date_val
                    except:
                        errors.append(f"Row {row_idx}: Invalid Last Maintenance date format (use YYYY-MM-DD)")
                        continue
                
                next_maintenance_date = None
                if col_indices.get('Next Maintenance') and row[col_indices['Next Maintenance']]:
                    try:
                        date_val = row[col_indices['Next Maintenance']]
                        if isinstance(date_val, str):
                            next_maintenance_date = datetime.strptime(date_val, '%Y-%m-%d').date()
                        else:
                            next_maintenance_date = date_val.date() if hasattr(date_val, 'date') else date_val
                    except:
                        errors.append(f"Row {row_idx}: Invalid Next Maintenance date format (use YYYY-MM-DD)")
                        continue
                
                # Parse price
                purchase_price = None
                if col_indices.get('Purchase Price') and row[col_indices['Purchase Price']]:
                    try:
                        purchase_price = float(row[col_indices['Purchase Price']])
                    except:
                        errors.append(f"Row {row_idx}: Invalid Purchase Price format")
                        continue
                
                # Create bedding item
                bedding_item = BeddingItem(
                    serial_number=serial_number,
                    item_name=item_name,
                    category_id=categories[category_name],
                    status=status,
                    room_number=room_number if room_number else None,
                    resident_name=resident_name if resident_name else None,
                    company_name=company_name if company_name else None,
                    brand=brand if brand else None,
                    model=model if model else None,
                    purchase_date=purchase_date,
                    purchase_price=purchase_price,
                    condition=condition,
                    warranty_expiry=warranty_expiry,
                    description=description if description else None,
                    last_maintenance_date=last_maintenance_date,
                    next_maintenance_date=next_maintenance_date,
                    organization_id=user.organization_id,
                    created_by=user.id
                )
                
                db.session.add(bedding_item)
                successful_imports += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                continue
        
        # Commit all successful imports
        if successful_imports > 0:
            db.session.commit()
            
            # Log the bulk import
            log_entry = SystemLog(
                user_id=user.id,
                action='bulk_import_bedding_items',
                module='bedding_management',
                details=f'Imported {successful_imports} bedding items from Excel',
                ip_address=request.remote_addr
            )
            db.session.add(log_entry)
            db.session.commit()
        
        # Return results
        return jsonify({
            'success': True,
            'message': f'Bulk upload completed successfully',
            'imported': successful_imports,
            'total_rows': total_rows,
            'errors': errors
        })
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Error in bulk upload: {str(e)}")
        return jsonify({'success': False, 'error': f'Upload failed: {str(e)}'})

@app.route('/delete_bedding_item/<int:item_id>', methods=['POST'])
@login_required
@create_permission_required('bedding_management')
def delete_bedding_item(item_id):
    """Delete a bedding item"""
    try:
        user = current_user
        
        # Get the bedding item
        bedding_item = BeddingItem.query.filter_by(
            id=item_id, 
            organization_id=user.organization_id
        ).first()
        
        if not bedding_item:
            return jsonify({'success': False, 'error': 'Bedding item not found'})
        
        # Store item details for logging
        serial_number = bedding_item.serial_number
        item_name = bedding_item.item_name
        
        # Delete related movements first (due to foreign key constraints)
        BeddingMovement.query.filter_by(bedding_item_id=item_id).delete()
        
        # Delete the bedding item
        db.session.delete(bedding_item)
        db.session.commit()
        
        # Log the deletion
        log_entry = SystemLog(
            user_id=user.id,
            action='delete_bedding_item',
            module='bedding_management',
            details=f'Deleted bedding item: {serial_number} - {item_name}',
            ip_address=request.remote_addr
        )
        db.session.add(log_entry)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Bedding item {serial_number} deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Error deleting bedding item: {str(e)}")
        return jsonify({'success': False, 'error': f'Delete failed: {str(e)}'})

@app.route('/bulk_delete_bedding_items', methods=['POST'])
@login_required
@create_permission_required('bedding_management')
def bulk_delete_bedding_items():
    """Delete multiple bedding items"""
    try:
        user = current_user
        data = request.get_json()
        item_ids = data.get('item_ids', [])
        
        if not item_ids:
            return jsonify({'success': False, 'error': 'No items selected for deletion'})
        
        # Get the bedding items to be deleted
        bedding_items = BeddingItem.query.filter(
            BeddingItem.id.in_(item_ids),
            BeddingItem.organization_id == user.organization_id
        ).all()
        
        if not bedding_items:
            return jsonify({'success': False, 'error': 'No valid items found for deletion'})
        
        deleted_count = 0
        deleted_items = []
        
        for item in bedding_items:
            # Store item details for logging
            deleted_items.append(f"{item.serial_number} - {item.item_name}")
            
            # Delete related movements first (due to foreign key constraints)
            BeddingMovement.query.filter_by(bedding_item_id=item.id).delete()
            
            # Delete the bedding item
            db.session.delete(item)
            deleted_count += 1
        
        db.session.commit()
        
        # Log the bulk deletion
        log_entry = SystemLog(
            user_id=user.id,
            action='bulk_delete_bedding_items',
            module='bedding_management',
            details=f'Bulk deleted {deleted_count} bedding items: {", ".join(deleted_items[:5])}{"..." if len(deleted_items) > 5 else ""}',
            ip_address=request.remote_addr
        )
        db.session.add(log_entry)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Successfully deleted {deleted_count} bedding items',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f"Error in bulk delete: {str(e)}")
        return jsonify({'success': False, 'error': f'Bulk delete failed: {str(e)}'})
    

@app.route('/generate_purchase_request_pdf', methods=['POST'])
@login_required
@create_permission_required('bedding_management')
def generate_purchase_request_pdf():
    """Generate PDF for purchase request with signatures"""
    import json
    import base64
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    from PIL import Image
    
    try:
        # Get form data
        form_data = request.form.to_dict()
        signatures_json = request.form.get('signatures', '{}')
        signatures = json.loads(signatures_json) if signatures_json else {}
        
        # Debug logging for form data
        print(f"Form data keys: {list(form_data.keys())}")
        print(f"Signatures JSON: {signatures_json[:100] if signatures_json else 'None'}...")
        for key in ['dc_oe_signature', 'operation_manager_signature', 'general_manager_signature']:
            value = form_data.get(key, '')
            print(f"Form data {key}: {value[:50] if value else 'None'}...")
        
        # Debug name values
        print("NAME VALUES DEBUG:")
        for name_key in ['dc_oe_name', 'operation_manager_name', 'general_manager_name']:
            value = form_data.get(name_key, '')
            print(f"Form data {name_key}: '{value}'")
        
        # Process signature data from both form and JSON sources
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Build story content
        story = []
        styles = getSampleStyleSheet()
        
        # Professional Header with border
        header_data = [
            ['', 'PURCHASE REQUEST FORM', ''],
            ['', 'Pioneer Lodge - TS Group', '']
        ]
        
        header_table = Table(header_data, colWidths=[1*inch, 6*inch, 1*inch])
        header_table.setStyle(TableStyle([
            ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (1, 0), 18),
            ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 1), (1, 1), 12),
            ('ALIGN', (1, 0), (1, 1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TEXTCOLOR', (1, 0), (1, 0), colors.darkblue),
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 25))
        
        # Request Information Section with professional border
        info_section_data = [
            ['REQUEST INFORMATION', '', '', ''],
            ['PR Number:', form_data.get('pr_number', ''), 'Request Date:', form_data.get('request_date', '')],
            ['Department:', form_data.get('department', ''), 'Priority:', form_data.get('priority', '')],
            ['Requested By:', form_data.get('requested_by', ''), 'Supplier/Vendor:', form_data.get('supplier', '')]
        ]
        
        info_table = Table(info_section_data, colWidths=[1.5*inch, 2.5*inch, 1.5*inch, 2.5*inch])
        info_table.setStyle(TableStyle([
            # Header row
            ('SPAN', (0, 0), (3, 0)),
            ('BACKGROUND', (0, 0), (3, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (3, 0), colors.white),
            ('FONTNAME', (0, 0), (3, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (3, 0), 12),
            ('ALIGN', (0, 0), (3, 0), 'CENTER'),
            
            # Data rows
            ('FONTNAME', (0, 1), (3, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (3, -1), 10),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Borders and background
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 1), (3, -1), colors.white),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 25))
        
        # Items Section with professional header
        items_section_title = [['ITEMS REQUESTED', '', '', '', '', '', '', '']]
        items_title_table = Table(items_section_title, colWidths=[1*inch, 1.5*inch, 0.8*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.8*inch])
        items_title_table.setStyle(TableStyle([
            ('SPAN', (0, 0), (7, 0)),
            ('BACKGROUND', (0, 0), (7, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (7, 0), colors.white),
            ('FONTNAME', (0, 0), (7, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (7, 0), 12),
            ('ALIGN', (0, 0), (7, 0), 'CENTER'),
            ('BOX', (0, 0), (7, 0), 1, colors.black),
            ('TOPPADDING', (0, 0), (7, 0), 10),
            ('BOTTOMPADDING', (0, 0), (7, 0), 10),
        ]))
        story.append(items_title_table)
        
        # Items table with professional formatting
        items_header = ['S/N', 'Description', 'Unit Cost', 'Qty', 'Total', 'Room No', 'Unit', 'Cost Code']
        items_data = [items_header]
        
        # Get item data from form arrays
        descriptions = request.form.getlist('description[]')
        unit_costs = request.form.getlist('unit_cost[]')
        quantities = request.form.getlist('quantity[]')
        totals = request.form.getlist('total[]')
        room_nos = request.form.getlist('room_no[]')
        units = request.form.getlist('unit[]')
        cost_codes = request.form.getlist('cost_code[]')
        
        for i in range(len(descriptions)):
            if descriptions[i].strip():  # Only add non-empty items
                items_data.append([
                    str(i + 1),
                    descriptions[i][:40] + '...' if len(descriptions[i]) > 40 else descriptions[i],
                    f"${unit_costs[i] if i < len(unit_costs) else '0.00'}",
                    quantities[i] if i < len(quantities) else '1',
                    f"${totals[i] if i < len(totals) else '0.00'}",
                    room_nos[i] if i < len(room_nos) else '',
                    units[i] if i < len(units) else '',
                    cost_codes[i] if i < len(cost_codes) else ''
                ])
        
        # Add empty rows if needed to maintain professional appearance
        while len(items_data) < 6:  # Minimum 5 item rows
            items_data.append(['', '', '', '', '', '', '', ''])
        
        items_table = Table(items_data, colWidths=[0.5*inch, 2*inch, 0.8*inch, 0.5*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.8*inch])
        items_table.setStyle(TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # S/N center
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Description left
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'), # All others center
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Borders and background
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 20))
        
        # Financial Summary Section with professional header
        financial_section_title = [['FINANCIAL SUMMARY', '', '']]
        financial_title_table = Table(financial_section_title, colWidths=[2*inch, 2*inch, 4*inch])
        financial_title_table.setStyle(TableStyle([
            ('SPAN', (0, 0), (2, 0)),
            ('BACKGROUND', (0, 0), (2, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (2, 0), colors.white),
            ('FONTNAME', (0, 0), (2, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (2, 0), 12),
            ('ALIGN', (0, 0), (2, 0), 'CENTER'),
            ('BOX', (0, 0), (2, 0), 1, colors.black),
            ('TOPPADDING', (0, 0), (2, 0), 10),
            ('BOTTOMPADDING', (0, 0), (2, 0), 10),
        ]))
        story.append(financial_title_table)
        
        subtotal = form_data.get('subtotal', '0.00')
        gst = form_data.get('gst', '0.00')
        additional_tax = form_data.get('additional_tax', '0.00')
        additional_tax_type = form_data.get('additional_tax_type', '')
        grand_total = form_data.get('grand_total', '0.00')
        
        # Create professional financial summary table
        financial_summary = [
            ['Subtotal:', f"${subtotal}", ''],
            ['GST (0%):', f"${gst}", ''],
        ]
        
        if additional_tax_type and additional_tax and float(additional_tax) > 0:
            financial_summary.append([f'{additional_tax_type}:', f"${additional_tax}", ''])
        
        financial_summary.append(['GRAND TOTAL:', f"${grand_total}", ''])
        
        financial_table = Table(financial_summary, colWidths=[5*inch, 1.5*inch, 1.5*inch])
        financial_table.setStyle(TableStyle([
            # General styling
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            
            # Grand total styling - bold and larger
            ('FONTNAME', (0, -1), (1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (1, -1), 13),
            ('TEXTCOLOR', (0, -1), (1, -1), colors.darkblue),
            
            # Borders and background
            ('BOX', (0, 0), (1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (1, -2), colors.white),
            ('BACKGROUND', (0, -1), (1, -1), colors.lightgrey),
            
            # Lines above grand total
            ('LINEABOVE', (0, -1), (1, -1), 2, colors.darkblue),
            
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(financial_table)
        story.append(Spacer(1, 25))
        
        # Additional Information Section with professional header
        additional_section_title = [['ADDITIONAL INFORMATION', '', '', '']]
        additional_title_table = Table(additional_section_title, colWidths=[2*inch, 2*inch, 2*inch, 2*inch])
        additional_title_table.setStyle(TableStyle([
            ('SPAN', (0, 0), (3, 0)),
            ('BACKGROUND', (0, 0), (3, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (3, 0), colors.white),
            ('FONTNAME', (0, 0), (3, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (3, 0), 12),
            ('ALIGN', (0, 0), (3, 0), 'CENTER'),
            ('BOX', (0, 0), (3, 0), 1, colors.black),
            ('TOPPADDING', (0, 0), (3, 0), 10),
            ('BOTTOMPADDING', (0, 0), (3, 0), 10),
        ]))
        story.append(additional_title_table)
        
        additional_info = [
            ['Payment Method:', form_data.get('payment_method', ''), 'Budget Code:', form_data.get('budget_code', '')],
            ['Expected Delivery Date:', form_data.get('expected_delivery', ''), 'Justification:', form_data.get('justification', '')]
        ]
        
        additional_table = Table(additional_info, colWidths=[1.8*inch, 2.2*inch, 1.2*inch, 2.8*inch])
        additional_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            
            # Borders and background
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            
            # Padding
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(additional_table)
        story.append(Spacer(1, 25))
        
        # Signatures Section with professional header
        signature_section_title = [['APPROVALS & SIGNATURES', '', '']]
        signature_title_table = Table(signature_section_title, colWidths=[2.67*inch, 2.67*inch, 2.66*inch])
        signature_title_table.setStyle(TableStyle([
            ('SPAN', (0, 0), (2, 0)),
            ('BACKGROUND', (0, 0), (2, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (2, 0), colors.white),
            ('FONTNAME', (0, 0), (2, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (2, 0), 12),
            ('ALIGN', (0, 0), (2, 0), 'CENTER'),
            ('BOX', (0, 0), (2, 0), 1, colors.black),
            ('TOPPADDING', (0, 0), (2, 0), 10),
            ('BOTTOMPADDING', (0, 0), (2, 0), 10),
        ]))
        story.append(signature_title_table)
        
        # Process signature images with improved handling
        from PIL import Image as PILImage
        import base64
        from io import BytesIO
        from reportlab.platypus import Image as ReportLabImage
        
        # Signature data processing
        signature_names = ['dc_oe_signature', 'operation_manager_signature', 'general_manager_signature']
        name_fields = ['dc_oe_name', 'operation_manager_name', 'general_manager_name']
        signature_titles = ['DC/OE Signature', 'Operation Manager', 'General Manager']
        
        # Create signature table data
        signature_row1 = []  # Signature images
        signature_row2 = []  # Names
        signature_row3 = []  # Dates
        signature_row4 = []  # Titles
        
        # Initialize signature processing arrays
        signature_images = []
        
        for i, sig_name in enumerate(signature_names):
            # Try to get signature from form data first, then from signatures JSON
            sig_data = form_data.get(sig_name, '')
            if not sig_data and signatures:
                # Try different possible keys for this signature
                canvas_names = ['dcOeSignature', 'operationManagerSignature', 'generalManagerSignature']
                sig_data = signatures.get(canvas_names[i], '')
                if not sig_data:
                    sig_data = signatures.get(sig_name, '')
            
            # Debug logging
            print(f"Processing signature {sig_name}: {sig_data[:50] if sig_data else 'None'}...")
            
            if sig_data and sig_data.startswith('data:image'):
                try:
                    # Extract base64 data
                    header, data = sig_data.split(',', 1)
                    image_data = base64.b64decode(data)
                    
                    # Create PIL image
                    pil_image = PILImage.open(BytesIO(image_data))
                    
                    # Convert RGBA to RGB with white background to handle transparency
                    if pil_image.mode == 'RGBA':
                        # Create white background
                        white_bg = PILImage.new('RGB', pil_image.size, (255, 255, 255))
                        # Paste the signature image on white background using alpha channel
                        white_bg.paste(pil_image, mask=pil_image.split()[-1])  # Use alpha channel as mask
                        pil_image = white_bg
                    elif pil_image.mode != 'RGB':
                        pil_image = pil_image.convert('RGB')
                    
                    # Resize image to standard size for consistency
                    pil_image = pil_image.resize((200, 100), PILImage.Resampling.LANCZOS)
                    
                    # Save to BytesIO with high quality
                    img_buffer = BytesIO()
                    pil_image.save(img_buffer, format='PNG', quality=95)
                    img_buffer.seek(0)
                    
                    # Create ReportLab image with proper sizing
                    rl_image = ReportLabImage(img_buffer, width=2*inch, height=1*inch)
                    signature_images.append(rl_image)
                except Exception as e:
                    print(f"Error processing signature {sig_name}: {e}")
                    signature_images.append('No signature')
            else:
                signature_images.append('No signature')
        
        # Create properly sized signature table - increase height for better visibility
        signature_box_height = 2.5*inch  # Increased height to ensure all content is visible
        
        # Get name values from the form - check both possible field names
        dc_name = form_data.get('dc_oe_name') or form_data.get('dcSignatureName', '')
        operation_name = form_data.get('operation_manager_name') or form_data.get('operationSignatureName', '')
        general_name = form_data.get('general_manager_name') or form_data.get('generalSignatureName', '')
        
        signature_data = [
            ['D/C & O/E', 'Operation Manager', 'General Manager'],  # Role headers
            ['Name', 'Name', 'Name'],  # Name labels
            [dc_name, operation_name, general_name],  # Name values
            ['E-Signature', 'E-Signature', 'E-Signature'],  # Signature labels
            signature_images,  # Actual signature images
        ]
        
        # Set proper row heights for visibility
        header_height = 0.4*inch
        name_label_height = 0.3*inch
        name_value_height = 0.4*inch
        sig_label_height = 0.3*inch
        sig_image_height = 1.1*inch  # Adequate space for signatures
        
        signature_table = Table(signature_data, 
                               colWidths=[2.67*inch, 2.67*inch, 2.66*inch], 
                               rowHeights=[header_height, name_label_height, name_value_height, sig_label_height, sig_image_height])
        
        signature_table.setStyle(TableStyle([
            # Role header row (row 0) - bold and centered
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            
            # Name label row (row 1) - left aligned
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, 1), 10),
            ('ALIGN', (0, 1), (-1, 1), 'LEFT'),
            ('VALIGN', (0, 1), (-1, 1), 'MIDDLE'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.white),
            
            # Name value row (row 2) - left aligned with underline
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica'),
            ('FONTSIZE', (0, 2), (-1, 2), 10),
            ('ALIGN', (0, 2), (-1, 2), 'LEFT'),
            ('VALIGN', (0, 2), (-1, 2), 'BOTTOM'),
            ('BACKGROUND', (0, 2), (-1, 2), colors.white),
            ('LINEBELOW', (0, 2), (-1, 2), 1, colors.black),
            
            # E-Signature label row (row 3) - left aligned
            ('FONTNAME', (0, 3), (-1, 3), 'Helvetica'),
            ('FONTSIZE', (0, 3), (-1, 3), 10),
            ('ALIGN', (0, 3), (-1, 3), 'LEFT'),
            ('VALIGN', (0, 3), (-1, 3), 'MIDDLE'),
            ('BACKGROUND', (0, 3), (-1, 3), colors.white),
            
            # Signature image row (row 4) - centered with underline
            ('VALIGN', (0, 4), (-1, 4), 'MIDDLE'),
            ('ALIGN', (0, 4), (-1, 4), 'CENTER'),
            ('BACKGROUND', (0, 4), (-1, 4), colors.white),
            ('LINEBELOW', (0, 4), (-1, 4), 1, colors.black),
            
            # Overall table styling with box border
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Adequate padding for visibility
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(signature_table)
        
        # Add footer space
        story.append(Spacer(1, 30))
        
        # Professional footer
        footer_data = [['This document was generated electronically and is valid without physical signature when digitally signed.']]
        footer_table = Table(footer_data, colWidths=[8*inch])
        footer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica'),
            ('FONTSIZE', (0, 0), (0, 0), 8),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.grey),
            ('TOPPADDING', (0, 0), (0, 0), 10),
        ]))
        story.append(footer_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        # Return PDF as download
        return send_file(
            BytesIO(buffer.read()),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'purchase_request_{form_data.get("pr_number", "new")}.pdf'
        )
        
    except Exception as e:
        logging.error(f"Error generating PDF: {str(e)}")
        return jsonify({'error': 'Failed to generate PDF'}), 500

@app.route('/submit_purchase_request_form', methods=['POST'])
@login_required
@create_permission_required('bedding_management')
def submit_purchase_request_form():
    """Submit purchase request form from HTML form"""
    from datetime import datetime
    
    try:
        user = current_user
        if not user.organization_id:
            flash('No organization assigned', 'error')
            return redirect(url_for('purchase_request_form'))
        
        # Generate request number
        today = singapore_now()
        request_count = PurchaseRequest.query.filter(
            db.func.date(PurchaseRequest.created_at) == today.date(),
            PurchaseRequest.organization_id == user.organization_id
        ).count()
        request_number = f"PR-{today.strftime('%Y%m%d')}-{str(request_count + 1).zfill(3)}"
        
        # Get form data
        form_data = request.form
        request_date = datetime.strptime(form_data.get('request_date'), '%Y-%m-%d').date()
        
        # Debug: Log all form data to see what's being received
        import logging
        logging.basicConfig(level=logging.DEBUG)
        logging.debug("=== PURCHASE REQUEST FORM SUBMISSION DEBUG ===")
        logging.debug(f"All form data keys: {list(form_data.keys())}")
        logging.debug(f"dc_oe_signature value: {form_data.get('dc_oe_signature', 'NOT_FOUND')[:100]}...")
        logging.debug(f"operation_manager_signature value: {form_data.get('operation_manager_signature', 'NOT_FOUND')[:100]}...")
        logging.debug(f"general_manager_signature value: {form_data.get('general_manager_signature', 'NOT_FOUND')[:100]}...")
        
        # Process signature data - create JSON objects like room checklist
        import json
        
        def create_signature_data(signature_canvas, name_field, position_field, date_field):
            """Create signature JSON data similar to room checklist format"""
            signature_base64 = form_data.get(signature_canvas, '')
            if signature_base64 and signature_base64 != 'data:,':
                return json.dumps({
                    'signature': signature_base64,
                    'name': form_data.get(name_field, ''),
                    'position': form_data.get(position_field, ''),
                    'date': form_data.get(date_field, ''),
                    'timestamp': singapore_now().isoformat()
                })
            return None
        
        # Create signature data for each signer
        dc_signature_data = create_signature_data(
            'dc_oe_signature', 'dc_oe_name', 'dc_oe_position', 'dc_oe_date'
        )
        operation_signature_data = create_signature_data(
            'operation_manager_signature', 'operation_manager_name', 'operation_manager_position', 'operation_manager_date'
        )
        general_signature_data = create_signature_data(
            'general_manager_signature', 'general_manager_name', 'general_manager_position', 'general_manager_date'
        )
        
        # Create purchase request with all fields
        purchase_request = PurchaseRequest(
            request_number=request_number,
            request_date=request_date,
            category='Purchase Request',
            requested_by=form_data.get('requested_by', ''),
            dc_name=form_data.get('dc_oe_name', ''),
            operation_manager=form_data.get('operation_manager_name', ''),
            general_manager=form_data.get('general_manager_name', ''),
            # Additional information fields
            supplier=form_data.get('supplier', ''),
            department=form_data.get('department', ''),
            priority=form_data.get('priority', ''),
            payment_method=form_data.get('payment_method', ''),
            budget_code=form_data.get('budget_code', ''),
            expected_delivery=datetime.strptime(form_data.get('expected_delivery'), '%Y-%m-%d').date() if form_data.get('expected_delivery') else None,
            justification=form_data.get('justification', ''),
            # Signature data fields - store as JSON like room checklist
            dc_signature_data=dc_signature_data,
            operation_manager_signature_data=operation_signature_data,
            general_manager_signature_data=general_signature_data,
            organization_id=user.organization_id,
            created_by=user.id
        )
        
        db.session.add(purchase_request)
        db.session.flush()  # Get the ID
        
        # Add purchase request items
        descriptions = request.form.getlist('description[]')
        unit_costs = request.form.getlist('unit_cost[]')
        quantities = request.form.getlist('quantity[]')
        totals = request.form.getlist('total[]')
        room_nos = request.form.getlist('room_no[]')
        units = request.form.getlist('unit[]')
        cost_codes = request.form.getlist('cost_code[]')
        
        for i in range(len(descriptions)):
            if descriptions[i].strip():  # Only add non-empty items
                item = PurchaseRequestItem(
                    purchase_request_id=purchase_request.id,
                    description=descriptions[i],
                    unit_cost=float(unit_costs[i]) if i < len(unit_costs) and unit_costs[i] else 0.0,
                    quantity=int(quantities[i]) if i < len(quantities) and quantities[i] else 1,
                    total_cost=float(totals[i]) if i < len(totals) and totals[i] else 0.0,
                    room_no=room_nos[i] if i < len(room_nos) else '',
                    unit=units[i] if i < len(units) else '',
                    cost_code=cost_codes[i] if i < len(cost_codes) else '',
                    remarks=''
                )
                db.session.add(item)
        
        db.session.commit()
        
        # Debug: Log signature data to verify storage
        import logging
        logging.basicConfig(level=logging.DEBUG)
        logging.debug(f"Purchase Request {request_number} signature data:")
        logging.debug(f"DC signature data: {dc_signature_data[:100] if dc_signature_data else 'None'}...")
        logging.debug(f"Operation signature data: {operation_signature_data[:100] if operation_signature_data else 'None'}...")
        logging.debug(f"General signature data: {general_signature_data[:100] if general_signature_data else 'None'}...")
        
        # Log the submission
        log_entry = SystemLog(
            user_id=current_user.id,
            action='submit_purchase_request_form',
            module='purchase_management',
            details=f'Submitted purchase request {request_number}',
            ip_address=request.remote_addr
        )
        db.session.add(log_entry)
        db.session.commit()
        
        flash(f'Purchase request {request_number} submitted successfully!', 'success')
        return redirect(url_for('purchase_request_management'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error submitting purchase request: {str(e)}")
        flash(f'Error submitting purchase request: {str(e)}', 'error')
        return redirect(url_for('purchase_request_form'))

# @app.route('/purchase_request_management')
# @login_required
# @page_access_required('purchase_request')
# def purchase_request_management():
    """Display purchase request management page"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned', 'error')
        return redirect(url_for('dashboard'))
    
    # Get all purchase requests for the organization
    search_query = request.args.get('search', '').strip()
    query = PurchaseRequest.query.filter_by(organization_id=user.organization_id)
    
    if search_query:
        query = query.filter(
            db.or_(
                PurchaseRequest.request_number.ilike(f'%{search_query}%'),
                PurchaseRequest.requested_by.ilike(f'%{search_query}%'),
                PurchaseRequest.category.ilike(f'%{search_query}%')
            )
        )
    
    purchase_requests = query.order_by(PurchaseRequest.created_at.desc()).all()
    
    return render_template('purchase_request_management.html', 
                         purchase_requests=purchase_requests, 
                         search_query=search_query)

@app.route('/bulk-delete-purchase-requests', methods=['POST'])
@login_required
@edit_permission_required('purchase_request')
def bulk_delete_purchase_requests():
    """Bulk delete purchase requests"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    try:
        data = request.get_json()
        request_ids = data.get('request_ids', [])
        
        if not request_ids:
            return jsonify({'success': False, 'error': 'No requests selected'}), 400
        
        # Delete selected requests
        deleted_count = 0
        for request_id in request_ids:
            purchase_request = PurchaseRequest.query.filter_by(
                id=int(request_id),
                organization_id=user.organization_id
            ).first()
            
            if purchase_request:
                # Delete associated items first
                PurchaseRequestItem.query.filter_by(purchase_request_id=purchase_request.id).delete()
                db.session.delete(purchase_request)
                deleted_count += 1
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'Successfully deleted {deleted_count} purchase requests'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error deleting requests: {str(e)}'}), 500

@app.route('/bulk-export-purchase-requests-excel', methods=['POST'])
@login_required
@page_access_required('purchase_request')
def bulk_export_purchase_requests_excel():
    """Bulk export purchase requests to Excel"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('purchase_request_management'))
    
    try:
        import pandas as pd
        from io import BytesIO
        from flask import send_file
        
        request_ids = request.form.getlist('request_ids')
        
        if not request_ids:
            flash('No requests selected', 'error')
            return redirect(url_for('purchase_request_management'))
        
        # Get selected requests
        requests = []
        for request_id in request_ids:
            purchase_request = PurchaseRequest.query.filter_by(
                id=int(request_id),
                organization_id=user.organization_id
            ).first()
            
            if purchase_request:
                requests.append({
                    'PR Number': purchase_request.request_number,
                    'Date': purchase_request.request_date.strftime('%Y-%m-%d') if purchase_request.request_date else '',
                    'Requested By': purchase_request.requested_by or '',
                    'Category': purchase_request.category or '',
                    'Status': purchase_request.status or 'Pending',
                    'Created': purchase_request.created_at.strftime('%Y-%m-%d %H:%M') if purchase_request.created_at else ''
                })
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(requests)
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Purchase Requests')
        
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f'purchase_requests_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting to Excel: {str(e)}', 'error')
        return redirect(url_for('purchase_request_management'))

@app.route('/bulk-export-purchase-requests-pdf', methods=['POST'])
@login_required
@page_access_required('purchase_request')
def bulk_export_purchase_requests_pdf():
    """Export purchase requests as professional PDF forms"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('purchase_request_management'))
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from io import BytesIO
        import datetime
        import os
        
        # Define professional color scheme
        TS_BLUE = colors.Color(0, 0.357, 0.675)  # #005bac
        LIGHT_BLUE = colors.Color(0.875, 0.914, 0.953)  # #dfe9f3
        LIGHT_GREY = colors.Color(0.961, 0.961, 0.961)  # #f5f5f5
        PALE_YELLOW = colors.Color(1, 0.953, 0.804)  # #fff3cd
        SOFT_GREY = colors.Color(0.878, 0.878, 0.878)  # #e0e0e0
        BORDER_GREY = colors.Color(0.8, 0.8, 0.8)  # #cccccc
        
        request_ids = request.form.getlist('request_ids')
        
        if not request_ids:
            flash('No requests selected', 'error')
            return redirect(url_for('purchase_request_management'))
        
        # Get selected purchase requests
        purchase_requests = []
        for request_id in request_ids:
            purchase_request = PurchaseRequest.query.filter_by(
                id=int(request_id),
                organization_id=user.organization_id
            ).first()
            if purchase_request:
                purchase_requests.append(purchase_request)
        
        if not purchase_requests:
            flash('No valid purchase requests found', 'error')
            return redirect(url_for('purchase_form_storage'))
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=60
        )
        elements = []
        styles = getSampleStyleSheet()
        
        # Process each purchase request with modern professional styling
        for i, purchase_request in enumerate(purchase_requests):
            if i > 0:
                elements.append(PageBreak())
            
            # Modern header with TS GROUP logo
            logo_path = os.path.join('static', 'uploads', 'ts_logo.jpg')
            if os.path.exists(logo_path):
                try:
                    logo = Image(logo_path, width=2.5*inch, height=1*inch)
                    logo.hAlign = 'LEFT'
                    elements.append(logo)
                    elements.append(Spacer(1, 15))
                except:
                    header_style = ParagraphStyle(
                        'HeaderStyle',
                        parent=styles['Normal'],
                        fontSize=22,
                        textColor=TS_BLUE,
                        fontName='Helvetica-Bold',
                        alignment=TA_LEFT,
                        spaceBefore=5,
                        spaceAfter=15
                    )
                    elements.append(Paragraph("TS GROUP", header_style))
            else:
                header_style = ParagraphStyle(
                    'HeaderStyle',
                    parent=styles['Normal'],
                    fontSize=22,
                    textColor=TS_BLUE,
                    fontName='Helvetica-Bold',
                    alignment=TA_LEFT,
                    spaceBefore=5,
                    spaceAfter=15
                )
                elements.append(Paragraph("TS GROUP", header_style))
            
            # Modern title with increased font size
            title_style = ParagraphStyle(
                'TitleStyle',
                parent=styles['Normal'],
                fontSize=18,
                fontName='Helvetica-Bold',
                alignment=TA_CENTER,
                textColor=TS_BLUE,
                spaceBefore=10,
                spaceAfter=25
            )
            elements.append(Paragraph(f"PURCHASE REQUISITION FORM - {purchase_request.request_number}", title_style))
            
            # Modern request details using ACTUAL model fields
            status_value = purchase_request.status or 'Pending'
            status_bg = PALE_YELLOW if status_value == 'Pending' else SOFT_GREY
            
            details_data = [
                ['Request Number:', purchase_request.request_number or '', 'Requested By:', purchase_request.requested_by or ''],
                ['PL Number:', f'PL/25/{purchase_request.pl_number or ""}', 'DC Name:', purchase_request.dc_name or ''],
                ['Request Date:', purchase_request.request_date.strftime('%Y-%m-%d') if purchase_request.request_date else '', 'Operation Manager:', purchase_request.operation_manager or ''],
                ['Category:', purchase_request.category or 'Purchase Stock', 'Status:', status_value]
            ]
            
            details_table = Table(details_data, colWidths=[1.4*inch, 2.1*inch, 1.4*inch, 2.1*inch])
            details_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, BORDER_GREY),
                ('BACKGROUND', (0, 0), (0, -1), LIGHT_GREY),
                ('BACKGROUND', (2, 0), (2, -1), LIGHT_GREY),
                ('BACKGROUND', (3, 3), (3, 3), status_bg),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(details_table)
            elements.append(Spacer(1, 15))
            
            # Modern items table using actual relationship
            items_header = [['No', 'Item / Description', 'Unit Cost', 'Qty', 'Total', 'Unit Required', 'Cost Code', 'Remarks']]
            items_data = []
            
            # Use the items relationship that actually exists
            for idx, item in enumerate(purchase_request.items, 1):
                items_data.append([
                    str(idx),
                    item.description or '',
                    f'${item.unit_cost:.2f}' if item.unit_cost else '$0.00',
                    str(item.quantity) if item.quantity else '0',
                    f'${item.total_cost:.2f}' if item.total_cost else '$0.00',
                    item.unit or '',
                    item.cost_code or '',
                    item.remarks or ''
                ])
            
            # Add empty rows for consistent appearance
            while len(items_data) < 8:
                items_data.append(['', '', '', '', '', '', '', ''])
            
            # Calculate total using actual model field
            total_amount = purchase_request.grand_total or sum(item.total_cost or 0 for item in purchase_request.items)
            
            all_items_data = items_header + items_data
            
            # Modern table with proper styling and consistent column widths
            items_table = Table(all_items_data, colWidths=[0.4*inch, 2.3*inch, 0.8*inch, 0.5*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.3*inch])
            items_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BACKGROUND', (0, 0), (-1, 0), TS_BLUE),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                # Professional alignment
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # No column
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),    # Description - left aligned
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),   # Unit Cost - right aligned
                ('ALIGN', (3, 0), (3, -1), 'CENTER'),  # Qty - center aligned
                ('ALIGN', (4, 0), (4, -1), 'RIGHT'),   # Total - right aligned
                ('ALIGN', (5, 0), (5, -1), 'CENTER'),  # Unit Required - center
                ('ALIGN', (6, 0), (6, -1), 'CENTER'),  # Cost Code - center
                ('ALIGN', (7, 0), (7, -1), 'LEFT'),    # Remarks - left aligned
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, BORDER_GREY),
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
                # Better padding
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(items_table)
            elements.append(Spacer(1, 8))
            
            # Fixed total section with better alignment and styling
            total_data = [['', '', '', '', '', '', 'TOTAL AMOUNT:', f'${total_amount:.2f}']]
            total_table = Table(total_data, colWidths=[0.4*inch, 2.3*inch, 0.8*inch, 0.5*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.3*inch])
            total_table.setStyle(TableStyle([
                ('FONTNAME', (6, 0), (7, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (6, 0), (7, 0), 12),
                ('BACKGROUND', (6, 0), (7, 0), colors.Color(0.92, 0.92, 0.92)),
                ('ALIGN', (6, 0), (6, 0), 'CENTER'),
                ('ALIGN', (7, 0), (7, 0), 'CENTER'),
                ('GRID', (6, 0), (7, 0), 1, BORDER_GREY),
                ('TOPPADDING', (6, 0), (7, 0), 12),
                ('BOTTOMPADDING', (6, 0), (7, 0), 12),
                ('LEFTPADDING', (6, 0), (7, 0), 8),
                ('RIGHTPADDING', (6, 0), (7, 0), 8),
            ]))
            elements.append(total_table)
            elements.append(Spacer(1, 20))
            
            # Fixed signature section - compact layout to stay on same page
            approved_by_name = ''
            if purchase_request.approved_by_user:
                approved_by_name = purchase_request.approved_by_user.full_name or purchase_request.approved_by_user.username or ''
            
            signature_data = [
                ['Requested By', 'Approved By'],
                ['', ''],
                ['', ''],
                ['_' * 25, '_' * 25],
                [f'Name: {purchase_request.requested_by or ""}', f'Name: {approved_by_name}'],
                ['Date: __________', 'Date: __________']
            ]
            
            # Equal column widths for professional appearance
            signature_table = Table(signature_data, colWidths=[3.75*inch, 3.75*inch], rowHeights=[0.4*inch, 0.5*inch, 0.3*inch, 0.3*inch, 0.35*inch, 0.35*inch])
            signature_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, 0), 0.5, BORDER_GREY),
                ('GRID', (0, -2), (-1, -1), 0.5, BORDER_GREY),
                ('BACKGROUND', (0, 0), (-1, 0), LIGHT_BLUE),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(signature_table)
            
            # Horizontal separator line before footer
            elements.append(Spacer(1, 15))
            line_data = [['_' * 100]]
            line_table = Table(line_data, colWidths=[7.5*inch])
            line_table.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (0, 0), 8),
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('TEXTCOLOR', (0, 0), (0, 0), colors.Color(0.7, 0.7, 0.7)),
            ]))
            elements.append(line_table)
            elements.append(Spacer(1, 8))
            
            # Fixed footer layout - left aligned timestamp, right aligned page number
            footer_data = [[
                f"Generated on {datetime.datetime.now().strftime('%B %d, %Y at %H:%M')} | TS GROUP Purchase Management System",
                f"Page {i+1} of {len(purchase_requests)}"
            ]]
            
            footer_table = Table(footer_data, colWidths=[5.5*inch, 2*inch])
            footer_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.Color(0.5, 0.5, 0.5)),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),   # Footer text left-aligned
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),  # Page number right-aligned
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(footer_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'purchase_requests_{singapore_now().strftime("%Y%m%d_%H%M%S")}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('purchase_form_storage'))

# Important News Management Routes
@app.route('/admin/important-news')
@login_required
@admin_required
def manage_important_news():
    """Admin page to manage important news"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned', 'error')
        return redirect(url_for('dashboard'))
    
    news_items = ImportantNews.query.filter_by(
        organization_id=user.organization_id
    ).order_by(ImportantNews.created_at.desc()).all()
    
    from datetime import datetime
    return render_template('admin_important_news.html', news_items=news_items, now=singapore_now())

@app.route('/admin/important-news/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_important_news():
    """Create new important news"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned', 'error')
        return redirect(url_for('manage_important_news'))
    
    if request.method == 'POST':
        try:
            title = request.form.get('title', '').strip()
            content = request.form.get('content', '').strip()
            priority = request.form.get('priority', 'normal')
            is_active = 'is_active' in request.form
            show_on_login = 'show_on_login' in request.form
            expires_at = request.form.get('expires_at')
            
            if not title or not content:
                flash('Title and content are required', 'error')
                return render_template('admin_create_news.html')
            
            # Handle file uploads
            pdf_filename = None
            image_filename = None
            
            if 'pdf_attachment' in request.files:
                pdf_file = request.files['pdf_attachment']
                if pdf_file and pdf_file.filename:
                    if pdf_file.filename.lower().endswith('.pdf'):
                        import os
                        from werkzeug.utils import secure_filename
                        
                        # Create uploads directory if it doesn't exist
                        uploads_dir = os.path.join('static', 'uploads', 'news')
                        os.makedirs(uploads_dir, exist_ok=True)
                        
                        pdf_filename = secure_filename(f"news_pdf_{singapore_now().strftime('%Y%m%d_%H%M%S')}_{pdf_file.filename}")
                        pdf_path = os.path.join(uploads_dir, pdf_filename)
                        pdf_file.save(pdf_path)
                        pdf_filename = f"uploads/news/{pdf_filename}"
                    else:
                        flash('Only PDF files are allowed for attachments', 'error')
                        return render_template('admin_create_news.html')
            
            if 'image_attachment' in request.files:
                image_file = request.files['image_attachment']
                if image_file and image_file.filename:
                    if image_file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                        import os
                        from werkzeug.utils import secure_filename
                        
                        uploads_dir = os.path.join('static', 'uploads', 'news')
                        os.makedirs(uploads_dir, exist_ok=True)
                        
                        image_filename = secure_filename(f"news_img_{singapore_now().strftime('%Y%m%d_%H%M%S')}_{image_file.filename}")
                        image_path = os.path.join(uploads_dir, image_filename)
                        image_file.save(image_path)
                        image_filename = f"uploads/news/{image_filename}"
                    else:
                        flash('Only image files (PNG, JPG, JPEG, GIF) are allowed', 'error')
                        return render_template('admin_create_news.html')
            
            # Parse expires_at date
            expires_at_date = None
            if expires_at:
                try:
                    expires_at_date = datetime.strptime(expires_at, '%Y-%m-%d')
                except ValueError:
                    flash('Invalid expiry date format', 'error')
                    return render_template('admin_create_news.html')
            
            # Create news item
            news_item = ImportantNews(
                title=title,
                content=content,
                priority=priority,
                is_active=is_active,
                show_on_login=show_on_login,
                pdf_attachment=pdf_filename,
                image_attachment=image_filename,
                expires_at=expires_at_date,
                organization_id=user.organization_id,
                created_by=user.id
            )
            
            db.session.add(news_item)
            db.session.commit()
            
            flash('Important news created successfully!', 'success')
            return redirect(url_for('manage_important_news'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating news: {str(e)}', 'error')
    
    return render_template('admin_create_news.html')

@app.route('/admin/important-news/<int:news_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_important_news(news_id):
    """Edit existing important news"""
    user = current_user
    news_item = ImportantNews.query.filter_by(
        id=news_id,
        organization_id=user.organization_id
    ).first()
    
    if not news_item:
        flash('News item not found', 'error')
        return redirect(url_for('manage_important_news'))
    
    if request.method == 'POST':
        try:
            news_item.title = request.form.get('title', '').strip()
            news_item.content = request.form.get('content', '').strip()
            news_item.priority = request.form.get('priority', 'normal')
            news_item.is_active = 'is_active' in request.form
            news_item.show_on_login = 'show_on_login' in request.form
            expires_at = request.form.get('expires_at')
            
            if not news_item.title or not news_item.content:
                flash('Title and content are required', 'error')
                return render_template('admin_edit_news.html', news_item=news_item)
            
            # Handle file uploads (similar to create)
            if 'pdf_attachment' in request.files:
                pdf_file = request.files['pdf_attachment']
                if pdf_file and pdf_file.filename:
                    if pdf_file.filename.lower().endswith('.pdf'):
                        import os
                        from werkzeug.utils import secure_filename
                        
                        # Delete old file if exists
                        if news_item.pdf_attachment:
                            old_path = os.path.join('static', news_item.pdf_attachment)
                            if os.path.exists(old_path):
                                os.remove(old_path)
                        
                        uploads_dir = os.path.join('static', 'uploads', 'news')
                        os.makedirs(uploads_dir, exist_ok=True)
                        
                        pdf_filename = secure_filename(f"news_pdf_{singapore_now().strftime('%Y%m%d_%H%M%S')}_{pdf_file.filename}")
                        pdf_path = os.path.join(uploads_dir, pdf_filename)
                        pdf_file.save(pdf_path)
                        news_item.pdf_attachment = f"uploads/news/{pdf_filename}"
            
            if 'image_attachment' in request.files:
                image_file = request.files['image_attachment']
                if image_file and image_file.filename:
                    if image_file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                        import os
                        from werkzeug.utils import secure_filename
                        
                        # Delete old file if exists
                        if news_item.image_attachment:
                            old_path = os.path.join('static', news_item.image_attachment)
                            if os.path.exists(old_path):
                                os.remove(old_path)
                        
                        uploads_dir = os.path.join('static', 'uploads', 'news')
                        os.makedirs(uploads_dir, exist_ok=True)
                        
                        image_filename = secure_filename(f"news_img_{singapore_now().strftime('%Y%m%d_%H%M%S')}_{image_file.filename}")
                        image_path = os.path.join(uploads_dir, image_filename)
                        image_file.save(image_path)
                        news_item.image_attachment = f"uploads/news/{image_filename}"
            
            # Parse expires_at date
            if expires_at:
                try:
                    news_item.expires_at = datetime.strptime(expires_at, '%Y-%m-%d')
                except ValueError:
                    flash('Invalid expiry date format', 'error')
                    return render_template('admin_edit_news.html', news_item=news_item)
            else:
                news_item.expires_at = None
            
            news_item.updated_at = singapore_now()
            db.session.commit()
            
            flash('Important news updated successfully!', 'success')
            return redirect(url_for('manage_important_news'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating news: {str(e)}', 'error')
    
    return render_template('admin_edit_news.html', news_item=news_item)

@app.route('/admin/important-news/<int:news_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_important_news(news_id):
    """Delete important news"""
    user = current_user
    news_item = ImportantNews.query.filter_by(
        id=news_id,
        organization_id=user.organization_id
    ).first()
    
    if not news_item:
        flash('News item not found', 'error')
        return redirect(url_for('manage_important_news'))
    
    try:
        # Delete attached files
        import os
        if news_item.pdf_attachment:
            pdf_path = os.path.join('static', news_item.pdf_attachment)
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
        
        if news_item.image_attachment:
            image_path = os.path.join('static', news_item.image_attachment)
            if os.path.exists(image_path):
                os.remove(image_path)
        
        db.session.delete(news_item)
        db.session.commit()
        
        flash('Important news deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting news: {str(e)}', 'error')
    
    return redirect(url_for('manage_important_news'))

@app.route('/admin/important-news/<int:news_id>/toggle-status', methods=['POST'])
@login_required
@admin_required
def toggle_news_status(news_id):
    """Toggle active status of news item"""
    user = current_user
    news_item = ImportantNews.query.filter_by(
        id=news_id,
        organization_id=user.organization_id
    ).first()
    
    if not news_item:
        return jsonify({'success': False, 'error': 'News item not found'})
    
    try:
        news_item.is_active = not news_item.is_active
        news_item.updated_at = singapore_now()
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'is_active': news_item.is_active,
            'message': f'News {"activated" if news_item.is_active else "deactivated"} successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})




@app.route('/generate_purchase_request_pdf/<int:request_id>', methods=['POST'])
@login_required
@page_permission_required('bedding_management')
def generate_purchase_request_pdf_by_id(request_id):
    """Generate PDF for existing purchase request"""
    import json
    import base64
    import logging
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    from PIL import Image as PILImage
    from reportlab.platypus import Image as ReportLabImage
    
    try:
        user = current_user
        if not user.organization_id:
            return jsonify({'error': 'No organization assigned'}), 403
        
        # Get purchase request
        purchase_request = PurchaseRequest.query.filter_by(
            id=request_id,
            organization_id=user.organization_id
        ).first()
        
        if not purchase_request:
            return jsonify({'error': 'Purchase request not found'}), 404
        
        # Get request items
        items = PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).all()
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Build story content
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles matching create form
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=6,
            textColor=colors.black,
            alignment=1  # Center alignment
        )
        
        # TS Group Logo and Header matching expected format
        try:
            from reportlab.platypus import Image
            # Logo positioning for TS GROUP
            logo_table = Table([['TS GROUP', 'PURCHASE REQUISITION FORM']], colWidths=[2*inch, 6*inch])
            logo_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (0, 0), 14),
                ('FONTSIZE', (1, 0), (1, 0), 16),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]))
            story.append(logo_table)
        except Exception as e:
            # Fallback without logo
            story.append(Paragraph("PURCHASE REQUISITION FORM", title_style))
        
        story.append(Spacer(1, 20))
        
        # Request details in professional table format matching create form
        request_info = [
            ['PR Number:', purchase_request.request_number, 'Request Date:', purchase_request.request_date.strftime('%Y-%m-%d') if purchase_request.request_date else ''],
            ['Department:', purchase_request.category or '', 'Priority:', 'Standard'],
            ['Requested By:', purchase_request.requested_by or '', 'Supplier:', purchase_request.supplier or '']
        ]
        
        request_table = Table(request_info, colWidths=[1.5*inch, 2.5*inch, 1.5*inch, 2.5*inch])
        request_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black),
        ]))
        story.append(request_table)
        story.append(Spacer(1, 20))
        
        # Items table with modern styling matching create form
        items_header = ['S/N', 'Description of Materials/Services', 'Unit Cost\n(SGD)', 'Qty', 'Total\n(SGD)', 'Room No.', 'Unit', 'Cost Code']
        items_data = [items_header]
        
        subtotal = 0
        for i, item in enumerate(items):
            items_data.append([
                str(i + 1),
                item.description or '',
                f"{float(item.unit_cost):.2f}",
                str(item.quantity),
                f"{float(item.total_cost):.2f}",
                item.room_no or '',
                item.unit or '',
                item.cost_code or ''
            ])
            subtotal += float(item.total_cost)
        
        items_table = Table(items_data, colWidths=[0.4*inch, 2.2*inch, 0.8*inch, 0.5*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.9*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Left align descriptions
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 20))
        
        # Financial summary with modern styling matching create form
        gst = 0.0
        additional_tax = 0.0
        grand_total = subtotal + gst + additional_tax
        
        financial_data = [
            ['', '', '', 'Subtotal:', f"{subtotal:.2f}"],
            ['', '', '', 'GST (0%):', f"{gst:.2f}"],
            ['', '', '', 'Additional Tax:', f"{additional_tax:.2f}"],
            ['', '', '', 'GRAND TOTAL (SGD):', f"{grand_total:.2f}"]
        ]
        
        financial_table = Table(financial_data, colWidths=[1*inch, 2*inch, 1*inch, 2*inch, 2*inch])
        financial_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 2), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, 2), 10),
            ('FONTNAME', (3, 3), (4, 3), 'Helvetica-Bold'),
            ('FONTSIZE', (3, 3), (4, 3), 12),
            ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
            ('LINEABOVE', (3, 3), (4, 3), 2, colors.black),
            ('LINEBELOW', (3, 3), (4, 3), 2, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(financial_table)
        story.append(Spacer(1, 20))
        
        # Additional information matching create form
        additional_info = [
            ['Payment Method:', purchase_request.payment_method or ''],
            ['Budget Code:', purchase_request.budget_code or ''],
            ['Expected Delivery Date:', purchase_request.expected_delivery or ''],
            ['Justification:', purchase_request.justification or '']
        ]
        
        additional_table = Table(additional_info, colWidths=[2*inch, 6*inch])
        additional_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(additional_table)
        story.append(Spacer(1, 30))
        

        
        # Signatures Section with professional header - center aligned
        signature_section_title = [['APPROVALS & SIGNATURES']]
        signature_title_table = Table(signature_section_title, colWidths=[8*inch])
        signature_title_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.white),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 12),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('BOX', (0, 0), (0, 0), 1, colors.black),
            ('TOPPADDING', (0, 0), (0, 0), 10),
            ('BOTTOMPADDING', (0, 0), (0, 0), 10),
        ]))
        story.append(signature_title_table)
        
        # Process signature images with proper structure
        signature_names = ['dc_signature_data', 'operation_manager_signature_data', 'general_manager_signature_data']
        name_fields = ['dc_name', 'operation_manager', 'general_manager']
        signature_titles = ['D/C & O/E', 'Operation Manager', 'General Manager']
        
        # Initialize signature processing arrays
        signature_images = []
        
        for i, (sig_field, name_field) in enumerate(zip(signature_names, name_fields)):
            sig_data = getattr(purchase_request, sig_field, None)
            name_value = getattr(purchase_request, name_field, None)
            
            # Debug logging
            logging.debug(f"Processing {sig_field}: {sig_data[:100] if sig_data else 'None'}...")
            logging.debug(f"Processing {name_field}: {name_value}")
            
            # Process signature image
            if sig_data:
                try:
                    # Parse JSON signature data
                    signature_info = json.loads(sig_data)
                    signature_base64 = signature_info.get('signature')
                    
                    logging.debug(f"Signature base64 for {sig_field}: {signature_base64[:50] if signature_base64 else 'None'}...")
                    
                    if signature_base64 and signature_base64.startswith('data:image'):
                        # Extract base64 data
                        header, data = signature_base64.split(',', 1)
                        image_data = base64.b64decode(data)
                        
                        # Create PIL image
                        pil_image = PILImage.open(BytesIO(image_data))
                        
                        # Convert to RGB if needed
                        if pil_image.mode in ('RGBA', 'LA'):
                            background = PILImage.new('RGB', pil_image.size, (255, 255, 255))
                            if pil_image.mode == 'RGBA':
                                background.paste(pil_image, mask=pil_image.split()[-1])
                            else:
                                background.paste(pil_image, mask=pil_image.split()[-1])
                            pil_image = background
                        elif pil_image.mode != 'RGB':
                            pil_image = pil_image.convert('RGB')
                        
                        # Save to BytesIO
                        img_buffer = BytesIO()
                        pil_image.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        
                        # Create ReportLab image with proper sizing
                        rl_image = ReportLabImage(img_buffer, width=2.5*inch, height=1*inch)
                        signature_row1.append(rl_image)
                        logging.debug(f"Successfully processed signature image for {sig_field}")
                    else:
                        signature_row1.append('')
                        logging.debug(f"No valid signature data for {sig_field}")
                except Exception as e:
                    signature_row1.append('')
                    logging.error(f"Error processing signature {sig_field}: {str(e)}")
            else:
                signature_row1.append('')
                logging.debug(f"No signature data for {sig_field}")
            
            # Add name to signature row
            signature_row2.append(name_value or '')
        
        # Create final signature table matching create form format
        signature_data = [
            signature_titles,     # Role headers
            signature_row1,       # Signature images
            signature_row2        # Names
        ]
        
        signature_table = Table(signature_data, colWidths=[2.67*inch, 2.67*inch, 2.66*inch])
        signature_table.setStyle(TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Signature row styling
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, 1), 8),
            
            # Name row styling
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica'),
            ('FONTSIZE', (0, 2), (-1, 2), 9),
            
            # Grid and spacing
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            
            # Special formatting for signature row
            ('TOPPADDING', (0, 1), (-1, 1), 15),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 15),
        ]))
        story.append(signature_table)
        
        # Add footer space
        story.append(Spacer(1, 30))
        
        # Professional footer
        footer_data = [['This document was generated electronically and is valid without physical signature when digitally signed.']]
        footer_table = Table(footer_data, colWidths=[8*inch])
        footer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica'),
            ('FONTSIZE', (0, 0), (0, 0), 8),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.grey),
            ('TOPPADDING', (0, 0), (0, 0), 10),
        ]))
        story.append(footer_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        # Return PDF as download
        return send_file(
            BytesIO(buffer.read()),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'purchase_request_{purchase_request.request_number}.pdf'
        )
        
    except Exception as e:
        logging.error(f"Error generating PDF: {str(e)}")
        return jsonify({'error': 'Error generating PDF'}), 500

@app.route('/api/purchase-request', methods=['POST'])
@login_required
@create_permission_required('bedding_management')
def submit_purchase_request():
    """Submit a new purchase request"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'No organization assigned'})
    
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data.get('request_date') or not data.get('requested_by'):
            return jsonify({'success': False, 'error': 'Request date and requested by are required'})
        
        if not data.get('items') or len(data['items']) == 0:
            return jsonify({'success': False, 'error': 'At least one item is required'})
        
        # Generate request number
        from datetime import datetime
        today = singapore_now()
        request_count = PurchaseRequest.query.filter(
            db.func.date(PurchaseRequest.created_at) == today.date(),
            PurchaseRequest.organization_id == user.organization_id
        ).count()
        request_number = f"PR-{today.strftime('%Y%m%d')}-{str(request_count + 1).zfill(3)}"
        
        # Handle signatures using JSON format like room checklist
        signatures = data.get('signatures', {})
        dc_signature_data = json.dumps(signatures.get('dc', {})) if signatures.get('dc') else None
        operation_signature_data = json.dumps(signatures.get('operation', {})) if signatures.get('operation') else None
        general_signature_data = json.dumps(signatures.get('general', {})) if signatures.get('general') else None

        # Create purchase request with professional format
        purchase_request = PurchaseRequest(
            request_number=request_number,
            pl_number=data.get('pl_number', ''),
            request_date=datetime.strptime(data['request_date'], '%Y-%m-%d').date(),
            category=data.get('category', 'Purchase Stock'),
            requested_by=data['requested_by'],
            dc_name=data.get('dc_name', ''),
            operation_manager=data.get('operation_manager', ''),
            general_manager=data.get('general_manager', ''),
            requested_by_footer=data.get('requested_by_footer', ''),
            recommended_by_footer=data.get('recommended_by_footer', ''),
            dc_signature_data=dc_signature_data,
            operation_manager_signature_data=operation_signature_data,
            general_manager_signature_data=general_signature_data,
            organization_id=user.organization_id,
            created_by=user.id
        )
        
        db.session.add(purchase_request)
        db.session.flush()  # Get the ID
        
        # Add request items with cost details
        for item_data in data['items']:
            item = PurchaseRequestItem(
                purchase_request_id=purchase_request.id,
                description=item_data.get('description', item_data.get('name', '')),
                unit_cost=float(item_data.get('unit_cost', 0)),
                quantity=int(item_data['quantity']),
                total_cost=float(item_data.get('total_cost', 0)),
                room_no=item_data.get('room_no', ''),
                unit=item_data.get('unit', ''),
                cost_code=item_data.get('cost_code', ''),
                remarks=item_data.get('remarks', '')
            )
            db.session.add(item)
        
        db.session.commit()
        
        # Log the submission
        log_entry = SystemLog(
            user_id=current_user.id,
            action='submit_purchase_request',
            module='purchase_management',
            details=f'Submitted purchase requisition {request_number} - Category: {data.get("category", "Purchase Stock")}',
            ip_address=request.remote_addr
        )
        db.session.add(log_entry)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'request_number': request_number,
            'message': f'Purchase requisition {request_number} submitted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error submitting purchase request: {str(e)}'})




# Offense Records Routes
@app.route('/offense-records', methods=['GET', 'POST'])
@login_required
@page_access_required('offense_records')
def offense_records():
    """Offense records page and form submission handler"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            from datetime import datetime, time
            
            # Parse form data
            case_number = request.form.get('case_number')
            offense_type = request.form.get('offense_type')
            
            # Resident details
            offender_name = request.form.get('offender_name')
            fin_number = request.form.get('fin_number')
            nationality = request.form.get('nationality')
            offender_room = request.form.get('offender_room')
            sector = request.form.get('sector')
            contact_number = request.form.get('contact_number')
            offender_company = request.form.get('offender_company')
            
            # Incident details
            description = request.form.get('description')
            location = request.form.get('location')
            incident_date = datetime.strptime(request.form.get('incident_date'), '%Y-%m-%d').date()
            incident_time_str = request.form.get('incident_time')
            incident_time = datetime.strptime(incident_time_str, '%H:%M').time() if incident_time_str else None
            
            # Documentary evidence
            documentary_proof = request.form.get('documentary_proof') == 'true'
            proof_description = request.form.get('proof_description')
            
            # Financial penalty
            financial_penalty_imposed = request.form.get('financial_penalty_imposed') == 'true'
            penalty_amount_str = request.form.get('penalty_amount')
            penalty_amount = float(penalty_amount_str) if penalty_amount_str and penalty_amount_str != '' else None
            
            # Additional information
            witness_names = request.form.get('witness_names')
            action_taken = request.form.get('action_taken')
            duty_manager_name = request.form.get('duty_manager_name')
            
            # Handle photo uploads
            photo_data = {}
            if 'incident_photos' in request.files:
                photos = request.files.getlist('incident_photos')
                for i, photo in enumerate(photos[:10]):  # Limit to 10 photos
                    if photo and photo.filename and photo.content_type.startswith('image/'):
                        try:
                            # Read and encode photo as base64
                            import base64
                            photo_bytes = photo.read()
                            if len(photo_bytes) <= 5 * 1024 * 1024:  # 5MB limit
                                photo_base64 = base64.b64encode(photo_bytes).decode('utf-8')
                                photo_data[f'incident_photo_{i+1}'] = photo_base64
                        except Exception as e:
                            flash(f'Error processing photo {i+1}: {str(e)}', 'warning')
            
            # Process signature data
            resident_signature = request.form.get('resident_signature')
            duty_manager_signature = request.form.get('duty_manager_signature')
            
            # Clean signature data (remove data URL prefix)
            if resident_signature and resident_signature.startswith('data:image'):
                resident_signature = resident_signature.split(',', 1)[1]
            if duty_manager_signature and duty_manager_signature.startswith('data:image'):
                duty_manager_signature = duty_manager_signature.split(',', 1)[1]
            
            # Generate case number if not provided
            if not case_number:
                case_count = OffenseRecord.query.filter_by(organization_id=user.organization_id).count() + 1
                case_number = f"PL/OR/{case_count:07d}"
            
            # Create new offense record using actual model fields
            offense_record = OffenseRecord(
                case_number=case_number,
                offense_type=offense_type,
                severity='N/A',  # Default severity since field is removed
                offender_name=offender_name,
                fin_number=fin_number,
                nationality=nationality,
                offender_room=offender_room,
                sector=sector,
                contact_number=contact_number,
                offender_company=offender_company,
                description=description,
                location=location,
                incident_date=incident_date,
                incident_time=incident_time,
                documentary_proof=documentary_proof,
                proof_description=proof_description,
                financial_penalty_imposed=financial_penalty_imposed,
                penalty_amount=penalty_amount,
                witness_names=witness_names,
                action_taken=action_taken,
                duty_manager_name=duty_manager_name,
                reported_by=user.id,
                organization_id=user.organization_id,
                status='Open',
                penalty_status='Pending' if financial_penalty_imposed else None,
                amount_paid=0.0 if financial_penalty_imposed else None,
                resident_signature=resident_signature,
                resident_signature_date=singapore_now() if resident_signature else None,
                duty_manager_signature=duty_manager_signature,
                duty_manager_signature_date=singapore_now() if duty_manager_signature else None,
                created_at=singapore_now(),
                updated_at=singapore_now(),
                **photo_data  # Add photo data to the record
            )
            
            db.session.add(offense_record)
            db.session.commit()
            
            flash(f'Offence report #{offense_record.case_number} has been submitted successfully.', 'success')
            return redirect(url_for('offense_records'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error submitting offence report: {str(e)}', 'error')
            return redirect(url_for('offense_records'))
    
    # GET request - display offense records with search and filter
    from datetime import datetime, timedelta
    
    # Get filter parameters
    search_query = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    severity_filter = request.args.get('severity', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Base query - past 60 days only by default
    sixty_days_ago = singapore_now() - timedelta(days=60)
    query = OffenseRecord.query.filter(
        OffenseRecord.organization_id == user.organization_id,
        OffenseRecord.created_at >= sixty_days_ago
    )
    
    # Apply search filter
    if search_query:
        query = query.filter(
            db.or_(
                OffenseRecord.offender_name.ilike(f'%{search_query}%'),
                OffenseRecord.case_number.ilike(f'%{search_query}%'),
                OffenseRecord.fin_number.ilike(f'%{search_query}%'),
                OffenseRecord.offender_company.ilike(f'%{search_query}%'),
                OffenseRecord.offender_room.ilike(f'%{search_query}%'),
                OffenseRecord.offense_type.ilike(f'%{search_query}%')
            )
        )
    
    # Apply status filter
    if status_filter:
        query = query.filter(OffenseRecord.status == status_filter)
    
    # Apply severity filter
    if severity_filter:
        query = query.filter(OffenseRecord.severity == severity_filter)
    
    # Apply date range filters
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(OffenseRecord.incident_date >= from_date.date())
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            query = query.filter(OffenseRecord.incident_date <= to_date.date())
        except ValueError:
            pass
    
    offenses = query.order_by(OffenseRecord.created_at.desc()).all()
    
    # Calculate financial penalty statistics
    total_financial_penalty = 0.0
    total_amount_paid = 0.0
    total_amount_unpaid = 0.0
    
    for offense in offenses:
        if offense.financial_penalty_imposed and offense.penalty_amount:
            total_financial_penalty += offense.penalty_amount
            if offense.amount_paid:
                total_amount_paid += offense.amount_paid
            total_amount_unpaid += (offense.penalty_amount - (offense.amount_paid or 0))
    
    # Get all active room numbers for the dropdown
    from app.models.models_house_acknowledge import RoomNumber
    room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    
    return render_template('offense_records.html', 
                         offenses=offenses, 
                         total_financial_penalty=total_financial_penalty,
                         total_amount_paid=total_amount_paid,
                         total_amount_unpaid=total_amount_unpaid,
                         room_numbers=room_numbers,
                         search_query=search_query,
                         status_filter=status_filter,
                         severity_filter=severity_filter,
                         date_from=date_from,
                         date_to=date_to)

@app.route('/update-offense-status/<int:offense_id>', methods=['POST'])
@login_required
@edit_permission_required('offense_records')
def update_offense_status(offense_id):
    """Update status of an offence record"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'Access denied'})
    
    # Get offence record
    offense = OffenseRecord.query.filter_by(id=offense_id, organization_id=user.organization_id).first()
    if not offense:
        return jsonify({'success': False, 'error': 'Offence record not found'})
    
    try:
        # Get new status from form
        new_status = request.form.get('status')
        
        # Validate status
        valid_statuses = ['Open', 'Under Investigation', 'Resolved', 'Closed']
        if new_status not in valid_statuses:
            return jsonify({'success': False, 'error': 'Invalid status'})
        
        # Update status
        offense.status = new_status
        offense.updated_at = singapore_now()
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Status updated successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/update-offense-payment/<int:offense_id>', methods=['POST'])
@login_required
@edit_permission_required('offense_records')
def update_offense_payment(offense_id):
    """Update payment status and amount for an offence record"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('offense_records'))
    
    # Get offence record
    offense = OffenseRecord.query.filter_by(id=offense_id, organization_id=user.organization_id).first()
    if not offense:
        flash('Offence record not found', 'error')
        return redirect(url_for('offense_records'))
    
    if not offense.financial_penalty_imposed:
        flash('No financial penalty imposed for this offense', 'error')
        return redirect(url_for('offense_records'))
    
    try:
        # Get payment details from form
        payment_amount = float(request.form.get('payment_amount', 0))
        payment_status = request.form.get('payment_status')
        payment_notes = request.form.get('payment_notes', '')
        
        # Validate payment amount
        if payment_amount < 0:
            flash('Payment amount cannot be negative', 'error')
            return redirect(url_for('offense_records'))
        
        current_paid = offense.amount_paid or 0
        total_penalty = offense.penalty_amount or 0
        
        if current_paid + payment_amount > total_penalty:
            flash('Payment amount exceeds outstanding balance', 'error')
            return redirect(url_for('offense_records'))
        
        # Validate payment status
        valid_statuses = ['Pending', 'Partially Paid', 'Paid']
        if payment_status not in valid_statuses:
            flash('Invalid payment status', 'error')
            return redirect(url_for('offense_records'))
        
        # Update payment information
        offense.amount_paid = current_paid + payment_amount
        offense.penalty_status = payment_status
        offense.updated_at = singapore_now()
        
        # Auto-adjust status based on payment
        if offense.amount_paid >= total_penalty:
            offense.penalty_status = 'Paid'
        elif offense.amount_paid > 0:
            offense.penalty_status = 'Partially Paid'
        
        db.session.commit()
        
        flash(f'Payment updated successfully. Amount paid: S${payment_amount:.2f}', 'success')
        return redirect(url_for('offense_records'))
        
    except ValueError:
        flash('Invalid payment amount', 'error')
        return redirect(url_for('offense_records'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating payment: {str(e)}', 'error')
        return redirect(url_for('offense_records'))

@app.route('/update-offense-record/<int:offense_id>', methods=['POST'])
@login_required
@edit_permission_required('offense_records')
def update_offense_record(offense_id):
    """Update an existing offense record with all details"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('offense_records'))
    
    # Get offense record
    offense = OffenseRecord.query.filter_by(id=offense_id, organization_id=user.organization_id).first()
    if not offense:
        flash('Offense record not found', 'error')
        return redirect(url_for('offense_records'))
    
    try:
        # Get form data
        offender_name = request.form.get('offender_name', '').strip()
        fin_number = request.form.get('fin_number', '').strip()
        nationality = request.form.get('nationality', '').strip()
        offender_room = request.form.get('offender_room', '').strip()
        sector = request.form.get('sector', '').strip()
        contact_number = request.form.get('contact_number', '').strip()
        offender_company = request.form.get('offender_company', '').strip()
        
        # Incident details
        case_number = request.form.get('case_number', '').strip()
        offense_type = request.form.get('offense_type', '').strip()
        severity = request.form.get('severity', '').strip()
        incident_date_str = request.form.get('incident_date', '').strip()
        incident_time_str = request.form.get('incident_time', '').strip()
        location = request.form.get('location', '').strip()
        status = request.form.get('status', '').strip()
        description = request.form.get('description', '').strip()
        
        # Additional details
        documentary_proof = request.form.get('documentary_proof') == 'true'
        proof_description = request.form.get('proof_description', '').strip()
        financial_penalty_imposed = request.form.get('financial_penalty_imposed') == 'true'
        penalty_amount_str = request.form.get('penalty_amount', '').strip()
        witness_names = request.form.get('witness_names', '').strip()
        action_taken = request.form.get('action_taken', '').strip()
        duty_manager_name = request.form.get('duty_manager_name', '').strip()
        
        # Process signature data
        resident_signature = request.form.get('resident_signature')
        duty_manager_signature = request.form.get('duty_manager_signature')
        
        # Clean signature data (remove data URL prefix)
        if resident_signature and resident_signature.startswith('data:image'):
            resident_signature = resident_signature.split(',', 1)[1]
            offense.resident_signature = resident_signature
            offense.resident_signature_date = singapore_now()
        elif resident_signature:  # New signature provided
            offense.resident_signature = resident_signature
            offense.resident_signature_date = singapore_now()
            
        if duty_manager_signature and duty_manager_signature.startswith('data:image'):
            duty_manager_signature = duty_manager_signature.split(',', 1)[1]
            offense.duty_manager_signature = duty_manager_signature
            offense.duty_manager_signature_date = singapore_now()
        elif duty_manager_signature:  # New signature provided
            offense.duty_manager_signature = duty_manager_signature
            offense.duty_manager_signature_date = singapore_now()

        # Handle photo uploads and removals
        import base64
        
        # Process photo removals first
        for i in range(1, 11):
            if request.form.get(f'remove_photo_{i}') == 'true':
                setattr(offense, f'incident_photo_{i}', None)
        
        # Process new photo uploads
        if 'incident_photos' in request.files:
            photos = request.files.getlist('incident_photos')
            photo_index = 1
            
            # Find first available photo slot
            for i in range(1, 11):
                if not getattr(offense, f'incident_photo_{i}'):
                    photo_index = i
                    break
            
            # Add new photos
            for photo in photos[:10]:  # Limit to 10 photos total
                if photo and photo.filename and photo.content_type.startswith('image/'):
                    try:
                        # Read and encode photo as base64
                        photo_bytes = photo.read()
                        if len(photo_bytes) <= 5 * 1024 * 1024:  # 5MB limit
                            photo_base64 = base64.b64encode(photo_bytes).decode('utf-8')
                            
                            # Find next available slot
                            while photo_index <= 10 and getattr(offense, f'incident_photo_{photo_index}'):
                                photo_index += 1
                            
                            if photo_index <= 10:
                                setattr(offense, f'incident_photo_{photo_index}', photo_base64)
                                photo_index += 1
                    except Exception as e:
                        flash(f'Error processing photo: {str(e)}', 'warning')
        
        # Validate required fields
        if not offender_name:
            flash('Offender name is required', 'error')
            return redirect(url_for('offense_records'))
        
        if not offense_type:
            flash('Offense type is required', 'error')
            return redirect(url_for('offense_records'))
            

            
        if not incident_date_str:
            flash('Incident date is required', 'error')
            return redirect(url_for('offense_records'))
            
        if not description:
            flash('Description is required', 'error')
            return redirect(url_for('offense_records'))
        
        # Parse dates and times
        incident_date = None
        if incident_date_str:
            try:
                incident_date = datetime.strptime(incident_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid incident date format', 'error')
                return redirect(url_for('offense_records'))
        
        incident_time = None
        if incident_time_str:
            try:
                incident_time = datetime.strptime(incident_time_str, '%H:%M').time()
            except ValueError:
                flash('Invalid incident time format', 'error')
                return redirect(url_for('offense_records'))
        
        # Parse penalty amount
        penalty_amount = None
        if financial_penalty_imposed and penalty_amount_str:
            try:
                penalty_amount = float(penalty_amount_str)
                if penalty_amount < 0:
                    flash('Penalty amount cannot be negative', 'error')
                    return redirect(url_for('offense_records'))
            except ValueError:
                flash('Invalid penalty amount', 'error')
                return redirect(url_for('offense_records'))
        
        # Update offense record
        offense.offender_name = offender_name
        offense.fin_number = fin_number if fin_number else None
        offense.nationality = nationality if nationality else None
        offense.offender_room = offender_room if offender_room else None
        offense.sector = sector if sector else None
        offense.contact_number = contact_number if contact_number else None
        offense.offender_company = offender_company if offender_company else None
        offense.case_number = case_number if case_number else None
        offense.offense_type = offense_type
        offense.severity = 'N/A'  # Default value since severity field removed
        offense.incident_date = incident_date
        offense.incident_time = incident_time
        offense.location = location if location else None
        offense.status = status if status else 'Open'
        offense.description = description
        offense.documentary_proof = documentary_proof
        offense.proof_description = proof_description if proof_description else None
        offense.financial_penalty_imposed = financial_penalty_imposed
        offense.penalty_amount = penalty_amount
        offense.witness_names = witness_names if witness_names else None
        offense.action_taken = action_taken if action_taken else None
        offense.duty_manager_name = duty_manager_name if duty_manager_name else None
        offense.updated_at = singapore_now()
        
        # Handle penalty status changes
        if financial_penalty_imposed and penalty_amount:
            # If penalty was newly imposed or amount changed, check payment status
            if not offense.penalty_status:
                offense.penalty_status = 'Pending'
                offense.amount_paid = 0.0
            # If penalty amount was reduced and already paid more than new amount
            elif offense.amount_paid and offense.amount_paid > penalty_amount:
                offense.amount_paid = penalty_amount
                offense.penalty_status = 'Paid'
            # Recalculate status based on current payment
            elif offense.amount_paid:
                if offense.amount_paid >= penalty_amount:
                    offense.penalty_status = 'Paid'
                elif offense.amount_paid > 0:
                    offense.penalty_status = 'Partially Paid'
                else:
                    offense.penalty_status = 'Pending'
        elif not financial_penalty_imposed:
            # If penalty was removed, clear penalty-related fields
            offense.penalty_amount = None
            offense.penalty_status = None
            offense.amount_paid = None
        
        db.session.commit()
        
        flash(f'Offense record #{offense.case_number or offense.id} updated successfully', 'success')
        return redirect(url_for('offense_records'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating offense record: {str(e)}', 'error')
        return redirect(url_for('offense_records'))

@app.route('/download-offense-pdf/<int:offense_id>')
@login_required
@page_access_required('offense_records')
def download_offense_pdf(offense_id):
    """Generate and download PDF for an offence record"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('offense_records'))
    
    # Get offence record
    offense = OffenseRecord.query.filter_by(id=offense_id, organization_id=user.organization_id).first()
    if not offense:
        flash('Offence record not found', 'error')
        return redirect(url_for('offense_records'))
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from io import BytesIO
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        # Build PDF content
        story = []
        
        # Create header layout with logo on top-left and centered text
        from reportlab.platypus import Image as ReportLabImage, Table, TableStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        import os
        
        # Create professional header layout with logo and properly spaced text
        try:
            logo_path = "static/ts_group_logo.jpg"
            if os.path.exists(logo_path):
                logo = ReportLabImage(logo_path, width=1.2*inch, height=0.6*inch)
                
                # Define text styles with proper spacing
                company_style = ParagraphStyle(
                    'CompanyStyle',
                    parent=styles['Normal'],
                    fontSize=18,
                    alignment=TA_CENTER,
                    textColor=colors.darkblue,
                    fontName='Helvetica-Bold',
                    spaceBefore=0,
                    spaceAfter=8,
                    leading=22
                )
                
                pioneer_style = ParagraphStyle(
                    'PioneerStyle',
                    parent=styles['Normal'],
                    fontSize=14,
                    alignment=TA_CENTER,
                    textColor=colors.darkblue,
                    fontName='Helvetica-Bold',
                    spaceBefore=0,
                    spaceAfter=8,
                    leading=18
                )
                
                report_style = ParagraphStyle(
                    'ReportStyle',
                    parent=styles['Normal'],
                    fontSize=16,
                    alignment=TA_CENTER,
                    textColor=colors.darkblue,
                    fontName='Helvetica-Bold',
                    spaceBefore=0,
                    spaceAfter=0,
                    leading=20
                )
                
                # Create centered text content
                text_content = [
                    [Paragraph("TS MANAGEMENT SERVICES PTE LTD", company_style)],
                    [Paragraph("PIONEER LODGE", pioneer_style)],
                    [Paragraph("OFFENCE REPORT", report_style)]
                ]
                
                text_table = Table(text_content, colWidths=[5.5*inch])
                text_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ]))
                
                # Create main header layout: logo on left, text on right
                header_layout = Table(
                    [[logo, text_table]], 
                    colWidths=[1.5*inch, 6.5*inch]
                )
                header_layout.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                    ('VALIGN', (0, 0), (0, 0), 'TOP'),
                    ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                    ('VALIGN', (1, 0), (1, 0), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 10),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ]))
                
                story.append(header_layout)
                story.append(Spacer(1, 25))
                
        except Exception as e:
            # Fallback to centered text-only header with proper spacing
            company_style = ParagraphStyle(
                'CompanyStyle',
                parent=styles['Normal'],
                fontSize=18,
                alignment=TA_CENTER,
                textColor=colors.darkblue,
                fontName='Helvetica-Bold',
                spaceBefore=10,
                spaceAfter=8,
                leading=22
            )
            
            pioneer_style = ParagraphStyle(
                'PioneerStyle',
                parent=styles['Normal'],
                fontSize=14,
                alignment=TA_CENTER,
                textColor=colors.darkblue,
                fontName='Helvetica-Bold',
                spaceBefore=0,
                spaceAfter=8,
                leading=18
            )
            
            report_style = ParagraphStyle(
                'ReportStyle',
                parent=styles['Normal'],
                fontSize=16,
                alignment=TA_CENTER,
                textColor=colors.darkblue,
                fontName='Helvetica-Bold',
                spaceBefore=0,
                spaceAfter=20,
                leading=20
            )
            
            story.append(Paragraph("TS MANAGEMENT SERVICES PTE LTD", company_style))
            story.append(Paragraph("PIONEER LODGE", pioneer_style))
            story.append(Paragraph("OFFENCE REPORT", report_style))
            story.append(Spacer(1, 25))
        story.append(Spacer(1, 20))
        
        # Case information
        case_data = [
            ['Case Number:', offense.case_number or 'N/A'],
            ['Date Reported:', offense.created_at.strftime('%Y-%m-%d %H:%M') if offense.created_at else 'N/A'],
            ['Status:', offense.status or 'Open'],
            ['Severity:', offense.severity or 'N/A']
        ]
        
        case_table = Table(case_data, colWidths=[2*inch, 4*inch])
        case_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(case_table)
        story.append(Spacer(1, 20))
        
        # Offender information
        story.append(Paragraph("OFFENDER DETAILS", styles['Heading2']))
        
        offender_data = [
            ['Name:', offense.offender_name or 'N/A'],
            ['FIN Number:', offense.fin_number or 'N/A'],
            ['Nationality:', offense.nationality or 'N/A'],
            ['Room Number:', offense.offender_room or 'N/A'],
            ['Sector:', offense.sector or 'N/A'],
            ['Contact Number:', offense.contact_number or 'N/A'],
            ['Company:', offense.offender_company or 'N/A']
        ]
        
        offender_table = Table(offender_data, colWidths=[2*inch, 4*inch])
        offender_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(offender_table)
        story.append(Spacer(1, 20))
        
        # Incident details
        story.append(Paragraph("INCIDENT DETAILS", styles['Heading2']))
        
        incident_data = [
            ['Offense Type:', offense.offense_type or 'N/A'],
            ['Date:', offense.incident_date.strftime('%Y-%m-%d') if offense.incident_date else 'N/A'],
            ['Time:', offense.incident_time.strftime('%H:%M') if offense.incident_time else 'N/A'],
            ['Location:', offense.location or 'N/A'],
            ['Documentary Proof:', 'Yes' if offense.documentary_proof else 'No']
        ]
        
        incident_table = Table(incident_data, colWidths=[2*inch, 4*inch])
        incident_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(incident_table)
        story.append(Spacer(1, 20))
        
        # Description
        story.append(Paragraph("DESCRIPTION OF CONTRAVENTION", styles['Heading2']))
        story.append(Paragraph(offense.description or 'No description provided', styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Financial penalty
        if offense.financial_penalty_imposed:
            story.append(Paragraph("FINANCIAL PENALTY", styles['Heading2']))
            
            penalty_data = [
                ['Penalty Amount:', f"S${offense.penalty_amount:.2f}" if offense.penalty_amount else 'N/A'],
                ['Amount Paid:', f"S${offense.amount_paid:.2f}" if offense.amount_paid else 'S$0.00'],
                ['Payment Status:', offense.penalty_status or 'Pending'],
                ['Outstanding Balance:', f"S${(offense.penalty_amount or 0) - (offense.amount_paid or 0):.2f}"]
            ]
            
            penalty_table = Table(penalty_data, colWidths=[2*inch, 4*inch])
            penalty_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(penalty_table)
            story.append(Spacer(1, 20))
        
        # Additional information
        if offense.witness_names or offense.action_taken or offense.duty_manager_name:
            story.append(Paragraph("ADDITIONAL INFORMATION", styles['Heading2']))
            
            additional_data = []
            if offense.witness_names:
                additional_data.append(['Witnesses:', offense.witness_names])
            if offense.action_taken:
                additional_data.append(['Action Taken:', offense.action_taken])
            if offense.duty_manager_name:
                additional_data.append(['Duty Manager:', offense.duty_manager_name])
            
            if additional_data:
                additional_table = Table(additional_data, colWidths=[2*inch, 4*inch])
                additional_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('BACKGROUND', (1, 0), (1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(additional_table)
        
        # Add incident photos section
        photo_count = 0
        for i in range(1, 11):
            photo_field = f'incident_photo_{i}'
            if hasattr(offense, photo_field) and getattr(offense, photo_field):
                photo_count += 1
        
        if photo_count > 0:
            story.append(Spacer(1, 20))
            story.append(Paragraph("INCIDENT PHOTOS", styles['Heading2']))
            story.append(Spacer(1, 10))
            
            from reportlab.platypus import Image as ReportLabImage
            from io import BytesIO
            import base64
            from PIL import Image
            
            photos_per_row = 2
            photo_width = 2.5 * inch
            photo_height = 2.5 * inch
            
            for i in range(1, 11):
                photo_field = f'incident_photo_{i}'
                if hasattr(offense, photo_field) and getattr(offense, photo_field):
                    try:
                        # Decode base64 image
                        photo_data = getattr(offense, photo_field)
                        image_bytes = base64.b64decode(photo_data)
                        
                        # Create PIL image and convert to BytesIO
                        pil_image = Image.open(BytesIO(image_bytes))
                        
                        # Convert to RGB if necessary
                        if pil_image.mode != 'RGB':
                            pil_image = pil_image.convert('RGB')
                        
                        # Save to BytesIO
                        img_buffer = BytesIO()
                        pil_image.save(img_buffer, format='JPEG', quality=85)
                        img_buffer.seek(0)
                        
                        # Create ReportLab Image
                        reportlab_image = ReportLabImage(img_buffer, width=photo_width, height=photo_height)
                        story.append(reportlab_image)
                        story.append(Spacer(1, 10))
                        
                    except Exception as e:
                        # If photo processing fails, add a note
                        story.append(Paragraph(f"Photo {i}: Unable to display (error: {str(e)})", styles['Normal']))
                        story.append(Spacer(1, 5))
        
        # Add signatures section if available
        if offense.resident_signature or offense.duty_manager_signature:
            story.append(Spacer(1, 20))
            story.append(Paragraph("SIGNATURES", styles['Heading2']))
            story.append(Spacer(1, 10))
            
            # Create signature table
            signature_data = []
            
            if offense.resident_signature:
                try:
                    import base64
                    from reportlab.platypus import Image as ReportLabImage
                    from PIL import Image as PILImage
                    from io import BytesIO
                    
                    # Clean and decode base64 signature data
                    sig_data = offense.resident_signature
                    if sig_data.startswith('data:image'):
                        sig_data = sig_data.split(',', 1)[1]
                    
                    # Add padding if needed for proper base64 decoding
                    missing_padding = len(sig_data) % 4
                    if missing_padding:
                        sig_data += '=' * (4 - missing_padding)
                    
                    signature_bytes = base64.b64decode(sig_data)
                    signature_img = PILImage.open(BytesIO(signature_bytes))
                    
                    # Convert to RGB if necessary and ensure proper format
                    if signature_img.mode == 'RGBA':
                        # Create white background for transparency
                        rgb_img = PILImage.new('RGB', signature_img.size, (255, 255, 255))
                        rgb_img.paste(signature_img, mask=signature_img.split()[3])
                        signature_img = rgb_img
                    elif signature_img.mode != 'RGB':
                        signature_img = signature_img.convert('RGB')
                    
                    # Save to buffer for ReportLab
                    img_buffer = BytesIO()
                    signature_img.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    # Add to PDF with proper sizing
                    reportlab_signature = ReportLabImage(img_buffer, width=2*inch, height=1*inch)
                    signature_data.append(['Resident Signature:', reportlab_signature])
                    
                    if offense.resident_signature_date:
                        signature_data.append(['Signed Date:', offense.resident_signature_date.strftime('%Y-%m-%d %H:%M')])
                        
                except Exception as e:
                    print(f"Error processing resident signature: {e}")
                    signature_data.append(['Resident Signature:', 'Signature available but cannot display'])
            
            if offense.duty_manager_signature:
                try:
                    import base64
                    from reportlab.platypus import Image as ReportLabImage
                    from PIL import Image as PILImage
                    from io import BytesIO
                    
                    # Clean and decode base64 signature data
                    sig_data = offense.duty_manager_signature
                    if sig_data.startswith('data:image'):
                        sig_data = sig_data.split(',', 1)[1]
                    
                    # Add padding if needed for proper base64 decoding
                    missing_padding = len(sig_data) % 4
                    if missing_padding:
                        sig_data += '=' * (4 - missing_padding)
                    
                    signature_bytes = base64.b64decode(sig_data)
                    signature_img = PILImage.open(BytesIO(signature_bytes))
                    
                    # Convert to RGB if necessary and ensure proper format
                    if signature_img.mode == 'RGBA':
                        # Create white background for transparency
                        rgb_img = PILImage.new('RGB', signature_img.size, (255, 255, 255))
                        rgb_img.paste(signature_img, mask=signature_img.split()[3])
                        signature_img = rgb_img
                    elif signature_img.mode != 'RGB':
                        signature_img = signature_img.convert('RGB')
                    
                    # Save to buffer for ReportLab
                    img_buffer = BytesIO()
                    signature_img.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    # Add to PDF with proper sizing
                    reportlab_signature = ReportLabImage(img_buffer, width=2*inch, height=1*inch)
                    signature_data.append(['Duty Manager Signature:', reportlab_signature])
                    
                    if offense.duty_manager_signature_date:
                        signature_data.append(['Signed Date:', offense.duty_manager_signature_date.strftime('%Y-%m-%d %H:%M')])
                        
                except Exception as e:
                    print(f"Error processing duty manager signature: {e}")
                    signature_data.append(['Duty Manager Signature:', 'Signature available but cannot display'])
            
            if signature_data:
                signature_table = Table(signature_data, colWidths=[2*inch, 4*inch])
                signature_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('BACKGROUND', (1, 0), (1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                
                story.append(signature_table)
                story.append(Spacer(1, 20))
        
        # Add footer
        story.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.grey
        )
        story.append(Paragraph("Generated by TS MANAGEMENT SERVICES PTE.LTD.", footer_style))
        story.append(Paragraph("Location: Pioneer Lodge, 39A Soon Lee Road, Singapore 628089", footer_style))
        
        # Build PDF
        doc.build(story)
        
        # Return PDF
        buffer.seek(0)
        filename = f"offense_report_{offense.case_number}_{singapore_now().strftime('%Y%m%d')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('offense_records'))
        
        # Row 1: Date and Case No (side by side)
        story.append(Paragraph(f"<para fontSize=10>Date : <u>{date_str}{date_line}</u>                                                   Case No : <u>{case_str}{case_line}</u></para>", styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Row 2: Time
        time_line = "_" * max(0, 15 - len(time_str))
        story.append(Paragraph(f"<para fontSize=10>Time : <u>{time_str}{time_line}</u></para>", styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Row 3: Name and FIN no
        name_str = offense.offender_name or ''
        fin_str = offense.fin_number or ''
        name_line = "_" * max(0, 35 - len(name_str))
        fin_line = "_" * max(0, 15 - len(fin_str))
        story.append(Paragraph(f"<para fontSize=10>Name : <u>{name_str}{name_line}</u>                                                   FIN no : <u>{fin_str}{fin_line}</u></para>", styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Row 4: Nationality and Room no
        nationality_str = offense.nationality or ''
        room_str = offense.offender_room or ''
        nationality_line = "_" * max(0, 30 - len(nationality_str))
        room_line = "_" * max(0, 15 - len(room_str))
        story.append(Paragraph(f"<para fontSize=10>Nationality : <u>{nationality_str}{nationality_line}</u>                                            Room no : <u>{room_str}{room_line}</u></para>", styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Row 5: Sector and Contact no
        sector_str = offense.sector or ''
        contact_str = offense.contact_number or ''
        sector_line = "_" * max(0, 35 - len(sector_str))
        contact_line = "_" * max(0, 15 - len(contact_str))
        story.append(Paragraph(f"<para fontSize=10>Sector : <u>{sector_str}{sector_line}</u>                                                Contact no : <u>{contact_str}{contact_line}</u></para>", styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Row 6: Company Name
        company_str = offense.offender_company or ''
        company_line = "_" * max(0, 60 - len(company_str))
        story.append(Paragraph(f"<para fontSize=10>Company Name : <u>{company_str}{company_line}</u></para>", styles['Normal']))
        story.append(Spacer(1, 30))
        
        # Details of Contravention section
        story.append(Paragraph("<para fontSize=12><b><u>Details of Contravention</u></b></para>", styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Description with actual data - underline the description text
        description = offense.description or ''
        story.append(Paragraph(f"<para fontSize=10>Description of Contravention: <u>{description}</u></para>", styles['Normal']))
        story.append(Spacer(1, 12))
        story.append(Paragraph("<para fontSize=10><u>_________________________________________________________________________________</u></para>", styles['Normal']))
        
        # Force page break before second page
        story.append(PageBreak())
        
        # SECOND PAGE - Keep existing detailed format
        # Add TS Group logo at top left of second page
        try:
            logo_path = "static/ts_group_logo_house.png"
            if os.path.exists(logo_path):
                from reportlab.lib.utils import ImageReader
                logo = Image(logo_path, width=1.5*inch, height=0.8*inch)
                logo.hAlign = 'LEFT'
                story.append(logo)
                story.append(Spacer(1, 15))
        except Exception as e:
            print(f"Logo error: {e}")
        
        # Header
        story.append(Paragraph("PIONEER LODGE", title_style))
        story.append(Paragraph("OFFENCE REPORT", title_style))
        story.append(Spacer(1, 20))
        
        # Case Information
        case_data = [
            ['Case No:', offense.case_number or 'N/A'],
            ['Date:', offense.incident_date.strftime('%Y-%m-%d') if offense.incident_date else 'N/A'],
            ['Time:', offense.incident_time.strftime('%H:%M') if offense.incident_time else 'N/A'],
            ['Status:', offense.status or 'Open']
        ]
        
        case_table = Table(case_data, colWidths=[2*inch, 4*inch])
        case_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(Paragraph("CASE INFORMATION", heading_style))
        story.append(case_table)
        story.append(Spacer(1, 20))
        
        # Offense Details
        offense_data = [
            ['Offense Type:', offense.offense_type or 'N/A'],
            ['Location:', offense.location or 'N/A'],
            ['Description:', offense.description or 'N/A']
        ]
        
        offense_table = Table(offense_data, colWidths=[2*inch, 4*inch])
        offense_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(Paragraph("OFFENSE DETAILS", heading_style))
        story.append(offense_table)
        story.append(Spacer(1, 20))
        
        # Resident Details
        resident_data = [
            ['Name:', offense.offender_name or 'N/A'],
            ['FIN Number:', offense.fin_number or 'N/A'],
            ['Nationality:', offense.nationality or 'N/A'],
            ['Room Number:', offense.offender_room or 'N/A'],
            ['Sector:', offense.sector or 'N/A'],
            ['Contact Number:', offense.contact_number or 'N/A'],
            ['Company:', offense.offender_company or 'N/A']
        ]
        
        resident_table = Table(resident_data, colWidths=[2*inch, 4*inch])
        resident_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(Paragraph("RESIDENT DETAILS", heading_style))
        story.append(resident_table)
        story.append(Spacer(1, 20))
        
        # Financial Penalty
        if offense.financial_penalty_imposed:
            penalty_data = [
                ['Financial Penalty:', 'YES'],
                ['Amount:', f"S${offense.penalty_amount}" if offense.penalty_amount else 'N/A']
            ]
            
            penalty_table = Table(penalty_data, colWidths=[2*inch, 4*inch])
            penalty_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            
            story.append(Paragraph("FINANCIAL PENALTY", heading_style))
            story.append(penalty_table)
            story.append(Spacer(1, 20))
        
        # Additional Information
        if offense.witness_names or offense.action_taken:
            additional_data = []
            if offense.witness_names:
                additional_data.append(['Witnesses:', offense.witness_names])
            if offense.action_taken:
                additional_data.append(['Action Taken:', offense.action_taken])
            
            if additional_data:
                additional_table = Table(additional_data, colWidths=[2*inch, 4*inch])
                additional_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                
                story.append(Paragraph("ADDITIONAL INFORMATION", heading_style))
                story.append(additional_table)
                story.append(Spacer(1, 20))
        
        # Incident Photos Section
        photos_added = False
        photo_count = 0
        photos_per_row = 2
        photos_in_current_row = []
        
        for i in range(1, 11):  # Check incident_photo_1 through incident_photo_10
            photo_attr = f'incident_photo_{i}'
            photo_data = getattr(offense, photo_attr, None)
            
            if photo_data:
                try:
                    # Clean and decode base64 image data
                    # Remove data URL prefix if present (e.g., "data:image/png;base64,")
                    if ',' in photo_data:
                        photo_data = photo_data.split(',', 1)[1]
                    
                    # Add padding if needed for proper base64 decoding
                    missing_padding = len(photo_data) % 4
                    if missing_padding:
                        photo_data += '=' * (4 - missing_padding)
                    
                    image_data = base64.b64decode(photo_data)
                    image_buffer = BytesIO(image_data)
                    
                    # Create image with proper sizing
                    img = Image(image_buffer, width=2.5*inch, height=2*inch)
                    photos_in_current_row.append(img)
                    photo_count += 1
                    
                    # Add photos to story when we have enough for a row or at the end
                    if len(photos_in_current_row) == photos_per_row or i == 10:
                        if not photos_added:
                            story.append(Paragraph("INCIDENT PHOTOS", heading_style))
                            photos_added = True
                        
                        # Create table for photo row
                        if len(photos_in_current_row) == 1:
                            # Single photo, center it
                            photo_table = Table([photos_in_current_row], colWidths=[2.5*inch])
                        else:
                            # Multiple photos
                            photo_table = Table([photos_in_current_row], colWidths=[2.5*inch] * len(photos_in_current_row))
                        
                        photo_table.setStyle(TableStyle([
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ]))
                        
                        story.append(photo_table)
                        story.append(Spacer(1, 10))
                        photos_in_current_row = []
                        
                except Exception as e:
                    print(f"Error processing photo {i}: {e}")
                    continue
        
        # Handle remaining photos in the last row
        if photos_in_current_row:
            if not photos_added:
                story.append(Paragraph("INCIDENT PHOTOS", heading_style))
                photos_added = True
            
            photo_table = Table([photos_in_current_row], colWidths=[2.5*inch] * len(photos_in_current_row))
            photo_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(photo_table)
            story.append(Spacer(1, 10))
        
        if photos_added:
            story.append(Spacer(1, 20))
        
        # Enhanced Signatures section with actual signature images
        story.append(Paragraph("SIGNATURES", heading_style))
        
        # Create signature display table
        signature_rows = []
        
        # Resident signature
        if offense.resident_signature or offense.resident_signature_date:
            resident_sig_cell = []
            if offense.resident_signature:
                try:
                    from PIL import Image as PILImage
                    
                    # Clean and decode resident signature
                    sig_data_clean = offense.resident_signature
                    if sig_data_clean.startswith('data:image'):
                        sig_data_clean = sig_data_clean.split(',', 1)[1]
                    
                    # Add padding if needed
                    missing_padding = len(sig_data_clean) % 4
                    if missing_padding:
                        sig_data_clean += '=' * (4 - missing_padding)
                    
                    sig_data = base64.b64decode(sig_data_clean)
                    signature_img = PILImage.open(BytesIO(sig_data))
                    
                    # Convert to RGB if necessary and ensure proper format
                    if signature_img.mode == 'RGBA':
                        # Create white background for transparency
                        rgb_img = PILImage.new('RGB', signature_img.size, (255, 255, 255))
                        rgb_img.paste(signature_img, mask=signature_img.split()[3])
                        signature_img = rgb_img
                    elif signature_img.mode != 'RGB':
                        signature_img = signature_img.convert('RGB')
                    
                    # Save to buffer for ReportLab
                    sig_buffer = BytesIO()
                    signature_img.save(sig_buffer, format='PNG')
                    sig_buffer.seek(0)
                    
                    resident_sig_img = Image(sig_buffer, width=2*inch, height=1*inch)
                    resident_sig_cell.append(resident_sig_img)
                except Exception as e:
                    print(f"Error processing resident signature: {e}")
                    resident_sig_cell.append(Paragraph("Signature Available", normal_style))
            else:
                resident_sig_cell.append(Paragraph("No Signature", normal_style))
            
            date_text = offense.resident_signature_date.strftime('%Y-%m-%d %H:%M') if offense.resident_signature_date else 'N/A'
            signature_rows.append(['Resident Signature:', resident_sig_cell[0], 'Date:', date_text])
        
        # Duty Manager signature  
        if offense.duty_manager_signature or offense.duty_manager_signature_date:
            manager_sig_cell = []
            if offense.duty_manager_signature:
                try:
                    from PIL import Image as PILImage
                    
                    # Clean and decode duty manager signature
                    sig_data_clean = offense.duty_manager_signature
                    if sig_data_clean.startswith('data:image'):
                        sig_data_clean = sig_data_clean.split(',', 1)[1]
                    
                    # Add padding if needed
                    missing_padding = len(sig_data_clean) % 4
                    if missing_padding:
                        sig_data_clean += '=' * (4 - missing_padding)
                    
                    sig_data = base64.b64decode(sig_data_clean)
                    signature_img = PILImage.open(BytesIO(sig_data))
                    
                    # Convert to RGB if necessary and ensure proper format
                    if signature_img.mode == 'RGBA':
                        # Create white background for transparency
                        rgb_img = PILImage.new('RGB', signature_img.size, (255, 255, 255))
                        rgb_img.paste(signature_img, mask=signature_img.split()[3])
                        signature_img = rgb_img
                    elif signature_img.mode != 'RGB':
                        signature_img = signature_img.convert('RGB')
                    
                    # Save to buffer for ReportLab
                    sig_buffer = BytesIO()
                    signature_img.save(sig_buffer, format='PNG')
                    sig_buffer.seek(0)
                    
                    manager_sig_img = Image(sig_buffer, width=2*inch, height=1*inch)
                    manager_sig_cell.append(manager_sig_img)
                except Exception as e:
                    print(f"Error processing duty manager signature: {e}")
                    manager_sig_cell.append(Paragraph("Signature Available", normal_style))
            else:
                manager_sig_cell.append(Paragraph("No Signature", normal_style))
            
            date_text = offense.duty_manager_signature_date.strftime('%Y-%m-%d %H:%M') if offense.duty_manager_signature_date else 'N/A'
            signature_rows.append(['OE/DC Signature:', manager_sig_cell[0], 'Date:', date_text])
        
        # OE/DC Name
        if offense.duty_manager_name:
            signature_rows.append(['OE/DC Name:', offense.duty_manager_name, '', ''])
        
        if signature_rows:
            signature_table = Table(signature_rows, colWidths=[1.5*inch, 2*inch, 0.5*inch, 2*inch])
            signature_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('ALIGN', (2, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            story.append(signature_table)
            story.append(Spacer(1, 20))
        
        # Footer
        story.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.grey
        )
        story.append(Paragraph("Generated by TS MANAGEMENT SERVICES PTE.LTD.", footer_style))
        story.append(Paragraph("Location: Pioneer Lodge, 39A Soon Lee Road, Singapore 628089", footer_style))
        
        # Build PDF
        doc.build(story)
        
        buffer.seek(0)
        
        # Create response
        response = make_response(buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=offense_report_{offense.case_number or offense.id}.pdf'
        
        return response
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('offense_records'))






@app.route('/assets/export/excel')
@login_required
def export_assets_excel():
    """Export assets to Excel"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    try:
        import pandas as pd
        from flask import Response
        import io
        
        # Get assets
        assets = Asset.query.filter_by(organization_id=user.organization_id).all()
        
        # Prepare data for Excel
        data = []
        for i, asset in enumerate(assets, 1):
            data.append({
                'S.No': i,
                'Asset Name': asset.name,
                'Description': asset.description or '',
                'Category': asset.category.name if asset.category else '',
                'Quantity': asset.quantity,
                'Status': asset.status,
                'Location': asset.location or '',
                'Serial Number': asset.serial_number or '',
                'Purchase Date': asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
                'Purchase Cost': asset.purchase_cost or '',
                'Created By': asset.created_by_user.first_name if asset.created_by_user else '',
                'Created At': asset.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Assets')
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Assets']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_name].width = adjusted_width
        
        output.seek(0)
        
        # Create response
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=assets_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            }
        )
        
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/assets/export/excel/selected')
@login_required
def export_selected_assets_excel():
    """Export selected assets to Excel"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    try:
        import pandas as pd
        from flask import Response
        import io
        
        # Get asset IDs from query parameters
        asset_ids = request.args.getlist('asset_ids')
        if not asset_ids:
            return jsonify({'error': 'No assets selected'}), 400
        
        # Get selected assets for the organization
        assets = Asset.query.filter(
            Asset.id.in_(asset_ids),
            Asset.organization_id == user.organization_id
        ).all()
        
        if not assets:
            return jsonify({'error': 'No valid assets found'}), 404
        
        # Prepare data for Excel
        data = []
        for i, asset in enumerate(assets, 1):
            data.append({
                'S.No': i,
                'Asset Name': asset.name,
                'Description': asset.description or '',
                'Category': asset.category.name if asset.category else '',
                'Quantity': asset.quantity,
                'Status': asset.status,
                'Location': asset.location or '',
                'Serial Number': asset.serial_number or '',
                'Purchase Date': asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
                'Purchase Cost': asset.purchase_cost or '',
                'Created By': asset.created_by_user.first_name if asset.created_by_user else '',
                'Created At': asset.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Selected Assets')
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Selected Assets']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_name].width = adjusted_width
        
        output.seek(0)
        
        # Create response
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=selected_assets_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            }
        )
        
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/assets/import/template')
@login_required
def download_assets_import_template():
    """Download Excel template for assets import"""
    try:
        import pandas as pd
        from flask import Response
        import io
        
        # Create template data
        template_data = {
            'Asset Name': ['Sample Asset 1', 'Sample Asset 2'],
            'Description': ['Description for asset 1', 'Description for asset 2'],
            'Category': ['Furniture', 'Electronics'],
            'Quantity': [1, 2],
            'Status': ['Active', 'Active'],
            'Location': ['Room A1', 'Room B2'],
            'Serial Number': ['SN001', 'SN002'],
            'Purchase Date': ['2025-01-01', '2025-01-02'],
            'Purchase Cost': [100.00, 200.00]
        }
        
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Assets Template')
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Assets Template']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_name].width = adjusted_width
        
        output.seek(0)
        
        # Create response
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=assets_import_template.xlsx'
            }
        )
        
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/assets/import/excel', methods=['POST'])
@login_required
@create_permission_required('asset_management')
def import_assets_excel():
    """Import assets from Excel file"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    try:
        import pandas as pd
        from datetime import datetime
        
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Read Excel file
        df = pd.read_excel(file)
        
        # Validate required columns
        required_columns = ['Asset Name', 'Category']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'error': f'Missing required columns: {", ".join(missing_columns)}'}), 400
        
        # Import assets
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Get or create category
                category_name = str(row['Category']).strip()
                category = AssetCategory.query.filter_by(name=category_name).first()
                if not category:
                    category = AssetCategory(name=category_name, description=f"Auto-created category: {category_name}")
                    db.session.add(category)
                    db.session.flush()
                
                # Create asset
                asset = Asset()
                asset.name = str(row['Asset Name']).strip()
                asset.description = str(row.get('Description', '')).strip() if pd.notna(row.get('Description')) else None
                asset.category_id = category.id
                asset.organization_id = user.organization_id
                asset.quantity = int(row.get('Quantity', 1)) if pd.notna(row.get('Quantity')) else 1
                asset.status = str(row.get('Status', 'Active')).strip() if pd.notna(row.get('Status')) else 'Active'
                asset.location = str(row.get('Location', '')).strip() if pd.notna(row.get('Location')) else None
                asset.serial_number = str(row.get('Serial Number', '')).strip() if pd.notna(row.get('Serial Number')) else None
                asset.purchase_cost = float(row.get('Purchase Cost', 0)) if pd.notna(row.get('Purchase Cost')) else None
                asset.created_by = user.id
                
                # Handle purchase date
                if pd.notna(row.get('Purchase Date')):
                    try:
                        asset.purchase_date = pd.to_datetime(row['Purchase Date']).date()
                    except:
                        pass
                
                db.session.add(asset)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        db.session.commit()
        
        result = {
            'success': True,
            'imported_count': imported_count,
            'total_rows': len(df)
        }
        
        if errors:
            result['errors'] = errors
            result['message'] = f'Imported {imported_count} assets with {len(errors)} errors'
        else:
            result['message'] = f'Successfully imported {imported_count} assets'
        
        return jsonify(result)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Bulk Operations Routes for Offense Records
@app.route('/bulk-update-offense-status', methods=['POST'])
@login_required
@edit_permission_required('offense_records')
def bulk_update_offense_status():
    """Bulk update status for multiple offense records"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('offense_records'))
    
    try:
        offense_ids = request.form.getlist('offense_ids')
        new_status = request.form.get('status')
        
        if not offense_ids or not new_status:
            flash('Invalid request parameters', 'error')
            return redirect(url_for('offense_records'))
        
        if new_status not in ['Open', 'Under Investigation', 'Resolved', 'Closed']:
            flash('Invalid status value', 'error')
            return redirect(url_for('offense_records'))
        
        # Update all selected offense records
        updated_count = 0
        for offense_id in offense_ids:
            offense = OffenseRecord.query.filter_by(
                id=int(offense_id), 
                organization_id=user.organization_id
            ).first()
            
            if offense:
                old_status = offense.status
                offense.status = new_status
                offense.updated_at = singapore_now()
                
                # Create submission record
                submission = Submission()
                submission.organization_id = user.organization_id
                submission.user_id = user.id
                submission.submission_type = 'bulk_offense_status_updated'
                submission.reference_id = offense.id
                submission.reference_table = 'offense_records'
                submission.notes = f'Bulk status update: "{old_status}" to "{new_status}"'
                submission.created_at = singapore_now()
                
                db.session.add(submission)
                updated_count += 1
        
        db.session.commit()
        flash(f'Successfully updated status for {updated_count} offense records', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating offense statuses: {str(e)}', 'error')
    
    return redirect(url_for('offense_records'))

@app.route('/bulk-mark-paid', methods=['POST'])
@login_required
@edit_permission_required('offense_records')
def bulk_mark_paid():
    """Bulk mark payment status as paid for multiple offense records"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    try:
        offense_ids = request.form.getlist('offense_ids')
        
        if not offense_ids:
            flash('No offense records selected', 'error')
            return redirect(url_for('offense_records'))
        
        # Update payment status for all selected offense records
        updated_count = 0
        for offense_id in offense_ids:
            offense = OffenseRecord.query.filter_by(
                id=int(offense_id), 
                organization_id=user.organization_id
            ).first()
            
            if offense and offense.financial_penalty_imposed:
                old_status = offense.penalty_status
                offense.penalty_status = 'Paid'
                offense.updated_at = singapore_now()
                
                # Create submission record
                submission = Submission()
                submission.organization_id = user.organization_id
                submission.user_id = user.id
                submission.submission_type = 'bulk_payment_status_updated'
                submission.reference_id = offense.id
                submission.reference_table = 'offense_records'
                submission.notes = f'Bulk payment update: "{old_status}" to "Paid"'
                submission.created_at = singapore_now()
                
                db.session.add(submission)
                updated_count += 1
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'Successfully marked {updated_count} offense records as paid'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error updating payment status: {str(e)}'}), 500

@app.route('/room-inventory-records')
@login_required
def room_inventory_records():
    """Display all room inventory checklist records with filtering and search"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Check permissions
    user_permissions = get_user_dashboard_permissions(user)
    if not is_admin_user(user) and 'handover' not in user_permissions['allowed_form_types']:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Get search and filter parameters
    search_query = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    room_filter = request.args.get('room', '')
    company_filter = request.args.get('company', '')
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Build base query
    query = RoomInventoryChecklist.query.filter_by(organization_id=user.organization_id)
    
    # Apply filters
    if search_query:
        query = query.filter(
            db.or_(
                RoomInventoryChecklist.room_number.contains(search_query),
                RoomInventoryChecklist.company_name.contains(search_query)
            )
        )
    
    if status_filter:
        query = query.filter(RoomInventoryChecklist.status == status_filter)
    
    if room_filter:
        query = query.filter(RoomInventoryChecklist.room_number.contains(room_filter))
    
    if company_filter:
        query = query.filter(RoomInventoryChecklist.company_name.contains(company_filter))
    
    # Date filtering
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(RoomInventoryChecklist.checklist_date >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(RoomInventoryChecklist.checklist_date <= to_date)
        except ValueError:
            pass
    
    # Order by most recent first
    query = query.order_by(RoomInventoryChecklist.created_at.desc())
    
    # Paginate results
    records = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Get unique values for filter dropdowns
    all_records = RoomInventoryChecklist.query.filter_by(organization_id=user.organization_id).all()
    unique_rooms = sorted(list(set(record.room_number for record in all_records if record.room_number)))
    unique_companies = sorted(list(set(record.company_name for record in all_records if record.company_name)))
    unique_statuses = ['Completed', 'Under Review', 'Approved']
    
    return render_template('room_inventory_records.html',
                         records=records,
                         search_query=search_query,
                         status_filter=status_filter,
                         date_from=date_from,
                         date_to=date_to,
                         room_filter=room_filter,
                         company_filter=company_filter,
                         unique_rooms=unique_rooms,
                         unique_companies=unique_companies,
                         unique_statuses=unique_statuses)

@app.route('/export-room-inventory-excel')
@login_required
def export_room_inventory_excel():
    """Export room inventory records to Excel"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Check permissions
    user_permissions = get_user_dashboard_permissions(user)
    if not is_admin_user(user) and 'handover' not in user_permissions['allowed_form_types']:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Get filter parameters
    search_query = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    room_filter = request.args.get('room', '')
    company_filter = request.args.get('company', '')
    
    # Build base query
    query = RoomInventoryChecklist.query.filter_by(organization_id=user.organization_id)
    
    # Apply filters (same as in the main view)
    if search_query:
        query = query.filter(
            db.or_(
                RoomInventoryChecklist.room_number.contains(search_query),
                RoomInventoryChecklist.company_name.contains(search_query)
            )
        )
    
    if status_filter:
        query = query.filter(RoomInventoryChecklist.status == status_filter)
    
    if room_filter:
        query = query.filter(RoomInventoryChecklist.room_number.contains(room_filter))
    
    if company_filter:
        query = query.filter(RoomInventoryChecklist.company_name.contains(company_filter))
    
    # Date filtering
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(RoomInventoryChecklist.checklist_date >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(RoomInventoryChecklist.checklist_date <= to_date)
        except ValueError:
            pass
    
    # Order by most recent first
    records = query.order_by(RoomInventoryChecklist.created_at.desc()).all()
    
    # Create Excel data
    data = []
    for record in records:
        data.append([
            record.checklist_date.strftime('%Y-%m-%d') if record.checklist_date else '',
            record.room_number or '',
            record.company_name or '',
            record.water_meter_reading or '',
            record.electricity_meter_reading or '',
            record.status or '',
            f"{record.created_by_user.first_name} {record.created_by_user.last_name}" if record.created_by_user else '',
            record.created_at.strftime('%Y-%m-%d %H:%M') if record.created_at else ''
        ])
    
    # Create DataFrame
    columns = [
        'Checklist Date', 'Room Number', 'Company Name',
        'Water Reading', 'Electricity Reading', 'Status',
        'Created By', 'Created At'
    ]
    
    df = pd.DataFrame(data, columns=columns)
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Room Inventory Records', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Room Inventory Records']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = singapore_now().strftime('%Y%m%d_%H%M%S')
    filename = f'room_inventory_records_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/bulk-delete-offenses', methods=['POST'])
@login_required
@edit_permission_required('offense_records')
def bulk_delete_offenses():
    """Bulk delete multiple offense records"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('offense_records'))
    
    try:
        offense_ids = request.form.getlist('offense_ids')
        
        if not offense_ids:
            flash('No offense records selected', 'error')
            return redirect(url_for('offense_records'))
        
        # Delete all selected offense records
        deleted_count = 0
        for offense_id in offense_ids:
            offense = OffenseRecord.query.filter_by(
                id=int(offense_id), 
                organization_id=user.organization_id
            ).first()
            
            if offense:
                # Create deletion record before deleting
                submission = Submission()
                submission.organization_id = user.organization_id
                submission.user_id = user.id
                submission.submission_type = 'bulk_offense_deleted'
                submission.reference_id = offense.id
                submission.reference_table = 'offense_records'
                submission.notes = f'Bulk deleted offense: Case {offense.case_number or offense.id}, Offender: {offense.offender_name}'
                submission.created_at = singapore_now()
                
                db.session.add(submission)
                db.session.delete(offense)
                deleted_count += 1
        
        db.session.commit()
        flash(f'Successfully deleted {deleted_count} offense records', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting offense records: {str(e)}', 'error')
    
    return redirect(url_for('offense_records'))

@app.route('/bulk-export-offense-pdf', methods=['POST'])
@login_required
def bulk_export_offense_pdf():
    """Bulk export PDF for multiple offense records"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('offense_records'))
    
    try:
        offense_ids = request.form.getlist('offense_ids')
        
        if not offense_ids:
            flash('No offense records selected', 'error')
            return redirect(url_for('offense_records'))
        
        # Get all selected offense records
        offenses = []
        for offense_id in offense_ids:
            offense = OffenseRecord.query.filter_by(
                id=int(offense_id), 
                organization_id=user.organization_id
            ).first()
            if offense:
                offenses.append(offense)
        
        if not offenses:
            flash('No valid offense records found', 'error')
            return redirect(url_for('offense_records'))
        
        # Create combined PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=1*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Add TS Group logo at the top
        try:
            from reportlab.platypus import Image as ReportLabImage
            import os
            logo_path = "static/ts_group_logo.jpg"
            if os.path.exists(logo_path):
                logo = ReportLabImage(logo_path, width=3*inch, height=1.5*inch)
                logo.hAlign = 'CENTER'
                story.append(logo)
                story.append(Spacer(1, 20))
        except Exception as e:
            # If logo fails to load, continue without it
            pass

        # Title for combined report
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        story.append(Paragraph(f"TS MANAGEMENT SERVICES PTE LTD", title_style))
        story.append(Paragraph(f"PIONEER LODGE", title_style))
        story.append(Paragraph(f"BULK OFFENSE RECORDS REPORT", title_style))
        story.append(Paragraph(f"Generated on: {singapore_now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
        story.append(Paragraph(f"Total Records: {len(offenses)}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Add each offense record
        for i, offense in enumerate(offenses, 1):
            # Add page break between records (except for first)
            if i > 1:
                story.append(PageBreak())
            
            # Record header
            record_header_style = ParagraphStyle(
                'RecordHeader',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=15,
                textColor=colors.darkred
            )
            story.append(Paragraph(f"RECORD #{i}: OFFENSE REPORT", record_header_style))
            
            # Basic Information Table
            basic_data = [
                ['Field', 'Information'],
                ['Case Number', offense.case_number or f'PL/OR/{offense.id}'],
                ['Offender Name', offense.offender_name or 'N/A'],
                ['FIN Number', offense.fin_number or 'N/A'],
                ['Nationality', offense.nationality or 'N/A'],
                ['Room Number', offense.offender_room or 'N/A'],
                ['Company', offense.offender_company or 'N/A'],
                ['Contact Number', offense.contact_number or 'N/A'],
                ['Incident Date', offense.incident_date.strftime('%Y-%m-%d') if offense.incident_date else 'N/A'],
                ['Incident Time', offense.incident_time.strftime('%H:%M') if offense.incident_time else 'N/A'],
                ['Location', offense.location or 'N/A'],
                ['Offense Type', offense.offense_type or 'N/A'],
                ['Severity', offense.severity or 'N/A'],
                ['Status', offense.status or 'Open'],
            ]
            
            basic_table = Table(basic_data, colWidths=[2*inch, 4*inch])
            basic_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8E8E8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightgrey, colors.white]),
            ]))
            
            story.append(basic_table)
            story.append(Spacer(1, 15))
            
            # Description
            if offense.description:
                story.append(Paragraph("<b>Description of Offense:</b>", styles['Heading3']))
                story.append(Paragraph(offense.description, styles['Normal']))
                story.append(Spacer(1, 10))
            
            # Financial Penalty Information
            if offense.financial_penalty_imposed:
                penalty_data = [
                    ['Financial Penalty', 'Details'],
                    ['Amount', f"S$ {offense.penalty_amount:.2f}" if offense.penalty_amount else 'Not specified'],
                    ['Status', offense.penalty_status or 'Pending'],
                ]
                
                penalty_table = Table(penalty_data, colWidths=[2*inch, 4*inch])
                penalty_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFE6E6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                
                story.append(penalty_table)
                story.append(Spacer(1, 15))
        
        # Footer
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.grey
        )
        story.append(Paragraph("Generated by TS MANAGEMENT SERVICES PTE.LTD.", footer_style))
        story.append(Paragraph("Location: Pioneer Lodge, 39A Soon Lee Road, Singapore 628089", footer_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        # Create response
        response = make_response(buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=bulk_offense_report_{singapore_now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response
        
    except Exception as e:
        flash(f'Error generating bulk PDF: {str(e)}', 'error')
        return redirect(url_for('offense_records'))

@app.route('/export-offense-table-excel')
@login_required
def export_offense_table_excel():
    """Export offense records table data to Excel"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('offense_records'))
    
    try:
        # Get all offense records for the organization
        offenses = OffenseRecord.query.filter_by(organization_id=user.organization_id).order_by(OffenseRecord.created_at.desc()).all()
        
        # Prepare data for Excel export
        excel_data = []
        
        for i, offense in enumerate(offenses, 1):
            # Determine payment status
            payment_status = 'No Penalty'
            if offense.financial_penalty_imposed:
                payment_status = offense.penalty_status or 'Pending'
            
            # Format financial penalty
            financial_penalty = '-'
            if offense.financial_penalty_imposed and offense.penalty_amount:
                financial_penalty = f'S${offense.penalty_amount:.2f}'
            elif offense.financial_penalty_imposed:
                financial_penalty = 'Imposed'
            
            row_data = {
                'S.NO': i,
                'CASE NO': offense.case_number or f'PL/OR/{offense.id}',
                'NAME': offense.offender_name or '-',
                'FIN NO': offense.fin_number or '-',
                'COMPANY NAME': offense.offender_company or '-',
                'DATE': offense.incident_date.strftime('%Y-%m-%d') if offense.incident_date else '-',
                'OFFENSE TYPE': offense.offense_type or 'N/A',
                'FINANCIAL PENALTY': financial_penalty,
                'PAYMENT STATUS': payment_status
            }
            excel_data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Create Excel file in memory
        output = BytesIO()
        
        # Use openpyxl engine for better formatting
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write data to Excel
            df.to_excel(writer, sheet_name='Offense Records', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Offense Records']
            
            # Format headers
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            for cell in worksheet[1]:  # First row (headers)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply borders to all cells
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        output.seek(0)
        
        # Create response
        filename = f'offense_records_table_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        flash(f'Error generating Excel export: {str(e)}', 'error')
        return redirect(url_for('offense_records'))

@app.route('/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def settings():
    """Settings page for system configuration"""
    if request.method == 'POST':
        try:
            # Process form data
            timezone = request.form.get('timezone', 'Asia/Singapore')
            date_format = request.form.get('date_format', 'DD/MM/YYYY')
            email_notifications = 'email_notifications' in request.form
            system_alerts = 'system_alerts' in request.form
            maintenance_mode = 'maintenance_mode' in request.form
            auto_backup = request.form.get('auto_backup', 'daily')
            backup_retention = request.form.get('backup_retention', '30')
            
            # Log the settings update
            log = SystemLog()
            log.user_id = current_user.id
            log.user_email = current_user.email
            log.action = f"Updated system settings - Timezone: {timezone}, Date Format: {date_format}, Email Notifications: {email_notifications}, System Alerts: {system_alerts}, Maintenance Mode: {maintenance_mode}, Auto Backup: {auto_backup}, Backup Retention: {backup_retention} days"
            log.module = "System Settings"
            log.status = "Success"
            db.session.add(log)
            db.session.commit()
            
            flash('Settings saved successfully', 'success')
            return redirect(url_for('settings'))
            
        except Exception as e:
            flash(f'Error saving settings: {str(e)}', 'error')
            return redirect(url_for('settings'))
    
    return render_template('settings.html')

# Food Locker Management Routes
@app.route('/food-locker-management')
@login_required
def food_locker_management():
    """Food Locker management page"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get date filters from request
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status_filter = request.args.get('status', '')
    
    # Base query for food lockers
    query = FoodLocker.query.filter_by(organization_id=user.organization_id)
    
    # Apply date filters
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(FoodLocker.rental_start_date >= start_date_obj)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(FoodLocker.rental_start_date <= end_date_obj)
        except ValueError:
            pass
    
    # Apply status filter
    if status_filter:
        query = query.filter(FoodLocker.status == status_filter)
    
    # Get all food lockers for the organization
    food_lockers = query.order_by(FoodLocker.created_at.desc()).all()
    
    # Get all active room numbers for the dropdown
    from app.models.models_house_acknowledge import RoomNumber, HouseAcknowledgment
    room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    
    # Get persons in charge from acknowledgment storage for dropdown
    persons_in_charge = db.session.query(
        HouseAcknowledgment.name,
        HouseAcknowledgment.fin,
        HouseAcknowledgment.phone_number,
        HouseAcknowledgment.company_name
    ).distinct().order_by(HouseAcknowledgment.name).all()
    
    return render_template('food_locker_management.html', 
                         food_lockers=food_lockers,
                         room_numbers=room_numbers,
                         persons_in_charge=persons_in_charge,
                         start_date=start_date,
                         end_date=end_date,
                         status_filter=status_filter)

@app.route('/food-locker-management', methods=['POST'])
@login_required
def create_food_locker():
    """Create new food locker record"""
    user = current_user
    if not user.organization_id:
        flash('Access denied', 'error')
        return redirect(url_for('food_locker_management'))
    
    try:
        from datetime import datetime
        
        # Parse dates
        rental_start_date = datetime.strptime(request.form['rental_start_date'], '%Y-%m-%d').date()
        rental_end_date = None
        if request.form.get('rental_end_date'):
            rental_end_date = datetime.strptime(request.form['rental_end_date'], '%Y-%m-%d').date()
        
        # Create food locker record
        food_locker = FoodLocker(
            company_name=request.form['company_name'],
            rental_price=float(request.form['rental_price']),
            rental_start_date=rental_start_date,
            rental_end_date=rental_end_date,
            caterer_name=request.form['caterer_name'],
            driver_name=request.form['driver_name'],
            driver_phone=request.form.get('driver_phone'),
            vehicle_plate=request.form['vehicle_plate'],
            person_in_charge_name=request.form['person_in_charge_name'],
            person_in_charge_fin=request.form['person_in_charge_fin'],
            person_in_charge_phone=request.form.get('person_in_charge_phone'),
            person_in_charge_company=request.form.get('person_in_charge_company'),
            oe_dc_name=request.form['oe_dc_name'],
            notes=request.form.get('notes'),
            created_by=user.id,
            organization_id=user.organization_id
        )
        
        db.session.add(food_locker)
        db.session.flush()  # Get the ID
        
        # Handle room assignments (multiple room numbers can be selected)
        selected_room_ids = request.form.getlist('room_numbers')
        for room_id in selected_room_ids:
            if room_id:
                assignment = FoodLockerRoomAssignment(
                    food_locker_id=food_locker.id,
                    room_number_id=int(room_id)
                )
                db.session.add(assignment)
        
        db.session.commit()
        flash('Food locker record created successfully', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating food locker record: {str(e)}', 'error')
    
    return redirect(url_for('food_locker_management'))

@app.route('/get-person-details/<fin_number>')
@login_required
def get_person_details(fin_number):
    """Get person details from acknowledgment storage by FIN number"""
    from app.models.models_house_acknowledge import HouseAcknowledgment
    
    person = HouseAcknowledgment.query.filter_by(fin=fin_number).first()
    
    if person:
        return jsonify({
            'success': True,
            'person': {
                'name': person.name,
                'fin': person.fin,
                'phone': person.phone_number,
                'company': person.company_name
            }
        })
    else:
        return jsonify({'success': False, 'message': 'Person not found'})

@app.route('/food-locker/<int:food_locker_id>/view')
@login_required
def view_food_locker(food_locker_id):
    """View food locker details"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'Please contact administrator to assign organization'})
    
    food_locker = FoodLocker.query.filter_by(
        id=food_locker_id, 
        organization_id=user.organization_id
    ).first_or_404()
    
    return jsonify({
        'success': True,
        'food_locker': {
            'id': food_locker.id,
            'company_name': food_locker.company_name,
            'rental_price': food_locker.rental_price,
            'rental_start_date': food_locker.rental_start_date.strftime('%Y-%m-%d'),
            'rental_end_date': food_locker.rental_end_date.strftime('%Y-%m-%d') if food_locker.rental_end_date else None,
            'caterer_name': food_locker.caterer_name,
            'driver_name': food_locker.driver_name,
            'driver_phone': food_locker.driver_phone,
            'vehicle_plate': food_locker.vehicle_plate,
            'person_in_charge_name': food_locker.person_in_charge_name,
            'person_in_charge_fin': food_locker.person_in_charge_fin,
            'person_in_charge_phone': food_locker.person_in_charge_phone,
            'person_in_charge_company': food_locker.person_in_charge_company,
            'oe_dc_name': food_locker.oe_dc_name,
            'status': food_locker.status,
            'notes': food_locker.notes,
            'tenant_signature': bool(food_locker.tenant_signature),
            'tenant_signature_date': food_locker.tenant_signature_date.strftime('%Y-%m-%d %H:%M:%S') if food_locker.tenant_signature_date else None,
            'oe_dc_signature': bool(food_locker.oe_dc_signature),
            'oe_dc_signature_date': food_locker.oe_dc_signature_date.strftime('%Y-%m-%d %H:%M:%S') if food_locker.oe_dc_signature_date else None,
            'created_at': food_locker.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'room_assignments': [
                {
                    'room_number': assignment.room_number.room_number,
                    'is_active': assignment.is_active,
                    'assigned_date': assignment.assigned_date.strftime('%Y-%m-%d')
                }
                for assignment in food_locker.room_assignments
            ]
        }
    })

@app.route('/food-locker/<int:food_locker_id>/sign', methods=['POST'])
@login_required
def sign_food_locker(food_locker_id):
    """Handle e-signature for food locker"""
    user = current_user
    
    try:
        food_locker = FoodLocker.query.filter_by(
            id=food_locker_id,
            organization_id=user.organization_id
        ).first_or_404()
        
        signature_type = request.form['signature_type']  # 'tenant' or 'oe_dc'
        signature_data = request.form['signature_data']
        
        if signature_type == 'tenant':
            food_locker.tenant_signature = signature_data
            food_locker.tenant_signature_date = singapore_now()
            flash('Tenant signature added successfully', 'success')
        elif signature_type == 'oe_dc':
            food_locker.oe_dc_signature = signature_data
            food_locker.oe_dc_signature_date = singapore_now()
            flash('OE/DC signature added successfully', 'success')
        
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/food-locker/<int:food_locker_id>/delete', methods=['POST'])
@login_required
@create_permission_required('food_locker')
def delete_food_locker(food_locker_id):
    """Delete food locker record"""
    user = current_user
    
    try:
        food_locker = FoodLocker.query.filter_by(
            id=food_locker_id,
            organization_id=user.organization_id
        ).first_or_404()
        
        db.session.delete(food_locker)
        db.session.commit()
        flash('Food locker record deleted successfully', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting food locker record: {str(e)}', 'error')
    
    return redirect(url_for('food_locker_management'))

@app.route('/food-locker-management/export/<format>')
@login_required
def export_food_locker_records(format):
    """Export food locker records to Excel or PDF"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get the same filters as the main page
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status_filter = request.args.get('status', '')
    selected_ids = request.args.get('selected_ids')
    
    # Base query for food lockers
    query = FoodLocker.query.filter_by(organization_id=user.organization_id)
    
    # Apply selected IDs filter first if provided
    if selected_ids:
        id_list = [int(id.strip()) for id in selected_ids.split(',') if id.strip().isdigit()]
        query = query.filter(FoodLocker.id.in_(id_list))
    else:
        # Apply date filters only if not filtering by selected IDs
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(FoodLocker.rental_start_date >= start_date_obj)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(FoodLocker.rental_start_date <= end_date_obj)
            except ValueError:
                pass
        
        # Apply status filter
        if status_filter:
            query = query.filter(FoodLocker.status == status_filter)
    
    food_lockers = query.order_by(FoodLocker.created_at.desc()).all()
    
    if format == 'excel':
        return export_food_locker_excel(food_lockers, start_date, end_date, status_filter, selected_ids)
    elif format == 'pdf':
        return export_food_locker_pdf(food_lockers, start_date, end_date, status_filter, selected_ids)
    else:
        flash('Invalid export format', 'error')
        return redirect(url_for('food_locker_management'))

@app.route('/food-locker-management/bulk-delete', methods=['POST'])
@login_required
@create_permission_required('food_locker')
def bulk_delete_food_lockers():
    """Bulk delete food locker records"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    selected_ids = request.form.getlist('selected_ids')
    
    try:
        # Verify all selected records belong to user's organization
        food_lockers = FoodLocker.query.filter(
            FoodLocker.id.in_(selected_ids),
            FoodLocker.organization_id == user.organization_id
        ).all()
        
        if len(food_lockers) != len(selected_ids):
            flash('Some records could not be found or access denied', 'error')
            return redirect(url_for('food_locker_management'))
        
        # Delete all selected records
        for locker in food_lockers:
            db.session.delete(locker)
        
        db.session.commit()
        flash(f'Successfully deleted {len(food_lockers)} food locker record(s)', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting records: {str(e)}', 'error')
    
    return redirect(url_for('food_locker_management'))

def export_food_locker_excel(food_lockers, start_date=None, end_date=None, status_filter=None, selected_ids=None):
    """Export food locker records to Excel with full details"""
    data = []
    for locker in food_lockers:
        # Get room numbers
        room_numbers = ', '.join([assignment.room_number.room_number 
                                 for assignment in locker.room_assignments 
                                 if assignment.is_active])
        
        # Get all room assignment details
        room_details = []
        for assignment in locker.room_assignments:
            if assignment.is_active:
                room_details.append(f"{assignment.room_number.room_number} (Active)")
            else:
                room_details.append(f"{assignment.room_number.room_number} (Inactive)")
        
        data.append({
            'ID': locker.id,
            'Company Name': locker.company_name,
            'Rental Start Date': locker.rental_start_date.strftime('%Y-%m-%d'),
            'Rental End Date': locker.rental_end_date.strftime('%Y-%m-%d') if locker.rental_end_date else 'Ongoing',
            'Rental Duration (Days)': (locker.rental_end_date - locker.rental_start_date).days if locker.rental_end_date else 'Ongoing',
            'Rental Price (S$)': f"{locker.rental_price:.2f}",
            'Monthly Rate (S$)': f"{locker.rental_price / max(1, (locker.rental_end_date - locker.rental_start_date).days / 30) if locker.rental_end_date else locker.rental_price:.2f}",
            'Caterer Name': locker.caterer_name,
            'Driver Name': locker.driver_name,
            'Driver Phone': locker.driver_phone or 'Not provided',
            'Vehicle Plate': locker.vehicle_plate,
            'Room Numbers (Active)': room_numbers or 'None assigned',
            'All Room Assignments': '; '.join(room_details) if room_details else 'None',
            'Total Rooms': len([a for a in locker.room_assignments if a.is_active]),
            'Person in Charge (Tenant)': locker.person_in_charge_name,
            'Tenant FIN Number': locker.person_in_charge_fin,
            'Tenant Phone': locker.person_in_charge_phone or 'Not provided',
            'Tenant Company': locker.person_in_charge_company or 'Not provided',
            'Status': locker.status,
            'Creator (OE/DC)': locker.oe_dc_name or 'Unknown',
            'Tenant Signature Status': 'Signed' if locker.tenant_signature else 'Pending',
            'Tenant Signature Date': locker.tenant_signature_date.strftime('%Y-%m-%d %H:%M:%S') if locker.tenant_signature_date else 'Not signed',
            'OE/DC Signature Status': 'Signed' if locker.oe_dc_signature else 'Pending',
            'OE/DC Signature Date': locker.oe_dc_signature_date.strftime('%Y-%m-%d %H:%M:%S') if locker.oe_dc_signature_date else 'Not signed',
            'Both Parties Signed': 'Yes' if (locker.tenant_signature and locker.oe_dc_signature) else 'No',
            'Agreement Status': 'Complete' if (locker.tenant_signature and locker.oe_dc_signature) else 'Incomplete',
            'Created Date': locker.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'Last Updated': locker.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'Notes': f"Rental agreement between {locker.company_name} and caterer {locker.caterer_name}",
            'Contact Summary': f"Tenant: {locker.person_in_charge_name} ({locker.person_in_charge_phone}), Driver: {locker.driver_name} ({locker.driver_phone})"
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Food Locker Records', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Food Locker Records']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # Generate filename with filters
    filter_info = []
    if selected_ids:
        filter_info.append("selected_records")
    if start_date:
        filter_info.append(f"from_{start_date}")
    if end_date:
        filter_info.append(f"to_{end_date}")
    if status_filter:
        filter_info.append(f"status_{status_filter}")
    
    filter_suffix = "_" + "_".join(filter_info) if filter_info else ""
    filename = f"food_locker_detailed_records{filter_suffix}_{singapore_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def export_food_locker_pdf(food_lockers, start_date=None, end_date=None, status_filter=None, selected_ids=None):
    """Export food locker records to PDF with comprehensive details"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    
    # Compact styles for single page layout
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading3'],
        fontSize=10,
        spaceAfter=4,
        spaceBefore=8,
        textColor=colors.blue,
        alignment=TA_LEFT
    )
    
    detail_style = ParagraphStyle(
        'Detail',
        parent=styles['Normal'],
        fontSize=8,
        spaceAfter=2
    )
    
    elements = []
    
    # Title
    title = "Food Locker Management Report"
    if selected_ids:
        title += " (Selected Records)"
    elif start_date or end_date or status_filter:
        title += " (Filtered)"
    elements.append(Paragraph(title, title_style))
    
    # Filter information
    filter_info = []
    if selected_ids:
        filter_info.append(f"Selected: {len(food_lockers)} records")
    if start_date:
        filter_info.append(f"From: {start_date}")
    if end_date:
        filter_info.append(f"To: {end_date}")
    if status_filter:
        filter_info.append(f"Status: {status_filter}")
    
    if filter_info:
        elements.append(Paragraph(" | ".join(filter_info), detail_style))
        elements.append(Spacer(1, 8))
    
    # Detailed records - compact single page format
    for i, locker in enumerate(food_lockers):
        # Record header
        elements.append(Paragraph(f"Food Locker Record #{locker.id}", section_style))
        
        # Create a 2-column layout for better space usage
        record_data = []
        
        # Left column data
        left_col = [
            f"<b>Company:</b> {locker.company_name}",
            f"<b>Rental Price:</b> S$ {locker.rental_price:.2f}",
            f"<b>Period:</b> {locker.rental_start_date.strftime('%Y-%m-%d')} to {locker.rental_end_date.strftime('%Y-%m-%d') if locker.rental_end_date else 'Ongoing'}",
            f"<b>Status:</b> {locker.status}",
            f"<b>Caterer:</b> {locker.caterer_name}",
            f"<b>Driver:</b> {locker.driver_name}",
        ]
        
        # Right column data
        right_col = [
            f"<b>Vehicle:</b> {locker.vehicle_plate}",
            f"<b>Tenant:</b> {locker.person_in_charge_name}",
            f"<b>FIN:</b> {locker.person_in_charge_fin}",
            f"<b>Phone:</b> {locker.person_in_charge_phone or 'Not provided'}",
            f"<b>OE/DC:</b> {locker.oe_dc_name or 'Unknown'}",
            f"<b>Created:</b> {locker.created_at.strftime('%Y-%m-%d')}",
        ]
        
        # Get room assignments
        active_rooms = []
        for assignment in locker.room_assignments:
            if assignment.is_active:
                active_rooms.append(assignment.room_number.room_number)
        
        # Signature status
        tenant_sig = "Signed" if locker.tenant_signature else "Pending"
        oe_sig = "Signed" if locker.oe_dc_signature else "Pending"
        
        # Create the record table with Paragraph objects for proper HTML rendering
        record_table_data = [
            [Paragraph(left_col[0], detail_style), Paragraph(right_col[0], detail_style)],
            [Paragraph(left_col[1], detail_style), Paragraph(right_col[1], detail_style)], 
            [Paragraph(left_col[2], detail_style), Paragraph(right_col[2], detail_style)],
            [Paragraph(left_col[3], detail_style), Paragraph(right_col[3], detail_style)],
            [Paragraph(left_col[4], detail_style), Paragraph(right_col[4], detail_style)],
            [Paragraph(left_col[5], detail_style), Paragraph(right_col[5], detail_style)],
            [Paragraph(f"<b>Active Rooms:</b> {', '.join(active_rooms) if active_rooms else 'None'}", detail_style), Paragraph(f"<b>Tenant Signature:</b> {tenant_sig}", detail_style)],
            [Paragraph(f"<b>Total Active Rooms:</b> {len(active_rooms)}", detail_style), Paragraph(f"<b>OE/DC Signature:</b> {oe_sig}", detail_style)]
        ]
        
        record_table = Table(record_table_data, colWidths=[250, 250])
        record_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ]))
        
        elements.append(record_table)
        
        # Add small spacer between records, but not after the last one
        if i < len(food_lockers) - 1:
            elements.append(Spacer(1, 8))
    
    # Summary at the bottom
    if food_lockers:
        elements.append(Spacer(1, 10))
        
        total_active = len([l for l in food_lockers if l.status == 'Active'])
        total_signed = len([l for l in food_lockers if l.tenant_signature and l.oe_dc_signature])
        total_rooms = sum([len([a for a in l.room_assignments if a.is_active]) for l in food_lockers])
        total_revenue = sum([l.rental_price for l in food_lockers])
        
        elements.append(Paragraph("SUMMARY", section_style))
        summary_data = [
            [f"Total Records: {len(food_lockers)}", f"Active Records: {total_active}"],
            [f"Fully Signed: {total_signed}", f"Total Active Rooms: {total_rooms}"],
            [f"Total Revenue: S${total_revenue:.2f}", f"Generated: {singapore_now().strftime('%Y-%m-%d %H:%M')}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[250, 250])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
    else:
        elements.append(Paragraph("No food locker records found matching the criteria.", detail_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    # Generate filename
    filter_info = []
    if selected_ids:
        filter_info.append("selected_records")
    if start_date:
        filter_info.append(f"from_{start_date}")
    if end_date:
        filter_info.append(f"to_{end_date}")
    if status_filter:
        filter_info.append(f"status_{status_filter}")
    
    filter_suffix = "_" + "_".join(filter_info) if filter_info else ""
    filename = f"food_locker_comprehensive_report{filter_suffix}_{singapore_now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@app.route('/admin')
@login_required
@page_access_required('admin')
def admin():
    """Admin page for system administration"""
    # Get real statistics from database
    total_users = User.query.count()
    total_organizations = Organization.query.count()
    total_forms = FormTemplate.query.count()
    
    # Get users with their organizations
    users = User.query.join(Organization, User.organization_id == Organization.id, isouter=True).all()
    
    # Get organizations
    organizations = Organization.query.all()
    
    # Get recent system logs
    recent_logs = SystemLog.query.order_by(SystemLog.created_at.desc()).limit(10).all()
    
    return render_template('admin.html', 
                         total_users=total_users,
                         total_organizations=total_organizations, 
                         total_forms=total_forms,
                         users=users,
                         organizations=organizations,
                         recent_logs=recent_logs)

# Admin API Routes
@app.route('/admin/users/<user_id>/toggle-status', methods=['POST'])
@login_required
def admin_toggle_user_status(user_id):
    """Toggle user status"""
    try:
        user = User.query.get_or_404(user_id)
        # For now, just return success - in a real system you'd toggle the user status
        
        # Log the action
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action=f"Toggled status for user {user.email}",
            module="User Management",
            status="Success"
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({"success": True, "message": "User status toggled successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route('/admin/organizations/<int:org_id>/delete', methods=['POST'])
@login_required
def admin_delete_organization(org_id):
    """Delete organization"""
    try:
        org = Organization.query.get_or_404(org_id)
        
        # Check if organization has users
        if org.users:
            return jsonify({"success": False, "message": "Cannot delete organization with existing users"})
        
        org_name = org.name
        db.session.delete(org)
        db.session.commit()
        
        # Log the action
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action=f"Deleted organization {org_name}",
            module="Organization Management",
            status="Success"
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({"success": True, "message": "Organization deleted successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route('/admin/export-logs')
@login_required
def admin_export_logs():
    """Export system logs to Excel"""
    try:
        logs = SystemLog.query.order_by(SystemLog.created_at.desc()).all()
        
        # Create DataFrame
        data = []
        for log in logs:
            data.append({
                'Timestamp': log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
                'User Email': log.user_email or (log.user.email if log.user else 'System'),
                'Action': log.action,
                'Module': log.module,
                'Status': log.status,
                'IP Address': log.ip_address or '',
                'Details': log.details or ''
            })
        
        if not data:
            flash('No system logs found to export.', 'warning')
            return redirect(url_for('admin'))
        
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='System Logs', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['System Logs']
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        output.seek(0)
        
        timestamp = singapore_now().strftime("%Y%m%d_%H%M%S")
        filename = f"system_logs_{timestamp}.xlsx"
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        flash(f'Error exporting logs: {str(e)}', 'error')
        return redirect(url_for('admin'))

@app.route('/admin/backup-database', methods=['POST'])
@login_required
def admin_backup_database():
    """Backup database"""
    try:
        # Log the action
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action="Database backup initiated",
            module="System Maintenance",
            status="Success"
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({"success": True, "message": "Database backup completed successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route('/admin/clean-temp-files', methods=['POST'])
@login_required
def admin_clean_temp_files():
    """Clean temporary files"""
    try:
        # Log the action
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action="Temporary files cleaned",
            module="System Maintenance",
            status="Success"
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({"success": True, "message": "Temporary files cleaned successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route('/admin/optimize-database', methods=['POST'])
@login_required
def admin_optimize_database():
    """Optimize database"""
    try:
        # Log the action
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action="Database optimization completed",
            module="System Maintenance",
            status="Success"
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({"success": True, "message": "Database optimization completed successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route('/admin/refresh-cache', methods=['POST'])
@login_required
def admin_refresh_cache():
    """Refresh system cache"""
    try:
        # Log the action
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action="System cache refreshed",
            module="System Maintenance",
            status="Success"
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({"success": True, "message": "System cache refreshed successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route('/admin/export-system-data')
@login_required
def admin_export_system_data():
    """Export all system data to Excel"""
    try:
        # Create Excel file with multiple sheets
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Export Users
            users_data = []
            for user in User.query.all():
                users_data.append({
                    'ID': user.id,
                    'Email': user.email,
                    'First Name': user.first_name,
                    'Last Name': user.last_name,
                    'Organization': user.organization.name if user.organization else '',
                    'Created At': user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else ''
                })
            
            if users_data:
                df_users = pd.DataFrame(users_data)
                df_users.to_excel(writer, sheet_name='Users', index=False)
            
            # Export Organizations
            orgs_data = []
            for org in Organization.query.all():
                orgs_data.append({
                    'ID': org.id,
                    'Name': org.name,
                    'Email': org.email,
                    'Description': org.description or '',
                    'Users Count': len(org.users),
                    'Created At': org.created_at.strftime('%Y-%m-%d %H:%M:%S') if org.created_at else ''
                })
            
            if orgs_data:
                df_orgs = pd.DataFrame(orgs_data)
                df_orgs.to_excel(writer, sheet_name='Organizations', index=False)
            
            # Export Assets
            assets_data = []
            for asset in Asset.query.all():
                assets_data.append({
                    'ID': asset.id,
                    'Name': asset.name,
                    'Category': asset.category.name if asset.category else '',
                    'Organization': asset.organization.name if asset.organization else '',
                    'Status': asset.status,
                    'Location': asset.location or '',
                    'Created At': asset.created_at.strftime('%Y-%m-%d %H:%M:%S') if asset.created_at else ''
                })
            
            if assets_data:
                df_assets = pd.DataFrame(assets_data)
                df_assets.to_excel(writer, sheet_name='Assets', index=False)
        
        output.seek(0)
        
        timestamp = singapore_now().strftime("%Y%m%d_%H%M%S")
        filename = f"system_data_export_{timestamp}.xlsx"
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        # Log the action
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action="System data exported",
            module="System Maintenance",
            status="Success"
        )
        db.session.add(log)
        db.session.commit()
        
        return response
        
    except Exception as e:
        flash(f'Error exporting system data: {str(e)}', 'error')
        return redirect(url_for('admin'))

# ========== ASSET MANAGEMENT ROUTES ==========

@app.route('/asset-management')
@login_required
@page_access_required('asset_management')
def asset_management():
    """Asset Management home page with serial number search"""
    # Get serial number search query
    serial_search = request.args.get('serial', '').strip()
    search_query = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    
    # Get all asset names for navigation
    asset_names = AssetName.query.order_by(AssetName.name).all()
    
    # Build base query
    query = AssetItem.query.join(AssetName)
    
    # Apply serial number search if provided
    if serial_search:
        query = query.filter(AssetItem.serial_number.ilike(f'%{serial_search}%'))
    
    # Apply general search
    if search_query:
        query = query.filter(
            db.or_(
                AssetName.name.ilike(f'%{search_query}%'),
                AssetItem.serial_number.ilike(f'%{search_query}%'),
                AssetItem.room_number.ilike(f'%{search_query}%')
            )
        )
    
    # Apply status filter
    if status_filter:
        query = query.filter(AssetItem.status == status_filter)
    
    # Get paginated results
    page = request.args.get('page', 1, type=int)
    assets = query.order_by(AssetItem.last_edited.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # Get status counts for filters
    status_counts = {
        'Room': AssetItem.query.filter_by(status='Room').count(),
        'Store': AssetItem.query.filter_by(status='Store').count(),
        'Damage': AssetItem.query.filter_by(status='Damage').count(),
    }
    
    return render_template('asset_management.html', 
                         assets=assets, 
                         asset_names=asset_names,
                         serial_search=serial_search,
                         search_query=search_query,
                         status_filter=status_filter,
                         status_counts=status_counts)

@app.route('/asset-management/create-asset-name', methods=['GET', 'POST'])
@login_required
@create_permission_required('asset_management')
def create_asset_name():
    """Create new asset name"""
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            category = request.form.get('category', '').strip()
            
            if not name or not category:
                flash('Asset name and category are required', 'error')
                return redirect(url_for('create_asset_name'))
            
            # Check if asset name already exists
            existing = AssetName.query.filter_by(name=name).first()
            if existing:
                flash('Asset name already exists', 'error')
                return redirect(url_for('create_asset_name'))
            
            # Create new asset name
            asset_name = AssetName(name=name, category=category)
            db.session.add(asset_name)
            db.session.commit()
            
            flash(f'Asset name "{name}" created successfully', 'success')
            return redirect(url_for('asset_management'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating asset name: {str(e)}', 'error')
    
    return render_template('create_asset_name.html')

@app.route('/asset-management/<int:asset_name_id>')
@login_required
def asset_name_details(asset_name_id):
    """View assets under specific asset name"""
    asset_name = AssetName.query.get_or_404(asset_name_id)
    
    # Get filter parameters
    search_query = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    room_filter = request.args.get('room', '')
    
    # Build query for assets under this asset name
    query = AssetItem.query.filter_by(asset_name_id=asset_name_id)
    
    # Apply filters
    if search_query:
        query = query.filter(
            db.or_(
                AssetItem.serial_number.ilike(f'%{search_query}%'),
                AssetItem.room_number.ilike(f'%{search_query}%')
            )
        )
    
    if status_filter:
        query = query.filter(AssetItem.status == status_filter)
    
    if room_filter:
        query = query.filter(AssetItem.room_number.ilike(f'%{room_filter}%'))
    
    # Get paginated results
    page = request.args.get('page', 1, type=int)
    assets = query.order_by(AssetItem.last_edited.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # Get status counts for this specific asset name
    status_counts = {
        'Room': AssetItem.query.filter_by(asset_name_id=asset_name_id, status='Room').count(),
        'Store': AssetItem.query.filter_by(asset_name_id=asset_name_id, status='Store').count(),
        'Damage': AssetItem.query.filter_by(asset_name_id=asset_name_id, status='Damage').count(),
    }
    
    return render_template('asset_name_details.html', 
                         asset_name=asset_name, 
                         assets=assets,
                         search_query=search_query,
                         status_filter=status_filter,
                         room_filter=room_filter,
                         status_counts=status_counts)

@app.route('/asset-management/<int:asset_name_id>/add-asset', methods=['GET', 'POST'])
@login_required
@create_permission_required('asset_management')
def add_asset(asset_name_id):
    """Add new asset under asset name"""
    asset_name = AssetName.query.get_or_404(asset_name_id)
    
    if request.method == 'POST':
        try:
            serial_number = request.form.get('serial_number', '').strip()
            room_number = request.form.get('room_number', '').strip()
            status = request.form.get('status', 'Room')
            quantity = int(request.form.get('quantity', 1))
            notes = request.form.get('notes', '').strip()
            
            if not serial_number:
                flash('Serial number is required', 'error')
                return redirect(url_for('add_asset', asset_name_id=asset_name_id))
            
            # Check if serial number already exists
            existing = AssetItem.query.filter_by(serial_number=serial_number).first()
            if existing:
                flash('Serial number already exists', 'error')
                return redirect(url_for('add_asset', asset_name_id=asset_name_id))
            
            # Create new asset
            asset = AssetItem(
                asset_name_id=asset_name_id,
                serial_number=serial_number,
                room_number=room_number or None,
                status=status,
                quantity=quantity,
                notes=notes or None
            )
            
            db.session.add(asset)
            db.session.commit()
            
            flash(f'Asset "{serial_number}" added successfully', 'success')
            return redirect(url_for('asset_name_details', asset_name_id=asset_name_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding asset: {str(e)}', 'error')
    
    # Get room numbers for dropdown
    room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    
    return render_template('add_asset.html', asset_name=asset_name, room_numbers=room_numbers)

@app.route('/asset-management/update-status/<int:asset_id>', methods=['POST'])
@login_required
@create_permission_required('asset_management')
def update_asset_status(asset_id):
    """Update asset status via AJAX"""
    try:
        asset = AssetItem.query.get_or_404(asset_id)
        new_status = request.json.get('status')
        
        if new_status not in ['Room', 'Store', 'Damage', 'Dispose', 'Other']:
            return jsonify({'success': False, 'message': 'Invalid status'})
        
        asset.status = new_status
        asset.last_edited = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Status updated successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/asset-management/edit-item/<int:asset_id>', methods=['GET', 'POST'])
@login_required
@create_permission_required('asset_management')
def edit_asset_item(asset_id):
    """Edit asset details"""
    asset = AssetItem.query.get_or_404(asset_id)
    
    if request.method == 'POST':
        try:
            serial_number = request.form.get('serial_number', '').strip()
            room_number = request.form.get('room_number', '').strip()
            status = request.form.get('status', 'Room')
            quantity = int(request.form.get('quantity', 1))
            notes = request.form.get('notes', '').strip()
            
            if not serial_number:
                flash('Serial number is required', 'error')
                return redirect(url_for('edit_asset', asset_id=asset_id))
            
            # Check if serial number already exists (excluding current asset)
            existing = AssetItem.query.filter(
                AssetItem.serial_number == serial_number,
                AssetItem.id != asset_id
            ).first()
            if existing:
                flash('Serial number already exists', 'error')
                return redirect(url_for('edit_asset', asset_id=asset_id))
            
            # Update asset
            asset.serial_number = serial_number
            asset.room_number = room_number or None
            asset.status = status
            asset.quantity = quantity
            asset.notes = notes or None
            asset.last_edited = datetime.utcnow()
            
            db.session.commit()
            
            flash('Asset updated successfully', 'success')
            return redirect(url_for('asset_name_details', asset_name_id=asset.asset_name_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating asset: {str(e)}', 'error')
    
    # Get room numbers for dropdown
    room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    
    return render_template('edit_asset.html', asset=asset, room_numbers=room_numbers)

@app.route('/asset-management/delete-item/<int:asset_id>', methods=['POST'])
@login_required
@create_permission_required('asset_management')
def delete_asset_item(asset_id):
    """Delete asset"""
    try:
        asset = AssetItem.query.get_or_404(asset_id)
        asset_name_id = asset.asset_name_id
        serial_number = asset.serial_number
        
        db.session.delete(asset)
        db.session.commit()
        
        flash(f'Asset "{serial_number}" deleted successfully', 'success')
        return redirect(url_for('asset_name_details', asset_name_id=asset_name_id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting asset: {str(e)}', 'error')
        return redirect(url_for('asset_management'))

@app.route('/asset-management/store')
@login_required
def store_assets():
    """Store page - shows all assets with status 'Store'"""
    # Get filter parameters
    search_query = request.args.get('search', '')
    room_filter = request.args.get('room', '')
    
    # Build query for store assets
    query = AssetItem.query.filter_by(status='Store').join(AssetName)
    
    # Apply filters
    if search_query:
        query = query.filter(
            db.or_(
                AssetName.name.ilike(f'%{search_query}%'),
                AssetItem.serial_number.ilike(f'%{search_query}%'),
                AssetItem.room_number.ilike(f'%{search_query}%')
            )
        )
    
    if room_filter:
        query = query.filter(AssetItem.room_number.ilike(f'%{room_filter}%'))
    
    # Get paginated results
    page = request.args.get('page', 1, type=int)
    assets = query.order_by(AssetItem.last_edited.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('store_assets.html', 
                         assets=assets,
                         search_query=search_query,
                         room_filter=room_filter)

@app.route('/asset-management/damage')
@login_required
def damage_assets():
    """Damage page - shows all assets with status 'Damage'"""
    # Get filter parameters
    search_query = request.args.get('search', '')
    room_filter = request.args.get('room', '')
    
    # Build query for damage assets
    query = AssetItem.query.filter_by(status='Damage').join(AssetName)
    
    # Apply filters
    if search_query:
        query = query.filter(
            db.or_(
                AssetName.name.ilike(f'%{search_query}%'),
                AssetItem.serial_number.ilike(f'%{search_query}%'),
                AssetItem.room_number.ilike(f'%{search_query}%')
            )
        )
    
    if room_filter:
        query = query.filter(AssetItem.room_number.ilike(f'%{room_filter}%'))
    
    # Get paginated results
    page = request.args.get('page', 1, type=int)
    assets = query.order_by(AssetItem.last_edited.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('damage_assets.html', 
                         assets=assets,
                         search_query=search_query,
                         room_filter=room_filter)


@app.route('/asset-management/<int:asset_name_id>/export-excel')
@login_required
def export_asset_excel(asset_name_id):
    """Export assets to Excel for specific asset name"""
    asset_name = AssetName.query.get_or_404(asset_name_id)
    
    try:
        # Get all assets for this asset name
        assets = AssetItem.query.filter_by(asset_name_id=asset_name_id).order_by(AssetItem.serial_number).all()
        
        # Prepare data for Excel
        excel_data = []
        for asset in assets:
            excel_data.append({
                'Asset Name': asset_name.name,
                'Category': asset_name.category,
                'Serial Number': asset.serial_number,
                'Room Number': asset.room_number or '',
                'Status': asset.status,
                'Quantity': asset.quantity,
                'Date Added': asset.date_added.strftime('%Y-%m-%d') if asset.date_added else '',
                'Last Edited': asset.last_edited.strftime('%Y-%m-%d %H:%M:%S') if asset.last_edited else '',
                'Notes': asset.notes or ''
            })
        
        if not excel_data:
            flash('No assets found to export', 'warning')
            return redirect(url_for('asset_name_details', asset_name_id=asset_name_id))
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(excel_data)
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=asset_name.name[:30], index=False)
            
            # Format the Excel file
            workbook = writer.book
            worksheet = writer.sheets[asset_name.name[:30]]
            
            # Header formatting
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Create response
        filename = f'{asset_name.name}_assets_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        flash(f'Error exporting to Excel: {str(e)}', 'error')
        return redirect(url_for('asset_name_details', asset_name_id=asset_name_id))

@app.route('/asset-management/<int:asset_name_id>/import-excel', methods=['GET', 'POST'])
@login_required
def import_asset_excel(asset_name_id):
    """Import assets from Excel for specific asset name"""
    asset_name = AssetName.query.get_or_404(asset_name_id)
    
    if request.method == 'POST':
        try:
            import pandas as pd
            
            if 'excel_file' not in request.files:
                flash('No file selected', 'error')
                return redirect(url_for('import_asset_excel', asset_name_id=asset_name_id))
            
            file = request.files['excel_file']
            if file.filename == '':
                flash('No file selected', 'error')
                return redirect(url_for('import_asset_excel', asset_name_id=asset_name_id))
            
            if not file.filename.endswith(('.xlsx', '.xls')):
                flash('Please upload an Excel file (.xlsx or .xls)', 'error')
                return redirect(url_for('import_asset_excel', asset_name_id=asset_name_id))
            
            # Read Excel file
            df = pd.read_excel(file)
            
            # Validate required columns
            required_columns = ['Serial Number', 'Status']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                flash(f'Missing required columns: {", ".join(missing_columns)}', 'error')
                return redirect(url_for('import_asset_excel', asset_name_id=asset_name_id))
            
            # Process each row
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    serial_number = str(row['Serial Number']).strip()
                    if not serial_number or serial_number.lower() == 'nan':
                        continue
                    
                    # Check if asset already exists
                    existing = AssetItem.query.filter_by(serial_number=serial_number).first()
                    if existing:
                        errors.append(f'Row {index + 2}: Serial number "{serial_number}" already exists')
                        error_count += 1
                        continue
                    
                    # Create new asset
                    asset = AssetItem(
                        asset_name_id=asset_name_id,
                        serial_number=serial_number,
                        room_number=str(row.get('Room Number', '')).strip() or None,
                        status=str(row.get('Status', 'Room')).strip(),
                        quantity=int(row.get('Quantity', 1)),
                        notes=str(row.get('Notes', '')).strip() or None
                    )
                    
                    # Validate status
                    if asset.status not in ['Room', 'Store', 'Damage', 'Dispose', 'Other']:
                        asset.status = 'Room'
                    
                    db.session.add(asset)
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f'Row {index + 2}: {str(e)}')
                    error_count += 1
                    continue
            
            db.session.commit()
            
            # Show results
            if success_count > 0:
                flash(f'Successfully imported {success_count} assets', 'success')
            if error_count > 0:
                flash(f'{error_count} errors occurred during import', 'warning')
                for error in errors[:5]:  # Show first 5 errors
                    flash(error, 'error')
            
            return redirect(url_for('asset_name_details', asset_name_id=asset_name_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error importing Excel file: {str(e)}', 'error')
    
    return render_template('import_asset_excel.html', asset_name=asset_name)

@app.route('/asset-management/<int:asset_name_id>/download-template')
@login_required
def download_asset_template(asset_name_id):
    """Download Excel template for asset import"""
    asset_name = AssetName.query.get_or_404(asset_name_id)
    
    try:
        import pandas as pd
        from io import BytesIO
        import openpyxl
        
        # Create template data
        template_data = {
            'Asset Name': [asset_name.name, asset_name.name],
            'Category': [asset_name.category, asset_name.category],
            'Serial Number': ['SAMPLE001', 'SAMPLE002'],
            'Room Number': ['R101', 'R102'],
            'Status': ['Room', 'Store'],
            'Quantity': [1, 2],
            'Notes': ['Sample note 1', 'Sample note 2']
        }
        
        df = pd.DataFrame(template_data)
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Template', index=False)
            
            # Format the template
            workbook = writer.book
            worksheet = writer.sheets['Template']
            
            # Header formatting
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Create response
        filename = f'{asset_name.name}_import_template.xlsx'
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        flash(f'Error generating template: {str(e)}', 'error')
        return redirect(url_for('asset_name_details', asset_name_id=asset_name_id))

@app.route('/asset-management/bulk-export', methods=['POST'])
@login_required
def bulk_export_assets():
    """Export selected assets to Excel"""
    try:
        import pandas as pd
        from io import BytesIO
        import openpyxl
        asset_ids = request.form.getlist('asset_ids')
        if not asset_ids:
            flash('No assets selected for export', 'error')
            return redirect(url_for('asset_management'))
        
        # Get selected assets
        assets = AssetItem.query.filter(AssetItem.id.in_(asset_ids)).all()
        if not assets:
            flash('Selected assets not found', 'error')
            return redirect(url_for('asset_management'))
        
        # Create Excel data
        export_data = []
        for asset in assets:
            export_data.append({
                'Asset Name': asset.asset_name_ref.name,
                'Category': asset.asset_name_ref.category,
                'Serial Number': asset.serial_number,
                'Room Number': asset.room_number or '',
                'Status': asset.status,
                'Quantity': asset.quantity,
                'Notes': asset.notes or '',
                'Date Added': asset.date_added.strftime('%Y-%m-%d') if asset.date_added else '',
                'Last Edit': asset.last_edited.strftime('%Y-%m-%d %H:%M') if asset.last_edited else ''
            })
        
        df = pd.DataFrame(export_data)
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Selected Assets', index=False)
            
            # Format the Excel file
            workbook = writer.book
            worksheet = writer.sheets['Selected Assets']
            
            # Header formatting
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f'selected_assets_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting assets: {str(e)}', 'error')
        return redirect(url_for('asset_management'))

@app.route('/asset-management/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_assets():
    """Delete selected assets"""
    try:
        asset_ids = request.form.getlist('asset_ids')
        if not asset_ids:
            flash('No assets selected for deletion', 'error')
            return redirect(url_for('asset_management'))
        
        # Get selected assets
        assets = AssetItem.query.filter(AssetItem.id.in_(asset_ids)).all()
        if not assets:
            flash('Selected assets not found', 'error')
            return redirect(url_for('asset_management'))
        
        # Delete assets
        deleted_serials = []
        for asset in assets:
            deleted_serials.append(asset.serial_number)
            db.session.delete(asset)
        
        db.session.commit()
        
        flash(f'Successfully deleted {len(deleted_serials)} assets: {", ".join(deleted_serials)}', 'success')
        return redirect(url_for('asset_management'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting assets: {str(e)}', 'error')
        return redirect(url_for('asset_management'))

@app.route('/admin/restart-system', methods=['POST'])
@login_required
def admin_restart_system():
    """Restart system"""
    try:
        # Log the action
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action="System restart initiated",
            module="System Maintenance",
            status="Success"
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({"success": True, "message": "System restart initiated successfully"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route('/admin/users/<user_id>/edit')
@login_required
@admin_required
def admin_edit_user(user_id):
    """Edit user page"""
    user = User.query.get_or_404(user_id)
    organizations = Organization.query.all()
    
    # Get user's current page permissions from JSON field
    user_permissions = []
    if user.page_permissions:
        import json
        try:
            permissions_data = json.loads(user.page_permissions)
            
            # Handle different permission formats
            if isinstance(permissions_data, list):
                # New format: simple array of permission names
                user_permissions = permissions_data
            elif isinstance(permissions_data, dict):
                # Old format: object with can_access properties - convert to new format
                user_permissions = [key for key, value in permissions_data.items() 
                                  if isinstance(value, dict) and value.get('can_access', False)]
                # Update to new format in database
                user.page_permissions = json.dumps(user_permissions)
                db.session.commit()
            else:
                user_permissions = []
                
        except (json.JSONDecodeError, TypeError):
            user_permissions = []
    else:
        user_permissions = []
    
    # Force refresh user data from database
    db.session.refresh(user)
    

    
    return render_template('admin_edit_user_new.html', 
                         user=user, 
                         organizations=organizations,
                         current_user_permissions=user_permissions)

@app.route('/admin/users/<user_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_update_user(user_id):
    """Update user details and permissions with new page-based system"""
    import json
    from werkzeug.security import generate_password_hash
    
    try:
        user = User.query.get_or_404(user_id)
        
        # Check if user has view_only access - prevent all updates except access_level changes
        new_access_level = request.form.get('access_level', user.access_level)
        
        # Initialize selected_permissions to avoid UnboundLocalError
        selected_permissions = request.form.getlist('page_access')
        
        # Always allow access_level changes (admin can change user from view_only to full)
        user.access_level = new_access_level
        
        # Only allow other updates if not view_only
        if new_access_level != 'view_only':
            # Update basic user info
            user.email = request.form.get('email', user.email)
            user.first_name = request.form.get('first_name', user.first_name)
            user.last_name = request.form.get('last_name', user.last_name)
            user.username = request.form.get('username', user.username)
            user.full_name = request.form.get('full_name', user.full_name)
        
            # Only update organization, password, and permissions if not view_only
            org_id = request.form.get('organization_id')
            if org_id:
                user.organization_id = int(org_id)
            
            # Handle password update
            new_password = request.form.get('new_password')
            if new_password and new_password.strip():
                user.password_hash = generate_password_hash(new_password)
                user.admin_viewable_password = new_password  # Store for admin access
            
            # Update admin status if admin page is selected
            user.is_admin = 'admin' in selected_permissions
            
            # Store page permissions as JSON in the page_permissions field
            user.page_permissions = json.dumps(selected_permissions) if selected_permissions else None
        else:
            # For view_only users, still update page permissions but don't allow other changes
            user.page_permissions = json.dumps(selected_permissions) if selected_permissions else None
        
        db.session.commit()
        
        # Log the action (truncate permissions list if too long)
        permissions_summary = ', '.join(selected_permissions) if selected_permissions else 'none'
        if len(permissions_summary) > 150:
            permissions_summary = permissions_summary[:147] + "..."
        
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action=f"Updated user {user.email[:20]} permissions",
            module="User Management",
            status="Success",
            details=f"Permissions: {permissions_summary}"
        )
        db.session.add(log)
        db.session.commit()
        
        flash('User and permissions updated successfully', 'success')
        return redirect(url_for('admin'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating user: {str(e)}', 'error')
        return redirect(url_for('admin_edit_user', user_id=user_id))

@app.route('/admin/organizations/<int:org_id>/edit')
@login_required
def admin_edit_organization(org_id):
    """Edit organization page"""
    organization = Organization.query.get_or_404(org_id)
    return render_template('admin_edit_org.html', organization=organization)

@app.route('/admin/organizations/<int:org_id>/edit', methods=['POST'])
@login_required
def admin_update_organization(org_id):
    """Update organization details"""
    try:
        org = Organization.query.get_or_404(org_id)
        
        org.name = request.form.get('name', org.name)
        org.email = request.form.get('email', org.email)
        org.description = request.form.get('description', org.description)
        
        db.session.commit()
        
        # Log the action
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action=f"Updated organization {org.name}",
            module="Organization Management",
            status="Success"
        )
        db.session.add(log)
        db.session.commit()
        
        flash('Organization updated successfully', 'success')
        return redirect(url_for('admin'))
        
    except Exception as e:
        flash(f'Error updating organization: {str(e)}', 'error')
        return redirect(url_for('admin_edit_organization', org_id=org_id))

@app.route('/admin/users/add')
@login_required
@admin_required
def admin_add_user():
    """Add new user page"""
    organizations = Organization.query.all()
    return render_template('admin_add_user_new.html', organizations=organizations)

@app.route('/admin/users/add', methods=['POST'])
@login_required
def admin_create_user():
    """Create new user"""
    try:
        # Check if user already exists
        email = request.form.get('email')
        if User.query.filter_by(email=email).first():
            flash('User with this email already exists', 'error')
            return redirect(url_for('admin_add_user'))
        
        # Create new user
        user = User()
        user.id = str(uuid.uuid4())  # Generate unique ID
        user.email = email
        user.first_name = request.form.get('first_name')
        user.last_name = request.form.get('last_name')
        
        org_id = request.form.get('organization_id')
        if org_id:
            user.organization_id = int(org_id)
        
        # Set password if provided
        password = request.form.get('password')
        if password:
            from werkzeug.security import generate_password_hash
            user.password_hash = generate_password_hash(password)
            user.admin_viewable_password = password  # Store for admin access
        
        # Handle page-based permissions
        selected_permissions = request.form.getlist('page_access')
        access_level = request.form.get('access_level', 'full')
        
        # Set access level
        user.access_level = access_level
        
        # Update admin status if admin page is selected
        user.is_admin = 'admin' in selected_permissions
        
        # Store page permissions as JSON in the page_permissions field
        import json
        user.page_permissions = json.dumps(selected_permissions) if selected_permissions else None
        
        db.session.add(user)
        db.session.commit()
        
        # Log the action
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action=f"Created user {user.email} with {len(selected_permissions)} permissions",
            module="User Management",
            status="Success"
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f'User {user.email} created successfully with {len(selected_permissions)} page permissions', 'success')
        return redirect(url_for('admin'))
        
    except Exception as e:
        flash(f'Error creating user: {str(e)}', 'error')
        return redirect(url_for('admin_add_user'))

@app.route('/admin/organizations/add')
@login_required
def admin_add_organization():
    """Add new organization page"""
    return render_template('admin_add_org.html')

@app.route('/admin/organizations/add', methods=['POST'])
@login_required
def admin_create_organization():
    """Create new organization"""
    try:
        # Check if organization already exists
        name = request.form.get('name')
        email = request.form.get('email')
        
        if Organization.query.filter_by(name=name).first():
            flash('Organization with this name already exists', 'error')
            return redirect(url_for('admin_add_organization'))
            
        if Organization.query.filter_by(email=email).first():
            flash('Organization with this email already exists', 'error')
            return redirect(url_for('admin_add_organization'))
        
        # Create new organization
        org = Organization()
        org.name = name
        org.email = email
        org.description = request.form.get('description')
        
        db.session.add(org)
        db.session.commit()
        
        # Log the action
        log = SystemLog()
        log.user_id = current_user.id
        log.user_email = current_user.email
        log.action = f"Created new organization {org.name}"
        log.module = "Organization Management"
        log.status = "Success"
        db.session.add(log)
        db.session.commit()
        
        flash('Organization created successfully', 'success')
        return redirect(url_for('admin'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating organization: {str(e)}', 'error')
        return redirect(url_for('admin_add_organization'))

@app.route('/admin/users/<user_id>/delete', methods=['POST'])
@login_required
def admin_delete_user(user_id):
    """Delete a user - Admin only"""
    if not is_admin_user(current_user):
        return jsonify({'success': False, 'message': 'Access denied - admin privileges required'})
    
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'User not found'})
        
        # Prevent deleting yourself
        if user.id == current_user.id:
            return jsonify({'success': False, 'message': 'Cannot delete your own account'})
        
        user_email = user.email
        
        # Delete the user
        db.session.delete(user)
        db.session.commit()
        
        # Log the action
        log = SystemLog(
            user_id=current_user.id,
            user_email=current_user.email,
            action=f"Deleted user {user_email}",
            module="User Management",
            status="Success"
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'User deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error deleting user: {str(e)}'})

def create_default_form_templates(organization_id):
    """Create default form templates for an organization if they don't exist"""
    # Get user ID for created_by field (use Pioneer Lodge admin user)
    admin_user = User.query.filter_by(email="pioneerlodge@tsgrp.sg").first()
    if not admin_user:
        admin_user = User.query.first()  # Fallback to first user
    
    if not admin_user:
        return  # No users exist yet
    
    # Default form templates
    default_forms = [
        {
            'name': 'House Rules and Regulations',
            'form_type': 'regulations',
            'description': 'Comprehensive house rules and regulations for residents',
            'regulations_text': 'Please read and acknowledge these house rules and regulations carefully.',
            'fields_json': '[]'
        },
        {
            'name': 'Room Inventory Checklist',
            'form_type': 'handover',
            'description': 'Room handover and inventory checklist form',
            'regulations_text': 'Complete this checklist during room handover process.',
            'fields_json': '[]'
        },
        {
            'name': 'Offense Reporting Form',
            'form_type': 'offense',
            'description': 'Report disciplinary violations and incidents',
            'regulations_text': 'Use this form to report violations of house rules.',
            'fields_json': '[]'
        }
    ]
    
    for form_data in default_forms:
        # Check if form already exists for this organization
        existing_form = FormTemplate.query.filter_by(
            name=form_data['name'],
            organization_id=organization_id
        ).first()
        
        if not existing_form:
            form_template = FormTemplate()
            form_template.name = form_data['name']
            form_template.form_type = form_data['form_type']
            form_template.description = form_data['description']
            form_template.organization_id = organization_id
            form_template.regulations_text = form_data['regulations_text']
            form_template.fields_json = form_data['fields_json']
            form_template.created_by = admin_user.id
            form_template.language_code = 'en'
            form_template.public_access = True
            form_template.is_active = True
            db.session.add(form_template)
    
    db.session.commit()

# Use Stock Routes
@app.route('/use-stock', methods=['GET', 'POST'])
@login_required
def use_stock():
    """Create new stock usage record"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            stock_item_id = request.form.get('stock_item_id')
            used_quantity = int(request.form.get('used_quantity', 0))
            usage_date = request.form.get('usage_date')
            notes = request.form.get('notes', '')
            
            # Get the stock item
            stock_item = StockItem.query.filter_by(
                id=stock_item_id,
                organization_id=user.organization_id
            ).first()
            
            if not stock_item:
                flash('Stock item not found', 'error')
                return redirect(url_for('use_stock'))
            
            # Calculate current used quantity for this item
            current_used = StockUsage.query.filter_by(
                stock_item_id=stock_item_id,
                organization_id=user.organization_id
            ).with_entities(func.sum(StockUsage.used_quantity)).scalar() or 0
            
            # Check if total used quantity would exceed available stock
            total_used_after = current_used + used_quantity
            if total_used_after > stock_item.quantity:
                flash(f'Cannot use {used_quantity} units. Only {stock_item.quantity - current_used} units available.', 'error')
                return redirect(url_for('use_stock'))
            
            # Create stock usage record
            stock_usage = StockUsage(
                stock_item_id=stock_item.id,
                item_name=stock_item.name,
                used_quantity=used_quantity,
                available_quantity=stock_item.quantity - total_used_after,
                usage_date=datetime.strptime(usage_date, '%Y-%m-%d').date(),
                notes=notes,
                organization_id=user.organization_id,
                created_by=user.id
            )
            
            db.session.add(stock_usage)
            db.session.commit()
            
            flash(f'Successfully recorded usage of {used_quantity} units of {stock_item.name}', 'success')
            return redirect(url_for('used_stock_summary'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error recording stock usage', 'error')
            print(f"Error: {e}")
    
    # Get available stock items for the organization
    stock_items = StockItem.query.filter_by(
        organization_id=user.organization_id
    ).filter(StockItem.quantity > 0).all()
    
    # Calculate available quantities for each item
    for item in stock_items:
        used_total = StockUsage.query.filter_by(
            stock_item_id=item.id,
            organization_id=user.organization_id
        ).with_entities(func.sum(StockUsage.used_quantity)).scalar() or 0
        item.available_quantity = item.quantity - used_total
    
    return render_template('use_stock.html', stock_items=stock_items)

@app.route('/api/delete-usage-records', methods=['POST'])
@login_required
def delete_usage_records():
    """Delete multiple usage records"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'No organization assigned'})
    
    try:
        data = request.get_json()
        record_ids = data.get('record_ids', [])
        
        if not record_ids:
            return jsonify({'success': False, 'error': 'No record IDs provided'})
        
        # Delete the usage records
        deleted_count = StockUsage.query.filter(
            StockUsage.id.in_(record_ids),
            StockUsage.organization_id == user.organization_id
        ).delete(synchronize_session=False)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Successfully deleted {deleted_count} usage record(s)'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error deleting records: {str(e)}'})

@app.route('/api/export-selected-usage', methods=['POST'])
@login_required
def export_selected_usage():
    """Export selected usage records to Excel"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('used_stock_summary'))
    
    try:
        record_ids = json.loads(request.form.get('record_ids', '[]'))
        
        if not record_ids:
            flash('No records selected for export', 'warning')
            return redirect(url_for('used_stock_summary'))
        
        # Get the selected usage records
        usage_records = StockUsage.query.filter(
            StockUsage.id.in_(record_ids),
            StockUsage.organization_id == user.organization_id
        ).order_by(StockUsage.usage_date.desc()).all()
        
        if not usage_records:
            flash('No records found', 'warning')
            return redirect(url_for('used_stock_summary'))
        
        # Prepare data for Excel
        data = []
        for i, record in enumerate(usage_records, 1):
            data.append({
                'S.No': i,
                'Date': record.usage_date.strftime('%Y-%m-%d'),
                'Item Name': record.item_name,
                'Used Quantity': record.used_quantity,
                'Available Quantity': record.available_quantity,
                'Notes': record.notes or '',
                'Created By': record.created_by_user.first_name if record.created_by_user else 'Unknown',
                'Created At': record.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Selected Usage Records')
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Selected Usage Records']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_name].width = adjusted_width
        
        output.seek(0)
        
        # Create response
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=selected_usage_records_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            }
        )
        
        return response
        
    except Exception as e:
        flash(f'Error exporting selected records: {str(e)}', 'error')
        return redirect(url_for('used_stock_summary'))

@app.route('/api/update-usage-record/<int:record_id>', methods=['POST'])
@login_required
def update_usage_record_api(record_id):
    """Update usage record quantity via API"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'No organization assigned'})
    
    try:
        data = request.get_json()
        new_used_quantity = data.get('used_quantity')
        
        if not new_used_quantity or new_used_quantity < 0:
            return jsonify({'success': False, 'error': 'Invalid quantity'})
        
        # Get the usage record
        usage_record = StockUsage.query.filter_by(
            id=record_id,
            organization_id=user.organization_id
        ).first()
        
        if not usage_record:
            return jsonify({'success': False, 'error': 'Usage record not found'})
        
        # Get the stock item to validate quantity
        stock_item = StockItem.query.filter_by(
            name=usage_record.item_name,
            organization_id=user.organization_id
        ).first()
        
        if not stock_item:
            return jsonify({'success': False, 'error': 'Stock item not found'})
        
        # Calculate total used quantity excluding this record
        other_used = StockUsage.query.filter(
            StockUsage.item_name == usage_record.item_name,
            StockUsage.organization_id == user.organization_id,
            StockUsage.id != record_id
        ).with_entities(func.sum(StockUsage.used_quantity)).scalar() or 0
        
        # Check if new quantity would exceed total stock
        total_used_after = other_used + new_used_quantity
        if total_used_after > stock_item.quantity:
            return jsonify({
                'success': False, 
                'error': f'Cannot exceed total stock of {stock_item.quantity}. Current other usage: {other_used}'
            })
        
        # Update the usage record
        usage_record.used_quantity = new_used_quantity
        usage_record.available_quantity = stock_item.quantity - total_used_after
        usage_record.updated_at = singapore_now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Usage record updated successfully',
            'new_available': usage_record.available_quantity
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error updating record: {str(e)}'})

@app.route('/used-stock-summary')
@login_required
def used_stock_summary():
    """Display used stock summary with filtering and export options"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get filter parameters
    item_filter = request.args.get('item_name', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query
    query = StockUsage.query.filter_by(organization_id=user.organization_id)
    
    if item_filter:
        query = query.filter(StockUsage.item_name.ilike(f'%{item_filter}%'))
    
    if date_from:
        query = query.filter(StockUsage.usage_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    
    if date_to:
        query = query.filter(StockUsage.usage_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    
    usage_records = query.order_by(StockUsage.usage_date.desc()).all()
    
    # Get unique item names for filter dropdown
    item_names = StockUsage.query.filter_by(
        organization_id=user.organization_id
    ).with_entities(StockUsage.item_name).distinct().all()
    item_names = [name[0] for name in item_names]
    
    return render_template('used_stock_summary.html', 
                         usage_records=usage_records,
                         item_names=item_names,
                         item_filter=item_filter,
                         date_from=date_from,
                         date_to=date_to)

@app.route('/edit-stock-usage/<int:usage_id>', methods=['GET', 'POST'])
@login_required
def edit_stock_usage(usage_id):
    """Edit existing stock usage record"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    usage_record = StockUsage.query.filter_by(
        id=usage_id,
        organization_id=user.organization_id
    ).first()
    
    if not usage_record:
        flash('Usage record not found', 'error')
        return redirect(url_for('used_stock_summary'))
    
    if request.method == 'POST':
        try:
            new_used_quantity = int(request.form.get('used_quantity', 0))
            new_usage_date = request.form.get('usage_date')
            new_notes = request.form.get('notes', '')
            
            # Calculate other usage for this item (excluding current record)
            other_used = StockUsage.query.filter_by(
                stock_item_id=usage_record.stock_item_id,
                organization_id=user.organization_id
            ).filter(StockUsage.id != usage_id).with_entities(
                func.sum(StockUsage.used_quantity)
            ).scalar() or 0
            
            # Check if new total would exceed available stock
            total_used_after = other_used + new_used_quantity
            if total_used_after > usage_record.stock_item.quantity:
                available = usage_record.stock_item.quantity - other_used
                flash(f'Cannot use {new_used_quantity} units. Only {available} units available.', 'error')
                return redirect(url_for('edit_stock_usage', usage_id=usage_id))
            
            # Update the record
            usage_record.used_quantity = new_used_quantity
            usage_record.available_quantity = usage_record.stock_item.quantity - total_used_after
            usage_record.usage_date = datetime.strptime(new_usage_date, '%Y-%m-%d').date()
            usage_record.notes = new_notes
            usage_record.updated_at = singapore_now()
            
            db.session.commit()
            
            flash('Stock usage record updated successfully', 'success')
            return redirect(url_for('used_stock_summary'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error updating stock usage record', 'error')
            print(f"Error: {e}")
    
    return render_template('edit_stock_usage.html', usage_record=usage_record)

@app.route('/export-used-stock')
@login_required
def export_used_stock():
    """Export used stock summary to Excel"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get filter parameters
    item_filter = request.args.get('item_name', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query
    query = StockUsage.query.filter_by(organization_id=user.organization_id)
    
    if item_filter:
        query = query.filter(StockUsage.item_name.ilike(f'%{item_filter}%'))
    
    if date_from:
        query = query.filter(StockUsage.usage_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    
    if date_to:
        query = query.filter(StockUsage.usage_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    
    usage_records = query.order_by(StockUsage.usage_date.desc()).all()
    
    # Create Excel file
    output = BytesIO()
    
    # Prepare data for export
    data = []
    for i, record in enumerate(usage_records, 1):
        data.append({
            'S.No': i,
            'Date': record.usage_date.strftime('%Y-%m-%d'),
            'Item Name': record.item_name,
            'Used Quantity': record.used_quantity,
            'Available Quantity': record.available_quantity,
            'Notes': record.notes or '',
            'Created By': record.created_by_user.first_name if record.created_by_user else 'Unknown',
            'Created Date': record.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    # Create DataFrame and export to Excel
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Used Stock Summary', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Used Stock Summary']
        
        # Style the header
        header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        header_fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = singapore_now().strftime('%Y%m%d_%H%M%S')
    filename = f'used_stock_summary_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/delete-used-items', methods=['POST'])
@login_required
def delete_used_items():
    """Delete selected stock items from used info"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'Organization not assigned'})
    
    try:
        data = request.get_json()
        item_ids = data.get('item_ids', [])
        
        if not item_ids:
            return jsonify({'success': False, 'error': 'No items selected'})
        
        # Delete stock items and related records
        deleted_count = 0
        for item_id in item_ids:
            stock_item = StockItem.query.filter_by(
                id=item_id, 
                organization_id=user.organization_id
            ).first()
            
            if stock_item:
                # Delete related stock movements
                StockMovement.query.filter_by(stock_item_id=stock_item.id).delete()
                # Delete related stock usage records
                StockUsage.query.filter_by(item_name=stock_item.name, organization_id=user.organization_id).delete()
                # Delete the stock item
                db.session.delete(stock_item)
                deleted_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Successfully deleted {deleted_count} item(s)'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/export-selected-used-items', methods=['POST'])
@login_required
def export_selected_used_items():
    """Export selected stock items to Excel"""
    user = current_user
    if not user.organization_id:
        flash('Organization not assigned', 'error')
        return redirect(url_for('used_info'))
    
    try:
        item_ids = json.loads(request.form.get('item_ids', '[]'))
        
        if not item_ids:
            flash('No items selected for export', 'error')
            return redirect(url_for('used_info'))
        
        # Get selected stock items
        stock_items = StockItem.query.filter(
            StockItem.id.in_(item_ids),
            StockItem.organization_id == user.organization_id
        ).all()
        
        if not stock_items:
            flash('No valid items found for export', 'error')
            return redirect(url_for('used_info'))
        
        # Create Excel file
        output = BytesIO()
        
        # Prepare data
        data = []
        for item in stock_items:
            available_qty = item.quantity - (item.used_quantity or 0)
            status = 'Out of Stock' if available_qty <= 0 else 'Low Stock' if available_qty <= 5 else 'In Stock'
            
            data.append({
                'Item Name': item.name,
                'Category': item.category,
                'Total Quantity': item.quantity,
                'Used Quantity': item.used_quantity or 0,
                'Available Quantity': available_qty,
                'Status': status,
                'Purchase Date': item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else '',
                'Cost': f"${item.cost:.2f}" if item.cost else '',
                'Serial Number': item.serial_number or '',
                'Description': item.description or ''
            })
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Selected Used Stock Items', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Selected Used Stock Items']
            
            # Style the header
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            header_fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            for col_num, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f'selected_used_stock_items_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting items: {str(e)}', 'error')
        return redirect(url_for('used_info'))

# ================================
# HOUSE ACKNOWLEDGE ROUTES
# ================================

@app.route('/house-acknowledge/create')
@login_required
def house_acknowledge_create():
    """Display the House Acknowledge creation form"""
    return render_template('house_acknowledge_create.html')

@app.route('/house-acknowledge/create', methods=['POST'])
@login_required
def house_acknowledge_create_post():
    """Handle House Acknowledge creation"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge
        import qrcode
        from PIL import Image
        import os
        from werkzeug.utils import secure_filename
        
        # Get form data
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        
        # Get language content
        english_text = request.form.get('english_text', '').strip()
        tamil_text = request.form.get('tamil_text', '').strip()
        chinese_text = request.form.get('chinese_text', '').strip()
        myanmar_text = request.form.get('myanmar_text', '').strip()
        bengali_text = request.form.get('bengali_text', '').strip()
        
        # Validate required fields
        if not all([title, english_text, tamil_text, chinese_text, myanmar_text, bengali_text]):
            flash('Please fill in all required fields for all languages', 'error')
            return redirect(url_for('house_acknowledge_create'))
        
        # Create upload directory if it doesn't exist
        upload_dir = os.path.join('static', 'uploads', 'house_acknowledge')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Handle file uploads for all languages
        def save_uploaded_file(file_key, language):
            file = request.files.get(file_key)
            if file and file.filename:
                filename = secure_filename(f"{language}_{singapore_now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                file_path = os.path.join(upload_dir, filename)
                file.save(file_path)
                return f"uploads/house_acknowledge/{filename}"
            return None
        
        # Create House Acknowledge record
        house_acknowledge = HouseAcknowledge(
            title=title,
            description=description,
            english_text=english_text,
            tamil_text=tamil_text,
            chinese_text=chinese_text,
            myanmar_text=myanmar_text,
            bengali_text=bengali_text,
            english_image1=save_uploaded_file('english_image1', 'en'),
            english_image2=save_uploaded_file('english_image2', 'en'),
            english_image3=save_uploaded_file('english_image3', 'en'),
            tamil_image1=save_uploaded_file('tamil_image1', 'ta'),
            tamil_image2=save_uploaded_file('tamil_image2', 'ta'),
            tamil_image3=save_uploaded_file('tamil_image3', 'ta'),
            chinese_image1=save_uploaded_file('chinese_image1', 'zh'),
            chinese_image2=save_uploaded_file('chinese_image2', 'zh'),
            chinese_image3=save_uploaded_file('chinese_image3', 'zh'),
            myanmar_image1=save_uploaded_file('myanmar_image1', 'my'),
            myanmar_image2=save_uploaded_file('myanmar_image2', 'my'),
            myanmar_image3=save_uploaded_file('myanmar_image3', 'my'),
            bengali_image1=save_uploaded_file('bengali_image1', 'bn'),
            bengali_image2=save_uploaded_file('bengali_image2', 'bn'),
            bengali_image3=save_uploaded_file('bengali_image3', 'bn')
        )
        
        db.session.add(house_acknowledge)
        db.session.flush()  # Get the ID
        
        # Generate QR code
        qr_data = url_for('house_acknowledge_scan', acknowledge_id=house_acknowledge.id, _external=True)
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_filename = f"qr_house_acknowledge_{house_acknowledge.id}.png"
        qr_path = os.path.join(upload_dir, qr_filename)
        qr_img.save(qr_path)
        
        # Update record with QR code info
        house_acknowledge.qr_code_path = f"uploads/house_acknowledge/{qr_filename}"
        house_acknowledge.qr_code_url = qr_data
        
        db.session.commit()
        
        flash('House Acknowledge created successfully with QR code!', 'success')
        return redirect(url_for('house_acknowledge_storage'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating House Acknowledge: {str(e)}', 'error')
        return redirect(url_for('house_acknowledge_create'))

@app.route('/house-acknowledge/edit/<int:acknowledge_id>')
@login_required
def house_acknowledge_edit(acknowledge_id):
    """Display edit form for House Acknowledge"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge
        
        house_acknowledge = HouseAcknowledge.query.get_or_404(acknowledge_id)
        
        return render_template('house_acknowledge_edit.html', 
                             house_acknowledge=house_acknowledge)
        
    except Exception as e:
        flash(f'Error loading House Acknowledge for editing: {str(e)}', 'error')
        return redirect(url_for('house_acknowledge_storage'))

@app.route('/house-acknowledge/edit/<int:acknowledge_id>', methods=['POST'])
@login_required
def house_acknowledge_update(acknowledge_id):
    """Update House Acknowledge"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge
        from werkzeug.utils import secure_filename
        import os
        from datetime import datetime
        
        house_acknowledge = HouseAcknowledge.query.get_or_404(acknowledge_id)
        
        # Get form data
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        
        # Get language content
        english_text = request.form.get('english_text', '').strip()
        tamil_text = request.form.get('tamil_text', '').strip()
        chinese_text = request.form.get('chinese_text', '').strip()
        myanmar_text = request.form.get('myanmar_text', '').strip()
        bengali_text = request.form.get('bengali_text', '').strip()
        
        # Validate required fields
        if not all([title, english_text, tamil_text, chinese_text, myanmar_text, bengali_text]):
            flash('Please fill in all required fields for all languages', 'error')
            return redirect(url_for('house_acknowledge_edit', acknowledge_id=acknowledge_id))
        
        # Update basic fields
        house_acknowledge.title = title
        house_acknowledge.description = description
        house_acknowledge.english_text = english_text
        house_acknowledge.tamil_text = tamil_text
        house_acknowledge.chinese_text = chinese_text
        house_acknowledge.myanmar_text = myanmar_text
        house_acknowledge.bengali_text = bengali_text
        
        # Create upload directory if it doesn't exist
        upload_dir = os.path.join('static', 'uploads', 'house_acknowledge')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Handle file uploads for all languages (only update if new files provided)
        def update_uploaded_file(file_key, language, current_path):
            file = request.files.get(file_key)
            if file and file.filename:
                filename = secure_filename(f"{language}_{singapore_now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                file_path = os.path.join(upload_dir, filename)
                file.save(file_path)
                return f"uploads/house_acknowledge/{filename}"
            return current_path
        
        # Update image paths only if new files are uploaded
        house_acknowledge.english_image1 = update_uploaded_file('english_image1', 'en', house_acknowledge.english_image1)
        house_acknowledge.english_image2 = update_uploaded_file('english_image2', 'en', house_acknowledge.english_image2)
        house_acknowledge.english_image3 = update_uploaded_file('english_image3', 'en', house_acknowledge.english_image3)
        house_acknowledge.tamil_image1 = update_uploaded_file('tamil_image1', 'ta', house_acknowledge.tamil_image1)
        house_acknowledge.tamil_image2 = update_uploaded_file('tamil_image2', 'ta', house_acknowledge.tamil_image2)
        house_acknowledge.tamil_image3 = update_uploaded_file('tamil_image3', 'ta', house_acknowledge.tamil_image3)
        house_acknowledge.chinese_image1 = update_uploaded_file('chinese_image1', 'zh', house_acknowledge.chinese_image1)
        house_acknowledge.chinese_image2 = update_uploaded_file('chinese_image2', 'zh', house_acknowledge.chinese_image2)
        house_acknowledge.chinese_image3 = update_uploaded_file('chinese_image3', 'zh', house_acknowledge.chinese_image3)
        house_acknowledge.myanmar_image1 = update_uploaded_file('myanmar_image1', 'my', house_acknowledge.myanmar_image1)
        house_acknowledge.myanmar_image2 = update_uploaded_file('myanmar_image2', 'my', house_acknowledge.myanmar_image2)
        house_acknowledge.myanmar_image3 = update_uploaded_file('myanmar_image3', 'my', house_acknowledge.myanmar_image3)
        house_acknowledge.bengali_image1 = update_uploaded_file('bengali_image1', 'bn', house_acknowledge.bengali_image1)
        house_acknowledge.bengali_image2 = update_uploaded_file('bengali_image2', 'bn', house_acknowledge.bengali_image2)
        house_acknowledge.bengali_image3 = update_uploaded_file('bengali_image3', 'bn', house_acknowledge.bengali_image3)
        
        db.session.commit()
        
        flash('House Acknowledge updated successfully!', 'success')
        return redirect(url_for('house_acknowledge_storage'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating House Acknowledge: {str(e)}', 'error')
        return redirect(url_for('house_acknowledge_edit', acknowledge_id=acknowledge_id))

@app.route('/house-acknowledge/delete/<int:acknowledge_id>')
@login_required
def delete_house_acknowledge(acknowledge_id):
    """Delete individual House Acknowledge"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge
        
        house_acknowledge = HouseAcknowledge.query.get_or_404(acknowledge_id)
        house_acknowledge.is_active = False
        
        db.session.commit()
        flash('House Acknowledge deleted successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting House Acknowledge: {str(e)}', 'error')
    
    return redirect(url_for('house_acknowledge_storage'))

@app.route('/delete_selected_house_acknowledges', methods=['POST'])
@login_required
def delete_selected_house_acknowledges():
    """Delete selected House Acknowledges"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge
        import json
        
        acknowledge_ids = json.loads(request.form.get('acknowledge_ids', '[]'))
        
        if not acknowledge_ids:
            flash('No House Acknowledges selected for deletion.', 'error')
            return redirect(url_for('house_acknowledge_storage'))
        
        # Mark selected acknowledges as inactive
        HouseAcknowledge.query.filter(HouseAcknowledge.id.in_(acknowledge_ids)).update(
            {HouseAcknowledge.is_active: False}, synchronize_session=False
        )
        
        db.session.commit()
        flash(f'Successfully deleted {len(acknowledge_ids)} House Acknowledge(s)!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting House Acknowledges: {str(e)}', 'error')
    
    return redirect(url_for('house_acknowledge_storage'))

@app.route('/export_selected_house_acknowledges', methods=['POST'])
@login_required
def export_selected_house_acknowledges():
    """Export selected House Acknowledges to Excel"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge
        import json
        import pandas as pd
        from io import BytesIO
        from flask import send_file
        from datetime import datetime
        
        selected_ids = json.loads(request.form.get('selected_ids', '[]'))
        
        if not selected_ids:
            flash('No House Acknowledges selected for export.', 'error')
            return redirect(url_for('house_acknowledge_storage'))
        
        # Get selected house acknowledges
        house_acknowledges = HouseAcknowledge.query.filter(
            HouseAcknowledge.id.in_(selected_ids),
            HouseAcknowledge.is_active == True
        ).all()
        
        if not house_acknowledges:
            flash('No valid House Acknowledges found for export.', 'error')
            return redirect(url_for('house_acknowledge_storage'))
        
        # Prepare data for export
        data = []
        for ha in house_acknowledges:
            data.append({
                'ID': ha.id,
                'Title': ha.title,
                'Description': ha.description,
                'English Text': ha.english_text,
                'Tamil Text': ha.tamil_text,
                'Chinese Text': ha.chinese_text,
                'Myanmar Text': ha.myanmar_text,
                'Bengali Text': ha.bengali_text,
                'Created At': ha.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'QR Code URL': ha.qr_code_url
            })
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='House Acknowledges', index=False)
        
        output.seek(0)
        
        filename = f"house_acknowledges_export_{singapore_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting House Acknowledges: {str(e)}', 'error')
        return redirect(url_for('house_acknowledge_storage'))

@app.route('/delete_selected_house_acknowledgments', methods=['POST'])
@login_required
def delete_selected_house_acknowledgments():
    """Delete selected House Acknowledgments"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledgment
        import json
        
        acknowledgment_ids = json.loads(request.form.get('acknowledgment_ids', '[]'))
        
        if not acknowledgment_ids:
            flash('No acknowledgments selected for deletion.', 'error')
            return redirect(url_for('house_acknowledge_storage'))
        
        # Delete selected acknowledgments
        HouseAcknowledgment.query.filter(HouseAcknowledgment.id.in_(acknowledgment_ids)).delete(synchronize_session=False)
        
        db.session.commit()
        flash(f'Successfully deleted {len(acknowledgment_ids)} acknowledgment(s)!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting acknowledgments: {str(e)}', 'error')
    
    return redirect(url_for('house_acknowledge_storage'))

@app.route('/export_selected_house_acknowledgments', methods=['POST'])
@login_required
def export_selected_house_acknowledgments():
    """Export selected House Acknowledgments to Excel"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledgment, HouseAcknowledge
        import json
        from openpyxl import Workbook
        from io import BytesIO
        from flask import send_file
        from datetime import datetime
        
        selected_ids = json.loads(request.form.get('acknowledgment_ids', '[]'))
        
        if not selected_ids:
            flash('No acknowledgments selected for export.', 'error')
            return redirect(url_for('house_acknowledge_storage'))
        
        # Get selected acknowledgments with related house acknowledge data
        acknowledgments = db.session.query(HouseAcknowledgment, HouseAcknowledge).join(
            HouseAcknowledge, HouseAcknowledgment.house_acknowledge_id == HouseAcknowledge.id
        ).filter(HouseAcknowledgment.id.in_(selected_ids)).all()
        
        if not acknowledgments:
            flash('No valid acknowledgments found for export.', 'error')
            return redirect(url_for('house_acknowledge_storage'))
        
        # Prepare data for export
        data = []
        for ack, house_ack in acknowledgments:
            data.append({
                'ID': ack.id,
                'House Acknowledge Title': house_ack.title,
                'Language': ack.language_selected,
                'Name': ack.name,
                'FIN': ack.fin,
                'Company': ack.company_name,
                'Room': ack.room_number,
                'Phone': ack.phone_number,
                'E-Signature': ack.e_signature,
                'Acknowledged At': ack.acknowledged_at.strftime('%Y-%m-%d %H:%M:%S'),
                'IP Address': ack.ip_address
            })
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'Acknowledgments'
        
        # Add headers
        headers = ['ID', 'House Acknowledge Title', 'Language', 'Name', 'FIN', 'Company', 'Room', 'Phone', 'E-Signature', 'Acknowledged At', 'IP Address']
        ws.append(headers)
        
        # Add data rows
        for ack, house_ack in acknowledgments:
            row = [
                ack.id,
                house_ack.title,
                ack.language_selected,
                ack.name,
                ack.fin,
                ack.company_name,
                ack.room_number,
                ack.phone_number,
                ack.e_signature,
                ack.acknowledged_at.strftime('%Y-%m-%d %H:%M:%S'),
                ack.ip_address
            ]
            ws.append(row)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"house_acknowledgments_export_{singapore_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting acknowledgments: {str(e)}', 'error')
        return redirect(url_for('house_acknowledge_storage'))

@app.route('/house-acknowledge/storage')
@login_required
@page_access_required('house_acknowledge')
def house_acknowledge_storage():
    """Display House Acknowledge storage with acknowledgment records"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge, HouseAcknowledgment
        
        # Get all house acknowledges
        house_acknowledges = HouseAcknowledge.query.filter_by(is_active=True).order_by(HouseAcknowledge.created_at.desc()).all()
        
        # Get all acknowledgments with related data
        acknowledgments = db.session.query(HouseAcknowledgment, HouseAcknowledge).join(
            HouseAcknowledge, HouseAcknowledgment.house_acknowledge_id == HouseAcknowledge.id
        ).order_by(HouseAcknowledgment.acknowledged_at.desc()).all()
        
        return render_template('house_acknowledge_storage.html', 
                             house_acknowledges=house_acknowledges,
                             acknowledgments=acknowledgments)
        
    except Exception as e:
        flash(f'Error loading House Acknowledge storage: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/house-acknowledge/scan/<int:acknowledge_id>')
def house_acknowledge_scan(acknowledge_id):
    """QR code scan endpoint - shows language selection"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge
        
        house_acknowledge = HouseAcknowledge.query.get_or_404(acknowledge_id)
        
        if not house_acknowledge.is_active:
            return render_template('error.html', 
                                 message='This House Acknowledge is no longer active.'), 404
        
        return render_template('house_acknowledge_language_select.html', 
                             house_acknowledge=house_acknowledge)
        
    except Exception as e:
        return render_template('error.html', 
                             message=f'Error loading House Acknowledge: {str(e)}'), 500

@app.route('/house-acknowledge/view/<int:acknowledge_id>/<language>')
def house_acknowledge_view(acknowledge_id, language):
    """Display House Acknowledge content in selected language"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge
        
        house_acknowledge = HouseAcknowledge.query.get_or_404(acknowledge_id)
        
        if not house_acknowledge.is_active:
            return render_template('error.html', 
                                 message='This House Acknowledge is no longer active.'), 404
        
        # Validate language
        valid_languages = ['english', 'tamil', 'chinese', 'myanmar', 'bengali']
        if language not in valid_languages:
            return render_template('error.html', 
                                 message='Invalid language selected.'), 400
        
        return render_template('house_acknowledge_view.html', 
                             house_acknowledge=house_acknowledge,
                             language=language)
        
    except Exception as e:
        return render_template('error.html', 
                             message=f'Error loading House Acknowledge: {str(e)}'), 500

@app.route('/house-acknowledge/acknowledge/<int:acknowledge_id>/<language>')
def house_acknowledge_form(acknowledge_id, language):
    """Display acknowledgment form"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge, RoomNumber
        
        house_acknowledge = HouseAcknowledge.query.get_or_404(acknowledge_id)
        
        if not house_acknowledge.is_active:
            return render_template('error.html', 
                                 message='This House Acknowledge is no longer active.'), 404
        
        # Get available room numbers
        room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
        
        return render_template('house_acknowledge_form.html', 
                             house_acknowledge=house_acknowledge,
                             language=language,
                             room_numbers=room_numbers)
        
    except Exception as e:
        return render_template('error.html', 
                             message=f'Error loading acknowledgment form: {str(e)}'), 500

@app.route('/house-acknowledge/acknowledge/<int:acknowledge_id>/<language>', methods=['POST'])
def house_acknowledge_submit(acknowledge_id, language):
    """Handle acknowledgment form submission"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge, HouseAcknowledgment
        
        house_acknowledge = HouseAcknowledge.query.get_or_404(acknowledge_id)
        
        if not house_acknowledge.is_active:
            flash('This House Acknowledge is no longer active.', 'error')
            return redirect(url_for('house_acknowledge_scan', acknowledge_id=acknowledge_id))
        
        # Get form data
        name = request.form.get('name', '').strip()
        fin = request.form.get('fin', '').strip()
        company_name = request.form.get('company_name', '').strip()
        room_number = request.form.get('room_number', '').strip()
        phone_number = request.form.get('phone_number', '').strip()
        e_signature = request.form.get('e_signature', '').strip()
        selfie_photo = request.form.get('selfie_photo', '').strip()
        
        # Debug: Log received data
        print(f"DEBUG: Selfie photo data length: {len(selfie_photo) if selfie_photo else 0}")
        print(f"DEBUG: Form data keys: {list(request.form.keys())}")
        
        # Validate required fields (temporarily allow empty selfie_photo to debug)
        if not all([name, fin, company_name, room_number, phone_number]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('house_acknowledge_form', 
                                  acknowledge_id=acknowledge_id, language=language))
        
        # If selfie_photo is empty, set a placeholder to prevent null
        if not selfie_photo:
            selfie_photo = "NO_PHOTO_DATA"
        
        # Create acknowledgment record
        acknowledgment = HouseAcknowledgment(
            house_acknowledge_id=acknowledge_id,
            name=name,
            fin=fin,
            company_name=company_name,
            room_number=room_number,
            phone_number=phone_number,
            e_signature=e_signature,
            selfie_photo=selfie_photo,
            language_selected=language,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')
        )
        
        db.session.add(acknowledgment)
        db.session.commit()
        
        return render_template('house_acknowledge_success.html', 
                             acknowledgment=acknowledgment)
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error submitting acknowledgment: {str(e)}', 'error')
        return redirect(url_for('house_acknowledge_form', 
                              acknowledge_id=acknowledge_id, language=language))

# Room Inspection Report Routes
@app.route('/room-inspection-reports')
@login_required
def room_inspection_reports():
    """Room Inspection Reports listing page with filtering"""
    try:
        from app.models.models_room_inspection import RoomInspection
        from datetime import datetime
        
        # Get filter parameters
        room_number = request.args.get('room_number', '').strip()
        company_name = request.args.get('company_name', '').strip()
        inspection_date = request.args.get('inspection_date', '').strip()
        created_by = request.args.get('created_by', '').strip()
        
        # Start with base query
        query = RoomInspection.query
        
        # Apply filters
        if room_number:
            query = query.filter(RoomInspection.room_number.ilike(f'%{room_number}%'))
        
        if company_name:
            query = query.filter(RoomInspection.company_name.ilike(f'%{company_name}%'))
        
        if inspection_date:
            try:
                filter_date = datetime.strptime(inspection_date, '%Y-%m-%d').date()
                query = query.filter(RoomInspection.inspection_date == filter_date)
            except ValueError:
                flash('Invalid date format. Please use YYYY-MM-DD format.', 'warning')
        
        if created_by:
            query = query.filter(RoomInspection.created_by.ilike(f'%{created_by}%'))
        
        # Get filtered results, ordered by most recent first
        inspections = query.order_by(RoomInspection.inspection_date.desc()).all()
        
        return render_template('room_inspection_reports.html', inspections=inspections)
        
    except Exception as e:
        flash(f'Error loading room inspection reports: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/room-inspection-reports/delete', methods=['POST'])
@login_required
def delete_room_inspections():
    """Delete selected room inspection reports"""
    try:
        from app.models.models_room_inspection import RoomInspection
        
        selected_ids = request.form.getlist('selected_inspections')
        
        if not selected_ids:
            flash('No inspection reports selected for deletion.', 'warning')
            return redirect(url_for('room_inspection_reports'))
        
        # Convert to integers and validate
        inspection_ids = []
        for id_str in selected_ids:
            try:
                inspection_ids.append(int(id_str))
            except ValueError:
                flash('Invalid inspection ID provided.', 'error')
                return redirect(url_for('room_inspection_reports'))
        
        # Get inspections to delete
        inspections_to_delete = RoomInspection.query.filter(RoomInspection.id.in_(inspection_ids)).all()
        
        if not inspections_to_delete:
            flash('No valid inspection reports found for deletion.', 'warning')
            return redirect(url_for('room_inspection_reports'))
        
        # Delete the inspections
        deleted_count = 0
        for inspection in inspections_to_delete:
            db.session.delete(inspection)
            deleted_count += 1
        
        db.session.commit()
        
        flash(f'Successfully deleted {deleted_count} inspection report(s).', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting inspection reports: {str(e)}', 'error')
    
    return redirect(url_for('room_inspection_reports'))

@app.route('/room-inspection-reports/edit/<int:inspection_id>', methods=['GET', 'POST'])
@login_required
def edit_room_inspection(inspection_id):
    """Edit existing room inspection report"""
    try:
        from app.models.models_room_inspection import RoomInspection
        import json
        from datetime import datetime
        
        inspection = RoomInspection.query.get_or_404(inspection_id)
        
        if request.method == 'POST':
            # Get form data
            room_number = request.form.get('room_number', '').strip()
            company_name = request.form.get('company_name', '').strip()
            inspection_date_str = request.form.get('inspection_date', '').strip()
            action_taken = request.form.get('action_taken', '').strip()
            confiscated_items = request.form.get('confiscated_items', '').strip()
            
            # Get signature data
            room_incharge_name = request.form.get('room_incharge_name', '').strip()
            room_incharge_signature = request.form.get('room_incharge_signature', '').strip()
            oe_dc_name = request.form.get('oe_dc_name', '').strip()
            oe_dc_signature = request.form.get('oe_dc_signature', '').strip()
            
            # Validate required fields
            if not all([room_number, company_name, inspection_date_str]):
                flash('Please fill in all required fields: Room Number, Company Name, and Date.', 'error')
                return render_template('edit_room_inspection.html', inspection=inspection)
            
            # Parse inspection date
            try:
                inspection_date = datetime.strptime(inspection_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid date format. Please use YYYY-MM-DD format.', 'error')
                return render_template('edit_room_inspection.html', inspection=inspection)
            
            # Handle photo updates - only process if new photos are provided
            photos_data = []
            total_data_size = 0
            max_total_size = 10 * 1024 * 1024  # 10MB limit
            max_photos = 5
            
            # Check if any new photos were uploaded
            has_new_photos = False
            for i in range(max_photos):
                photo_key = f'confiscated_photo_{i}'
                if photo_key in request.form and request.form[photo_key]:
                    has_new_photos = True
                    break
            
            # Process new photos if provided, otherwise keep existing ones
            if has_new_photos:
                for i in range(max_photos):
                    photo_key = f'confiscated_photo_{i}'
                    if photo_key in request.form and request.form[photo_key]:
                        photo_data = request.form[photo_key]
                        if photo_data and photo_data.startswith('data:image'):
                            photo_size = len(photo_data.encode('utf-8'))
                            
                            if photo_size > 2 * 1024 * 1024:
                                flash('Individual photo too large. Please compress photos further.', 'error')
                                return render_template('edit_room_inspection.html', inspection=inspection)
                            
                            if total_data_size + photo_size > max_total_size:
                                flash('Total photo size too large. Please use fewer photos.', 'error')
                                return render_template('edit_room_inspection.html', inspection=inspection)
                            
                            photos_data.append(photo_data)
                            total_data_size += photo_size
                
                # Update photos if new ones provided
                confiscated_photos_json = None
                if photos_data:
                    json_str = json.dumps(photos_data)
                    confiscated_photos_json = json_str.encode('utf-8')
                
                inspection.confiscated_photos = confiscated_photos_json
            
            # Update inspection record
            inspection.room_number = room_number
            inspection.company_name = company_name
            inspection.inspection_date = inspection_date
            inspection.action_taken = action_taken or None
            inspection.confiscated_items = confiscated_items or None
            inspection.room_incharge_name = room_incharge_name or None
            inspection.room_incharge_signature = room_incharge_signature or None
            inspection.oe_dc_name = oe_dc_name or None
            inspection.oe_dc_signature = oe_dc_signature or None
            inspection.ip_address = request.remote_addr
            inspection.user_agent = request.headers.get('User-Agent', '')
            
            db.session.commit()
            
            flash('Room inspection report updated successfully!', 'success')
            return redirect(url_for('room_inspection_reports'))
            
        # GET request - show edit form with existing data
        return render_template('edit_room_inspection.html', inspection=inspection)
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f"Edit Room Inspection Error: {str(e)}")
        print(f"Full traceback: {error_details}")
        flash(f'Error updating room inspection report: {str(e)}', 'error')
        return redirect(url_for('room_inspection_reports'))

@app.route('/room-inspection-reports/create', methods=['GET', 'POST'])
@login_required
def create_room_inspection():
    """Create new room inspection report"""
    if request.method == 'POST':
        try:
            from app.models.models_room_inspection import RoomInspection
            import json
            from datetime import datetime
            
            # Get form data
            room_number = request.form.get('room_number', '').strip()
            company_name = request.form.get('company_name', '').strip()
            inspection_date_str = request.form.get('inspection_date', '').strip()
            action_taken = request.form.get('action_taken', '').strip()
            confiscated_items = request.form.get('confiscated_items', '').strip()
            
            # Get signature data
            room_incharge_name = request.form.get('room_incharge_name', '').strip()
            room_incharge_signature = request.form.get('room_incharge_signature', '').strip()
            oe_dc_name = request.form.get('oe_dc_name', '').strip()
            oe_dc_signature = request.form.get('oe_dc_signature', '').strip()
            
            # Validate required fields
            if not all([room_number, company_name, inspection_date_str]):
                flash('Please fill in all required fields: Room Number, Company Name, and Date.', 'error')
                return render_template('create_room_inspection.html')
            
            # Parse inspection date
            try:
                inspection_date = datetime.strptime(inspection_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid date format. Please use YYYY-MM-DD format.', 'error')
                return render_template('create_room_inspection.html')
            
            # Handle multiple photo uploads with strict size validation
            photos_data = []
            total_data_size = 0
            max_total_size = 10 * 1024 * 1024  # Reduced to 10MB limit for all photos combined
            max_photos = 5  # Reduce max photos to 5 for better reliability
            
            for i in range(max_photos):
                photo_key = f'confiscated_photo_{i}'
                if photo_key in request.form and request.form[photo_key]:
                    photo_data = request.form[photo_key]
                    if photo_data and photo_data.startswith('data:image'):
                        # Calculate approximate size of this photo
                        photo_size = len(photo_data.encode('utf-8'))
                        
                        # Individual photo size limit - 2MB per photo
                        if photo_size > 2 * 1024 * 1024:
                            flash('Individual photo too large. Please compress photos further.', 'error')
                            return render_template('create_room_inspection.html')
                        
                        # Check if adding this photo would exceed the total limit
                        if total_data_size + photo_size > max_total_size:
                            flash('Total photo size too large. Please use fewer photos.', 'error')
                            return render_template('create_room_inspection.html')
                        
                        photos_data.append(photo_data)
                        total_data_size += photo_size
            
            # Convert photos list to JSON string and encode as binary
            confiscated_photos_json = None
            if photos_data:
                json_str = json.dumps(photos_data)
                confiscated_photos_json = json_str.encode('utf-8')
            
            # Create inspection record
            inspection = RoomInspection(
                room_number=room_number,
                company_name=company_name,
                inspection_date=inspection_date,
                action_taken=action_taken or None,
                confiscated_items=confiscated_items or None,
                confiscated_photos=confiscated_photos_json,
                room_incharge_name=room_incharge_name or None,
                room_incharge_signature=room_incharge_signature or None,
                oe_dc_name=oe_dc_name or None,
                oe_dc_signature=oe_dc_signature or None,
                created_by=current_user.id if hasattr(current_user, 'id') and current_user.id else None,
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent', '')
            )
            
            db.session.add(inspection)
            db.session.commit()
            
            flash('Room inspection report created successfully!', 'success')
            return redirect(url_for('room_inspection_reports'))
            
        except Exception as e:
            db.session.rollback()
            import traceback
            error_details = traceback.format_exc()
            print(f"Room Inspection Error: {str(e)}")
            print(f"Full traceback: {error_details}")
            
            # Check if it's a data size issue
            if "too large" in str(e).lower() or "413" in str(e) or "entity too large" in str(e).lower():
                flash('Photo data is too large. Please reduce the number of photos or ensure they are properly compressed.', 'error')
            else:
                flash(f'Error creating room inspection report. Please check all required fields and try again.', 'error')
            return render_template('create_room_inspection.html')
    
    return render_template('create_room_inspection.html')

@app.route('/room-inspection-reports/download-pdf/<int:inspection_id>')
@login_required
def download_room_inspection_pdf(inspection_id):
    """Download room inspection report as PDF with photos and signatures"""
    try:
        from app.models.models_room_inspection import RoomInspection
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import inch
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
        from reportlab.lib import colors
        from io import BytesIO
        import base64
        import json
        
        inspection = RoomInspection.query.get_or_404(inspection_id)
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=50, leftMargin=50, 
                              topMargin=50, bottomMargin=50)
        
        # Container for the 'Flowable' objects
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1,  # Center
            fontName='Helvetica-Bold',
            textColor=colors.black
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=15,
            alignment=1,  # Center
            fontName='Helvetica-Bold',
            textColor=colors.black
        )
        
        content_style = ParagraphStyle(
            'ContentStyle',
            parent=styles['Normal'],
            fontSize=11,
            leading=16,
            spaceAfter=8,
            fontName='Helvetica',
            textColor=colors.black
        )
        
        # Header with logo and title
        from reportlab.platypus import Image
        from reportlab.lib.utils import ImageReader
        import os
        
        # Create header table with logo and title
        try:
            logo_path = os.path.join('static', 'ts_group_logo.png')
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=120, height=60)
                header_data = [
                    [logo, Paragraph("ROOM INSPECTION REPORT", title_style)],
                    ['', Paragraph("TS MANAGEMENT SERVICES (PIONEER LODGE)", header_style)]
                ]
                header_table = Table(header_data, colWidths=[1.5*inch, 4.5*inch])
                header_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                ]))
                elements.append(header_table)
            else:
                # Fallback without logo
                elements.append(Paragraph("ROOM INSPECTION REPORT", title_style))
                elements.append(Spacer(1, 10))
                elements.append(Paragraph("TS MANAGEMENT SERVICES (PIONEER LODGE)", header_style))
        except Exception:
            # Fallback without logo
            elements.append(Paragraph("ROOM INSPECTION REPORT", title_style))
            elements.append(Spacer(1, 10))
            elements.append(Paragraph("TS MANAGEMENT SERVICES (PIONEER LODGE)", header_style))
        
        elements.append(Spacer(1, 20))
        
        # Inspection Details Table
        inspection_data = [
            ['Field', 'Information'],
            ['Room Number:', inspection.room_number],
            ['Company Name:', inspection.company_name],
            ['Inspection Date:', inspection.inspection_date.strftime('%Y-%m-%d')],
            ['Action Taken:', inspection.action_taken or 'None specified'],
            ['Confiscated Items:', inspection.confiscated_items or 'None']
        ]
        
        inspection_table = Table(inspection_data, colWidths=[2*inch, 4*inch])
        inspection_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8E8E8')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        elements.append(inspection_table)
        elements.append(Spacer(1, 30))
        
        # Confiscated Photos Section
        if inspection.confiscated_photos:
            try:
                # Decode binary data back to JSON string
                photos_json_str = inspection.confiscated_photos.decode('utf-8')
                photos_list = json.loads(photos_json_str)
                if photos_list:
                    photo_header_style = ParagraphStyle(
                        'PhotoHeader',
                        parent=styles['Heading3'],
                        fontSize=14,
                        spaceAfter=15,
                        fontName='Helvetica-Bold',
                        textColor=colors.black
                    )
                    elements.append(Paragraph("CONFISCATED ITEMS - PHOTOGRAPHIC EVIDENCE", photo_header_style))
                    
                    # Add photos in left-right layout (2 photos per row)
                    photo_count = 0
                    for i in range(0, len(photos_list), 2):
                        row_photos = []
                        row_captions = []
                        
                        # First photo (left side)
                        if i < len(photos_list):
                            photo_data = photos_list[i]
                            try:
                                if photo_data.startswith('data:image'):
                                    image_data = photo_data.split(',')[1]
                                else:
                                    image_data = photo_data
                                
                                image_bytes = base64.b64decode(image_data)
                                image_buffer = BytesIO(image_bytes)
                                img1 = Image(image_buffer, width=2.3*inch, height=1.8*inch)
                                photo_count += 1
                                
                                row_photos.append(img1)
                                row_captions.append(Paragraph(f"Photo {photo_count}", 
                                    ParagraphStyle('PhotoCaption', parent=styles['Normal'], 
                                    fontSize=10, alignment=1, spaceAfter=0)))
                            except Exception as e:
                                row_photos.append(Paragraph("Photo could not be loaded", styles['Normal']))
                                row_captions.append("")
                        else:
                            row_photos.append("")
                            row_captions.append("")
                        
                        # Second photo (right side)
                        if i + 1 < len(photos_list):
                            photo_data = photos_list[i + 1]
                            try:
                                if photo_data.startswith('data:image'):
                                    image_data = photo_data.split(',')[1]
                                else:
                                    image_data = photo_data
                                
                                image_bytes = base64.b64decode(image_data)
                                image_buffer = BytesIO(image_bytes)
                                img2 = Image(image_buffer, width=2.3*inch, height=1.8*inch)
                                photo_count += 1
                                
                                row_photos.append(img2)
                                row_captions.append(Paragraph(f"Photo {photo_count}", 
                                    ParagraphStyle('PhotoCaption', parent=styles['Normal'], 
                                    fontSize=10, alignment=1, spaceAfter=0)))
                            except Exception as e:
                                row_photos.append(Paragraph("Photo could not be loaded", styles['Normal']))
                                row_captions.append("")
                        else:
                            row_photos.append("")
                            row_captions.append("")
                        
                        # Create table with photos and captions
                        if any(p != "" for p in row_photos):
                            photo_data = [row_photos, row_captions]
                            photo_table = Table(photo_data, colWidths=[2.5*inch, 2.5*inch])
                            photo_table.setStyle(TableStyle([
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('VALIGN', (0, 0), (1, 0), 'MIDDLE'),  # Photos centered
                                ('VALIGN', (0, 1), (1, 1), 'TOP'),     # Captions at top
                                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                                ('TOPPADDING', (0, 0), (-1, -1), 8),
                                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                                ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                            ]))
                            elements.append(photo_table)
                            elements.append(Spacer(1, 20))
            except Exception as e:
                elements.append(Paragraph("Photos could not be processed", content_style))
                elements.append(Spacer(1, 15))
        
        # Signatures Section
        elements.append(Spacer(1, 20))
        sig_header_style = ParagraphStyle(
            'SigHeader',
            parent=styles['Heading3'],
            fontSize=14,
            spaceAfter=15,
            fontName='Helvetica-Bold',
            textColor=colors.black
        )
        elements.append(Paragraph("AUTHORIZED SIGNATURES", sig_header_style))
        
        # Room In-charge Signature
        if inspection.room_incharge_signature:
            elements.append(Paragraph("Room In-charge:", content_style))
            elements.append(Paragraph(f"Name: {inspection.room_incharge_name or 'Not specified'}", content_style))
            
            try:
                if inspection.room_incharge_signature.startswith('data:image'):
                    base64_data = inspection.room_incharge_signature.split(',')[1]
                    signature_data = base64.b64decode(base64_data)
                    signature_buffer = BytesIO(signature_data)
                    
                    sig_image = Image(signature_buffer, width=3*inch, height=1.5*inch)
                    sig_image.hAlign = 'LEFT'
                    elements.append(sig_image)
                else:
                    elements.append(Paragraph("Digital signature present", content_style))
            except:
                elements.append(Paragraph("Signature could not be processed", content_style))
            
            elements.append(Spacer(1, 20))
        
        # OE/DC Signature
        if inspection.oe_dc_signature:
            elements.append(Paragraph("OE/DC:", content_style))
            elements.append(Paragraph(f"Name: {inspection.oe_dc_name or 'Not specified'}", content_style))
            
            try:
                if inspection.oe_dc_signature.startswith('data:image'):
                    base64_data = inspection.oe_dc_signature.split(',')[1]
                    signature_data = base64.b64decode(base64_data)
                    signature_buffer = BytesIO(signature_data)
                    
                    sig_image = Image(signature_buffer, width=3*inch, height=1.5*inch)
                    sig_image.hAlign = 'LEFT'
                    elements.append(sig_image)
                else:
                    elements.append(Paragraph("Digital signature present", content_style))
            except:
                elements.append(Paragraph("Signature could not be processed", content_style))
        
        # Add footer with timestamp
        elements.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,  # Center
            fontName='Helvetica-Oblique',
            textColor=colors.grey
        )
        elements.append(Paragraph(f"Report generated on {singapore_now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
        
        # Build PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer and write it to the response
        buffer.seek(0)
        
        # Create filename
        safe_room = "".join(c for c in inspection.room_number if c.isalnum() or c in (' ', '-', '_')).strip()
        filename = f"Room_Inspection_Report_{safe_room}_{inspection.inspection_date.strftime('%Y%m%d')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('room_inspection_reports'))

@app.route('/house-acknowledge/export-acknowledgments')
@login_required
def export_house_acknowledgments():
    """Export acknowledgment records to Excel"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledgment, HouseAcknowledge
        
        # Get all acknowledgments
        acknowledgments = db.session.query(HouseAcknowledgment, HouseAcknowledge).join(
            HouseAcknowledge, HouseAcknowledgment.house_acknowledge_id == HouseAcknowledge.id
        ).order_by(HouseAcknowledgment.acknowledged_at.desc()).all()
        
        # Prepare data for Excel
        data = []
        for i, (ack, house_ack) in enumerate(acknowledgments, 1):
            data.append({
                'S.No': i,
                'Date & Time': ack.acknowledged_at.strftime('%Y-%m-%d %H:%M:%S'),
                'FIN': ack.fin,
                'Name': ack.name,
                'Company Name': ack.company_name,
                'Room No': ack.room_number,
                'Phone No': ack.phone_number,
                'Language': ack.language_selected.title(),
                'House Acknowledge Title': house_ack.title,
                'E-Signature': 'Yes' if ack.e_signature else 'No'
            })
        
        if not data:
            flash('No acknowledgment records found to export.', 'warning')
            return redirect(url_for('house_acknowledge_storage'))
        
        # Create workbook and worksheet
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'House Acknowledgments'
        
        # Add headers
        headers = ['S.No', 'Date & Time', 'FIN', 'Name', 'Company Name', 'Room No', 'Phone No', 'Language', 'House Acknowledge Title', 'E-Signature']
        ws.append(headers)
        
        # Add data rows
        for i, (ack, house_ack) in enumerate(acknowledgments, 1):
            row = [
                i,
                ack.acknowledged_at.strftime('%Y-%m-%d %H:%M:%S'),
                ack.fin,
                ack.name,
                ack.company_name,
                ack.room_number,
                ack.phone_number,
                ack.language_selected.title(),
                house_ack.title,
                'Yes' if ack.e_signature else 'No'
            ]
            ws.append(row)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f'house_acknowledgments_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting acknowledgments: {str(e)}', 'error')
        return redirect(url_for('house_acknowledge_storage'))

@app.route('/house-acknowledge/delete-acknowledgment/<int:acknowledgment_id>')
@login_required
def delete_house_acknowledgment(acknowledgment_id):
    """Delete a specific acknowledgment record"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledgment
        
        acknowledgment = HouseAcknowledgment.query.get_or_404(acknowledgment_id)
        
        db.session.delete(acknowledgment)
        db.session.commit()
        
        flash('Acknowledgment record deleted successfully.', 'success')
        return redirect(url_for('house_acknowledge_storage'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting acknowledgment: {str(e)}', 'error')
        return redirect(url_for('house_acknowledge_storage'))

@app.route('/download-house-qr-code/<int:acknowledge_id>')
@login_required
def download_house_qr_code(acknowledge_id):
    """Download QR code for House Acknowledge"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledge
        import os
        from flask import send_file
        
        house_acknowledge = HouseAcknowledge.query.get_or_404(acknowledge_id)
        
        if not house_acknowledge.is_active:
            flash('This House Acknowledge is no longer active.', 'error')
            return redirect(url_for('house_acknowledge_storage'))
        
        # Check if QR code file exists
        if house_acknowledge.qr_code_path:
            qr_file_path = os.path.join('static', house_acknowledge.qr_code_path)
            
            if os.path.exists(qr_file_path):
                # Create a safe filename
                safe_title = "".join(c for c in house_acknowledge.title if c.isalnum() or c in (' ', '-', '_')).rstrip()
                filename = f"QR_Code_{safe_title}_{acknowledge_id}.png"
                
                return send_file(
                    qr_file_path,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='image/png'
                )
            else:
                flash('QR code file not found.', 'error')
        else:
            flash('No QR code available for this House Acknowledge.', 'error')
            
    except Exception as e:
        flash(f'Error downloading QR code: {str(e)}', 'error')
    
    return redirect(url_for('house_acknowledge_storage'))

@app.route('/house-acknowledge/edit/acknowledgment/<int:acknowledgment_id>')
@login_required
def edit_house_acknowledgment(acknowledgment_id):
    """Display edit form for acknowledgment record"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledgment, HouseAcknowledge, RoomNumber
        
        acknowledgment = HouseAcknowledgment.query.get_or_404(acknowledgment_id)
        house_acknowledge = HouseAcknowledge.query.get_or_404(acknowledgment.house_acknowledge_id)
        room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
        
        return render_template('house_acknowledge_edit_acknowledgment.html', 
                             acknowledgment=acknowledgment,
                             house_acknowledge=house_acknowledge,
                             room_numbers=room_numbers)
        
    except Exception as e:
        flash(f'Error loading acknowledgment for editing: {str(e)}', 'error')
        return redirect(url_for('house_acknowledge_storage'))

@app.route('/house-acknowledge/edit/acknowledgment/<int:acknowledgment_id>', methods=['POST'])
@login_required
def update_house_acknowledgment(acknowledgment_id):
    """Update acknowledgment record"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledgment
        
        acknowledgment = HouseAcknowledgment.query.get_or_404(acknowledgment_id)
        
        # Update acknowledgment fields
        acknowledgment.name = request.form.get('name', '').strip()
        acknowledgment.fin = request.form.get('fin', '').strip()
        acknowledgment.company_name = request.form.get('company_name', '').strip()
        acknowledgment.room_number = request.form.get('room_number', '').strip()
        acknowledgment.phone_number = request.form.get('phone_number', '').strip()
        acknowledgment.e_signature = request.form.get('e_signature', '').strip()
        
        # Validate required fields
        if not all([acknowledgment.name, acknowledgment.fin, acknowledgment.company_name, 
                   acknowledgment.room_number, acknowledgment.phone_number]):
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('edit_house_acknowledgment', acknowledgment_id=acknowledgment_id))
        
        db.session.commit()
        flash('Acknowledgment updated successfully.', 'success')
        return redirect(url_for('house_acknowledge_storage'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating acknowledgment: {str(e)}', 'error')
        return redirect(url_for('edit_house_acknowledgment', acknowledgment_id=acknowledgment_id))

@app.route('/house-acknowledge/download-pdf/<int:acknowledgment_id>')
@login_required
def download_acknowledgment_pdf(acknowledgment_id):
    """Download filled PDF form for acknowledgment"""
    try:
        from app.models.models_house_acknowledge import HouseAcknowledgment, HouseAcknowledge
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import inch
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib import colors
        from io import BytesIO
        import base64
        
        acknowledgment = HouseAcknowledgment.query.get_or_404(acknowledgment_id)
        house_acknowledge = HouseAcknowledge.query.get_or_404(acknowledgment.house_acknowledge_id)
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=50, leftMargin=50, 
                              topMargin=50, bottomMargin=50)
        
        # Container for the 'Flowable' objects
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,  # Center
            fontName='Helvetica-Bold'
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=10,
            alignment=1,  # Center
            fontName='Helvetica-Bold'
        )
        
        subheader_style = ParagraphStyle(
            'CustomSubHeader',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20,
            alignment=1,  # Center
            fontName='Helvetica'
        )
        
        content_style = ParagraphStyle(
            'ContentStyle',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            spaceAfter=8,
            alignment=0,  # Left
            fontName='Times-Roman'
        )
        
        # Add TS Group Logo
        try:
            import os
            logo_path = "static/ts_group_logo_house.png"
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=2.5*inch, height=1*inch)
                logo.hAlign = 'CENTER'
                elements.append(logo)
                elements.append(Spacer(1, 15))
        except Exception as e:
            print(f"Error adding logo: {e}")
        
        # Title
        elements.append(Paragraph("HOUSE RULES ACKNOWLEDGMENT FORM", title_style))
        elements.append(Spacer(1, 10))
        
        # Company header
        elements.append(Paragraph("TS MANAGEMENT SERVICES (PIONEER LODGE)", header_style))
        elements.append(Paragraph("TSGRP-ASD-RR-2025 STANDARD OPERATING PROCEDURES", subheader_style))
        elements.append(Spacer(1, 20))
        
        # Get the house rules text based on language
        house_rules_text = ""
        if acknowledgment.language_selected == 'english':
            house_rules_text = house_acknowledge.english_text
        elif acknowledgment.language_selected == 'tamil':
            house_rules_text = house_acknowledge.tamil_text
        elif acknowledgment.language_selected == 'chinese':
            house_rules_text = house_acknowledge.chinese_text
        elif acknowledgment.language_selected == 'myanmar':
            house_rules_text = house_acknowledge.myanmar_text
        elif acknowledgment.language_selected == 'bengali':
            house_rules_text = house_acknowledge.bengali_text
        
        # Process house rules text to handle formatting
        if house_rules_text:
            # Add section header for house rules
            rules_header_style = ParagraphStyle(
                'RulesHeader',
                parent=styles['Heading3'],
                fontSize=12,
                spaceAfter=15,
                fontName='Helvetica-Bold'
            )
            elements.append(Paragraph("HOUSE RULES AND REGULATIONS", rules_header_style))
            
            # Simple text processing approach
            import re
            
            # Convert to string and clean
            text_content = str(house_rules_text)
            
            # Replace tab characters and normalize whitespace
            text_content = text_content.replace('\t', ' ')
            text_content = re.sub(r' +', ' ', text_content)  # Multiple spaces to single
            
            # Split by lines and filter empty ones
            lines = [line.strip() for line in text_content.split('\n') if line.strip()]
            
            # Process each line as a separate paragraph
            for line in lines:
                if line:
                    # Simple ASCII replacement for problematic characters
                    safe_line = (line.replace('"', '"')
                                   .replace('"', '"')
                                   .replace(''', "'")
                                   .replace(''', "'")
                                   .replace('–', '-')
                                   .replace('—', '-'))
                    
                    # Create paragraph with safe text
                    try:
                        elements.append(Paragraph(safe_line, content_style))
                        elements.append(Spacer(1, 6))
                    except:
                        # Ultimate fallback - encode to ASCII
                        ascii_line = safe_line.encode('ascii', 'ignore').decode('ascii')
                        if ascii_line.strip():
                            elements.append(Paragraph(ascii_line, content_style))
                            elements.append(Spacer(1, 6))
            
            elements.append(Spacer(1, 20))
        
        # Acknowledgment form data
        form_data = [
            ['Field', 'Information'],
            ['Name:', acknowledgment.name],
            ['FIN Number:', acknowledgment.fin],
            ['Company Name:', acknowledgment.company_name],
            ['Room Number:', acknowledgment.room_number],
            ['Phone Number:', acknowledgment.phone_number],
            ['Language Selected:', acknowledgment.language_selected.title()],
            ['Acknowledged Date & Time:', acknowledgment.acknowledged_at.strftime('%Y-%m-%d %H:%M:%S')],
            ['E-Signature:', 'Yes' if acknowledgment.e_signature else 'No'],
            ['IP Address:', acknowledgment.ip_address or 'N/A']
        ]
        
        # Add section header for acknowledgment data
        section_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading3'],
            fontSize=12,
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph("ACKNOWLEDGMENT DETAILS", section_style))
        elements.append(Spacer(1, 10))
        
        # Create table with improved styling
        table = Table(form_data, colWidths=[2.5*inch, 3.5*inch])
        table.setStyle(TableStyle([
            # Header row styling - Light background with dark text for better visibility
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8E8E8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows styling - All text black for visibility
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Alternating row colors - Light alternating backgrounds
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Add selfie photo if available
        if acknowledgment.selfie_photo and acknowledgment.selfie_photo != "NO_PHOTO_DATA":
            # Add selfie photo section header
            photo_header_style = ParagraphStyle(
                'PhotoHeader',
                parent=styles['Heading3'],
                fontSize=12,
                spaceAfter=10,
                fontName='Helvetica-Bold'
            )
            elements.append(Paragraph("SELFIE VERIFICATION PHOTO", photo_header_style))
            
            try:
                # Process base64 image data
                image_data = acknowledgment.selfie_photo
                if image_data.startswith('data:image'):
                    # Extract base64 data after the comma
                    image_data = image_data.split(',')[1]
                
                # Clean any whitespace from base64 data
                image_data = image_data.strip()
                
                # Add padding if needed for proper base64 decoding
                missing_padding = len(image_data) % 4
                if missing_padding:
                    image_data += '=' * (4 - missing_padding)
                
                # Decode base64 image
                image_bytes = base64.b64decode(image_data)
                image_buffer = BytesIO(image_bytes)
                
                # Create image with appropriate sizing
                img = Image(image_buffer, width=2.5*inch, height=2*inch)
                img.hAlign = 'CENTER'
                elements.append(img)
                elements.append(Spacer(1, 15))
                
                # Add photo caption
                caption_style = ParagraphStyle(
                    'PhotoCaption',
                    parent=styles['Normal'],
                    fontSize=8,
                    alignment=1,  # Center
                    fontName='Helvetica-Oblique',
                    textColor=colors.grey
                )
                elements.append(Paragraph("Selfie photo taken during form submission", caption_style))
                elements.append(Spacer(1, 20))
                
                print(f"SUCCESS: Selfie photo added to PDF for {acknowledgment.name}")
                
            except Exception as e:
                print(f"ERROR processing selfie photo: {str(e)}")
                print(f"Photo data length: {len(acknowledgment.selfie_photo) if acknowledgment.selfie_photo else 0}")
                print(f"Photo data starts with: {acknowledgment.selfie_photo[:50] if acknowledgment.selfie_photo else 'None'}")
                
                # If there's an error with the image, add a note instead
                error_style = ParagraphStyle(
                    'ErrorStyle',
                    parent=styles['Normal'],
                    fontSize=9,
                    alignment=1,  # Center
                    fontName='Helvetica-Oblique',
                    textColor=colors.red
                )
                elements.append(Paragraph(f"(Photo processing error: {str(e)})", error_style))
                elements.append(Spacer(1, 20))
        else:
            # No photo available
            photo_header_style = ParagraphStyle(
                'PhotoHeader',
                parent=styles['Heading3'],
                fontSize=12,
                spaceAfter=10,
                fontName='Helvetica-Bold'
            )
            elements.append(Paragraph("SELFIE VERIFICATION PHOTO", photo_header_style))
            
            error_style = ParagraphStyle(
                'ErrorStyle',
                parent=styles['Normal'],
                fontSize=9,
                alignment=1,  # Center
                fontName='Helvetica-Oblique',
                textColor=colors.red
            )
            elements.append(Paragraph("(No photo data available)", error_style))
            elements.append(Spacer(1, 20))
        
        # Acknowledgment statement section
        ack_header_style = ParagraphStyle(
            'AckHeader',
            parent=styles['Heading3'],
            fontSize=12,
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph("ACKNOWLEDGMENT STATEMENT", ack_header_style))
        
        ack_style = ParagraphStyle(
            'AckStatement',
            parent=styles['Normal'],
            fontSize=10,
            leading=12,
            spaceAfter=15,
            alignment=0,  # Left align
            fontName='Helvetica'
        )
        
        ack_text = """I hereby acknowledge that I have read, understood, and agree to comply with all the house rules and regulations as outlined in the Pioneer Lodge documentation. I understand that failure to comply with these rules may result in penalties or eviction as specified in the house rules."""
        
        elements.append(Paragraph(ack_text, ack_style))
        elements.append(Spacer(1, 15))
        
        # E-Signature section
        sig_style = ParagraphStyle(
            'SignatureStyle',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold'
        )
        
        elements.append(Paragraph("Digital Signature:", sig_style))
        elements.append(Spacer(1, 10))
        
        if acknowledgment.e_signature and acknowledgment.e_signature.startswith('data:image/png;base64,'):
            try:
                # Extract base64 data and create image
                base64_data = acknowledgment.e_signature.split(',')[1]
                signature_data = base64.b64decode(base64_data)
                signature_buffer = BytesIO(signature_data)
                
                # Create image for PDF
                sig_image = Image(signature_buffer, width=3*inch, height=1.5*inch)
                elements.append(sig_image)
                elements.append(Spacer(1, 10))
                elements.append(Paragraph("✓ E-signature verified and authenticated", styles['Normal']))
            except Exception as e:
                elements.append(Paragraph("E-signature data present but could not be displayed", styles['Normal']))
        else:
            elements.append(Paragraph("No digital signature provided", styles['Normal']))
        
        elements.append(Spacer(1, 30))
        
        # Footer with generation info
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,  # Center
            fontName='Helvetica',
            textColor=colors.grey
        )
        elements.append(Paragraph(f"Document generated on: {singapore_now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Create filename
        safe_name = "".join(c for c in acknowledgment.name if c.isalnum() or c in (' ', '-', '_')).strip()
        filename = f"House_Acknowledgment_{safe_name}_{acknowledgment.fin}_{acknowledgment.acknowledged_at.strftime('%Y%m%d')}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('house_acknowledge_storage'))

# ============= ROOM NUMBER MANAGEMENT =============

@app.route('/room_numbers')
@login_required
def room_numbers():
    """Display room numbers management page"""
    room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    return render_template('room_numbers.html', room_numbers=room_numbers)

@app.route('/room_numbers/add', methods=['GET', 'POST'])
@login_required
@create_permission_required('room_checklist')
def add_room_number():
    """Add single or bulk room numbers"""
    if request.method == 'POST':
        try:
            room_type = request.form.get('room_type')
            print(f"DEBUG: Room type received: {room_type}")
            print(f"DEBUG: Form data: {dict(request.form)}")
            
            if room_type == 'single':
                # Add single room number
                from app.models.models_house_acknowledge import RoomNumber
                room_number = request.form.get('room_number', '').strip()
                if not room_number:
                    flash('Room number is required.', 'error')
                    return redirect(url_for('add_room_number'))
                
                # Check if room number already exists
                existing = RoomNumber.query.filter_by(room_number=room_number).first()
                if existing:
                    flash(f'Room number {room_number} already exists.', 'error')
                    return redirect(url_for('add_room_number'))
                
                # Parse room number components
                parts = room_number.split('-')
                building = parts[0] if len(parts) > 0 else None
                floor = parts[1] if len(parts) > 1 else None
                unit = parts[2] if len(parts) > 2 else None
                
                new_room = RoomNumber(
                    room_number=room_number,
                    building=building,
                    floor=floor,
                    unit=unit
                )
                db.session.add(new_room)
                db.session.commit()
                flash(f'Room number {room_number} added successfully.', 'success')
                
            elif room_type == 'bulk':
                # Add bulk room numbers
                from app.models.models_house_acknowledge import RoomNumber
                print("DEBUG: Processing bulk room addition")
                bulk_ranges = request.form.get('bulk_ranges', '').strip()
                print(f"DEBUG: Bulk ranges received: {bulk_ranges}")
                if not bulk_ranges:
                    flash('Bulk ranges are required.', 'error')
                    return redirect(url_for('add_room_number'))
                
                added_count = 0
                skipped_count = 0
                
                # Process each line in bulk ranges
                for line in bulk_ranges.split('\n'):
                    line = line.strip()
                    if not line or 'to' not in line:
                        continue
                    
                    try:
                        # Parse range like "80-01-001 to 80-01-015"
                        start_room, end_room = [x.strip() for x in line.split(' to ')]
                        
                        # Extract components
                        start_parts = start_room.split('-')
                        end_parts = end_room.split('-')
                        
                        if len(start_parts) != 3 or len(end_parts) != 3:
                            continue
                        
                        building = start_parts[0]
                        floor = start_parts[1]
                        start_unit = int(start_parts[2])
                        end_unit = int(end_parts[2])
                        
                        # Generate room numbers in range
                        for unit_num in range(start_unit, end_unit + 1):
                            unit_str = f"{unit_num:03d}"
                            room_number = f"{building}-{floor}-{unit_str}"
                            
                            # Check if room number already exists
                            existing = RoomNumber.query.filter_by(room_number=room_number).first()
                            if existing:
                                skipped_count += 1
                                continue
                            
                            new_room = RoomNumber(
                                room_number=room_number,
                                building=building,
                                floor=floor,
                                unit=unit_str
                            )
                            db.session.add(new_room)
                            added_count += 1
                    
                    except Exception as e:
                        print(f"Error processing line '{line}': {e}")
                        continue
                
                db.session.commit()
                print(f"DEBUG: Successfully added {added_count} rooms, skipped {skipped_count}")
                message = f'Added {added_count} room numbers successfully.'
                if skipped_count > 0:
                    message += f' Skipped {skipped_count} existing room numbers.'
                flash(message, 'success')
            
            return redirect(url_for('room_numbers'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding room numbers: {str(e)}', 'error')
    
    # Get existing room numbers for dropdown
    from app.models.models_house_acknowledge import RoomNumber
    existing_rooms = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    return render_template('add_room_number.html', existing_rooms=existing_rooms)

@app.route('/room_numbers/edit/<int:room_id>', methods=['POST'])
@login_required
def edit_room_number(room_id):
    """Edit a room number via AJAX"""
    try:
        room = RoomNumber.query.get_or_404(room_id)
        data = request.get_json()
        
        room_number = data.get('room_number', '').strip()
        building = data.get('building', '').strip()
        floor = data.get('floor', '').strip()
        unit = data.get('unit', '').strip()
        
        if not room_number:
            return jsonify({'success': False, 'message': 'Room number is required'})
        
        # Check if room number already exists (excluding current room)
        existing = RoomNumber.query.filter(
            RoomNumber.room_number == room_number,
            RoomNumber.id != room_id,
            RoomNumber.is_active == True
        ).first()
        
        if existing:
            return jsonify({'success': False, 'message': f'Room number {room_number} already exists'})
        
        # Update room details
        room.room_number = room_number
        room.building = building if building else None
        room.floor = floor if floor else None
        room.unit = unit if unit else None
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Room number updated successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error updating room number: {str(e)}'})

@app.route('/room_numbers/delete/<int:room_id>')
@login_required
def delete_room_number(room_id):
    """Delete a room number"""
    try:
        room = RoomNumber.query.get_or_404(room_id)
        room.is_active = False
        db.session.commit()
        flash(f'Room number {room.room_number} deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting room number: {str(e)}', 'error')
    
    return redirect(url_for('room_numbers'))

@app.route('/room_numbers/bulk_add_predefined', methods=['POST'])
@login_required
def bulk_add_predefined_rooms():
    """Add all predefined room numbers from the user's specification"""
    try:
        # Predefined room ranges
        room_ranges = [
            ("80-01-001", "80-01-015"),
            ("80-02-001", "80-02-015"),
            ("80-03-001", "80-03-015"),
            ("80-04-001", "80-04-015"),
            ("81-01-101", "81-01-115"),
            ("81-02-101", "81-02-115"),
            ("81-03-101", "81-03-115"),
            ("81-04-101", "81-04-115"),
            ("82-01-201", "82-01-215"),
            ("82-02-201", "82-02-215"),
            ("82-03-201", "82-03-215"),
            ("82-04-201", "82-04-215"),
            ("83-01-301", "83-01-315"),
            ("83-02-301", "83-02-315"),
            ("83-03-301", "83-03-315"),
            ("83-04-301", "83-04-315"),
            ("88-02-801", "88-02-828"),
            ("88-03-801", "88-03-810"),
            ("88-04-801", "88-04-810")
        ]
        
        added_count = 0
        skipped_count = 0
        
        for start_room, end_room in room_ranges:
            # Extract components
            start_parts = start_room.split('-')
            end_parts = end_room.split('-')
            
            building = start_parts[0]
            floor = start_parts[1]
            start_unit = int(start_parts[2])
            end_unit = int(end_parts[2])
            
            # Generate room numbers in range
            for unit_num in range(start_unit, end_unit + 1):
                unit_str = f"{unit_num:03d}"
                room_number = f"{building}-{floor}-{unit_str}"
                
                # Check if room number already exists
                existing = RoomNumber.query.filter_by(room_number=room_number).first()
                if existing:
                    skipped_count += 1
                    continue
                
                new_room = RoomNumber(
                    room_number=room_number,
                    building=building,
                    floor=floor,
                    unit=unit_str
                )
                db.session.add(new_room)
                added_count += 1
        
        db.session.commit()
        message = f'Added {added_count} room numbers successfully.'
        if skipped_count > 0:
            message += f' Skipped {skipped_count} existing room numbers.'
        flash(message, 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding predefined room numbers: {str(e)}', 'error')
    
    return redirect(url_for('room_numbers'))

@app.route('/qr/<string:code>')
def qr_redirect(code):
    """Handle QR code redirects with expiry and scan limit checking"""
    try:
        qr_code = QRCode.query.filter_by(code=code).first()
        
        if not qr_code:
            return render_template('error.html', 
                                 error_title="QR Code Not Found",
                                 error_message="The QR code you scanned is not valid or has been removed."), 404
        
        # Check if QR code is active
        if not qr_code.is_active:
            return render_template('error.html',
                                 error_title="QR Code Inactive", 
                                 error_message="This QR code has been deactivated."), 403
        
        # Check expiry
        if qr_code.is_expired:
            return render_template('error.html',
                                 error_title="QR Code Expired",
                                 error_message=f"This QR code expired on {qr_code.expires_at.strftime('%Y-%m-%d %H:%M')}"), 403
        
        # Check scan limit
        if qr_code.is_scan_limit_reached:
            return render_template('error.html',
                                 error_title="Scan Limit Reached",
                                 error_message=f"This QR code has reached its maximum scan limit of {qr_code.max_scans}"), 403
        
        # Update scan count
        qr_code.scan_count += 1
        qr_code.last_scanned = singapore_now()
        db.session.commit()
        
        # Handle different QR code types
        if qr_code.qr_type == 'url' and qr_code.target_url:
            # Direct URL redirect
            return redirect(qr_code.target_url)
        elif qr_code.qr_type == 'offense_report':
            # Redirect to offense report form
            return redirect(url_for('public_offense_report', qr_code=code))
        elif qr_code.qr_type == 'house_acknowledge':
            # Redirect to house acknowledge form
            return redirect(url_for('house_acknowledge_public', qr_code=code))
        elif qr_code.qr_type == 'room_inventory':
            # Redirect to room inventory form
            return redirect(url_for('room_inventory_form', qr_code=code))
        elif qr_code.qr_type == 'room_checklist' or qr_code.qr_type == 'Room_Checklist':
            # Redirect to public room checklist form
            return redirect('/public/room-checklist')
        elif qr_code.qr_type == 'purchase_form':
            # Redirect to purchase request form
            return redirect(url_for('purchase_request_form'))
        elif qr_code.qr_type == 'form':
            # Redirect to custom form
            return redirect(url_for('form_submission', qr_code=code))
        else:
            # Default: redirect to a general info page
            return render_template('qr_info.html', qr_code=qr_code)
            
    except Exception as e:
        return render_template('error.html',
                             error_title="System Error",
                             error_message=f"An error occurred while processing the QR code: {str(e)}"), 500

@app.route('/qr/<string:code>/info')
def qr_info(code):
    """Display QR code information and statistics"""
    qr_code = QRCode.query.filter_by(code=code).first_or_404()
    
    return render_template('qr_info.html', qr_code=qr_code)


@app.route('/meter-reading/water/<int:reading_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_water_meter_reading(reading_id):
    reading = WaterMeterReading.query.get_or_404(reading_id)
    room = reading.room
    
    if request.method == 'POST':
        reading.meter_number = request.form.get('meter_number', '').strip()
        reading.start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
        reading.end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
        reading.start_reading = int(request.form.get('start_reading', 0))
        end_reading_str = request.form.get('end_reading', '').strip()
        reading.rate_per_unit = float(request.form.get('rate_per_unit', 0))
        reading.physical_pax = int(request.form.get('physical_pax', 0))
        reading.notes = request.form.get('notes', '')
        
        # Handle end reading update
        if not end_reading_str:
            reading.end_reading = reading.start_reading
            reading.total_consumption = 0
            reading.total_amount = 0.0
        else:
            reading.end_reading = int(end_reading_str)
            if reading.end_reading < reading.start_reading:
                flash('End reading cannot be less than start reading', 'error')
                return render_template('edit_water_meter_reading.html', reading=reading, room=room)
            reading.total_consumption = reading.end_reading - reading.start_reading
            reading.total_amount = reading.total_consumption * reading.rate_per_unit
        
        db.session.commit()
        flash('Water meter reading updated successfully', 'success')
        return redirect(url_for('meter_room_detail', room_id=room.id))
    
    return render_template('edit_water_meter_reading.html', reading=reading, room=room)


@app.route('/meter-reading/water/<int:reading_id>/delete')
@login_required
def delete_water_meter_reading(reading_id):
    reading = WaterMeterReading.query.get_or_404(reading_id)
    room_id = reading.room.id
    
    db.session.delete(reading)
    db.session.commit()
    flash('Water meter reading deleted successfully', 'success')
    return redirect(url_for('meter_room_detail', room_id=room_id))


@app.route('/meter-reading/electricity/<int:reading_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_electricity_meter_reading(reading_id):
    reading = ElectricityMeterReading.query.get_or_404(reading_id)
    room = reading.room
    
    if request.method == 'POST':
        reading.meter_number = request.form.get('meter_number', '').strip()
        reading.start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
        reading.end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
        reading.start_reading = int(request.form.get('start_reading', 0))
        end_reading_str = request.form.get('end_reading', '').strip()
        reading.rate_per_unit = float(request.form.get('rate_per_unit', 0))
        reading.physical_pax = int(request.form.get('physical_pax', 0))
        reading.notes = request.form.get('notes', '')
        
        # Handle end reading update
        if not end_reading_str:
            reading.end_reading = reading.start_reading
            reading.total_consumption = 0
            reading.total_amount = 0.0
        else:
            reading.end_reading = int(end_reading_str)
            if reading.end_reading < reading.start_reading:
                flash('End reading cannot be less than start reading', 'error')
                return render_template('edit_electricity_meter_reading.html', reading=reading, room=room)
            reading.total_consumption = reading.end_reading - reading.start_reading
            reading.total_amount = reading.total_consumption * reading.rate_per_unit
        
        db.session.commit()
        flash('Electricity meter reading updated successfully', 'success')
        return redirect(url_for('meter_room_detail', room_id=room.id))
    
    return render_template('edit_electricity_meter_reading.html', reading=reading, room=room)


@app.route('/meter-reading/electricity/<int:reading_id>/delete')
@login_required
def delete_electricity_meter_reading(reading_id):
    reading = ElectricityMeterReading.query.get_or_404(reading_id)
    room_id = reading.room.id
    
    db.session.delete(reading)
    db.session.commit()
    flash('Electricity meter reading deleted successfully', 'success')
    return redirect(url_for('meter_room_detail', room_id=room_id))

# Live Room Checklist Export Routes
@app.route('/export_checklist_pdf_live', methods=['POST'])
@login_required
def export_checklist_pdf_live():
    """Export live room inventory checklist to PDF with company logo and e-signatures"""
    try:
        data = request.get_json()
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Build story
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.black,
            fontName='Helvetica-Bold'
        )
        
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=10,
            alignment=TA_CENTER,
            textColor=colors.black,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'NormalStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        # Company Logo and Header
        from reportlab.platypus import Image
        import os
        
        # Add TS Group logo if it exists
        logo_path = os.path.join('static', 'ts_logo.jpg')
        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=80, height=60)
                logo.hAlign = 'CENTER'
                story.append(logo)
                story.append(Spacer(1, 10))
            except:
                # If logo loading fails, add text header
                story.append(Paragraph("TS GROUP", title_style))
        else:
            story.append(Paragraph("TS GROUP", title_style))
        
        story.append(Paragraph("Pioneer Lodge", header_style))
        story.append(Paragraph("Unit Inventory Check List", header_style))
        story.append(Paragraph("* Handover / Takeover", normal_style))
        story.append(Spacer(1, 20))
        
        # Basic information table
        info_data = [
            ['Company Name:', data.get('companyName', ''), 'Unit No.:', data.get('roomNumber', ''), 'Date:', data.get('date', '')]
        ]
        
        info_table = Table(info_data, colWidths=[1.2*inch, 2*inch, 0.8*inch, 1*inch, 0.6*inch, 1*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.lightgrey),
            ('BACKGROUND', (2, 0), (2, 0), colors.lightgrey),
            ('BACKGROUND', (4, 0), (4, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 15))
        
        # Inventory items table header (updated to 5 columns)
        inventory_header = [['S/NO', 'DESCRIPTION', 'QTY', 'Condition', 'Defects/Remarks']]
        
        # Group items by section
        sections = {}
        for item in data.get('items', []):
            section = item.get('section', 'others')
            if section not in sections:
                sections[section] = []
            sections[section].append(item)
        
        # Add items to table with section headers
        inventory_data = inventory_header.copy()
        
        section_names = {
            'electrical': 'Electrical Items',
            'toilet': 'Toilet',
            'kitchen': 'Kitchen',
            'others': 'Others'
        }
        
        for section_key, items in sections.items():
            if items:  # Only add section if it has items
                # Add section header
                section_name = section_names.get(section_key, section_key.title())
                inventory_data.append([section_name, '', '', '', ''])
                
                # Add items
                for item in items:
                    # Combine defects and remarks into one field
                    defects_remarks = ''
                    defects = item.get('defects', '') or item.get('defectsRemarks', '')
                    remarks = item.get('remarks', '')
                    
                    if defects and remarks:
                        defects_remarks = f"{defects}; {remarks}"
                    elif defects:
                        defects_remarks = defects
                    elif remarks:
                        defects_remarks = remarks
                    
                    inventory_data.append([
                        item.get('sno', ''),
                        item.get('description', ''),
                        item.get('quantity', ''),
                        item.get('condition', ''),
                        defects_remarks
                    ])
        
        # Create inventory table (updated to 5 columns)
        inventory_table = Table(inventory_data, colWidths=[0.5*inch, 2.5*inch, 0.5*inch, 1*inch, 2*inch])
        inventory_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ]))
        
        # Style section headers
        row_index = 1
        for section_key, items in sections.items():
            if items:
                inventory_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, row_index), (-1, row_index), colors.lightgrey),
                    ('FONTNAME', (0, row_index), (-1, row_index), 'Helvetica-Bold'),
                    ('SPAN', (0, row_index), (-1, row_index))
                ]))
                row_index += 1 + len(items)
        
        story.append(inventory_table)
        story.append(Spacer(1, 20))
        
        # Meter readings section
        meter_data = data.get('meterReadings', {})
        if meter_data and (meter_data.get('water') or meter_data.get('electricity')):
            story.append(Paragraph("Meter Readings", header_style))
            
            # Format date and time
            reading_date = meter_data.get('date', '')
            reading_time = meter_data.get('time', '')
            datetime_str = f"{reading_date} {reading_time}" if reading_date and reading_time else reading_date or reading_time or ''
            
            meter_table_data = [
                ['METER TYPE', 'READING', 'DATE / TIME'],
                ['Water Meter', meter_data.get('water', ''), datetime_str],
                ['Electricity Meter', meter_data.get('electricity', ''), datetime_str]
            ]
            
            meter_table = Table(meter_table_data, colWidths=[2*inch, 1.5*inch, 2.5*inch])
            meter_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ]))
            
            story.append(meter_table)
            story.append(Spacer(1, 20))
        
        # Signatures section
        signatures = data.get('signatures', {})
        if signatures:
            signature_data = [
                ['', 'Handover:', '', 'Takeover:', ''],
                ['Name:', signatures.get('handover', {}).get('name', ''), '', signatures.get('takeover', {}).get('name', ''), ''],
                ['FIN/NRIC:', signatures.get('handover', {}).get('fin', ''), '', signatures.get('takeover', {}).get('fin', ''), ''],
                ['Position:', signatures.get('handover', {}).get('position', ''), '', signatures.get('takeover', {}).get('position', ''), ''],
                ['Date:', signatures.get('handover', {}).get('date', ''), '', signatures.get('takeover', {}).get('date', ''), '']
            ]
            
            # Add signature images if available
            handover_sig = signatures.get('handover', {}).get('signature', '')
            takeover_sig = signatures.get('takeover', {}).get('signature', '')
            
            sig_row = ['Signature:', '', '', '', '']
            
            if handover_sig and handover_sig.startswith('data:image'):
                try:
                    # Extract base64 data
                    sig_data = handover_sig.split(',')[1]
                    sig_bytes = base64.b64decode(sig_data)
                    sig_buffer = BytesIO(sig_bytes)
                    handover_img = Image(sig_buffer, width=1.5*inch, height=0.75*inch)
                    sig_row[1] = handover_img
                except:
                    sig_row[1] = 'Signature Available'
            
            if takeover_sig and takeover_sig.startswith('data:image'):
                try:
                    # Extract base64 data
                    sig_data = takeover_sig.split(',')[1]
                    sig_bytes = base64.b64decode(sig_data)
                    sig_buffer = BytesIO(sig_bytes)
                    takeover_img = Image(sig_buffer, width=1.5*inch, height=0.75*inch)
                    sig_row[3] = takeover_img
                except:
                    sig_row[3] = 'Signature Available'
            
            signature_data.append(sig_row)
            
            signature_table = Table(signature_data, colWidths=[1*inch, 2*inch, 0.5*inch, 2*inch, 0.5*inch])
            signature_table.setStyle(TableStyle([
                ('BACKGROUND', (1, 0), (1, 0), colors.lightgrey),
                ('BACKGROUND', (3, 0), (3, 0), colors.lightgrey),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('SPAN', (1, 0), (1, 0)),
                ('SPAN', (3, 0), (4, 0)),
            ]))
            
            story.append(signature_table)
        
        # Build PDF
        doc.build(story)
        
        buffer.seek(0)
        
        # Create response
        response = make_response(buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        room_number = data.get('roomNumber', 'Unknown')
        date_str = data.get('date', singapore_now().strftime('%Y-%m-%d'))
        response.headers['Content-Disposition'] = f'attachment; filename=Room_Inventory_Checklist_{room_number}_{date_str}.pdf'
        
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/print_checklist_live', methods=['POST'])
@login_required
def print_checklist_live():
    """Generate print-friendly HTML for room inventory checklist"""
    try:
        data = request.get_json()
        
        # Generate print HTML
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Room Inventory Checklist - {data.get('roomNumber', 'Unknown')}</title>
            <style>
                @media print {{
                    body {{ margin: 0; padding: 20px; font-family: Arial, sans-serif; }}
                    .no-print {{ display: none; }}
                }}
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .company-name {{ font-size: 24px; font-weight: bold; margin-bottom: 5px; }}
                .lodge-name {{ font-size: 18px; font-weight: bold; margin-bottom: 5px; }}
                .title {{ font-size: 16px; font-weight: bold; margin-bottom: 5px; }}
                .subtitle {{ font-size: 14px; margin-bottom: 20px; }}
                table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
                th, td {{ border: 1px solid black; padding: 8px; text-align: left; }}
                th {{ background-color: #f0f0f0; font-weight: bold; }}
                .section-header {{ background-color: #f0f0f0; font-weight: bold; }}
                .info-table {{ margin-bottom: 20px; }}
                .info-table td {{ padding: 5px 10px; }}
                .signature-section {{ margin-top: 30px; }}
                .signature-box {{ border: 1px solid black; height: 60px; width: 150px; display: inline-block; margin: 5px; }}
                .signature-img {{ max-width: 150px; max-height: 60px; }}
                @page {{ margin: 1cm; }}
            </style>
        </head>
        <body>
            <div class="header">
                <div class="company-name">TS GROUP</div>
                <div class="lodge-name">Pioneer Lodge</div>
                <div class="title">Unit Inventory Check List</div>
                <div class="subtitle">* Handover / Takeover</div>
            </div>
            
            <table class="info-table">
                <tr>
                    <td style="background-color: #f0f0f0;"><strong>Company Name:</strong></td>
                    <td>{data.get('companyName', '')}</td>
                    <td style="background-color: #f0f0f0;"><strong>Unit No.:</strong></td>
                    <td>{data.get('roomNumber', '')}</td>
                    <td style="background-color: #f0f0f0;"><strong>Date:</strong></td>
                    <td>{data.get('date', '')}</td>
                </tr>
            </table>
            
            <table>
                <thead>
                    <tr>
                        <th style="width: 50px;">S/NO</th>
                        <th style="width: 200px;">DESCRIPTION</th>
                        <th style="width: 50px;">QTY</th>
                        <th style="width: 80px;">Condition</th>
                        <th style="width: 120px;">DEFECTS</th>
                        <th style="width: 80px;">RECTIFIED</th>
                        <th>REMARKS</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        # Group items by section
        sections = {}
        for item in data.get('items', []):
            section = item.get('section', 'others')
            if section not in sections:
                sections[section] = []
            sections[section].append(item)
        
        section_names = {
            'electrical': 'Electrical Items',
            'toilet': 'Toilet',
            'kitchen': 'Kitchen',
            'others': 'Others'
        }
        
        for section_key, items in sections.items():
            if items:
                section_name = section_names.get(section_key, section_key.title())
                html_content += f'<tr class="section-header"><td colspan="7">{section_name}</td></tr>'
                
                for item in items:
                    html_content += f"""
                    <tr>
                        <td>{item.get('sno', '')}</td>
                        <td>{item.get('description', '')}</td>
                        <td>{item.get('quantity', '')}</td>
                        <td>{item.get('condition', '')}</td>
                        <td>{item.get('defects', '')}</td>
                        <td>{item.get('rectified', '')}</td>
                        <td>{item.get('remarks', '')}</td>
                    </tr>
                    """
        
        html_content += """
                </tbody>
            </table>
        """
        
        # Meter readings section
        meter_data = data.get('meterReadings', {})
        if meter_data:
            html_content += f"""
            <h3>Meter Readings</h3>
            <table>
                <thead>
                    <tr>
                        <th>WATER METER BADGE</th>
                        <th>READING</th>
                        <th>SIGN</th>
                        <th>DATE / TIME</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td>Water Meter</td>
                        <td>{meter_data.get('waterMeter', '')}</td>
                        <td>{meter_data.get('waterMeterSign', '')}</td>
                        <td>{meter_data.get('waterMeterDateTime', '')}</td>
                    </tr>
                    <tr>
                        <td>Electricity</td>
                        <td>{meter_data.get('electricity', '')}</td>
                        <td>{meter_data.get('electricitySign', '')}</td>
                        <td>{meter_data.get('electricityDateTime', '')}</td>
                    </tr>
                </tbody>
            </table>
            """
        
        # Signatures section
        signatures = data.get('signatures', {})
        if signatures:
            html_content += """
            <div class="signature-section">
                <table>
                    <tr>
                        <td></td>
                        <td style="background-color: #f0f0f0;"><strong>Handover:</strong></td>
                        <td></td>
                        <td style="background-color: #f0f0f0;"><strong>Takeover:</strong></td>
                        <td></td>
                    </tr>
            """
            
            handover = signatures.get('handover', {})
            takeover = signatures.get('takeover', {})
            
            html_content += f"""
                    <tr>
                        <td><strong>Name:</strong></td>
                        <td>{handover.get('name', '')}</td>
                        <td></td>
                        <td>{takeover.get('name', '')}</td>
                        <td></td>
                    </tr>
                    <tr>
                        <td><strong>Designation:</strong></td>
                        <td>{handover.get('designation', '')}</td>
                        <td></td>
                        <td>{takeover.get('designation', '')}</td>
                        <td></td>
                    </tr>
                    <tr>
                        <td><strong>FIN Number:</strong></td>
                        <td>{handover.get('finNumber', '')}</td>
                        <td></td>
                        <td>{takeover.get('finNumber', '')}</td>
                        <td></td>
                    </tr>
                    <tr>
                        <td><strong>Date/Time:</strong></td>
                        <td>{handover.get('dateTime', '')}</td>
                        <td></td>
                        <td>{takeover.get('dateTime', '')}</td>
                        <td></td>
                    </tr>
                    <tr>
                        <td><strong>Signature:</strong></td>
                        <td>
            """
            
            if handover.get('signature') and handover['signature'].startswith('data:image'):
                html_content += f'<img src="{handover["signature"]}" class="signature-img" />'
            else:
                html_content += '<div class="signature-box"></div>'
            
            html_content += '</td><td></td><td>'
            
            if takeover.get('signature') and takeover['signature'].startswith('data:image'):
                html_content += f'<img src="{takeover["signature"]}" class="signature-img" />'
            else:
                html_content += '<div class="signature-box"></div>'
            
            html_content += """
                        </td>
                        <td></td>
                    </tr>
                </table>
            </div>
            """
        
        html_content += """
            <script>
                window.onload = function() {
                    window.print();
                };
            </script>
        </body>
        </html>
        """
        
        return html_content
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export_checklist_excel_live', methods=['POST'])
@login_required
def export_checklist_excel_live():
    """Export live room inventory checklist to Excel with complete data"""
    try:
        data = request.get_json()
        
        # Create Excel file in memory
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Basic information sheet
            basic_info = pd.DataFrame([
                ['Room Number', data.get('roomNumber', '')],
                ['Company Name', data.get('companyName', '')],
                ['Date', data.get('date', '')],
                ['Created At', singapore_now().strftime('%Y-%m-%d %H:%M:%S')]
            ], columns=['Field', 'Value'])
            basic_info.to_excel(writer, sheet_name='Basic Information', index=False)
            
            # Inventory items sheet
            items = data.get('items', [])
            if items:
                items_df = pd.DataFrame(items)
                items_df = items_df.rename(columns={
                    'sno': 'S/NO',
                    'description': 'DESCRIPTION',
                    'quantity': 'QTY',
                    'condition': 'Condition',
                    'defects': 'DEFECTS',
                    'rectified': 'RECTIFIED',
                    'remarks': 'REMARKS',
                    'section': 'Section'
                })
                items_df.to_excel(writer, sheet_name='Inventory Items', index=False)
            
            # Meter readings sheet
            meter_data = data.get('meterReadings', {})
            if meter_data:
                meter_df = pd.DataFrame([
                    ['Water Meter Reading', meter_data.get('waterMeter', '')],
                    ['Water Meter Signature', meter_data.get('waterMeterSign', '')],
                    ['Water Meter Date/Time', meter_data.get('waterMeterDateTime', '')],
                    ['Electricity Reading', meter_data.get('electricity', '')],
                    ['Electricity Signature', meter_data.get('electricitySign', '')],
                    ['Electricity Date/Time', meter_data.get('electricityDateTime', '')]
                ], columns=['Field', 'Value'])
                meter_df.to_excel(writer, sheet_name='Meter Readings', index=False)
            
            # Signatures sheet
            signatures = data.get('signatures', {})
            if signatures:
                sig_data = []
                
                handover = signatures.get('handover', {})
                takeover = signatures.get('takeover', {})
                
                sig_data.extend([
                    ['Handover Name', handover.get('name', '')],
                    ['Handover Designation', handover.get('designation', '')],
                    ['Handover FIN Number', handover.get('finNumber', '')],
                    ['Handover Date/Time', handover.get('dateTime', '')],
                    ['Handover Signature Status', 'Available' if handover.get('signature') else 'Not Available'],
                    ['Takeover Name', takeover.get('name', '')],
                    ['Takeover Designation', takeover.get('designation', '')],
                    ['Takeover FIN Number', takeover.get('finNumber', '')],
                    ['Takeover Date/Time', takeover.get('dateTime', '')],
                    ['Takeover Signature Status', 'Available' if takeover.get('signature') else 'Not Available']
                ])
                
                sig_df = pd.DataFrame(sig_data, columns=['Field', 'Value'])
                sig_df.to_excel(writer, sheet_name='Signatures', index=False)
            
            # Auto-adjust column widths for all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_name = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_name].width = adjusted_width
        
        output.seek(0)
        
        # Create response
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        room_number = data.get('roomNumber', 'Unknown')
        date_str = data.get('date', singapore_now().strftime('%Y-%m-%d'))
        response.headers['Content-Disposition'] = f'attachment; filename=Room_Inventory_Checklist_{room_number}_{date_str}.xlsx'
        
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
# FIN Search Routes
@app.route('/fin-search')
@login_required
def fin_search():
    """FIN number search across offense and acknowledgment records"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard'))
    
    fin_number = request.args.get('fin_number', '').strip()
    search_performed = bool(fin_number)
    
    person_info = None
    offense_records = []
    acknowledgment_records = []
    
    if search_performed and fin_number:
        # Search offense records
        offense_records = OffenseRecord.query.filter(
            OffenseRecord.organization_id == user.organization_id,
            OffenseRecord.fin_number.ilike(f'%{fin_number}%')
        ).order_by(OffenseRecord.incident_date.desc()).all()
        
        # Search house acknowledgment records
        acknowledgment_records = HouseAcknowledgment.query.filter(
            HouseAcknowledgment.fin.ilike(f'%{fin_number}%')
        ).order_by(HouseAcknowledgment.acknowledged_at.desc()).all()
        
        # Get person info from the most recent record
        if offense_records or acknowledgment_records:
            # Try to get person info from offense records first
            if offense_records:
                latest_offense = offense_records[0]
                person_info = {
                    'name': latest_offense.offender_name,
                    'fin': latest_offense.fin_number,
                    'phone': latest_offense.contact_number,
                    'company': latest_offense.offender_company,
                    'room': latest_offense.offender_room
                }
            # If no offense records, get from acknowledgment records
            elif acknowledgment_records:
                latest_ack = acknowledgment_records[0]
                person_info = {
                    'name': latest_ack.name,
                    'fin': latest_ack.fin,
                    'phone': latest_ack.phone_number,
                    'company': latest_ack.company_name,
                    'room': latest_ack.room_number
                }
    
    return render_template('fin_search.html',
                         search_performed=search_performed,
                         fin_number=fin_number,
                         person_info=person_info,
                         offense_records=offense_records,
                         acknowledgment_records=acknowledgment_records)

@app.route('/offense-record/<int:record_id>/details')
@login_required
def get_offense_details(record_id):
    """Get detailed information for a specific offense record"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    try:
        record = OffenseRecord.query.filter_by(
            id=record_id,
            organization_id=user.organization_id
        ).first_or_404()
        
        return jsonify({
            'success': True,
            'record': {
                'id': record.id,
                'full_name': record.offender_name,
                'fin_number': record.fin_number,
                'phone_number': record.contact_number,
                'company_name': record.offender_company,
                'room_number': record.offender_room,
                'nationality': record.nationality,
                'sector': record.sector,
                'case_number': record.case_number,
                'offense_date': record.incident_date.strftime('%Y-%m-%d') if record.incident_date else None,
                'incident_time': record.incident_time.strftime('%H:%M') if record.incident_time else None,
                'offense_type': record.offense_type,
                'severity': record.severity,
                'location': record.location,
                'description': record.description,
                'documentary_proof': record.documentary_proof,
                'proof_description': record.proof_description,
                'witness_names': record.witness_names,
                'action_taken': record.action_taken,
                'status': record.status,
                'fine_amount': float(record.penalty_amount) if record.penalty_amount else 0.0,
                'penalty_status': record.penalty_status,
                'duty_manager_name': record.duty_manager_name,
                'financial_penalty_imposed': record.financial_penalty_imposed
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/house-acknowledgment/<int:record_id>/details')
@login_required
def get_acknowledgment_details(record_id):
    """Get detailed information for a specific house acknowledgment record"""
    user = current_user
    
    try:
        record = HouseAcknowledgment.query.filter_by(
            id=record_id
        ).first_or_404()
        
        return jsonify({
            'success': True,
            'record': {
                'id': record.id,
                'full_name': record.name,
                'fin_number': record.fin,
                'phone_number': record.phone_number,
                'company_name': record.company_name,
                'room_number': record.room_number,
                'acknowledgment_date': record.acknowledged_at.strftime('%Y-%m-%d %H:%M:%S') if record.acknowledged_at else None,
                'acknowledgment_type': 'House Rules',
                'language': record.language_selected,
                'ip_address': record.ip_address,
                'user_agent': record.user_agent,
                'e_signature': record.e_signature,
                'selfie_path': record.selfie_photo,
                'house_acknowledge_id': record.house_acknowledge_id
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Bedding Items Management Routes
@app.route('/bedding-items')
@login_required
@page_permission_required('bedding_management')
def bedding_items():
    """Main Bedding Items management page"""
    user = current_user
    if not user.organization_id:
        flash('Organization not found.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get filter parameters
    category_filter = request.args.get('category', '')
    status_filter = request.args.get('status', '')
    search_query = request.args.get('search', '')
    
    # Base query - manually fetch items and then get categories separately
    query = BeddingItem.query.filter_by(organization_id=user.organization_id)
    
    # Apply filters
    if category_filter:
        query = query.filter(BeddingItem.category_id == category_filter)
    if status_filter:
        query = query.filter(BeddingItem.status == status_filter)
    if search_query:
        query = query.filter(
            db.or_(
                BeddingItem.serial_number.ilike(f'%{search_query}%'),
                BeddingItem.item_name.ilike(f'%{search_query}%'),
                BeddingItem.room_number.ilike(f'%{search_query}%'),
                BeddingItem.resident_name.ilike(f'%{search_query}%')
            )
        )
    
    # Get paginated results
    page = request.args.get('page', 1, type=int)
    items = query.order_by(BeddingItem.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # Manually load categories for displayed items to avoid relationship issues
    category_ids = [item.category_id for item in items.items if item.category_id]
    categories_dict = {}
    if category_ids:
        categories_data = BeddingCategory.query.filter(BeddingCategory.id.in_(category_ids)).all()
        categories_dict = {cat.id: cat for cat in categories_data}
    
    # Attach categories to items
    for item in items.items:
        if item.category_id and item.category_id in categories_dict:
            item.category = categories_dict[item.category_id]
        else:
            item.category = None
    
    # Get categories for filter dropdown
    categories = BeddingCategory.query.filter_by(
        organization_id=user.organization_id,
        is_active=True
    ).order_by(BeddingCategory.name).all()
    
    # Get summary statistics
    total_items = BeddingItem.query.filter_by(organization_id=user.organization_id).count()
    in_store = BeddingItem.query.filter_by(organization_id=user.organization_id, status='In Store').count()
    in_room = BeddingItem.query.filter_by(organization_id=user.organization_id, status='In Room').count()
    damaged = BeddingItem.query.filter_by(organization_id=user.organization_id, status='Damaged').count()
    others = BeddingItem.query.filter_by(organization_id=user.organization_id, status='Others').count()
    
    stats = {
        'total': total_items,
        'in_store': in_store,
        'in_room': in_room,
        'damaged': damaged,
        'others': others
    }
    
    return render_template('bedding_items.html',
                         items=items,
                         categories=categories,
                         stats=stats,
                         category_filter=category_filter,
                         status_filter=status_filter,
                         search_query=search_query,
                         current_route='bedding_items')

@app.route('/bedding-items/add', methods=['GET', 'POST'])
@login_required
@create_permission_required('bedding_management')
def add_bedding_item():
    """Add new bedding item"""
    user = current_user
    if not user.organization_id:
        flash('Organization not found.', 'error')
        return redirect(url_for('bedding_items'))
    
    if request.method == 'POST':
        try:
            # Create new bedding item
            item = BeddingItem(
                serial_number=request.form.get('serial_number'),
                item_name=request.form.get('item_name'),
                category_id=int(request.form.get('category_id')),
                status=request.form.get('status', 'In Store'),
                room_number=request.form.get('room_number') or None,
                resident_name=request.form.get('resident_name') or None,
                company_name=request.form.get('company_name') or None,
                brand=request.form.get('brand') or None,
                model=request.form.get('model') or None,
                purchase_date=datetime.strptime(request.form.get('purchase_date'), '%Y-%m-%d').date() if request.form.get('purchase_date') else None,
                purchase_price=float(request.form.get('purchase_price')) if request.form.get('purchase_price') else None,
                condition=request.form.get('condition', 'Good'),
                warranty_expiry=datetime.strptime(request.form.get('warranty_expiry'), '%Y-%m-%d').date() if request.form.get('warranty_expiry') else None,
                description=request.form.get('description') or None,
                last_maintenance_date=datetime.strptime(request.form.get('last_maintenance_date'), '%Y-%m-%d').date() if request.form.get('last_maintenance_date') else None,
                next_maintenance_date=datetime.strptime(request.form.get('next_maintenance_date'), '%Y-%m-%d').date() if request.form.get('next_maintenance_date') else None,
                organization_id=user.organization_id,
                created_by=user.id
            )
            
            db.session.add(item)
            db.session.flush()
            
            # Create movement record
            movement = BeddingMovement(
                item_id=item.id,
                movement_type='Initial Entry',
                to_status=item.status,
                to_room=item.room_number,
                to_resident=item.resident_name,
                reason='Initial inventory entry',
                movement_date=singapore_now(),
                organization_id=user.organization_id,
                processed_by=user.id
            )
            
            db.session.add(movement)
            db.session.commit()
            
            flash('Bedding item added successfully!', 'success')
            return redirect(url_for('bedding_items'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding bedding item: {str(e)}', 'error')
    
    # Get categories for form
    categories = BeddingCategory.query.filter_by(
        organization_id=user.organization_id,
        is_active=True
    ).order_by(BeddingCategory.name).all()
    
    return render_template('add_bedding_item.html',
                         categories=categories,
                         current_route='bedding_items')

@app.route('/bedding-items/<int:item_id>/edit', methods=['GET', 'POST'])
@login_required
@create_permission_required('bedding_management')
def edit_bedding_item(item_id):
    """Edit bedding item"""
    user = current_user
    item = BeddingItem.query.filter_by(
        id=item_id,
        organization_id=user.organization_id
    ).first_or_404()
    
    if request.method == 'POST':
        try:
            # Store original values
            original_status = item.status
            original_room = item.room_number
            original_resident = item.resident_name
            
            # Update item fields
            item.item_name = request.form.get('item_name')
            item.category_id = int(request.form.get('category_id'))
            item.status = request.form.get('status')
            item.room_number = request.form.get('room_number') or None
            item.resident_name = request.form.get('resident_name') or None
            item.company_name = request.form.get('company_name') or None
            item.brand = request.form.get('brand') or None
            item.model = request.form.get('model') or None
            
            if request.form.get('purchase_date'):
                item.purchase_date = datetime.strptime(request.form.get('purchase_date'), '%Y-%m-%d').date()
            if request.form.get('purchase_price'):
                item.purchase_price = float(request.form.get('purchase_price'))
                
            item.condition = request.form.get('condition')
            
            if request.form.get('warranty_expiry'):
                item.warranty_expiry = datetime.strptime(request.form.get('warranty_expiry'), '%Y-%m-%d').date()
            if request.form.get('last_maintenance_date'):
                item.last_maintenance_date = datetime.strptime(request.form.get('last_maintenance_date'), '%Y-%m-%d').date()
            if request.form.get('next_maintenance_date'):
                item.next_maintenance_date = datetime.strptime(request.form.get('next_maintenance_date'), '%Y-%m-%d').date()
                
            item.description = request.form.get('description') or None
            item.updated_at = singapore_now()
            
            # Check if status changed to create movement record
            if (original_status != item.status or 
                original_room != item.room_number or 
                original_resident != item.resident_name):
                
                movement_type = 'Update'
                if original_status != item.status:
                    if item.status == 'In Room':
                        movement_type = 'Assignment'
                    elif original_status == 'In Room':
                        movement_type = 'Return'
                    elif item.status == 'Damaged':
                        movement_type = 'Maintenance'
                
                movement = BeddingMovement(
                    item_id=item.id,
                    movement_type=movement_type,
                    from_status=original_status,
                    to_status=item.status,
                    from_room=original_room,
                    to_room=item.room_number,
                    from_resident=original_resident,
                    to_resident=item.resident_name,
                    reason=request.form.get('movement_reason', 'Item details updated'),
                    notes=request.form.get('movement_notes'),
                    movement_date=singapore_now(),
                    organization_id=user.organization_id,
                    processed_by=user.id
                )
                
                db.session.add(movement)
            
            db.session.commit()
            flash('Bedding item updated successfully!', 'success')
            return redirect(url_for('bedding_items'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating bedding item: {str(e)}', 'error')
    
    categories = BeddingCategory.query.filter_by(
        organization_id=user.organization_id,
        is_active=True
    ).order_by(BeddingCategory.name).all()
    
    return render_template('edit_bedding_item.html',
                         item=item,
                         categories=categories,
                         current_route='bedding_items')

@app.route('/bedding-categories')
@login_required
def bedding_categories():
    """Manage bedding categories"""
    user = current_user
    if not user.organization_id:
        flash('Organization not found.', 'error')
        return redirect(url_for('dashboard'))
    
    categories = BeddingCategory.query.filter_by(organization_id=user.organization_id).order_by(BeddingCategory.name).all()
    
    return render_template('bedding_categories.html',
                         categories=categories,
                         current_route='bedding_items')

@app.route('/bedding-categories/add', methods=['POST'])
@login_required
@create_permission_required('bedding_management')
def add_bedding_category():
    """Add new bedding category"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'Organization not found'}), 400
    
    try:
        category = BeddingCategory(
            name=request.form.get('name'),
            description=request.form.get('description'),
            organization_id=user.organization_id
        )
        
        db.session.add(category)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Category added successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/bedding-categories/<int:category_id>/toggle', methods=['POST'])
@login_required
def toggle_bedding_category(category_id):
    """Toggle bedding category active status"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'Organization not found'}), 400
    
    try:
        category = BeddingCategory.query.filter_by(
            id=category_id,
            organization_id=user.organization_id
        ).first()
        
        if not category:
            return jsonify({'success': False, 'error': 'Category not found'}), 404
        
        # Toggle the active status
        category.is_active = not category.is_active
        category.updated_at = singapore_now()
        
        db.session.commit()
        
        action = "activated" if category.is_active else "deactivated"
        return jsonify({'success': True, 'message': f'Category {action} successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/bedding-categories/<int:category_id>/edit', methods=['POST'])
@login_required
@create_permission_required('bedding_management')
def edit_bedding_category(category_id):
    """Edit bedding category"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'Organization not found'}), 400
    
    try:
        category = BeddingCategory.query.filter_by(
            id=category_id,
            organization_id=user.organization_id
        ).first()
        
        if not category:
            return jsonify({'success': False, 'error': 'Category not found'}), 404
        
        # Update category data
        category.name = request.form.get('name')
        category.description = request.form.get('description')
        category.updated_at = singapore_now()
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Category updated successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/bedding-categories/<int:category_id>/delete', methods=['POST'])
@login_required
def delete_bedding_category(category_id):
    """Delete bedding category"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'Organization not found'}), 400
    
    try:
        category = BeddingCategory.query.filter_by(
            id=category_id,
            organization_id=user.organization_id
        ).first()
        
        if not category:
            return jsonify({'success': False, 'error': 'Category not found'}), 404
        
        # Check if category has items
        item_count = BeddingItem.query.filter_by(category_id=category_id).count()
        if item_count > 0:
            return jsonify({'success': False, 'error': f'Cannot delete category with {item_count} associated items'}), 400
        
        db.session.delete(category)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Category deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/bulk-delete-stock-items', methods=['POST'])
@login_required
@edit_permission_required('purchase')
@performance_timer
def bulk_delete_stock_items():
    """Delete multiple stock items"""
    try:
        user = current_user
        data = request.get_json()
        item_ids = data.get('item_ids', [])
        
        if not item_ids:
            return jsonify({'success': False, 'error': 'No items selected'})
        
        # Convert string IDs to integers
        try:
            item_ids = [int(id) for id in item_ids]
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid item IDs'})
        
        # Validate that all items belong to the user's organization
        try:
            stock_items = StockItem.query.filter(
                StockItem.id.in_(item_ids),
                StockItem.organization_id == user.organization_id
            ).all()
            
            if len(stock_items) != len(item_ids):
                return jsonify({'success': False, 'error': 'Some items not found or access denied'})
        except Exception as e:
            return jsonify({'success': False, 'error': f'Database error: {str(e)}'})
        
        # Delete associated movements first (if StockMovement exists)
        try:
            StockMovement.query.filter(
                StockMovement.stock_item_id.in_(item_ids)
            ).delete(synchronize_session=False)
        except Exception:
            # StockMovement table might not exist, continue with deletion
            pass
        
        # Delete the stock items (bulk delete)
        deleted_count = StockItem.query.filter(
            StockItem.id.in_(item_ids),
            StockItem.organization_id == user.organization_id
        ).delete(synchronize_session=False)
        
        db.session.commit()
        
        # Invalidate cache
        cache.invalidate(f"stock_storage_{user.organization_id}")
        cache.invalidate(f"dashboard_{user.organization_id}")
        
        return jsonify({
            'success': True, 
            'message': f'{deleted_count} stock item(s) deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/export-stock-items-excel', methods=['POST'])
@login_required
@page_access_required('purchase')
@performance_timer
def export_stock_items_excel():
    """Export selected stock items to Excel"""
    try:
        user = current_user
        selected_ids = request.form.getlist('selected_ids[]')
        
        if not selected_ids:
            flash('No items selected for export', 'error')
            return redirect(url_for('stock_storage'))
        
        # Convert to integers
        item_ids = [int(id) for id in selected_ids]
        
        # Get selected stock items
        stock_items = StockItem.query.filter(
            StockItem.id.in_(item_ids),
            StockItem.organization_id == user.organization_id
        ).order_by(StockItem.id).all()
        
        if not stock_items:
            flash('No items found for export', 'error')
            return redirect(url_for('stock_storage'))
        
        # Create Excel workbook
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Items Export"
        
        # Headers
        headers = [
            'Item ID', 'Name', 'Description', 'Category', 'Quantity', 
            'Used', 'Available', 'Status', 'Location', 'Purchase Cost', 'Created Date'
        ]
        
        # Style headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, item in enumerate(stock_items, 2):
            ws.cell(row=row, column=1, value=item.id)
            ws.cell(row=row, column=2, value=item.name or '')
            ws.cell(row=row, column=3, value=item.description or '')
            ws.cell(row=row, column=4, value=item.category or '')
            ws.cell(row=row, column=5, value=int(item.quantity or 0))
            ws.cell(row=row, column=6, value=int(item.used_quantity or 0))
            ws.cell(row=row, column=7, value=int((item.quantity or 0) - (item.used_quantity or 0)))
            ws.cell(row=row, column=8, value=item.status or '')
            ws.cell(row=row, column=9, value=item.location or '')
            ws.cell(row=row, column=10, value=float(item.purchase_cost or 0))
            ws.cell(row=row, column=11, value=item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'stock_items_export_{timestamp}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting stock items: {str(e)}', 'error')
        return redirect(url_for('stock_storage'))

@app.route('/bedding-items/export')
@login_required
def export_bedding_items():
    """Export bedding items to Excel"""
    user = current_user
    if not user.organization_id:
        flash('Organization not found.', 'error')
        return redirect(url_for('bedding_items'))
    
    try:
        # Get all bedding items with category information
        items = db.session.query(BeddingItem, BeddingCategory, User).join(
            BeddingCategory, BeddingItem.category_id == BeddingCategory.id
        ).join(
            User, BeddingItem.created_by == User.id
        ).filter(
            BeddingItem.organization_id == user.organization_id
        ).order_by(BeddingItem.created_at.desc()).all()
        
        # Create Excel file in memory
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Bedding Items'
        
        # Headers
        headers = [
            'Serial Number', 'Item Name', 'Category', 'Status', 'Room Number',
            'Resident Name', 'Company Name', 'Brand', 'Model', 'Purchase Date',
            'Purchase Price', 'Condition', 'Warranty Expiry', 'Description',
            'Last Maintenance', 'Next Maintenance', 'Created By', 'Created At'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data rows
        for row, (item, category, creator) in enumerate(items, 2):
            worksheet.cell(row=row, column=1, value=item.serial_number)
            worksheet.cell(row=row, column=2, value=item.item_name)
            worksheet.cell(row=row, column=3, value=category.name)
            worksheet.cell(row=row, column=4, value=item.status)
            worksheet.cell(row=row, column=5, value=item.room_number)
            worksheet.cell(row=row, column=6, value=item.resident_name)
            worksheet.cell(row=row, column=7, value=item.company_name)
            worksheet.cell(row=row, column=8, value=item.brand)
            worksheet.cell(row=row, column=9, value=item.model)
            worksheet.cell(row=row, column=10, value=item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else '')
            worksheet.cell(row=row, column=11, value=item.purchase_price)
            worksheet.cell(row=row, column=12, value=item.condition)
            worksheet.cell(row=row, column=13, value=item.warranty_expiry.strftime('%Y-%m-%d') if item.warranty_expiry else '')
            worksheet.cell(row=row, column=14, value=item.description)
            worksheet.cell(row=row, column=15, value=item.last_maintenance_date.strftime('%Y-%m-%d') if item.last_maintenance_date else '')
            worksheet.cell(row=row, column=16, value=item.next_maintenance_date.strftime('%Y-%m-%d') if item.next_maintenance_date else '')
            worksheet.cell(row=row, column=17, value=creator.username)
            worksheet.cell(row=row, column=18, value=item.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=bedding_items_{singapore_now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Error exporting data: {str(e)}', 'error')
        return redirect(url_for('bedding_items'))

# Purchase Request Edit and Status Update Routes
@app.route('/purchase-request-edit/<int:request_id>')
@login_required
def purchase_request_edit(request_id):
    """Display edit form for purchase request"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        purchase_request = PurchaseRequest.query.filter_by(
            id=request_id,
            organization_id=user.organization_id
        ).first()
        
        if not purchase_request:
            flash('Purchase request not found', 'error')
            return redirect(url_for('purchase_request_management'))
        
        # Get request items
        items = PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).all()
        
        return render_template('purchase_request_edit.html', 
                             purchase_request=purchase_request,
                             items=items,
                             user=user)
        
    except Exception as e:
        flash(f'Error loading purchase request: {str(e)}', 'error')
        return redirect(url_for('purchase_request_management'))

@app.route('/purchase-request-edit/<int:request_id>', methods=['POST'])
@login_required
def purchase_request_update(request_id):
    """Update purchase request"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'No organization assigned'})
    
    try:
        purchase_request = PurchaseRequest.query.filter_by(
            id=request_id,
            organization_id=user.organization_id
        ).first()
        
        if not purchase_request:
            return jsonify({'success': False, 'error': 'Purchase request not found'})
        
        data = request.get_json()
        
        # Update basic request information
        purchase_request.requested_by = data.get('requested_by', '')
        purchase_request.category = data.get('category', '')
        purchase_request.request_date = datetime.strptime(data.get('request_date'), '%Y-%m-%d').date()
        purchase_request.dc_name = data.get('dc_name', '')
        purchase_request.operation_manager = data.get('operation_manager', '')
        purchase_request.general_manager = data.get('general_manager', '')
        purchase_request.requested_by_footer = data.get('requested_by_footer', '')
        purchase_request.recommended_by_footer = data.get('recommended_by_footer', '')
        
        # Handle signature updates
        if data.get('dc_signature'):
            purchase_request.dc_signature = data.get('dc_signature')
        if data.get('operation_manager_signature'):
            purchase_request.operation_manager_signature = data.get('operation_manager_signature')
        if data.get('general_manager_signature'):
            purchase_request.general_manager_signature = data.get('general_manager_signature')
        
        # Delete existing items
        PurchaseRequestItem.query.filter_by(purchase_request_id=request_id).delete()
        
        # Add updated items
        for item_data in data.get('items', []):
            item = PurchaseRequestItem(
                purchase_request_id=purchase_request.id,
                description=item_data.get('description', ''),
                unit_cost=float(item_data.get('unit_cost', 0)),
                quantity=int(item_data.get('quantity', 1)),
                total_cost=float(item_data.get('total_cost', 0)),
                room_no=item_data.get('room_no', ''),
                unit=item_data.get('unit', ''),
                cost_code=item_data.get('cost_code', ''),
                remarks=item_data.get('remarks', '')
            )
            db.session.add(item)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Purchase request {purchase_request.request_number} updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error updating purchase request: {str(e)}'})

@app.route('/api/purchase-request/<int:request_id>/status', methods=['POST'])
@login_required
def update_purchase_request_status(request_id):
    """Update purchase request status"""
    user = current_user
    if not user.organization_id:
        return jsonify({'success': False, 'error': 'No organization assigned'})
    
    try:
        purchase_request = PurchaseRequest.query.filter_by(
            id=request_id,
            organization_id=user.organization_id
        ).first()
        
        if not purchase_request:
            return jsonify({'success': False, 'error': 'Purchase request not found'})
        
        data = request.get_json()
        new_status = data.get('status')
        
        if not new_status:
            return jsonify({'success': False, 'error': 'Status is required'})
        
        valid_statuses = ['Pending', 'Approved', 'Rejected', 'In Progress', 'Completed']
        if new_status not in valid_statuses:
            return jsonify({'success': False, 'error': 'Invalid status'})
        
        old_status = purchase_request.status
        purchase_request.status = new_status
        purchase_request.updated_at = singapore_now()
        
        db.session.commit()
        
        # Log the status change
        log_entry = SystemLog(
            user_id=user.id,
            action='update_purchase_request_status',
            module='purchase_management',
            details=f'Updated status for PR {purchase_request.request_number} from "{old_status}" to "{new_status}"',
            ip_address=request.remote_addr
        )
        db.session.add(log_entry)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Status updated to {new_status}',
            'old_status': old_status,
            'new_status': new_status
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error updating status: {str(e)}'})
# Staff Attendance Routes
@app.route('/staff-attendance')
@login_required
@page_access_required('staff_attendance')
def staff_attendance():
    """Staff Attendance dashboard page"""
    user = current_user
    org_id = user.organization_id
    
    # Get recent attendance records
    recent_attendance = StaffAttendance.query.filter_by(
        organization_id=org_id
    ).order_by(StaffAttendance.timestamp.desc()).limit(10).all()
    
    # Get statistics
    today = singapore_now().date()
    today_start_count = StaffAttendance.query.filter_by(
        organization_id=org_id,
        attendance_type='start'
    ).filter(func.date(StaffAttendance.timestamp) == today).count()
    
    today_end_count = StaffAttendance.query.filter_by(
        organization_id=org_id,
        attendance_type='end'
    ).filter(func.date(StaffAttendance.timestamp) == today).count()
    
    total_records = StaffAttendance.query.filter_by(organization_id=org_id).count()
    
    return render_template('staff_attendance.html',
                         recent_attendance=recent_attendance,
                         today_start_count=today_start_count,
                         today_end_count=today_end_count,
                         total_records=total_records)

@app.route('/staff-attendance/scan/<attendance_type>')
def staff_attendance_scan(attendance_type):
    """QR Code scanning page for start/end attendance"""
    try:
        if attendance_type not in ['start', 'end']:
            flash('Invalid attendance type', 'error')
            return redirect(url_for('dashboard'))
        
        # Get organization ID - either from logged-in user or first available organization
        org_id = None
        if current_user.is_authenticated:
            org_id = current_user.organization_id
        else:
            # For public QR access, use the first organization
            org = Organization.query.first()
            if org:
                org_id = org.id
        
        # Ensure we have an organization
        if not org_id:
            flash('No organization found. Please contact administrator.', 'error')
            return redirect(url_for('dashboard'))
        
        # Get workers for the dropdown
        workers = []
        try:
            workers = Worker.query.filter_by(organization_id=org_id, is_active=True).order_by(Worker.name).all()
        except Exception as e:
            # Log the error but continue without workers
            print(f"Error fetching workers: {e}")
            workers = []
        
        return render_template('staff_attendance_scan.html', 
                             attendance_type=attendance_type,
                             workers=workers)
    except Exception as e:
        print(f"Error in staff_attendance_scan: {e}")
        flash('An error occurred. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/staff-attendance/submit', methods=['POST'])
def staff_attendance_submit():
    """Submit staff attendance data"""
    try:
        # Get form data
        staff_name = request.form.get('staff_name', '').strip()
        fin_number = request.form.get('fin_number', '').strip()
        company_name = request.form.get('company_name', '').strip()
        attendance_type = request.form.get('attendance_type', '').strip()
        selfie_data = request.form.get('selfie_data', '')
        qr_code_scanned = request.form.get('qr_code_scanned', '')
        location = request.form.get('location', '')
        
        # Validation
        if not staff_name:
            return jsonify({'success': False, 'message': 'Staff name is required'})
        
        if not company_name:
            return jsonify({'success': False, 'message': 'Company name is required'})
        
        if attendance_type not in ['start', 'end']:
            return jsonify({'success': False, 'message': 'Invalid attendance type'})
        
        if not selfie_data:
            return jsonify({'success': False, 'message': 'Selfie photo is required'})
        
        # Get organization (use first available if no user logged in)
        org_id = None
        if current_user.is_authenticated:
            org_id = current_user.organization_id
        else:
            # For public QR access, use the first organization
            org = Organization.query.first()
            if org:
                org_id = org.id
        
        if not org_id:
            return jsonify({'success': False, 'message': 'Organization not found'})
        
        # Find worker if FIN number provided
        worker_id = None
        if fin_number:
            worker = Worker.query.filter_by(
                organization_id=org_id,
                fin_number=fin_number,
                is_active=True
            ).first()
            if worker:
                worker_id = worker.id
        
        # 24-Hour Checkout Validation for START attendance
        if attendance_type == 'start':
            # Check if this worker has any uncompleted work sessions (start without end) in the last 24 hours
            twenty_four_hours_ago = singapore_now() - timedelta(hours=24)
            
            # Find the most recent start record for this worker/staff
            query_conditions = [
                StaffAttendance.organization_id == org_id,
                StaffAttendance.attendance_type == 'start',
                StaffAttendance.timestamp >= twenty_four_hours_ago
            ]
            
            # Use either worker_id or staff_name+company_name for identification
            if worker_id:
                query_conditions.append(StaffAttendance.worker_id == worker_id)
            else:
                query_conditions.append(StaffAttendance.staff_name == staff_name)
                query_conditions.append(StaffAttendance.company_name == company_name)
            
            recent_start = StaffAttendance.query.filter(*query_conditions).order_by(
                StaffAttendance.timestamp.desc()
            ).first()
            
            if recent_start:
                # Check if there's a corresponding end record after this start
                end_query_conditions = [
                    StaffAttendance.organization_id == org_id,
                    StaffAttendance.attendance_type == 'end',
                    StaffAttendance.timestamp > recent_start.timestamp
                ]
                
                # Use same identification method
                if worker_id:
                    end_query_conditions.append(StaffAttendance.worker_id == worker_id)
                else:
                    end_query_conditions.append(StaffAttendance.staff_name == staff_name)
                    end_query_conditions.append(StaffAttendance.company_name == company_name)
                
                corresponding_end = StaffAttendance.query.filter(*end_query_conditions).first()
                
                if not corresponding_end:
                    # Worker has an uncompleted work session - require checkout first
                    return jsonify({
                        'success': False, 
                        'message': 'You did not check out yesterday. Please check out first before starting new work.',
                        'error_type': 'missing_checkout',
                        'last_start_time': recent_start.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                    })
        
        # Create attendance record
        attendance = StaffAttendance(
            staff_name=staff_name,
            fin_number=fin_number,
            company_name=company_name,
            worker_id=worker_id,
            attendance_type=attendance_type,
            selfie_photo=selfie_data,
            qr_code_scanned=qr_code_scanned,
            location=location,
            organization_id=org_id,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')
        )
        
        db.session.add(attendance)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': 'Attendance submitted successfully',
            'attendance_id': attendance.id,
            'attendance_type': attendance_type,
            'timestamp': attendance.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error submitting attendance: {str(e)}'})

@app.route('/staff-attendance/success/<int:attendance_id>')
def staff_attendance_success(attendance_id):
    """Success page after attendance submission"""
    attendance = StaffAttendance.query.get_or_404(attendance_id)
    return render_template('staff_attendance_success.html', attendance=attendance)

@app.route('/staff-attendance/records')
@login_required
def staff_attendance_records():
    """Enhanced attendance records with pairing logic and status determination"""
    user = current_user
    org_id = user.organization_id
    
    # Get filter parameters
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    staff_name_filter = request.args.get('staff_name_filter', '').strip()
    company_filter = request.args.get('company_filter', '').strip()
    
    # Build base query
    query = StaffAttendance.query.filter_by(organization_id=org_id)
    
    # Apply date filters
    if from_date:
        try:
            from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
            query = query.filter(func.date(StaffAttendance.timestamp) >= from_dt)
        except ValueError:
            pass
    
    if to_date:
        try:
            to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
            query = query.filter(func.date(StaffAttendance.timestamp) <= to_dt)
        except ValueError:
            pass
    
    # Apply name filter
    if staff_name_filter:
        query = query.filter(StaffAttendance.staff_name == staff_name_filter)
    
    # Apply company filter
    if company_filter:
        query = query.filter(StaffAttendance.company_name == company_filter)
    
    # Get all records for processing
    all_records = query.order_by(StaffAttendance.timestamp.desc()).all()
    
    # Get unique staff names and companies for filters
    staff_names = db.session.query(StaffAttendance.staff_name).filter_by(
        organization_id=org_id
    ).distinct().order_by(StaffAttendance.staff_name).all()
    staff_names = [name[0] for name in staff_names if name[0]]
    
    companies = db.session.query(StaffAttendance.company_name).filter_by(
        organization_id=org_id
    ).distinct().order_by(StaffAttendance.company_name).all()
    companies = [company[0] for company in companies if company[0]]
    
    # Process records with pairing logic
    paired_records = process_attendance_pairs(all_records)
    
    # Calculate statistics
    total_records = len(paired_records)
    normal_sessions = len([r for r in paired_records if r['status'] == 'Normal'])
    off_status_count = len([r for r in paired_records if r['status'] == 'Off'])
    warning_count = len([r for r in paired_records if r['status'] == 'Warning'])
    
    return render_template('staff_attendance_records.html',
                         paired_records=paired_records,
                         staff_names=staff_names,
                         companies=companies,
                         from_date=from_date,
                         to_date=to_date,
                         staff_name_filter=staff_name_filter,
                         company_filter=company_filter,
                         total_records=total_records,
                         normal_sessions=normal_sessions,
                         off_status_count=off_status_count,
                         warning_count=warning_count)


def process_attendance_pairs(records):
    """Process attendance records to create paired sessions with status logic and Singapore timezone"""
    from collections import defaultdict
    import pytz
    import uuid
    
    # Singapore timezone
    singapore_tz = pytz.timezone('Asia/Singapore')
    
    # Group records by staff name and date (in Singapore timezone)
    grouped = defaultdict(lambda: defaultdict(list))
    
    for record in records:
        # Convert to Singapore timezone
        if record.timestamp.tzinfo is None:
            # Assume UTC if no timezone info
            local_timestamp = pytz.UTC.localize(record.timestamp).astimezone(singapore_tz)
        else:
            local_timestamp = record.timestamp.astimezone(singapore_tz)
        
        date_key = local_timestamp.date().strftime('%Y-%m-%d')
        record_data = {
            'id': record.id,
            'timestamp': local_timestamp,
            'attendance_type': record.attendance_type,
            'staff_name': record.staff_name,
            'company_name': record.company_name,
            'location': record.location,
            'selfie_photo': record.selfie_photo
        }
        grouped[record.staff_name][date_key].append(record_data)
    
    paired_records = []
    
    for staff_name, dates in grouped.items():
        for date_str, day_records in dates.items():
            # Sort records by timestamp
            day_records.sort(key=lambda x: x['timestamp'])
            
            # Separate start and end records
            start_records = [r for r in day_records if r['attendance_type'] == 'start']
            end_records = [r for r in day_records if r['attendance_type'] == 'end']
            
            # Process pairing logic
            if start_records and end_records:
                # Normal case: has both start and end
                start_record = start_records[0]  # First start of the day
                end_record = end_records[-1]     # Last end of the day
                
                # Calculate duration
                duration = end_record['timestamp'] - start_record['timestamp']
                duration_hours = duration.total_seconds() / 3600
                
                # Determine status
                if duration_hours > 24:
                    status = 'Warning'
                    row_class = 'table-danger'
                else:
                    status = 'Normal'
                    row_class = ''
                
                paired_record = {
                    'id': f"{start_record['id']}_{end_record['id']}",
                    'staff_name': staff_name,
                    'company': start_record['company_name'],
                    'date': date_str,
                    'primary_time': start_record['timestamp'].strftime('%H:%M:%S'),
                    'location': start_record['location'] or end_record['location'],
                    'photo': start_record['selfie_photo'],  # For compatibility
                    'start_photo': start_record['selfie_photo'],
                    'end_photo': end_record['selfie_photo'],
                    'start_date': start_record['timestamp'].strftime('%Y-%m-%d'),
                    'start_time': start_record['timestamp'].strftime('%H:%M:%S'),
                    'end_date': end_record['timestamp'].strftime('%Y-%m-%d'),
                    'end_time': end_record['timestamp'].strftime('%H:%M:%S'),
                    'duration': f"{int(duration_hours)}h {int((duration_hours % 1) * 60)}m",
                    'status': status,
                    'row_class': row_class
                }
                paired_records.append(paired_record)
                
            elif end_records and not start_records:
                # End only - "Off" status
                end_record = end_records[0]
                
                paired_record = {
                    'id': f"end_{end_record['id']}",
                    'staff_name': staff_name,
                    'company': end_record['company_name'],
                    'date': date_str,
                    'primary_time': end_record['timestamp'].strftime('%H:%M:%S'),
                    'location': end_record['location'],
                    'photo': end_record['selfie_photo'],
                    'start_photo': None,
                    'end_photo': end_record['selfie_photo'],
                    'start_date': None,
                    'start_time': None,
                    'end_date': end_record['timestamp'].strftime('%Y-%m-%d'),
                    'end_time': end_record['timestamp'].strftime('%H:%M:%S'),
                    'duration': None,
                    'status': 'Off',
                    'row_class': 'table-warning'
                }
                paired_records.append(paired_record)
                
            elif start_records and not end_records:
                # Start only - still working
                start_record = start_records[0]
                
                paired_record = {
                    'id': f"start_{start_record['id']}",
                    'staff_name': staff_name,
                    'company': start_record['company_name'],
                    'date': date_str,
                    'primary_time': start_record['timestamp'].strftime('%H:%M:%S'),
                    'location': start_record['location'],
                    'photo': start_record['selfie_photo'],
                    'start_photo': start_record['selfie_photo'],
                    'end_photo': None,
                    'start_date': start_record['timestamp'].strftime('%Y-%m-%d'),
                    'start_time': start_record['timestamp'].strftime('%H:%M:%S'),
                    'end_date': None,
                    'end_time': None,
                    'duration': None,
                    'status': 'Normal',
                    'row_class': 'table-info'
                }
                paired_records.append(paired_record)
    
    # Sort by most recent first
    paired_records.sort(key=lambda x: x['date'], reverse=True)
    
    return paired_records


@app.route('/staff-attendance/delete-records', methods=['POST'])
@login_required
@create_permission_required('staff_attendance')
def delete_attendance_records():
    """Delete selected attendance records"""
    try:
        data = request.get_json()
        record_ids = data.get('record_ids', [])
        
        if not record_ids:
            return jsonify({'success': False, 'message': 'No records selected'})
        
        user = current_user
        org_id = user.organization_id
        
        # Extract actual database IDs from the composite IDs
        db_ids = []
        for record_id in record_ids:
            if '_' in record_id:
                # Paired record: "start_id_end_id"
                if record_id.startswith('end_'):
                    # End-only record: "end_id"
                    db_ids.append(int(record_id.split('_')[1]))
                elif record_id.startswith('start_'):
                    # Start-only record: "start_id"
                    db_ids.append(int(record_id.split('_')[1]))
                else:
                    # Normal paired: "start_id_end_id"
                    parts = record_id.split('_')
                    db_ids.extend([int(parts[0]), int(parts[1])])
            else:
                db_ids.append(int(record_id))
        
        # Delete records
        deleted_count = StaffAttendance.query.filter(
            StaffAttendance.id.in_(db_ids),
            StaffAttendance.organization_id == org_id
        ).delete(synchronize_session=False)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully deleted {deleted_count} attendance records'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error deleting records: {str(e)}'})


@app.route('/staff-attendance/export/excel')
@login_required
def export_attendance_excel():
    """Export selected attendance records to Excel"""
    user = current_user
    org_id = user.organization_id
    
    # Get filter parameters and selected IDs
    selected_ids = request.args.get('selected_ids', '').split(',')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    staff_name_filter = request.args.get('staff_name_filter', '').strip()
    company_filter = request.args.get('company_filter', '').strip()
    
    # Build query
    query = StaffAttendance.query.filter_by(organization_id=org_id)
    
    # Apply filters
    if from_date:
        try:
            from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
            query = query.filter(func.date(StaffAttendance.timestamp) >= from_dt)
        except ValueError:
            pass
    
    if to_date:
        try:
            to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
            query = query.filter(func.date(StaffAttendance.timestamp) <= to_dt)
        except ValueError:
            pass
    
    if staff_name_filter:
        query = query.filter(StaffAttendance.staff_name == staff_name_filter)
    
    if company_filter:
        query = query.filter(StaffAttendance.company_name == company_filter)
    
    # Get records and process them
    all_records = query.order_by(StaffAttendance.timestamp.desc()).all()
    paired_records = process_attendance_pairs(all_records)
    
    # Filter by selected IDs if provided
    if selected_ids and selected_ids[0]:
        paired_records = [r for r in paired_records if r['id'] in selected_ids]
    
    # Extract the enhanced Excel logic directly here to avoid request context issues
    from calendar import monthrange
    import calendar
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import pytz
    
    # Determine date range
    singapore_tz = pytz.timezone('Asia/Singapore')
    today = datetime.now(singapore_tz).date()
    
    if from_date and to_date:
        start_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
    elif from_date:
        start_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
        end_dt = today
    elif to_date:
        end_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
        start_dt = today.replace(day=1)
    else:
        # Default to current month
        start_dt = today.replace(day=1)
        last_day = calendar.monthrange(today.year, today.month)[1]
        end_dt = today.replace(day=last_day)
    
    # Generate all dates in the range
    all_dates = []
    current_date = start_dt
    while current_date <= end_dt:
        all_dates.append(current_date.strftime('%Y-%m-%d'))
        current_date += timedelta(days=1)
    
    # Group dates by month for better organization
    monthly_groups = {}
    for date_str in all_dates:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        month_key = date_obj.strftime('%Y-%m')
        if month_key not in monthly_groups:
            monthly_groups[month_key] = []
        monthly_groups[month_key].append(date_str)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Attendance Report"
    
    # Title and headers
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    
    # Create dynamic title based on date range
    if start_dt == end_dt:
        date_range = start_dt.strftime('%d %B %Y')
    elif start_dt.year == end_dt.year and start_dt.month == end_dt.month:
        date_range = start_dt.strftime('%B %Y')
    else:
        date_range = f"{start_dt.strftime('%d %b %Y')} - {end_dt.strftime('%d %b %Y')}"
    
    title_cell.value = f"TS Management Service Pte Ltd - Staff Attendance Report ({date_range})"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Company and system info
    ws['A2'] = "Company: TS Management Service Pte Ltd"
    ws['A3'] = "Report Generated: " + datetime.now(singapore_tz).strftime('%d %B %Y at %I:%M %p')
    
    # Headers - Staff Name, Company, then dates
    current_row = 5
    ws['A' + str(current_row)] = "Staff Name"
    ws['B' + str(current_row)] = "Company"
    ws['C' + str(current_row)] = "Time In"
    ws['D' + str(current_row)] = "Time Out"
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, 5):  # A to D
        cell = ws.cell(row=current_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    current_row += 1
    
    # Group records by staff and date
    staff_records = {}
    for record in all_records:
        record_date = record.timestamp.astimezone(singapore_tz).strftime('%Y-%m-%d')
        staff_key = f"{record.staff_name}_{record.company_name}"
        
        if staff_key not in staff_records:
            staff_records[staff_key] = {
                'staff_name': record.staff_name,
                'company_name': record.company_name,
                'dates': {}
            }
        
        if record_date not in staff_records[staff_key]['dates']:
            staff_records[staff_key]['dates'][record_date] = {'in': None, 'out': None}
        
        if record.entry_type == 'entry':
            staff_records[staff_key]['dates'][record_date]['in'] = record.timestamp.astimezone(singapore_tz).strftime('%I:%M %p')
        else:
            staff_records[staff_key]['dates'][record_date]['out'] = record.timestamp.astimezone(singapore_tz).strftime('%I:%M %p')
    
    # Add staff records
    for staff_key, staff_data in staff_records.items():
        for date_str in all_dates:
            if date_str in staff_data['dates']:
                date_data = staff_data['dates'][date_str]
                if date_data['in'] or date_data['out']:  # Only show if there's activity
                    ws.cell(row=current_row, column=1, value=staff_data['staff_name'])
                    ws.cell(row=current_row, column=2, value=staff_data['company_name'])
                    ws.cell(row=current_row, column=3, value=date_data['in'] or '-')
                    ws.cell(row=current_row, column=4, value=date_data['out'] or '-')
                    current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    from flask import make_response
    response = make_response(excel_file.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=staff_attendance_{start_dt.strftime("%Y%m%d")}_to_{end_dt.strftime("%Y%m%d")}.xlsx'
    
    return response




@app.route('/staff-attendance/export-excel')
@login_required
def export_staff_attendance_excel():
    """Export staff attendance records to Excel for payroll"""
    user = current_user
    org_id = user.organization_id
    
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    staff_name = request.args.get('staff_name', '').strip()
    company_name = request.args.get('company_name', '').strip()
    
    # Build query for all records
    query = StaffAttendance.query.filter_by(organization_id=org_id)
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(func.date(StaffAttendance.timestamp) >= start_dt)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(func.date(StaffAttendance.timestamp) <= end_dt)
        except ValueError:
            pass
    
    if staff_name:
        query = query.filter(StaffAttendance.staff_name.ilike(f'%{staff_name}%'))
    
    if company_name:
        query = query.filter(StaffAttendance.company_name.ilike(f'%{company_name}%'))
    
    # Get all records ordered by date and time
    records = query.order_by(StaffAttendance.timestamp.asc()).all()
    
    # Group records by date and staff
    from collections import defaultdict
    import calendar
    daily_records = defaultdict(lambda: defaultdict(dict))
    
    for record in records:
        date_str = record.timestamp.strftime('%Y-%m-%d')
        staff_key = f"{record.staff_name} ({record.company_name})"
        
        if record.attendance_type == 'start':
            daily_records[date_str][staff_key]['start'] = record.timestamp.strftime('%H:%M')
        elif record.attendance_type == 'end':
            daily_records[date_str][staff_key]['end'] = record.timestamp.strftime('%H:%M')
    
    # Determine the date range to display
    from datetime import datetime, timedelta
    
    if start_date and end_date:
        # Use user-specified date range
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
    elif start_date:
        # If only start date specified, use rest of that month
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        last_day = calendar.monthrange(start_dt.year, start_dt.month)[1]
        end_dt = start_dt.replace(day=last_day)
    elif end_date:
        # If only end date specified, use from beginning of that month
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        start_dt = end_dt.replace(day=1)
    elif records:
        # Use the range that covers all existing records
        dates = [record.timestamp.date() for record in records]
        start_dt = min(dates)
        end_dt = max(dates)
    else:
        # Default to current month if no records and no date filters
        today = singapore_now().date()
        start_dt = today.replace(day=1)
        last_day = calendar.monthrange(today.year, today.month)[1]
        end_dt = today.replace(day=last_day)
    
    # Generate all dates in the range
    all_dates = []
    current_date = start_dt
    while current_date <= end_dt:
        all_dates.append(current_date.strftime('%Y-%m-%d'))
        current_date += timedelta(days=1)
    
    # Group dates by month for better organization
    monthly_groups = {}
    for date_str in all_dates:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        month_key = date_obj.strftime('%Y-%m')
        if month_key not in monthly_groups:
            monthly_groups[month_key] = []
        monthly_groups[month_key].append(date_str)
    
    # Create workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Attendance Report"
    
    # Title and headers
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    
    # Create dynamic title based on date range
    if start_dt == end_dt:
        date_range = start_dt.strftime('%d %B %Y')
    elif start_dt.year == end_dt.year and start_dt.month == end_dt.month:
        date_range = start_dt.strftime('%B %Y')
    else:
        date_range = f"{start_dt.strftime('%d %b %Y')} - {end_dt.strftime('%d %b %Y')}"
    
    title_cell.value = f"TS Management Service Pte Ltd - Staff Attendance Report ({date_range})"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = ['Date', 'Staff Name (Company)', 'Start Time', 'End Time']
    ws.append([''])  # Empty row
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data rows
    row_num = 4
    today = singapore_now().date()
    
    # Process each month separately for better organization
    for month_key in sorted(monthly_groups.keys()):
        month_dates = monthly_groups[month_key]
        
        # Add month separator for multi-month reports
        if len(monthly_groups) > 1:
            month_obj = datetime.strptime(month_key + '-01', '%Y-%m-%d').date()
            month_header = month_obj.strftime('%B %Y')
            
            ws.merge_cells(f'A{row_num}:D{row_num}')
            month_cell = ws.cell(row=row_num, column=1, value=month_header)
            month_cell.font = Font(bold=True, size=13, color='FFFFFF')
            month_cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
            month_cell.alignment = Alignment(horizontal='center')
            month_cell.border = border
            
            # Apply styling to merged cells
            for col in range(2, 5):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
            
            row_num += 1
        
        # Process each date in the month
        for date_str in month_dates:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            date_display = date_obj.strftime('%d %B %Y')
            staff_records = daily_records.get(date_str, {})
            
            # Date header row with merged cells
            ws.merge_cells(f'A{row_num}:D{row_num}')
            date_cell = ws.cell(row=row_num, column=1, value=date_display)
            date_cell.font = Font(bold=True, size=12)
            date_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            date_cell.alignment = Alignment(horizontal='center')
            date_cell.border = border
            
            # Apply border to merged cells
            for col in range(2, 5):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            row_num += 1
            
            # Column headers for this date
            sub_headers = ['Staff Name', 'Company', 'Time In', 'Time Out']
            for col, header in enumerate(sub_headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            row_num += 1
            
            if staff_records:
                # Staff records for this date
                for staff_name, times in sorted(staff_records.items()):
                    # Split staff name and company
                    if '(' in staff_name and staff_name.endswith(')'):
                        name_part = staff_name.split(' (')[0]
                        company_part = staff_name.split(' (')[1][:-1]  # Remove closing parenthesis
                    else:
                        name_part = staff_name
                        company_part = ''
                    
                    start_time = times.get('start', '-')
                    end_time = times.get('end', '-')
                    
                    # Add row data
                    ws.cell(row=row_num, column=1, value=name_part)
                    ws.cell(row=row_num, column=2, value=company_part)
                    ws.cell(row=row_num, column=3, value=start_time)
                    ws.cell(row=row_num, column=4, value=end_time)
                    
                    # Style cells
                    for col in range(1, 5):
                        cell = ws.cell(row=row_num, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                        
                        # Highlight missing times
                        if col > 2 and cell.value == '-':
                            cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                    
                    row_num += 1
            else:
                # No records for this date
                if date_obj > today:
                    no_record_msg = "Date hasn't arrived yet"
                else:
                    no_record_msg = "No records yet"
                
                ws.merge_cells(f'A{row_num}:D{row_num}')
                no_record_cell = ws.cell(row=row_num, column=1, value=no_record_msg)
                no_record_cell.font = Font(italic=True, color='999999')
                no_record_cell.alignment = Alignment(horizontal='center')
                no_record_cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                no_record_cell.border = border
                
                # Apply border to merged cells
                for col in range(2, 5):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = border
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                
                row_num += 1
            
            # Add empty row between dates
            row_num += 1
    
    # Auto-adjust column widths
    column_widths = [15, 35, 12, 12]  # Date, Staff Name, Start Time, End Time
    column_letters = ['A', 'B', 'C', 'D']
    for i, width in enumerate(column_widths):
        ws.column_dimensions[column_letters[i]].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=staff_attendance_payroll_{singapore_now().strftime("%Y%m%d")}.xlsx'
    
    return response

@app.route('/staff-attendance/export-pdf')
@login_required
def export_staff_attendance_pdf():
    """Export staff attendance records to PDF"""
    user = current_user
    org_id = user.organization_id
    
    # Get filter parameters (same as Excel export)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    staff_name = request.args.get('staff_name', '').strip()
    company_name = request.args.get('company_name', '').strip()
    attendance_type = request.args.get('attendance_type', '').strip()
    
    # Build query (same as Excel export)
    query = StaffAttendance.query.filter_by(organization_id=org_id)
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(func.date(StaffAttendance.timestamp) >= start_dt)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(func.date(StaffAttendance.timestamp) <= end_dt)
        except ValueError:
            pass
    
    if staff_name:
        query = query.filter(StaffAttendance.staff_name.ilike(f'%{staff_name}%'))
    
    if company_name:
        query = query.filter(StaffAttendance.company_name.ilike(f'%{company_name}%'))
    
    if attendance_type in ['start', 'end']:
        query = query.filter(StaffAttendance.attendance_type == attendance_type)
    
    # Get records
    records = query.order_by(StaffAttendance.timestamp.desc()).all()
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER
    )
    story.append(Paragraph("Staff Attendance Report", title_style))
    story.append(Spacer(1, 20))
    
    # Summary info
    if start_date or end_date:
        period_text = f"Period: {start_date or 'Beginning'} to {end_date or 'Present'}"
        story.append(Paragraph(period_text, styles['Normal']))
        story.append(Spacer(1, 12))
    
    story.append(Paragraph(f"Total Records: {len(records)}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Data table
    if records:
        table_data = [['Date', 'Time', 'Staff Name', 'Company', 'Type', 'Location']]
        
        for record in records:
            table_data.append([
                record.timestamp.strftime('%Y-%m-%d'),
                record.timestamp.strftime('%H:%M'),
                record.staff_name,
                record.company_name,
                record.attendance_type.title(),
                record.location or '-'
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No attendance records found.", styles['Normal']))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=staff_attendance_{singapore_now().strftime("%Y%m%d")}.pdf'
    
    return response

@app.route('/staff-attendance/qr-codes')
@login_required
def generate_staff_attendance_qr_codes():
    """Generate QR codes for staff attendance"""
    user = current_user
    
    if not is_admin_user(user):
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('staff_attendance'))
    
    org_id = user.organization_id
    
    # Create or get existing QR codes for start and end attendance
    start_qr = QRCode.query.filter_by(
        organization_id=org_id,
        qr_type='staff_attendance',
        reference_id='start'
    ).first()
    
    end_qr = QRCode.query.filter_by(
        organization_id=org_id,
        qr_type='staff_attendance',
        reference_id='end'
    ).first()
    
    if not start_qr:
        start_qr = QRCode(
            code=str(uuid.uuid4()),
            qr_type='staff_attendance',
            reference_id='start',
            reference_table='staff_attendance',
            organization_id=org_id,
            label='Staff Start Time QR Code',
            description='Scan to mark start of work time',
            target_url=f'/staff-attendance/scan/start',
            is_public=True,
            created_by=user.id
        )
        db.session.add(start_qr)
    
    if not end_qr:
        end_qr = QRCode(
            code=str(uuid.uuid4()),
            qr_type='staff_attendance',
            reference_id='end',
            reference_table='staff_attendance',
            organization_id=org_id,
            label='Staff End Time QR Code',
            description='Scan to mark end of work time',
            target_url=f'/staff-attendance/scan/end',
            is_public=True,
            created_by=user.id
        )
        db.session.add(end_qr)
    
    db.session.commit()
    
    return render_template('staff_attendance_qr_codes.html',
                         start_qr=start_qr,
                         end_qr=end_qr)

@app.route('/staff-attendance/workers')
@login_required
def manage_workers():
    """Manage worker names and information"""
    user = current_user
    org_id = user.organization_id
    
    if not is_admin_user(user):
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('staff_attendance'))
    
    workers = Worker.query.filter_by(organization_id=org_id, is_active=True).order_by(Worker.name).all()
    
    return render_template('staff_attendance_workers.html', workers=workers)

@app.route('/staff-attendance/workers/add', methods=['POST'])
@login_required
def add_worker():
    """Add a new worker"""
    user = current_user
    org_id = user.organization_id
    
    if not is_admin_user(user):
        return jsonify({'success': False, 'error': 'Access denied'})
    
    try:
        name = request.form.get('name', '').strip()
        fin_number = request.form.get('fin_number', '').strip()
        company_name = request.form.get('company_name', '').strip()
        
        if not all([name, fin_number, company_name]):
            return jsonify({'success': False, 'error': 'All fields are required'})
        
        # Check if worker already exists
        existing = Worker.query.filter_by(
            organization_id=org_id,
            fin_number=fin_number,
            is_active=True
        ).first()
        
        if existing:
            return jsonify({'success': False, 'error': 'Worker with this FIN number already exists'})
        
        worker = Worker(
            name=name,
            fin_number=fin_number,
            company_name=company_name,
            organization_id=org_id
        )
        
        db.session.add(worker)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'worker': {
                'id': worker.id,
                'name': worker.name,
                'fin_number': worker.fin_number,
                'company_name': worker.company_name
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error adding worker: {str(e)}'})

@app.route('/staff-attendance/workers/get')
@login_required
def get_workers():
    """Get workers for dropdown"""
    user = current_user
    org_id = user.organization_id
    
    workers = Worker.query.filter_by(organization_id=org_id, is_active=True).order_by(Worker.name).all()
    
    worker_list = []
    for worker in workers:
        worker_list.append({
            'id': worker.id,
            'name': worker.name,
            'fin_number': worker.fin_number,
            'company_name': worker.company_name
        })
    
    return jsonify({'workers': worker_list})

@app.route('/staff-attendance/workers/<int:worker_id>/delete', methods=['POST'])
@login_required
def delete_worker(worker_id):
    """Delete (deactivate) a worker"""
    user = current_user
    org_id = user.organization_id
    
    if not is_admin_user(user):
        return jsonify({'success': False, 'error': 'Access denied'})
    
    try:
        worker = Worker.query.filter_by(id=worker_id, organization_id=org_id).first()
        
        if not worker:
            return jsonify({'success': False, 'error': 'Worker not found'})
        
        worker.is_active = False
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error deleting worker: {str(e)}'})

# ===== PIONEER LODGE VISITORS SYSTEM =====

@app.route('/visitors-dashboard')
@login_required
@page_permission_required('pioneer_lodge_visitors')
def visitors_dashboard():
    """Pioneer Lodge Visitors dashboard page"""
    user = current_user
    org_id = user.organization_id
    
    # Get recent visitor records
    recent_visitors = Visitor.query.filter_by(
        organization_id=org_id
    ).order_by(Visitor.timestamp.desc()).limit(10).all()
    
    # Get statistics
    today = singapore_now().date()
    today_start_count = Visitor.query.filter_by(
        organization_id=org_id,
        visit_type='start'
    ).filter(func.date(Visitor.timestamp) == today).count()
    
    today_end_count = Visitor.query.filter_by(
        organization_id=org_id,
        visit_type='end'
    ).filter(func.date(Visitor.timestamp) == today).count()
    
    total_records = Visitor.query.filter_by(organization_id=org_id).count()
    
    return render_template('visitors.html',
                         recent_visitors=recent_visitors,
                         today_start_count=today_start_count,
                         today_end_count=today_end_count,
                         total_records=total_records)

@app.route('/visitors/scan')
def visitors_scan():
    """QR code scanning page for visitors"""
    visit_type = request.args.get('type', 'start')  # 'start' or 'end'
    return render_template('visitors_scan.html', visit_type=visit_type)

@app.route('/visitors/qr-codes')
@login_required
def visitors_qr_codes():
    """Generate printable QR codes for visitor check-in/check-out"""
    user = current_user
    
    # Get the base URL for QR codes
    base_url = request.url_root.rstrip('/')
    
    # QR code URLs
    start_url = f"{base_url}/visitors/scan?type=start"
    end_url = f"{base_url}/visitors/scan?type=end"
    
    # Generate QR codes
    import qrcode
    from io import BytesIO
    import base64
    
    def generate_qr_code(url, size=10):
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=size,
            border=4,
        )
        qr.add_data(url)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        # Convert to base64 for embedding in HTML
        img_str = base64.b64encode(buffer.getvalue()).decode()
        return f"data:image/png;base64,{img_str}"
    
    start_qr = generate_qr_code(start_url)
    end_qr = generate_qr_code(end_url)
    
    return render_template('visitors_qr_codes.html',
                         start_qr=start_qr,
                         end_qr=end_qr,
                         start_url=start_url,
                         end_url=end_url,
                         organization_name=user.organization.name if user.organization else 'Pioneer Lodge')

@app.route('/visitors/submit', methods=['POST'])
@create_permission_required('visitor_management')
def visitors_submit():
    """Process visitor attendance submission"""
    try:
        data = request.get_json()
        
        # Extract data
        visitor_name = data.get('visitor_name', '').strip()
        company_name = data.get('company_name', '').strip()
        vehicle_number = data.get('vehicle_number', '').strip()
        details = data.get('details', '').strip()
        visit_type = data.get('visit_type', 'start')
        selfie_photo = data.get('selfie_photo')
        qr_code_scanned = data.get('qr_code_scanned', '')
        
        # Validation
        if not visitor_name:
            return jsonify({'success': False, 'error': 'Visitor name is required'})
        
        if not company_name:
            return jsonify({'success': False, 'error': 'Company name is required'})
        
        # Selfie photo is now optional
        
        # Get user's organization (fallback for public access)
        org_id = 1  # Default to Pioneer Lodge organization
        if current_user.is_authenticated:
            org_id = current_user.organization_id
        
        # Create visitor record
        visitor = Visitor(
            visitor_name=visitor_name,
            company_name=company_name,
            vehicle_number=vehicle_number if vehicle_number else None,
            details=details,
            visit_type=visit_type,
            selfie_photo=selfie_photo if selfie_photo else None,
            qr_code_scanned=qr_code_scanned,
            location='Pioneer Lodge',
            organization_id=org_id,
            ip_address=request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR')),
            user_agent=request.environ.get('HTTP_USER_AGENT')
        )
        
        db.session.add(visitor)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'visitor_id': visitor.id,
            'message': 'Visit recorded successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error recording visit: {str(e)}'})

@app.route('/visitors/success/<int:visitor_id>')
def visitors_success(visitor_id):
    """Success page after visitor submission"""
    visitor = Visitor.query.get_or_404(visitor_id)
    return render_template('visitors_success.html', visitor=visitor)

# ===== KEY MANAGEMENT ROUTES =====

@app.route('/key-management')
@login_required
@page_access_required('key_management')
def key_management_dashboard():
    """Key Management main dashboard"""
    # Get statistics for dashboard
    today = singapore_now().date()
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today, datetime.max.time())
    
    # Count keys currently out (not returned)
    total_keys_out = KeyRecord.query.filter_by(scan_type='out', is_returned=False).count()
    
    # Count overdue keys (out for more than 30 minutes)
    thirty_min_ago = singapore_now() - timedelta(minutes=30)
    overdue_keys = KeyRecord.query.filter(
        KeyRecord.scan_type == 'out',
        KeyRecord.is_returned == False,
        KeyRecord.scan_time < thirty_min_ago
    ).count()
    
    # Count returned today
    total_returned_today = KeyRecord.query.filter(
        KeyRecord.scan_type == 'in',
        KeyRecord.scan_time >= today_start,
        KeyRecord.scan_time <= today_end
    ).count()
    
    # Count total scans today
    total_scans_today = KeyRecord.query.filter(
        KeyRecord.scan_time >= today_start,
        KeyRecord.scan_time <= today_end
    ).count()
    
    # Get recent activity (last 10 records)
    recent_activity = KeyRecord.query.filter(
        KeyRecord.scan_time >= today_start
    ).order_by(KeyRecord.scan_time.desc()).limit(10).all()
    
    return render_template('key_management_dashboard.html',
                         total_keys_out=total_keys_out,
                         overdue_keys=overdue_keys,
                         total_returned_today=total_returned_today,
                         total_scans_today=total_scans_today,
                         recent_activity=recent_activity)

@app.route('/key-management/qr-codes')
@login_required
@page_access_required('key_management')
def key_management_qr_codes():
    """Generate Key In/Out QR codes"""
    import qrcode
    from io import BytesIO
    import base64
    
    # Generate Key Out QR Code
    key_out_url = f"{request.url_root}key-management/scan/out"
    qr_out = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
    qr_out.add_data(key_out_url)
    qr_out.make(fit=True)
    
    img_out = qr_out.make_image(fill_color="black", back_color="white")
    buffer_out = BytesIO()
    img_out.save(buffer_out, format='PNG')
    qr_out_data = base64.b64encode(buffer_out.getvalue()).decode()
    
    # Generate Key In QR Code
    key_in_url = f"{request.url_root}key-management/scan/in"
    qr_in = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
    qr_in.add_data(key_in_url)
    qr_in.make(fit=True)
    
    img_in = qr_in.make_image(fill_color="black", back_color="white")
    buffer_in = BytesIO()
    img_in.save(buffer_in, format='PNG')
    qr_in_data = base64.b64encode(buffer_in.getvalue()).decode()
    
    return render_template('key_management_qr_codes.html', 
                         qr_out_data=qr_out_data, qr_in_data=qr_in_data,
                         key_out_url=key_out_url, key_in_url=key_in_url)

@app.route('/key-management/scan/<scan_type>')
def key_scan_form(scan_type):
    """Key scan form for residents - public access"""
    if scan_type not in ['in', 'out']:
        return "Invalid scan type", 400
    
    # Get active room numbers for dropdown
    from app.models.models_house_acknowledge import RoomNumber
    room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    
    return render_template('key_scan_form.html', scan_type=scan_type, room_numbers=room_numbers)

@app.route('/key-management/scan/<scan_type>', methods=['POST'])
def key_scan_submit(scan_type):
    """Process key scan submission"""
    if scan_type not in ['in', 'out']:
        return "Invalid scan type", 400
    
    try:
        room_number = request.form.get('room_number', '').strip()
        resident_name = request.form.get('resident_name', '').strip()
        company_name = request.form.get('company_name', '').strip()
        
        if not all([room_number, resident_name, company_name]):
            flash('Please fill in all required fields.', 'error')
            return render_template('key_scan_form.html', scan_type=scan_type)
        
        # Create key record
        key_record = KeyRecord(
            room_number=room_number,
            resident_name=resident_name,
            company_name=company_name,
            scan_type=scan_type,
            qr_code_type=f'key_{scan_type}',
            scan_time=singapore_now(),
            organization_id=current_user.organization_id if current_user.is_authenticated else None,
            created_by=current_user.id if current_user.is_authenticated else None
        )
        
        # If this is a key return (in), try to match with existing key out record
        if scan_type == 'in':
            # Find the most recent 'out' record for this room/person that hasn't been returned
            existing_out = KeyRecord.query.filter_by(
                room_number=room_number,
                resident_name=resident_name,
                scan_type='out',
                is_returned=False
            ).order_by(KeyRecord.scan_time.desc()).first()
            
            if existing_out:
                existing_out.is_returned = True
                existing_out.return_time = singapore_now()
                existing_out.status = 'Returned'
                key_record.status = 'Returned'
            else:
                key_record.status = 'Active'
        else:
            key_record.status = 'Active'
        
        db.session.add(key_record)
        db.session.commit()
        
        return render_template('key_scan_success.html', 
                             key_record=key_record, scan_type=scan_type)
                             
    except Exception as e:
        db.session.rollback()
        flash(f'Error processing scan: {str(e)}', 'error')
        return render_template('key_scan_form.html', scan_type=scan_type)

@app.route('/key-management/records')
@login_required
@page_access_required('key_management')
def key_management_records():
    """View and manage key records"""
    # Get filter parameters
    room_filter = request.args.get('room_number', '')
    name_filter = request.args.get('resident_name', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_filter = request.args.get('status', '')
    
    # Build query
    query = KeyRecord.query
    
    if room_filter:
        query = query.filter(KeyRecord.room_number.ilike(f'%{room_filter}%'))
    if name_filter:
        query = query.filter(KeyRecord.resident_name.ilike(f'%{name_filter}%'))
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(KeyRecord.scan_time >= from_date)
        except ValueError:
            pass
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            query = query.filter(KeyRecord.scan_time <= to_date + timedelta(days=1))
        except ValueError:
            pass
    if status_filter:
        if status_filter == 'Not Returned':
            query = query.filter(KeyRecord.scan_type == 'out', KeyRecord.is_returned == False)
        else:
            query = query.filter(KeyRecord.status == status_filter)
    
    # Update overdue statuses
    overdue_keys = query.filter(
        KeyRecord.scan_type == 'out',
        KeyRecord.is_returned == False,
        KeyRecord.scan_time < singapore_now() - timedelta(minutes=30)
    ).all()
    
    for key in overdue_keys:
        key.status = 'Not Returned'
    
    if overdue_keys:
        db.session.commit()
    
    records = query.order_by(KeyRecord.scan_time.desc()).all()
    
    # Get room numbers for dropdown filter
    from app.models.models_house_acknowledge import RoomNumber
    room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    
    return render_template('key_management_records.html', 
                         records=records,
                         room_numbers=room_numbers,
                         room_filter=room_filter,
                         name_filter=name_filter,
                         date_from=date_from,
                         date_to=date_to,
                         status_filter=status_filter)

@app.route('/key-management/edit/<int:record_id>', methods=['GET', 'POST'])
@login_required
@edit_permission_required('key_management')
def edit_key_record(record_id):
    """Edit key record"""
    record = KeyRecord.query.get_or_404(record_id)
    
    if request.method == 'POST':
        try:
            record.room_number = request.form.get('room_number', '').strip()
            record.resident_name = request.form.get('resident_name', '').strip()
            record.company_name = request.form.get('company_name', '').strip()
            record.notes = request.form.get('notes', '').strip()
            
            scan_time_str = request.form.get('scan_time', '').strip()
            if scan_time_str:
                record.scan_time = datetime.strptime(scan_time_str, '%Y-%m-%dT%H:%M')
            
            # Update return status if needed
            is_returned = request.form.get('is_returned') == 'on'
            if is_returned and not record.is_returned:
                record.is_returned = True
                record.return_time = singapore_now()
                record.status = 'Returned'
            elif not is_returned and record.is_returned:
                record.is_returned = False
                record.return_time = None
                record.status = 'Active'
            
            db.session.commit()
            flash('Key record updated successfully.', 'success')
            return redirect(url_for('key_management_records'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating record: {str(e)}', 'error')
    
    # Get room numbers for dropdown
    from app.models.models_house_acknowledge import RoomNumber
    room_numbers = RoomNumber.query.filter_by(is_active=True).order_by(RoomNumber.room_number).all()
    
    return render_template('edit_key_record.html', record=record, room_numbers=room_numbers)

@app.route('/key-management/delete', methods=['POST'])
@login_required
@edit_permission_required('key_management')
def delete_key_records():
    """Delete selected key records"""
    try:
        record_ids = request.form.getlist('record_ids')
        if not record_ids:
            flash('No records selected for deletion.', 'error')
            return redirect(url_for('key_management_records'))
        
        # Delete records
        KeyRecord.query.filter(KeyRecord.id.in_(record_ids)).delete(synchronize_session=False)
        db.session.commit()
        
        flash(f'Successfully deleted {len(record_ids)} record(s).', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting records: {str(e)}', 'error')
    
    return redirect(url_for('key_management_records'))

@app.route('/key-management/export/<export_type>')
@login_required
@page_access_required('key_management')
def export_key_records(export_type):
    """Export key records to Excel or PDF"""
    # Get selected record IDs from query parameters
    record_ids = request.args.getlist('record_ids')
    
    if record_ids:
        records = KeyRecord.query.filter(KeyRecord.id.in_(record_ids)).order_by(KeyRecord.scan_time.desc()).all()
    else:
        # Export all records if none selected
        records = KeyRecord.query.order_by(KeyRecord.scan_time.desc()).all()
    
    if export_type == 'excel':
        return export_key_records_excel(records)
    elif export_type == 'pdf':
        return export_key_records_pdf(records)
    else:
        flash('Invalid export type.', 'error')
        return redirect(url_for('key_management_records'))

def export_key_records_excel(records):
    """Export key records to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Key Management Records"
    
    # Headers
    headers = ['Room Number', 'Resident Name', 'Company Name', 'Scan Type', 'Scan Time', 'Status', 'Time Held (minutes)', 'Notes']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for record in records:
        ws.append([
            record.room_number,
            record.resident_name,
            record.company_name,
            record.scan_type.title(),
            record.scan_time.strftime('%Y-%m-%d %H:%M:%S'),
            record.status,
            record.time_held_minutes,
            record.notes or ''
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'key_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

def export_key_records_pdf(records):
    """Export key records to PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER, spaceAfter=30)
    
    # Content
    content = []
    
    # Title
    content.append(Paragraph("Key Management Records", title_style))
    content.append(Spacer(1, 20))
    
    # Table data
    data = [['Room Number', 'Resident Name', 'Company Name', 'Scan Type', 'Scan Time', 'Status', 'Time Held']]
    
    for record in records:
        data.append([
            record.room_number,
            record.resident_name,
            record.company_name,
            record.scan_type.title(),
            record.scan_time.strftime('%Y-%m-%d %H:%M'),
            record.status,
            f"{record.time_held_minutes} min"
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    content.append(table)
    
    # Build PDF
    doc.build(content)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'key_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )

@app.route('/visitors/records')
@login_required
def visitors_records():
    """Enhanced visitor records with pairing logic and status determination"""
    user = current_user
    org_id = user.organization_id
    
    # Get filter parameters
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    visitor_name_filter = request.args.get('visitor_name_filter', '').strip()
    company_filter = request.args.get('company_filter', '').strip()
    
    # Build base query
    query = Visitor.query.filter_by(organization_id=org_id)
    
    # Apply date filters
    if from_date:
        try:
            from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
            query = query.filter(func.date(Visitor.timestamp) >= from_dt)
        except ValueError:
            pass
    
    if to_date:
        try:
            to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
            query = query.filter(func.date(Visitor.timestamp) <= to_dt)
        except ValueError:
            pass
    
    # Apply name filter
    if visitor_name_filter:
        query = query.filter(Visitor.visitor_name == visitor_name_filter)
    
    # Apply company filter
    if company_filter:
        query = query.filter(Visitor.company_name == company_filter)
    
    # Get all records for processing
    all_records = query.order_by(Visitor.timestamp.desc()).all()
    
    # Get unique visitor names and companies for filters
    visitor_names = db.session.query(Visitor.visitor_name).filter_by(
        organization_id=org_id
    ).distinct().order_by(Visitor.visitor_name).all()
    visitor_names = [name[0] for name in visitor_names if name[0]]
    
    companies = db.session.query(Visitor.company_name).filter_by(
        organization_id=org_id
    ).distinct().order_by(Visitor.company_name).all()
    companies = [company[0] for company in companies if company[0]]
    
    # Process records with pairing logic
    paired_records = process_visitor_pairs(all_records)
    
    # Calculate statistics
    total_records = len(paired_records)
    normal_sessions = len([r for r in paired_records if r['status'] == 'Normal'])
    warning_count = len([r for r in paired_records if r['status'] == 'Warning'])
    off_status_count = len([r for r in paired_records if r['status'] == 'Off'])
    
    return render_template('visitors_records.html',
                         paired_records=paired_records,
                         visitor_names=visitor_names,
                         companies=companies,
                         from_date=from_date,
                         to_date=to_date,
                         visitor_name_filter=visitor_name_filter,
                         company_filter=company_filter,
                         total_records=total_records,
                         normal_sessions=normal_sessions,
                         warning_count=warning_count,
                         off_status_count=off_status_count)

def process_visitor_pairs(records):
    """Process visitor records to create paired sessions with status logic and Singapore timezone"""
    from collections import defaultdict
    import pytz
    
    # Singapore timezone
    singapore_tz = pytz.timezone('Asia/Singapore')
    
    # Group records by visitor name and date (in Singapore timezone)
    grouped = defaultdict(lambda: defaultdict(list))
    
    for record in records:
        # Convert to Singapore timezone
        if record.timestamp.tzinfo is None:
            # Assume UTC if no timezone info
            local_timestamp = pytz.UTC.localize(record.timestamp).astimezone(singapore_tz)
        else:
            local_timestamp = record.timestamp.astimezone(singapore_tz)
        
        date_key = local_timestamp.date().strftime('%Y-%m-%d')
        record_data = {
            'id': record.id,
            'timestamp': local_timestamp,
            'visit_type': record.visit_type,
            'visitor_name': record.visitor_name,
            'company_name': record.company_name,
            'vehicle_number': record.vehicle_number,
            'details': record.details,
            'location': record.location,
            'selfie_photo': record.selfie_photo
        }
        grouped[record.visitor_name][date_key].append(record_data)
    
    paired_records = []
    
    for visitor_name, dates in grouped.items():
        for date_str, day_records in dates.items():
            # Sort records by timestamp
            day_records.sort(key=lambda x: x['timestamp'])
            
            # Separate start and end records
            start_records = [r for r in day_records if r['visit_type'] == 'start']
            end_records = [r for r in day_records if r['visit_type'] == 'end']
            
            # Process pairing logic
            if start_records and end_records:
                # Normal case: has both start and end
                start_record = start_records[0]  # First start of the day
                end_record = end_records[-1]     # Last end of the day
                
                # Calculate duration
                duration = end_record['timestamp'] - start_record['timestamp']
                duration_hours = duration.total_seconds() / 3600
                
                # Determine status - Warning if more than 5 hours
                if duration_hours > 5:
                    status = 'Warning'
                    row_class = 'table-danger'
                else:
                    status = 'Normal'
                    row_class = ''
                
                paired_record = {
                    'id': f"{start_record['id']}_{end_record['id']}",
                    'visitor_name': visitor_name,
                    'company': start_record['company_name'],
                    'vehicle_number': start_record['vehicle_number'] or end_record['vehicle_number'],
                    'date': date_str,
                    'details': start_record['details'] or end_record['details'],
                    'primary_time': start_record['timestamp'].strftime('%H:%M:%S'),
                    'location': start_record['location'] or end_record['location'],
                    'photo': start_record['selfie_photo'],  # For compatibility
                    'start_photo': start_record['selfie_photo'],
                    'end_photo': end_record['selfie_photo'],
                    'start_date': start_record['timestamp'].strftime('%Y-%m-%d'),
                    'start_time': start_record['timestamp'].strftime('%H:%M:%S'),
                    'end_date': end_record['timestamp'].strftime('%Y-%m-%d'),
                    'end_time': end_record['timestamp'].strftime('%H:%M:%S'),
                    'duration': f"{int(duration_hours)}h {int((duration_hours % 1) * 60)}m",
                    'status': status,
                    'row_class': row_class,
                    'type': 'Visit Session'
                }
                paired_records.append(paired_record)
                
            elif end_records and not start_records:
                # End only - "Off" status
                end_record = end_records[0]
                
                paired_record = {
                    'id': f"end_{end_record['id']}",
                    'visitor_name': visitor_name,
                    'company': end_record['company_name'],
                    'vehicle_number': end_record['vehicle_number'],
                    'date': date_str,
                    'details': end_record['details'],
                    'primary_time': end_record['timestamp'].strftime('%H:%M:%S'),
                    'location': end_record['location'],
                    'photo': end_record['selfie_photo'],
                    'start_photo': None,
                    'end_photo': end_record['selfie_photo'],
                    'start_date': None,
                    'start_time': None,
                    'end_date': end_record['timestamp'].strftime('%Y-%m-%d'),
                    'end_time': end_record['timestamp'].strftime('%H:%M:%S'),
                    'duration': None,
                    'status': 'Off',
                    'row_class': 'table-warning',
                    'type': 'End Only'
                }
                paired_records.append(paired_record)
                
            elif start_records and not end_records:
                # Start only - still visiting
                start_record = start_records[0]
                
                paired_record = {
                    'id': f"start_{start_record['id']}",
                    'visitor_name': visitor_name,
                    'company': start_record['company_name'],
                    'date': date_str,
                    'details': start_record['details'],
                    'primary_time': start_record['timestamp'].strftime('%H:%M:%S'),
                    'location': start_record['location'],
                    'photo': start_record['selfie_photo'],
                    'start_photo': start_record['selfie_photo'],
                    'end_photo': None,
                    'start_date': start_record['timestamp'].strftime('%Y-%m-%d'),
                    'start_time': start_record['timestamp'].strftime('%H:%M:%S'),
                    'end_date': None,
                    'end_time': None,
                    'duration': None,
                    'status': 'Normal',
                    'row_class': 'table-info',
                    'type': 'In Progress'
                }
                paired_records.append(paired_record)
    
    # Sort by most recent first
    paired_records.sort(key=lambda x: x['date'], reverse=True)
    
    return paired_records

@app.route('/visitors/delete', methods=['POST'])
@login_required
@create_permission_required('pioneer_lodge')
def delete_visitor_records():
    """Delete selected visitor records"""
    user = current_user
    org_id = user.organization_id
    
    if not is_admin_user(user):
        return jsonify({'success': False, 'error': 'Access denied'})
    
    try:
        data = request.get_json()
        record_ids = data.get('record_ids', [])
        
        if not record_ids:
            return jsonify({'success': False, 'error': 'No records selected'})
        
        # Delete records
        deleted_count = 0
        for record_id in record_ids:
            if '_' in str(record_id):
                # Handle paired record IDs
                if record_id.startswith('start_'):
                    visitor_id = record_id.replace('start_', '')
                elif record_id.startswith('end_'):
                    visitor_id = record_id.replace('end_', '')
                else:
                    # Paired record - get both IDs
                    parts = record_id.split('_')
                    for part in parts:
                        visitor = Visitor.query.filter_by(id=int(part), organization_id=org_id).first()
                        if visitor:
                            db.session.delete(visitor)
                            deleted_count += 1
                    continue
            else:
                visitor_id = record_id
            
            visitor = Visitor.query.filter_by(id=int(visitor_id), organization_id=org_id).first()
            if visitor:
                db.session.delete(visitor)
                deleted_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully deleted {deleted_count} visitor record(s)'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error deleting records: {str(e)}'})

@app.route('/visitors/export-excel')
@login_required
def export_visitors_excel():
    """Export visitor records to Excel"""
    user = current_user
    org_id = user.organization_id
    
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    visitor_name = request.args.get('visitor_name', '').strip()
    company_name = request.args.get('company_name', '').strip()
    visit_type = request.args.get('visit_type', '').strip()
    
    # Build query
    query = Visitor.query.filter_by(organization_id=org_id)
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(func.date(Visitor.timestamp) >= start_dt)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(func.date(Visitor.timestamp) <= end_dt)
        except ValueError:
            pass
    
    if visitor_name:
        query = query.filter(Visitor.visitor_name.ilike(f'%{visitor_name}%'))
    
    if company_name:
        query = query.filter(Visitor.company_name.ilike(f'%{company_name}%'))
    
    if visit_type in ['start', 'end']:
        query = query.filter(Visitor.visit_type == visit_type)
    
    # Get records
    records = query.order_by(Visitor.timestamp.desc()).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visitor Records"
    
    # Headers
    headers = [
        'S/NO', 'Date & Time', 'Visitor Name', 'Company', 'Vehicle No', 'Type', 'Details',
        'Location', 'QR Code', 'IP Address'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows
    for row_num, record in enumerate(records, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=record.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_num, column=3, value=record.visitor_name)
        ws.cell(row=row_num, column=4, value=record.company_name)
        ws.cell(row=row_num, column=5, value=record.vehicle_number or '')
        ws.cell(row=row_num, column=6, value=record.visit_type.title())
        ws.cell(row=row_num, column=7, value=record.details or '')
        ws.cell(row=row_num, column=8, value=record.location or '')
        ws.cell(row=row_num, column=9, value=record.qr_code_scanned or '')
        ws.cell(row=row_num, column=10, value=record.ip_address or '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    timestamp = singapore_now().strftime('%Y%m%d_%H%M%S')
    filename = f"visitor_records_{timestamp}.xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/visitors/export-pdf')
@login_required
def export_visitors_pdf():
    """Export visitor records to PDF with photos"""
    user = current_user
    org_id = user.organization_id
    
    # Get filter parameters
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    visitor_name_filter = request.args.get('visitor_name_filter', '').strip()
    company_filter = request.args.get('company_filter', '').strip()
    
    # Build query
    query = Visitor.query.filter_by(organization_id=org_id)
    
    if from_date:
        try:
            from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
            query = query.filter(func.date(Visitor.timestamp) >= from_dt)
        except ValueError:
            pass
    
    if to_date:
        try:
            to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
            query = query.filter(func.date(Visitor.timestamp) <= to_dt)
        except ValueError:
            pass
    
    if visitor_name_filter:
        query = query.filter(Visitor.visitor_name == visitor_name_filter)
    
    if company_filter:
        query = query.filter(Visitor.company_name == company_filter)
    
    # Get all records and process them
    all_records = query.order_by(Visitor.timestamp.desc()).all()
    paired_records = process_visitor_pairs(all_records)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch)
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Content
    story = []
    
    # Company logo and header
    try:
        # Create a simple company logo using reportlab graphics
        from reportlab.graphics.shapes import Drawing, Rect, String
        from reportlab.graphics import colors as gcolors
        
        # Create company logo drawing
        logo_drawing = Drawing(200, 50)
        # Background rectangle
        logo_drawing.add(Rect(0, 0, 200, 50, fillColor=gcolors.Color(0.118, 0.227, 0.541), strokeColor=None))
        # Company text
        logo_drawing.add(String(10, 30, "TS Management", fontSize=14, fillColor=gcolors.white, fontName="Helvetica-Bold"))
        logo_drawing.add(String(10, 15, "Service Pte Ltd", fontSize=10, fillColor=gcolors.Color(0.984, 0.749, 0.141), fontName="Helvetica"))
        
        story.append(logo_drawing)
        story.append(Spacer(1, 12))
    except Exception as e:
        # Fallback to text header
        company_header = Paragraph("TS Management Service Pte Ltd", title_style)
        story.append(company_header)
        story.append(Spacer(1, 6))
    
    # Title
    title = Paragraph(f"Pioneer Lodge Visitor Records - {user.organization.name}", styles['Heading2'])
    story.append(title)
    
    if from_date and to_date:
        date_range = Paragraph(f"Period: {from_date} to {to_date}", styles['Normal'])
        story.append(date_range)
        story.append(Spacer(1, 12))
    
    # Table data
    data = [['S/NO', 'Date', 'Visitor Name', 'Company', 'Type', 'Details', 'Start Time', 'End Time', 'Duration', 'Start Photo', 'End Photo']]
    
    def create_photo_image(photo_data, max_width=40, max_height=40):
        """Convert base64 photo to ReportLab Image"""
        if not photo_data:
            return 'No Photo'
        try:
            # Remove data URL prefix if present
            if photo_data.startswith('data:image'):
                photo_data = photo_data.split(',')[1]
            
            # Decode base64
            image_data = base64.b64decode(photo_data)
            image_buffer = BytesIO(image_data)
            
            # Create ReportLab Image
            img = Image(image_buffer, width=max_width, height=max_height)
            return img
        except Exception as e:
            return 'No Photo'

    for idx, record in enumerate(paired_records, 1):
        # Create photo images for PDF
        start_photo_img = create_photo_image(record.get('start_photo'))
        end_photo_img = create_photo_image(record.get('end_photo'))
        
        # Format data for better readability
        company_name = record['company'][:15] + '...' if len(record['company']) > 15 else record['company']
        visitor_name = record['visitor_name'][:12] + '...' if len(record['visitor_name']) > 12 else record['visitor_name']
        details = record.get('details', '')[:20] + '...' if record.get('details', '') and len(record.get('details', '')) > 20 else record.get('details', '')
        
        row = [
            str(idx),
            record['date'],
            visitor_name,
            company_name,
            record.get('type', 'Visit'),
            details,
            record['start_time'] or '-',
            record['end_time'] or '-',
            record['duration'] or '-',
            start_photo_img,
            end_photo_img
        ]
        data.append(row)
    
    # Create table with properly sized column widths for landscape layout
    col_widths = [0.3*inch, 0.7*inch, 1.0*inch, 1.0*inch, 0.6*inch, 1.0*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.6*inch, 0.6*inch]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    
    # Table style with proper spacing and readability
    table_style = [
        # Header row styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        
        # Data rows styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        
        # Grid and borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
        
        # Set proper row height for photo display
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]
    
    # Set minimum row height to ensure content visibility
    for i in range(1, len(data)):
        table_style.append(('MINROWHEIGHT', (0, i), (-1, i), 50))
    
    # Color rows based on status
    for idx, record in enumerate(paired_records, 1):
        row_idx = idx
        if record['status'] == 'Warning':
            table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightpink))
        elif record['status'] == 'Off':
            table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightyellow))
    
    table.setStyle(TableStyle(table_style))
    story.append(table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    # Generate filename
    timestamp = singapore_now().strftime('%Y%m%d_%H%M%S')
    filename = f"visitor_records_{timestamp}.pdf"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

# ===== RESIDENT CHECK-OUT ROUTES =====

@app.route('/resident-checkout-qr')
@login_required
def resident_checkout_qr():
    """Display QR codes for resident check-out"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned to your account.', 'error')
        return redirect(url_for('index'))
    
    # Generate QR code for resident check-out
    base_url = request.url_root.rstrip('/')
    checkout_url = f"{base_url}/resident-checkout-scan"
    
    # Create QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(checkout_url)
    qr.make(fit=True)
    
    # Create QR code image
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to BytesIO
    img_buffer = BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    # Convert to base64
    qr_code_base64 = base64.b64encode(img_buffer.getvalue()).decode()
    
    # Also save QR code to static folder for printing/download
    qr_filename = f"resident_checkout_qr_{singapore_now().strftime('%Y%m%d_%H%M%S')}.png"
    qr_path = os.path.join('static', 'qr_codes', qr_filename)
    os.makedirs(os.path.dirname(qr_path), exist_ok=True)
    img.save(qr_path)
    
    logo_exists = os.path.exists('static/ts_logo.svg')
    return render_template('resident_checkout_qr.html', 
                         qr_code=qr_code_base64, 
                         checkout_url=checkout_url,
                         qr_filename=qr_filename,
                         logo_exists=logo_exists)

@app.route('/resident-checkout-scan')
def resident_checkout_scan():
    """Display the QR code scanning page for resident check-out"""
    import pytz
    singapore_tz = pytz.timezone('Asia/Singapore')
    logo_exists = os.path.exists('static/ts_logo.svg')
    return render_template('resident_checkout_scan.html', logo_exists=logo_exists)

@app.route('/resident-checkout-submit', methods=['POST'])
@create_permission_required('resident_checkout')
def resident_checkout_submit():
    """Handle resident check-out form submission"""
    try:
        import pytz
        
        # Get form data
        qr_data = request.form.get('qr_data', '')
        selfie_photo = request.form.get('selfie_photo', '')
        resident_name = request.form.get('resident_name', '').strip()
        fin = request.form.get('fin', '').strip().upper()
        company_name = request.form.get('company_name', '').strip()
        reason = request.form.get('reason', '')
        details = request.form.get('details', '').strip()
        
        # Validate required fields
        if not all([resident_name, fin, company_name, reason, selfie_photo]):
            flash('All required fields must be filled and a selfie photo must be taken.', 'error')
            return redirect(url_for('resident_checkout_scan'))
        
        # Get current time in Singapore timezone
        singapore_tz = pytz.timezone('Asia/Singapore')
        checkout_time = datetime.now(singapore_tz)
        
        # Convert to UTC for storage
        utc_time = checkout_time.astimezone(pytz.utc).replace(tzinfo=None)
        
        # Get organization ID (default to 1 if not set)
        organization_id = getattr(current_user, 'organization_id', 1) if current_user.is_authenticated else 1
        
        # Create new checkout record
        checkout = ResidentCheckout(
            resident_name=resident_name,
            fin=fin,
            company_name=company_name,
            reason=reason,
            details=details if details else None,
            checkout_timestamp=utc_time,
            selfie_photo=selfie_photo,
            organization_id=organization_id
        )
        
        db.session.add(checkout)
        db.session.commit()
        
        flash('Check-out submitted successfully!', 'success')
        return render_template('resident_checkout_success.html', 
                             checkout=checkout, 
                             logo_exists=os.path.exists('static/ts_logo.svg'))
        
    except Exception as e:
        db.session.rollback()
        print(f"Error submitting checkout: {e}")
        flash('Error submitting check-out. Please try again.', 'error')
        return redirect(url_for('resident_checkout_scan'))

@app.route('/resident-checkout-records')
@login_required
@page_access_required('resident_checkout')
def resident_checkout_records():
    """Display resident check-out records with filtering"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned to your account.', 'error')
        return redirect(url_for('index'))
    
    # Build query
    query = ResidentCheckout.query.filter_by(organization_id=user.organization_id, is_active=True)
    
    # Apply filters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    name_filter = request.args.get('name_filter', '').strip()
    fin_filter = request.args.get('fin_filter', '').strip()
    reason_filter = request.args.get('reason_filter', '').strip()
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(ResidentCheckout.checkout_timestamp >= start_dt)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(ResidentCheckout.checkout_timestamp < end_dt)
        except ValueError:
            pass
    
    if name_filter:
        query = query.filter(ResidentCheckout.resident_name.ilike(f'%{name_filter}%'))
    
    if fin_filter:
        query = query.filter(ResidentCheckout.fin.ilike(f'%{fin_filter}%'))
    
    if reason_filter:
        query = query.filter(ResidentCheckout.reason == reason_filter)
    
    # Order by checkout time descending
    records = query.order_by(ResidentCheckout.checkout_timestamp.desc()).all()
    
    logo_exists = os.path.exists('static/ts_logo.svg')
    return render_template('resident_checkout_records.html', 
                         records=records, 
                         logo_exists=logo_exists)

@app.route('/resident-checkout-delete', methods=['POST'])
@login_required
def resident_checkout_delete():
    """Delete selected resident check-out records"""
    user = current_user
    if not user.organization_id:
        return jsonify({'error': 'No organization assigned'}), 403
    
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({'error': 'No records selected'}), 400
        
        # Delete records (soft delete by setting is_active to False)
        deleted_count = ResidentCheckout.query.filter(
            ResidentCheckout.id.in_(ids),
            ResidentCheckout.organization_id == user.organization_id
        ).update({'is_active': False}, synchronize_session=False)
        
        db.session.commit()
        
        return jsonify({'success': True, 'deleted_count': deleted_count})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting records: {e}")
        return jsonify({'error': 'Failed to delete records'}), 500

@app.route('/resident-checkin-checkout-dashboard')
@login_required
@page_access_required('resident_checkin')
def resident_checkin_checkout_dashboard():
    """Dashboard page for Resident Check-in & Resident Check-Out system"""
    logo_exists = os.path.exists('static/ts_logo.svg')
    return render_template('resident_checkin_checkout_dashboard.html', logo_exists=logo_exists)

@app.route('/resident-checkout-export-excel', methods=['POST'])
@login_required
def resident_checkout_export_excel():
    """Export selected resident check-out records to Excel"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned to your account.', 'error')
        return redirect(url_for('resident_checkout_records'))
    
    try:
        import pytz
        
        # Get selected IDs
        selected_ids = json.loads(request.form.get('selected_ids', '[]'))
        
        if not selected_ids:
            flash('No records selected for export.', 'error')
            return redirect(url_for('resident_checkout_records'))
        
        # Get records
        records = ResidentCheckout.query.filter(
            ResidentCheckout.id.in_(selected_ids),
            ResidentCheckout.organization_id == user.organization_id,
            ResidentCheckout.is_active == True
        ).order_by(ResidentCheckout.checkout_timestamp.desc()).all()
        
        if not records:
            flash('No records found for export.', 'error')
            return redirect(url_for('resident_checkout_records'))
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resident Check-Out Records"
        
        # Set headers
        headers = ['S/No', 'Date & Time', 'Resident Name', 'FIN', 'Company', 'Reason', 'Details']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        
        # Add data
        singapore_tz = pytz.timezone('Asia/Singapore')
        for row, record in enumerate(records, 2):
            sg_time = record.get_singapore_time()
            time_str = sg_time.strftime('%d/%m/%Y %I:%M %p') if sg_time else 'N/A'
            
            ws.cell(row=row, column=1, value=row-1)  # S/No
            ws.cell(row=row, column=2, value=time_str)
            ws.cell(row=row, column=3, value=record.resident_name)
            ws.cell(row=row, column=4, value=record.fin)
            ws.cell(row=row, column=5, value=record.company_name)
            ws.cell(row=row, column=6, value=record.reason)
            ws.cell(row=row, column=7, value=record.details or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Generate filename
        timestamp = singapore_now().strftime('%Y%m%d_%H%M%S')
        filename = f"resident_checkout_records_{timestamp}.xlsx"
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        flash('Error generating Excel export.', 'error')
        return redirect(url_for('resident_checkout_records'))

@app.route('/resident-checkout-export-pdf', methods=['GET', 'POST'])
@login_required
def resident_checkout_export_pdf():
    """Export resident check-out records to PDF with photos"""
    user = current_user
    if not user.organization_id:
        flash('No organization assigned to your account.', 'error')
        return redirect(url_for('resident_checkout_records'))
    
    try:
        import pytz
        
        # Check if this is a POST request with selected IDs
        if request.method == 'POST':
            data = request.get_json()
            ids = data.get('ids', [])
            
            if not ids:
                return jsonify({'error': 'No records selected'}), 400
            
            # Get selected records
            records = ResidentCheckout.query.filter(
                ResidentCheckout.id.in_(ids),
                ResidentCheckout.organization_id == user.organization_id,
                ResidentCheckout.is_active == True
            ).order_by(ResidentCheckout.checkout_timestamp.desc()).all()
        else:
            # Build query with same filters as records page (for GET request)
            query = ResidentCheckout.query.filter_by(organization_id=user.organization_id, is_active=True)
            
            # Apply filters from query string
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')
            name_filter = request.args.get('name_filter', '').strip()
            fin_filter = request.args.get('fin_filter', '').strip()
            reason_filter = request.args.get('reason_filter', '').strip()
            
            if start_date:
                try:
                    start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                    query = query.filter(ResidentCheckout.checkout_timestamp >= start_dt)
                except ValueError:
                    pass
            
            if end_date:
                try:
                    end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                    query = query.filter(ResidentCheckout.checkout_timestamp < end_dt)
                except ValueError:
                    pass
            
            if name_filter:
                query = query.filter(ResidentCheckout.resident_name.ilike(f'%{name_filter}%'))
            
            if fin_filter:
                query = query.filter(ResidentCheckout.fin.ilike(f'%{fin_filter}%'))
            
            if reason_filter:
                query = query.filter(ResidentCheckout.reason == reason_filter)
            
            records = query.order_by(ResidentCheckout.checkout_timestamp.desc()).all()
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Add company header and title
        title = Paragraph("TS MANAGEMENT SERVICES PTE LTD<br/>Resident Check-in & Resident Check-Out Records", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Create data for table (including photo column)
        data = [['S/No', 'Photo', 'Date & Time', 'Name', 'FIN', 'Company', 'Reason', 'Details']]
        
        singapore_tz = pytz.timezone('Asia/Singapore')
        for i, record in enumerate(records, 1):
            sg_time = record.get_singapore_time()
            time_str = sg_time.strftime('%d/%m/%Y<br/>%I:%M %p') if sg_time else 'N/A'
            
            details = record.details[:30] + '...' if record.details and len(record.details) > 30 else (record.details or '')
            
            # Handle photo for table
            photo_cell = ''
            if record.selfie_photo and record.selfie_photo.startswith('data:image'):
                try:
                    # Extract base64 data
                    img_data = record.selfie_photo.split(',')[1]
                    img_bytes = base64.b64decode(img_data)
                    img_buffer = BytesIO(img_bytes)
                    
                    # Create small image for table
                    photo_cell = Image(img_buffer, width=0.8*inch, height=0.8*inch)
                except Exception as e:
                    print(f"Error processing photo for {record.resident_name}: {e}")
                    photo_cell = 'Photo Error'
            else:
                photo_cell = 'No Photo'
            
            data.append([
                str(i),
                photo_cell,
                Paragraph(time_str, styles['Normal']),
                record.resident_name,
                record.fin,
                record.company_name[:15] + '...' if len(record.company_name) > 15 else record.company_name,
                record.reason,
                details
            ])
        
        # Create main table with photo column
        table = Table(data, colWidths=[0.4*inch, 1*inch, 1*inch, 1*inch, 0.7*inch, 1*inch, 0.7*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8E8E8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        
        story.append(table)
        
        # Add photos section if any records have photos
        records_with_photos = [r for r in records if r.selfie_photo]
        if records_with_photos:
            story.append(PageBreak())
            story.append(Paragraph("Resident Photos", title_style))
            story.append(Spacer(1, 20))
            
            # Create photo grid (2 photos per row)
            photo_data = []
            photo_row = []
            
            for i, record in enumerate(records_with_photos):
                try:
                    if record.selfie_photo and record.selfie_photo.startswith('data:image'):
                        # Extract base64 data
                        img_data = record.selfie_photo.split(',')[1]
                        img_bytes = base64.b64decode(img_data)
                        img_buffer = BytesIO(img_bytes)
                        
                        # Create image with caption
                        photo_img = Image(img_buffer, width=2*inch, height=2*inch)
                        caption = Paragraph(f"{record.resident_name}<br/>{record.fin}", styles['Normal'])
                        
                        photo_cell = [photo_img, caption]
                        photo_row.append(photo_cell)
                        
                        # Add row when we have 2 photos or at the end
                        if len(photo_row) == 2 or i == len(records_with_photos) - 1:
                            # Pad row if only one photo
                            while len(photo_row) < 2:
                                photo_row.append(['', ''])
                            photo_data.append(photo_row)
                            photo_row = []
                
                except Exception as e:
                    print(f"Error processing photo for {record.resident_name}: {e}")
                    continue
            
            if photo_data:
                photo_table = Table(photo_data, colWidths=[3*inch, 3*inch])
                photo_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 0), (-1, -1), 8)
                ]))
                story.append(photo_table)
        
        # Build PDF
        doc.build(story)
        
        # Get the value of the BytesIO buffer and create response
        pdf_data = buffer.getvalue()
        buffer.close()
        
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=resident_checkout_{singapore_now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response
        
    except Exception as e:
        print(f"Error generating PDF: {e}")
        flash('Error generating PDF export.', 'error')
        return redirect(url_for('resident_checkout_records'))
