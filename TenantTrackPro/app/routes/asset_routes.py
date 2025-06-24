"""Asset management routes for inventory and tracking"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import current_user, login_required
from datetime import datetime
from functools import wraps
from io import BytesIO
import openpyxl
from app_main import db
from models import Asset, AssetCategory, User, Organization, Submission
from app.models.models_asset import AssetName, AssetItem

# Create Blueprint
asset_bp = Blueprint('asset', __name__)

def is_admin_user(user):
    """Check if user is an admin"""
    if user.role == 'admin':
        return True
    if hasattr(user, 'role') and user.role == 'admin':
        return True
    return False

def admin_required(f):
    """Decorator to require admin privileges"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('other.login'))
        if not is_admin_user(current_user):
            flash('Admin access required', 'error')
            return redirect(url_for('dashboard.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@asset_bp.route('/asset-management')
@login_required
def asset_management():
    """Asset management dashboard"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    # Get assets for the user's organization
    assets = Asset.query.filter_by(organization_id=user.organization_id).all()
    categories = AssetCategory.query.filter_by(organization_id=user.organization_id).all()
    
    # Get asset statistics
    total_assets = len(assets)
    active_assets = len([a for a in assets if a.status == 'Active'])
    
    return render_template('assets/asset_management.html',
                         assets=assets,
                         categories=categories,
                         total_assets=total_assets,
                         active_assets=active_assets,
                         user=user)

@asset_bp.route('/assets')
@login_required
def assets_list():
    """List all assets"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    assets = Asset.query.filter_by(organization_id=user.organization_id).all()
    return render_template('assets/assets.html', assets=assets, user=user)

@asset_bp.route('/assets/new')
@login_required
def new_asset():
    """Create new asset form"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    categories = AssetCategory.query.filter_by(organization_id=user.organization_id).all()
    return render_template('assets/add_asset.html', categories=categories, user=user)

@asset_bp.route('/assets/create', methods=['POST'])
@login_required
def create_asset():
    """Create new asset"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    try:
        name = request.form.get('name')
        category_id = request.form.get('category_id')
        serial_number = request.form.get('serial_number')
        description = request.form.get('description')
        location = request.form.get('location')
        status = request.form.get('status', 'Active')
        
        if not name:
            flash('Asset name is required', 'error')
            return redirect(url_for('asset.new_asset'))
        
        new_asset = Asset(
            name=name,
            category_id=int(category_id) if category_id else None,
            serial_number=serial_number,
            description=description,
            location=location,
            status=status,
            organization_id=user.organization_id,
            created_by=user.id,
            created_at=datetime.now()
        )
        
        db.session.add(new_asset)
        db.session.commit()
        
        # Create submission record
        submission = Submission(
            organization_id=user.organization_id,
            user_id=user.id,
            submission_type='asset_created',
            reference_id=new_asset.id,
            notes=f'Created asset: {name}'
        )
        db.session.add(submission)
        db.session.commit()
        
        flash(f'Asset "{name}" created successfully!', 'success')
        return redirect(url_for('asset.asset_management'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating asset: {str(e)}', 'error')
        return redirect(url_for('asset.new_asset'))

@asset_bp.route('/assets/<int:asset_id>/edit')
@login_required
def edit_asset(asset_id):
    """Edit asset form"""
    user = current_user
    asset = Asset.query.filter_by(id=asset_id, organization_id=user.organization_id).first()
    
    if not asset:
        flash('Asset not found', 'error')
        return redirect(url_for('asset.asset_management'))
    
    categories = AssetCategory.query.filter_by(organization_id=user.organization_id).all()
    return render_template('assets/edit_asset.html', asset=asset, categories=categories, user=user)

@asset_bp.route('/assets/<int:asset_id>/update', methods=['POST'])
@login_required
def update_asset(asset_id):
    """Update asset"""
    user = current_user
    asset = Asset.query.filter_by(id=asset_id, organization_id=user.organization_id).first()
    
    if not asset:
        flash('Asset not found', 'error')
        return redirect(url_for('asset.asset_management'))
    
    try:
        asset.name = request.form.get('name')
        asset.category_id = int(request.form.get('category_id')) if request.form.get('category_id') else None
        asset.serial_number = request.form.get('serial_number')
        asset.description = request.form.get('description')
        asset.location = request.form.get('location')
        asset.status = request.form.get('status')
        
        db.session.commit()
        
        # Create submission record
        submission = Submission(
            organization_id=user.organization_id,
            user_id=user.id,
            submission_type='asset_updated',
            reference_id=asset.id,
            notes=f'Updated asset: {asset.name}'
        )
        db.session.add(submission)
        db.session.commit()
        
        flash(f'Asset "{asset.name}" updated successfully!', 'success')
        return redirect(url_for('asset.asset_management'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating asset: {str(e)}', 'error')
        return redirect(url_for('asset.edit_asset', asset_id=asset_id))

@asset_bp.route('/assets/<int:asset_id>/delete', methods=['POST'])
@login_required
def delete_asset(asset_id):
    """Delete asset"""
    user = current_user
    asset = Asset.query.filter_by(id=asset_id, organization_id=user.organization_id).first()
    
    if not asset:
        flash('Asset not found', 'error')
        return redirect(url_for('asset.asset_management'))
    
    try:
        asset_name = asset.name
        
        # Create submission record before deletion
        submission = Submission(
            organization_id=user.organization_id,
            user_id=user.id,
            submission_type='asset_deleted',
            reference_id=asset.id,
            notes=f'Deleted asset: {asset_name}'
        )
        db.session.add(submission)
        
        db.session.delete(asset)
        db.session.commit()
        flash(f'Asset "{asset_name}" deleted successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting asset: {str(e)}', 'error')
    
    return redirect(url_for('asset.asset_management'))

@asset_bp.route('/asset-categories')
@login_required
@admin_required
def asset_categories():
    """Manage asset categories"""
    user = current_user
    categories = AssetCategory.query.filter_by(organization_id=user.organization_id).all()
    return render_template('assets/asset_categories.html', categories=categories, user=user)

@asset_bp.route('/assets/export/excel')
@login_required
def export_assets_excel():
    """Export assets to Excel"""
    user = current_user
    if not user.organization_id:
        flash('Please contact administrator to assign organization', 'warning')
        return redirect(url_for('dashboard.dashboard'))
    
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assets"
        
        # Headers
        headers = ['ID', 'Name', 'Category', 'Serial Number', 'Description', 'Location', 'Status', 'Created Date']
        ws.append(headers)
        
        # Get assets
        assets = Asset.query.filter_by(organization_id=user.organization_id).all()
        
        for asset in assets:
            ws.append([
                asset.id,
                asset.name,
                asset.category.name if asset.category else '',
                asset.serial_number or '',
                asset.description or '',
                asset.location or '',
                asset.status,
                asset.created_at.strftime('%Y-%m-%d %H:%M:%S') if asset.created_at else ''
            ])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'assets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        flash(f'Error exporting assets: {str(e)}', 'error')
        return redirect(url_for('asset.asset_management'))

@asset_bp.route('/assets/import/excel', methods=['GET', 'POST'])
@login_required
@admin_required
def import_assets_excel():
    """Import assets from Excel"""
    user = current_user
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('asset.import_assets_excel'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('asset.import_assets_excel'))
        
        try:
            # Read Excel file
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            imported_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Check if first column (name) has value
                    name = str(row[0])
                    category_name = str(row[1]) if row[1] else None
                    serial_number = str(row[2]) if row[2] else None
                    description = str(row[3]) if row[3] else None
                    location = str(row[4]) if row[4] else None
                    status = str(row[5]) if row[5] else 'Active'
                    
                    # Find or create category
                    category = None
                    if category_name:
                        category = AssetCategory.query.filter_by(
                            name=category_name,
                            organization_id=user.organization_id
                        ).first()
                        if not category:
                            category = AssetCategory(
                                name=category_name,
                                organization_id=user.organization_id
                            )
                            db.session.add(category)
                            db.session.flush()
                    
                    # Create asset
                    asset = Asset(
                        name=name,
                        category_id=category.id if category else None,
                        serial_number=serial_number,
                        description=description,
                        location=location,
                        status=status,
                        organization_id=user.organization_id,
                        created_by=user.id,
                        created_at=datetime.now()
                    )
                    
                    db.session.add(asset)
                    imported_count += 1
            
            db.session.commit()
            flash(f'Successfully imported {imported_count} assets!', 'success')
            return redirect(url_for('asset.asset_management'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error importing assets: {str(e)}', 'error')
    
    return render_template('assets/import_asset_excel.html', user=user)